from __future__ import annotations

from contextlib import contextmanager
from contextvars import ContextVar
from datetime import date, datetime, timedelta
import json
import math
from pathlib import Path
import re
import sqlite3
from typing import Any, Iterator

from app.import_mapping import confirm_cases_dry_run_plan, confirm_preflight_report, mapping_preview
from app.settings import get_settings


_current_actor: ContextVar[str] = ContextVar("current_actor", default="system")


def set_current_actor(actor: str) -> None:
    """記錄目前請求的操作者，供 write_audit_log 使用（由 API 層每個請求綁定）。"""
    _current_actor.set(actor)


_owner_scope: ContextVar[str | None] = ContextVar("owner_scope", default=None)
_owner_display_name: ContextVar[str | None] = ContextVar("owner_display_name", default=None)

# 匯入模式（使用者拍板 2026-07-31／待你決定 A2）：匯入進來的是「已經在跑的舊案子」，
# 不是新申請——不該落成「草稿＋TMP- 暫時號」再要人一筆筆補號，直接算已成立（approved）
# 並在當年度配正式流水號。只在匯入路徑打開，一般 UI 新建案件不受影響（仍走申請→複核）。
_import_mode: ContextVar[bool] = ContextVar("import_mode", default=False)


@contextmanager
def import_mode() -> Iterator[None]:
    """匯入期間建立的案件直接視為已成立。用 contextmanager 確保離開就還原，
    不會讓後續一般操作誤沾到匯入語意。"""
    token = _import_mode.set(True)
    try:
        yield
    finally:
        _import_mode.reset(token)
# 限縮方式：owner＝比對案件負責帳號（承辦）；group＝比對案件所屬組別（組長看本組）
_scope_kind: ContextVar[str] = ContextVar("scope_kind", default="owner")


def set_owner_scope(scope: str | None, kind: str = "owner") -> None:
    """設定資料列可視範圍。scope 為 None＝看全部（部長/CIO/助理）。

    kind='owner'（承辦）：scope 是帳號，只看 owner 屬此帳號的案件及其關聯資料。
    kind='group'（組長）：scope 是組別名稱，看本組所有承辦的案件——組長要管整組，
      但看不到別組（使用者拍板 2026-07-30）。組長自己也可能是承辦（很多簽呈他自己做），
      所以不是「只看自己的」，而是「看本組的」，自己送的案自然也在裡面。
    """
    _owner_scope.set(scope)
    _scope_kind.set(kind)


def set_owner_display_name(name: str | None) -> None:
    """設定承辦者的顯示名稱（人名），供專案「依負責人隔離」比對用——專案 owner 欄存的是
    人名（可能用「/」列多個共同負責人，如「令狐沖/黃蓉」），不是登入帳號，跟案件的
    owner=帳號 是不同比對基準，分開存。"""
    _owner_display_name.set(name)


def _scope_where(table: str, scope: str, alias: str = "") -> tuple[str, list[Any]]:
    """把 table 限縮到 scope 擁有案件範圍的 (WHERE 片段, 參數)。
    alias：呼叫端 SQL 若把該表取了別名（如 `FROM projects t`）就傳進來，
    避免欄名跟 JOIN 進來的其他表同名欄位（如 projects.owner 撞 cases.owner）產生
    SQLite「ambiguous column name」500 錯；大多數呼叫端（如 list_rows）沒有 JOIN、
    表名本身就是唯一來源，維持不傳（欄名不加前綴）即可。"""
    prefix = f"{alias}." if alias else ""
    # 承辦比帳號、組長比組別（兩者都只認「案件」這一層，其他表都靠 case_id 掛過來）
    case_col = "group_name" if _scope_kind.get() == "group" else "owner"
    owned = f"SELECT id FROM cases WHERE {case_col} = ?"
    if table == "cases":
        return f"{prefix}{case_col} = ?", [scope]
    if table in ("contracts", "signoffs", "purchases"):
        # 這些靠 case_id 掛在案件上 → 依案件歸屬隔離（承辦只看自己案件下的）。
        # 預算(budgets) 不在此列：是全公司共享資料，不管誰負責、大家都看得到。
        return f"{prefix}case_id IN ({owned})", [scope]
    if table == "payments":
        return f"{prefix}contract_id IN (SELECT id FROM contracts WHERE case_id IN ({owned}))", [scope]
    if table == "documents":
        return (
            f"({prefix}case_id IN ({owned}) OR {prefix}contract_id IN "
            f"(SELECT id FROM contracts WHERE case_id IN ({owned})))",
            [scope, scope],
        )
    if table == "projects":
        # 組長：專案跟著案件的組別走（本組的案子底下的專案都看得到），不看專案負責人。
        if _scope_kind.get() == "group":
            return f"{prefix}case_id IN ({owned})", [scope]
        # 承辦：專案依負責人隔離（使用者拍板）：一人負責只有那人看得到；「/」列多個共同負責人時，
        # 名字有列在裡面的都看得到，沒列的看不到。用字串邊界比對（"/"+owner+"/" LIKE
        # "%/name/%"）避免子字串誤判（如「王小明」誤配到「王小明志」）。
        # 沒有顯示名稱（如內建示範帳號 ap01~04 的顯示名稱是角色而非真人名）就一律看不到，
        # 而非退回看全部——符合「只有列進去的人才看得到」的規則。
        name = _owner_display_name.get()
        if not name:
            return "0", []
        return f"('/' || {prefix}owner || '/') LIKE ?", [f"%/{name}/%"]
    return "", []


def _row_in_scope(conn: sqlite3.Connection, table: str, row_id: int, scope: str) -> bool:
    """該列是否在 scope(承辦) 的可視範圍內。用於寫入(改/停用/刪)前的越權防護。"""
    where, params = _scope_where(table, scope)
    if not where:
        return True
    return conn.execute(
        f"SELECT 1 FROM {table} WHERE id = ? AND ({where}) LIMIT 1", [row_id, *params]
    ).fetchone() is not None


SCHEMA = """
CREATE TABLE IF NOT EXISTS cases (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    case_code TEXT NOT NULL UNIQUE,
    title TEXT NOT NULL,
    owner TEXT NOT NULL DEFAULT '',
    status TEXT NOT NULL DEFAULT 'draft',
    amount REAL NOT NULL DEFAULT 0,
    risk_level TEXT NOT NULL DEFAULT 'normal',
    note TEXT NOT NULL DEFAULT '',
    next_step TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS contracts (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    contract_code TEXT NOT NULL UNIQUE,
    contract_name TEXT NOT NULL,
    vendor_name TEXT NOT NULL DEFAULT '',
    amount REAL NOT NULL DEFAULT 0,
    status TEXT NOT NULL DEFAULT 'active',
    case_id INTEGER,
    purchase_id INTEGER,
    end_date TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS payments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    contract_id INTEGER NOT NULL,
    payment_month TEXT NOT NULL,
    payment_amount REAL NOT NULL,
    invoice_status TEXT NOT NULL DEFAULT 'not_received',
    status TEXT NOT NULL DEFAULT 'pending',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

-- 合約費用調整紀錄（需求書 §10，原始案例：中華電信板橋機櫃電力費調整）。
-- 同一份合約中途改金額（機櫃增減、電費調價）不是新合約、也不能直接把 contracts.amount 蓋掉——
-- 蓋掉就答不出「什麼時候、為什麼、從多少調到多少、誰調的」。改成每次調整留一筆，
-- contracts.amount 永遠是「調整後的現值」，歷史在這張表。
CREATE TABLE IF NOT EXISTS contract_adjustments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    contract_id INTEGER NOT NULL,
    effective_date TEXT NOT NULL DEFAULT '',
    old_amount REAL NOT NULL DEFAULT 0,
    new_amount REAL NOT NULL DEFAULT 0,
    delta REAL NOT NULL DEFAULT 0,
    reason TEXT NOT NULL DEFAULT '',
    note TEXT NOT NULL DEFAULT '',
    created_by TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

-- 預計付款排程（Payment Schedule）＝合約約定「應該」付的排程（第幾期、預計金額、預計付款日）。
-- 需求書 §8：與「實際費用 Expense＝payments 表」分開但關聯，避免同一筆金額在 Dashboard 重複計算。
-- 一份合約可有多筆排程；method 決定金額怎麼算（固定金額/固定期數/週期/里程碑%）。
CREATE TABLE IF NOT EXISTS payment_schedules (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    contract_id INTEGER NOT NULL,
    case_id INTEGER,
    seq INTEGER NOT NULL DEFAULT 0,
    label TEXT NOT NULL DEFAULT '',
    method TEXT NOT NULL DEFAULT 'fixed',
    planned_amount REAL NOT NULL DEFAULT 0,
    percent REAL NOT NULL DEFAULT 0,
    due_date TEXT NOT NULL DEFAULT '',
    status TEXT NOT NULL DEFAULT 'planned',
    note TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS documents (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    file_name TEXT NOT NULL,
    document_type TEXT NOT NULL DEFAULT 'other',
    source_note TEXT NOT NULL DEFAULT '',
    status TEXT NOT NULL DEFAULT 'active',
    case_id INTEGER,
    contract_id INTEGER,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS budgets (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    budget_code TEXT NOT NULL UNIQUE,
    category TEXT NOT NULL DEFAULT '',
    unit_name TEXT NOT NULL DEFAULT '',
    fiscal_year TEXT NOT NULL DEFAULT '',
    amount REAL NOT NULL DEFAULT 0,
    status TEXT NOT NULL DEFAULT 'active',
    case_id INTEGER,
    note TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS projects (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_code TEXT NOT NULL UNIQUE,
    project_name TEXT NOT NULL,
    source TEXT NOT NULL DEFAULT '',
    necessity TEXT NOT NULL DEFAULT '',
    progress REAL NOT NULL DEFAULT 0,
    owner TEXT NOT NULL DEFAULT '',
    status TEXT NOT NULL DEFAULT 'active',
    case_id INTEGER,
    note TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS signoffs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    signoff_code TEXT NOT NULL UNIQUE,
    subject TEXT NOT NULL,
    applicant TEXT NOT NULL DEFAULT '',
    amount REAL NOT NULL DEFAULT 0,
    status TEXT NOT NULL DEFAULT 'draft',
    sign_date TEXT NOT NULL DEFAULT '',
    case_id INTEGER,
    note TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS purchases (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    purchase_code TEXT NOT NULL UNIQUE,
    item_name TEXT NOT NULL,
    vendor_name TEXT NOT NULL DEFAULT '',
    quantity REAL NOT NULL DEFAULT 0,
    amount REAL NOT NULL DEFAULT 0,
    status TEXT NOT NULL DEFAULT 'pending',
    case_id INTEGER,
    signoff_id INTEGER,
    note TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

-- ── 費用模組三層（黃助理 0803 附件一）──────────────────────────────
-- 第一層 費用主檔：一份合約可以有多個費用主檔；沒有合約的費用（例行性費用）直接建在這裡。
-- 「合約總費用(含稅)」是第二層所有費用區段加總的檢核基準——對不起來就不准確認排程。
CREATE TABLE IF NOT EXISTS expense_masters (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    contract_id INTEGER,                             -- 空＝無合約費用
    case_id INTEGER,
    expense_name TEXT NOT NULL,                      -- 有合約時帶合約名稱，無合約時人工填費用名稱
    vendor_name TEXT NOT NULL DEFAULT '',
    vendor_tax_id TEXT NOT NULL DEFAULT '',
    start_date TEXT NOT NULL DEFAULT '',             -- 無合約時停用（不得輸入）
    end_date TEXT NOT NULL DEFAULT '',
    total_amount REAL NOT NULL DEFAULT 0,            -- 含稅；有合約時由合約帶入且不得改
    modes TEXT NOT NULL DEFAULT '',                  -- 可複選：milestone,periodic,commitment
    signoff_ref TEXT NOT NULL DEFAULT '',            -- 簽呈／請購編號（沒有時填原因）
    signoff_none_reason TEXT NOT NULL DEFAULT '',
    owner TEXT NOT NULL DEFAULT '',                  -- 承辦人
    note TEXT NOT NULL DEFAULT '',
    status TEXT NOT NULL DEFAULT 'active',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

-- 第二層之一 費用區段：選了幾種模式就有幾個區段（混合型）；各區段金額合計＝第一層總費用
CREATE TABLE IF NOT EXISTS expense_sections (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    expense_id INTEGER NOT NULL,
    mode TEXT NOT NULL,                              -- milestone / periodic / commitment
    section_name TEXT NOT NULL DEFAULT '',
    section_amount REAL NOT NULL DEFAULT 0,
    price_method TEXT NOT NULL DEFAULT '',           -- 里程碑：percent 依比例 / fixed 固定金額
    periods INTEGER NOT NULL DEFAULT 0,              -- 總期數
    frequency TEXT NOT NULL DEFAULT '',              -- 定期費用：monthly/quarterly/semi/yearly
    period_start TEXT NOT NULL DEFAULT '',           -- 費用期間起日
    period_end TEXT NOT NULL DEFAULT '',
    first_amount REAL NOT NULL DEFAULT 0,            -- 第一期費用（後續各期預設值）
    first_month TEXT NOT NULL DEFAULT '',            -- 第一期費用年月 YYYY-MM
    first_due_date TEXT NOT NULL DEFAULT '',         -- 第一期預計應付日（後續順延基準）
    note TEXT NOT NULL DEFAULT '',
    status TEXT NOT NULL DEFAULT 'draft',            -- draft 草稿 / confirmed 已確認
    confirmed_by TEXT NOT NULL DEFAULT '',
    confirmed_at TEXT NOT NULL DEFAULT '',
    version INTEGER NOT NULL DEFAULT 1,              -- 已確認後重新編輯＝新版本，舊版留著
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

-- 第二層之二 排程明細：里程碑逐期人工填；定期費用由系統依頻率推算後可逐期修正
CREATE TABLE IF NOT EXISTS expense_schedules (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    section_id INTEGER NOT NULL,
    seq INTEGER NOT NULL DEFAULT 0,                  -- 第 N 期
    milestone_name TEXT NOT NULL DEFAULT '',         -- 簽約款/交付款/驗收款/自訂（僅里程碑）
    custom_name TEXT NOT NULL DEFAULT '',            -- 選「自訂」時必填
    percent REAL NOT NULL DEFAULT 0,                 -- 比例計價時的各期比例
    planned_amount REAL NOT NULL DEFAULT 0,          -- 應付費用
    expense_month TEXT NOT NULL DEFAULT '',          -- 費用年月 YYYY-MM
    billing_start TEXT NOT NULL DEFAULT '',          -- 計費期間
    billing_end TEXT NOT NULL DEFAULT '',
    due_date TEXT NOT NULL DEFAULT '',               -- 預計應付日／預計發生日
    note TEXT NOT NULL DEFAULT '',
    manual_adjusted INTEGER NOT NULL DEFAULT 0,      -- 人工調過系統算出來的值
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

-- 第三層之一 最低承諾金額的實際費用明細（助理 0803 附件一第六節）：
-- 承諾金額只是門檻，每期實際用了多少要在這裡登錄，系統再回寫承諾達成情形。
CREATE TABLE IF NOT EXISTS expense_actuals (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    section_id INTEGER NOT NULL,
    schedule_id INTEGER NOT NULL,                    -- 對應第二層某一期排程
    commit_period INTEGER NOT NULL DEFAULT 0,        -- 所屬承諾期別（系統帶）
    usage_amount REAL NOT NULL DEFAULT 0,            -- 當期使用／應付金額（依認列基礎）
    billing_start TEXT NOT NULL DEFAULT '',
    billing_end TEXT NOT NULL DEFAULT '',
    description TEXT NOT NULL DEFAULT '',
    adjust_amount REAL NOT NULL DEFAULT 0,           -- 折讓／退款／前期調整，可正可負
    adjust_reason TEXT NOT NULL DEFAULT '',          -- 調整金額不為 0 時必填
    recognized_amount REAL NOT NULL DEFAULT 0,       -- 系統算：使用金額＋調整金額
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

-- 第三層之二 請款／核銷（助理 0803）：一次作業只對一筆排程＋一張發票，不得複選。
CREATE TABLE IF NOT EXISTS expense_settlements (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    expense_id INTEGER NOT NULL,
    section_id INTEGER NOT NULL,
    schedule_id INTEGER NOT NULL,                    -- 單選一筆已確認的排程
    actual_id INTEGER,                               -- 最低承諾模式：關聯一筆實際費用明細
    settle_month TEXT NOT NULL DEFAULT '',           -- 費用核銷月份（系統帶、可調）
    billing_start TEXT NOT NULL DEFAULT '',
    billing_end TEXT NOT NULL DEFAULT '',
    vendor_name TEXT NOT NULL DEFAULT '',
    vendor_tax_id TEXT NOT NULL DEFAULT '',
    invoice_date TEXT NOT NULL DEFAULT '',
    invoice_no TEXT NOT NULL DEFAULT '',
    claim_amount REAL NOT NULL DEFAULT 0,            -- 廠商本次請款金額
    progress TEXT NOT NULL DEFAULT 'invoice_pending',
    confirmed INTEGER NOT NULL DEFAULT 0,            -- 承辦「確認完成」→ 通知核銷者
    settler TEXT NOT NULL DEFAULT '',                -- 核銷者
    signoff_no TEXT NOT NULL DEFAULT '',             -- 費用核銷簽呈編號
    doc_ref TEXT NOT NULL DEFAULT '',                -- 請款文件
    diff_reason TEXT NOT NULL DEFAULT '',            -- 請款差異不為 0 時必填
    note TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS settings (
    key TEXT PRIMARY KEY,
    value TEXT NOT NULL DEFAULT '',
    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS users (
    username TEXT PRIMARY KEY,
    role_code TEXT NOT NULL,
    display_name TEXT NOT NULL DEFAULT '',
    email TEXT NOT NULL DEFAULT '',
    password_hash TEXT NOT NULL DEFAULT '',
    disabled INTEGER NOT NULL DEFAULT 0,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS audit_logs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    table_name TEXT NOT NULL,
    row_id INTEGER NOT NULL,
    action TEXT NOT NULL,
    before_json TEXT,
    after_json TEXT,
    actor TEXT NOT NULL DEFAULT 'local-dev',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS import_batches (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    source_name TEXT NOT NULL,
    status TEXT NOT NULL DEFAULT 'created',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS import_rows (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    batch_id INTEGER NOT NULL,
    row_number INTEGER NOT NULL,
    raw_json TEXT NOT NULL,
    status TEXT NOT NULL DEFAULT 'staged',
    error_message TEXT,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS project_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id INTEGER NOT NULL,
    seq INTEGER NOT NULL DEFAULT 0,
    item_name TEXT NOT NULL DEFAULT '',
    owner TEXT NOT NULL DEFAULT '',
    start_date TEXT NOT NULL DEFAULT '',
    end_date TEXT NOT NULL DEFAULT '',
    exec_status TEXT NOT NULL DEFAULT '',
    sub_total INTEGER NOT NULL DEFAULT 0,
    sub_done INTEGER NOT NULL DEFAULT 0,
    progress REAL NOT NULL DEFAULT 0,
    rag TEXT NOT NULL DEFAULT '',
    risk_note TEXT NOT NULL DEFAULT '',
    decision_needed TEXT NOT NULL DEFAULT '',
    support_needed TEXT NOT NULL DEFAULT '',
    duration_days TEXT NOT NULL DEFAULT '',
    status TEXT NOT NULL DEFAULT 'active',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

-- 工作項底下的子項目（使用者 2026-08-12：「子項總數怎不能繼續追下去」）。
-- 原本 project_items 只有 sub_total／sub_done 兩個數字欄位，填了 3/3 也看不出那三項是什麼。
-- 拆了子項之後，那兩個數字改由這張表算出來，不再手填。
CREATE TABLE IF NOT EXISTS project_subitems (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    item_id INTEGER NOT NULL,
    seq INTEGER NOT NULL DEFAULT 0,
    name TEXT NOT NULL DEFAULT '',
    owner TEXT NOT NULL DEFAULT '',
    start_date TEXT NOT NULL DEFAULT '',
    end_date TEXT NOT NULL DEFAULT '',
    done INTEGER NOT NULL DEFAULT 0,
    note TEXT NOT NULL DEFAULT '',
    status TEXT NOT NULL DEFAULT 'active',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

-- WBS 展延紀錄（第三次回饋 8.4「申請展延日期」）：工作項逾期需展延結束日時，
-- 保留原日期及展延歷程，不直接覆蓋——跟 §10 合約調整（contract_adjustments）同一個 pattern。
-- project_items.end_date 永遠是「現在的結束日」，「什麼時候、為什麼、從哪天展延到哪天」查這張表。
CREATE TABLE IF NOT EXISTS project_item_extensions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    item_id INTEGER NOT NULL,
    old_end_date TEXT NOT NULL DEFAULT '',
    new_end_date TEXT NOT NULL DEFAULT '',
    reason TEXT NOT NULL DEFAULT '',
    note TEXT NOT NULL DEFAULT '',
    created_by TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS budget_allocations (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    budget_id INTEGER NOT NULL,
    seq INTEGER NOT NULL DEFAULT 0,
    unit_code TEXT NOT NULL DEFAULT '',
    unit_name TEXT NOT NULL DEFAULT '',
    share_pct REAL NOT NULL DEFAULT 0,
    amount REAL NOT NULL DEFAULT 0,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

-- 年度費用明細（L3）：一個預算項目在「某年度×某期間」的金額；
-- 全年度＝同年各期間加總、費用差異＝與相鄰前一年比，皆讀取時動態算、不存死。
CREATE TABLE IF NOT EXISTS budget_periods (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    budget_id INTEGER NOT NULL,
    fiscal_year TEXT NOT NULL DEFAULT '',
    period TEXT NOT NULL DEFAULT '',
    amount REAL NOT NULL DEFAULT 0,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

-- 年度費用比較的「每年備註」（L3）：給主管/助理寫差異說明、決策註記。一預算一年一筆。
CREATE TABLE IF NOT EXISTS budget_year_notes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    budget_id INTEGER NOT NULL,
    fiscal_year TEXT NOT NULL DEFAULT '',
    note TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    UNIQUE(budget_id, fiscal_year)
);

CREATE TABLE IF NOT EXISTS unit_headcounts (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    unit_code TEXT NOT NULL DEFAULT '',
    unit_name TEXT NOT NULL DEFAULT '',
    headcount INTEGER NOT NULL DEFAULT 0,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

-- 單位主檔（Step 2）：每個真實單位一個永不變的內部編號；代號/名稱都是別名。
-- 非破壞式：原始資料(budget_allocations/unit_headcounts)不動，讀取時經別名認到同一主檔。
CREATE TABLE IF NOT EXISTS unit_master (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    canonical_code TEXT NOT NULL DEFAULT '',
    canonical_name TEXT NOT NULL DEFAULT '',
    status TEXT NOT NULL DEFAULT 'active',
    note TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS personnel_master (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL DEFAULT '',
    status TEXT NOT NULL DEFAULT 'active',
    note TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS unit_aliases (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    master_id INTEGER NOT NULL,
    alias_code TEXT NOT NULL DEFAULT '',
    alias_name TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    UNIQUE(alias_code, alias_name)
);

-- 單位裁決紀錄（防呆＋後悔藥）：每次合併/分開都留誰、何時、為什麼，
-- 並記 undo_ops（每個別名的前一個歸屬）以便逐筆復原。原始資料本就不動，這裡是決策層的可逆帳。
CREATE TABLE IF NOT EXISTS unit_decisions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    action TEXT NOT NULL DEFAULT '',
    reason TEXT NOT NULL DEFAULT '',
    actor TEXT NOT NULL DEFAULT '',
    detail_json TEXT NOT NULL DEFAULT '',
    undo_ops_json TEXT NOT NULL DEFAULT '',
    undone INTEGER NOT NULL DEFAULT 0,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

-- 按類別分攤基準（Phase 2）：一個「類別」（台股功能/複委託功能/台複共用…）底下，
-- 各單位的百分比（來源＝資訊架構部費用分攤表『對照』表 NEW 欄，每類加總=100%）。
CREATE TABLE IF NOT EXISTS category_shares (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    category TEXT NOT NULL DEFAULT '',
    unit_code TEXT NOT NULL DEFAULT '',
    unit_name TEXT NOT NULL DEFAULT '',
    share_pct REAL NOT NULL DEFAULT 0,
    source_file TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    UNIQUE(category, unit_code, unit_name)
);

-- 名稱歸納（比照單位主檔）：把「中華電信/中華電」這種同一實體的不同寫法歸成一個。
-- kind＝case(案件名)/project(專案名)/vendor(廠商名)。canonical＝以誰為準；別名皆對到它。
CREATE TABLE IF NOT EXISTS name_master (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    kind TEXT NOT NULL DEFAULT '',
    canonical_name TEXT NOT NULL DEFAULT '',
    note TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    UNIQUE(kind, canonical_name)
);

CREATE TABLE IF NOT EXISTS name_aliases (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    master_id INTEGER NOT NULL,
    kind TEXT NOT NULL DEFAULT '',
    alias_name TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    UNIQUE(kind, alias_name)
);

CREATE TABLE IF NOT EXISTS name_decisions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    kind TEXT NOT NULL DEFAULT '',
    action TEXT NOT NULL DEFAULT '',
    reason TEXT NOT NULL DEFAULT '',
    actor TEXT NOT NULL DEFAULT '',
    detail_json TEXT NOT NULL DEFAULT '',
    undo_ops_json TEXT NOT NULL DEFAULT '',
    undone INTEGER NOT NULL DEFAULT 0,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);
"""


STATUS_VALUES: dict[str, dict[str, set[str]]] = {
    "cases": {
        # 需求書 §4 的審核結果：核准新案(approved)／退回補件(returned)／併入既有案(merged)／
        # 拒絕建立(rejected)。被拒絕與被併走的都留著（不是停用、也不是刪除），申請紀錄查得到。
        # 需求書 §4 完整狀態機：暫存→待審核→(退回補件)→已核准→進行中→(暫停)→已結案／已取消。
        # merged/rejected 是審核的兩個終點，disabled 是舊的「停用」（保留給既有資料）。
        "status": {"draft", "pending_review", "reviewing", "approved", "disabled",
                   "returned", "rejected", "merged",
                   "in_progress", "paused", "closed", "cancelled"},
        # 第一個重要判斷：占不占用年度預算。空＝舊資料還沒分流過。
        "budget_type": {"", "in_budget", "out_budget"},
        "expense_kind": {"", "expense", "capex"},   # 費用 / 資本支出
    },
    "contracts": {
        "status": {"active", "reviewing", "closed", "disabled"},
        # 合約性質（助理 0803 用語：新購／續約／增購附屬）＝與舊約的關係。
        # 空＝新購；renew續約 / addon增購附屬 / merge整併（後三者都要指 parent_contract_id）。
        # 與 contract_type（採購/維護/租賃…）不同：那是「買什麼」，這是「跟哪份舊約的關係」。
        "relation_type": {"", "renew", "addon", "merge"},
        # 提前結束的原因：選了就轉灰燈（不再按到期日催）。空＝仍在正常存續。
        "end_reason": {"", "merged", "not_renew"},
    },
    "payments": {
        "invoice_status": {"not_received", "received", "verified"},
        "status": {"pending", "scheduled", "closed", "disabled"},
    },
    "documents": {
        "status": {"active", "reviewing", "archived", "disabled"},
    },
    "budgets": {
        "status": {"active", "closed", "disabled"},
    },
    "projects": {
        "status": {"active", "completed", "paused", "disabled"},
    },
    "signoffs": {
        "status": {"draft", "submitted", "approved", "rejected", "disabled"},
    },
    "expense_masters": {
        "status": {"active", "closed", "disabled"},
    },
    "expense_sections": {
        # 助理 0803：里程碑／定期費用／最低承諾金額，可複選成混合型
        "mode": {"milestone", "periodic", "commitment"},
        "price_method": {"", "percent", "fixed"},
        "frequency": {"", "monthly", "quarterly", "semi", "yearly"},
        "status": {"draft", "confirmed"},
        # 最低承諾金額：後續各期承諾金額怎麼來、達成率用哪個金額算
        "next_amount_rule": {"", "same", "growth", "manual"},
        "achievement_basis": {"", "usage", "payable"},   # 使用金額／應付金額
    },
    "expense_settlements": {
        # 助理 0803 第六節的處理進度五態，預設「發票尚未收到」
        "progress": {"invoice_pending", "ready_to_sign", "signing", "approved", "submitted"},
    },
    "purchases": {
        "status": {"pending", "ordered", "arrived", "closed", "disabled"},
    },
}


def dict_row(cursor: sqlite3.Cursor, row: sqlite3.Row) -> dict[str, Any]:
    return {column[0]: row[index] for index, column in enumerate(cursor.description)}


@contextmanager
def connect() -> Iterator[sqlite3.Connection]:
    settings = get_settings()
    db_path = Path(settings.database_path)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.row_factory = dict_row
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def initialize_database() -> None:
    with connect() as conn:
        conn.executescript(SCHEMA)
        ensure_column(conn, "documents", "status", "TEXT NOT NULL DEFAULT 'active'")
        ensure_column(conn, "audit_logs", "actor", "TEXT NOT NULL DEFAULT 'local-dev'")
        ensure_column(conn, "cases", "note", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "cases", "next_step", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "cases", "due_date", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "cases", "created_by", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "cases", "approved_by", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "cases", "approved_at", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "contracts", "end_date", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "projects", "due_date", "TEXT NOT NULL DEFAULT ''")
        # 專案：對齊真實 Excel 欄位
        ensure_column(conn, "projects", "level", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "projects", "progress_planned", "REAL NOT NULL DEFAULT 0")
        ensure_column(conn, "projects", "rag_status", "TEXT NOT NULL DEFAULT ''")
        # 專案進度總表：起訖日（供甘特／落後天數計算）
        ensure_column(conn, "projects", "start_date", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "projects", "end_date", "TEXT NOT NULL DEFAULT ''")
        # 需求書 §6 專案主檔：廠商、是否跨子公司（金控/集團合作案是主管與處長的關注條件）
        ensure_column(conn, "projects", "vendor_name", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "projects", "cross_company", "TEXT NOT NULL DEFAULT ''")
        # 助理 0803 附件二第三點：專案建立時先問「涉及請購或合約？」——
        # 是 → 自動排標準採購流程的工作項；否 → 由同仁自己建需要的工作項。
        ensure_column(conn, "projects", "involves_procurement", "INTEGER NOT NULL DEFAULT 0")
        # 費用模組第二層「最低承諾金額」模式（助理 0803 附件一 5.3）：
        # 承諾金額是費用管理門檻，不是每期一次付清；實際發生的費用在第三層登錄，
        # 系統再回頭算各期達成率、未達差額與超額轉入。
        ensure_column(conn, "expense_sections", "commit_span_months", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "expense_sections", "next_amount_rule", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "expense_sections", "growth_pct", "REAL NOT NULL DEFAULT 0")
        ensure_column(conn, "expense_sections", "carry_over", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "expense_sections", "achievement_basis", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "expense_sections", "shortfall_action", "TEXT NOT NULL DEFAULT ''")
        # 排程明細屬於第幾個承諾期（最低承諾金額才有值）
        ensure_column(conn, "expense_schedules", "commit_period", "INTEGER NOT NULL DEFAULT 0")
        # 已封存的舊版費用區段（重新編輯時整段複製留存）：清單要看得到，但金額不能再被算一次——
        # 不標的話總費用檢核與「排程總額」會把同一段算兩遍。
        ensure_column(conn, "expense_sections", "archived", "INTEGER NOT NULL DEFAULT 0")
        # WBS 燈號是「人工指定」還是「系統自動判的」——不分開的話，自動判出來的值存進去之後
        # 會被誤認為人工指定，之後改子項目數就再也不會重算（做完了還掛著黃燈）。
        # 需求書 §6：「燈號可由系統判斷，也保留人工調整」，所以兩種都要留得住。
        ensure_column(conn, "project_items", "rag_manual", "INTEGER NOT NULL DEFAULT 0")
        # 預算共同費用分攤：尾數承擔單位（整數化後湊不齊的尾數歸給哪個單位；空＝自動抓填寫部門）
        ensure_column(conn, "budgets", "remainder_unit_code", "TEXT NOT NULL DEFAULT ''")
        # 預算分攤方法：fixed(固定金額) / headcount(按人數) / category(按類別，Phase 2)
        ensure_column(conn, "budgets", "alloc_method", "TEXT NOT NULL DEFAULT 'fixed'")
        ensure_column(conn, "budgets", "alloc_category_kind", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "budgets", "alloc_category", "TEXT NOT NULL DEFAULT ''")
        # 預算項目 metadata（L3）：對齊 Excel 的「費用內容／填寫部門／預估人員」
        ensure_column(conn, "budgets", "expense_detail", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "budgets", "fill_dept", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "budgets", "estimator", "TEXT NOT NULL DEFAULT ''")
        # 記匯入來源檔名，讓單位撞名清單能指回是哪個 Excel（人類追資料來源）
        ensure_column(conn, "budget_allocations", "source_file", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "unit_headcounts", "source_file", "TEXT NOT NULL DEFAULT ''")
        # 簽呈附件參照：貼簽呈系統連結或檔案位置（勾稽用，不存 PDF 本身）
        ensure_column(conn, "signoffs", "attachment_ref", "TEXT NOT NULL DEFAULT ''")
        # 人員歸屬組別（主機組/資料庫組/網路組…）：案件的「負責人」要能依組別過濾。
        # 組別本身是可維護的選項（不同單位組織不一樣），不寫死在程式裡。
        ensure_column(conn, "personnel_master", "group_name", "TEXT NOT NULL DEFAULT ''")
        # 助理 2026-08-13 回報「人員＋組別＋EMAIL 沒填好就沒辦法繼續測」：
        # 通知原本只查得到「登入帳號」的 email，但核銷者／負責人這些欄位存的是人員主檔的
        # 人名，寄信時對不到收件者。email 直接掛在人員主檔上，人名就找得到信箱。
        ensure_column(conn, "personnel_master", "email", "TEXT NOT NULL DEFAULT ''")
        # 帳號的管轄組別：組長要能由管理員指派管哪一組（內建 ap05 寫在程式裡，
        # 後台自建的組長帳號靠這個欄位）。非組長角色留空即可。
        ensure_column(conn, "users", "group_name", "TEXT NOT NULL DEFAULT ''")
        # 系統編號：案件領「所屬年度＋四位流水號」，各階段共用此尾碼做跨階段勾稽
        ensure_column(conn, "cases", "fiscal_year", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "cases", "seq", "INTEGER NOT NULL DEFAULT 0")
        # 助理回饋（2026-07-29）案件申請表單欄位：組別、預算內/外、費用or資本支出、預算名目、
        # 案件來源、案件說明。「預算內/外」是第一個重要判斷——決定要不要占用年度預算，
        # 也決定預算名目是「從匯入的預算表挑」還是「自己填」。
        ensure_column(conn, "cases", "group_name", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "cases", "budget_type", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "cases", "expense_kind", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "cases", "budget_item", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "cases", "source", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "cases", "description", "TEXT NOT NULL DEFAULT ''")
        # 需求書 §4 審核關卡：申請階段只有暫時號（temp_seq），核准才配正式 seq；
        # 退回補件／拒絕建立要留原因；併入既有案要記併到哪一件（不能只是把申請刪掉）。
        ensure_column(conn, "cases", "temp_seq", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "cases", "review_note", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "cases", "merged_into_case_id", "INTEGER")
        # 需求書 §4：已結案案件可重新開啟，但「必須記錄重新開啟人、時間與原因」。
        # 這裡存最後一次重開；完整歷史在 audit_logs（同一件案可能重開多次）。
        ensure_column(conn, "cases", "reopened_by", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "cases", "reopened_at", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "cases", "reopen_reason", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "cases", "status_note", "TEXT NOT NULL DEFAULT ''")   # 暫停/取消原因
        # Excel 來源勾稽：記匯入來源檔＋原始列號，清單顯示 📎 讓人回 Excel 核對
        ensure_column(conn, "cases", "source_file", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "cases", "source_row", "INTEGER NOT NULL DEFAULT 0")
        # 付款(核銷)：對齊真實費用整合表欄位
        for col in ("item", "settle_no", "ref_no", "period", "billing_period",
                    "settled_by", "vendor", "approval_level", "owner", "owner_email"):
            ensure_column(conn, "payments", col, "TEXT NOT NULL DEFAULT ''")
        # 核銷編號流水號：核銷各年獨立遞增，組成「功能碼+西元年+四位流水號」(12碼無連字號)
        ensure_column(conn, "payments", "settle_seq", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "payments", "net_amount", "REAL NOT NULL DEFAULT 0")
        ensure_column(conn, "payments", "tax_amount", "REAL NOT NULL DEFAULT 0")
        # §8 預計/實際分離：實際費用(payments)可回指它所履行的預計付款排程(可留白＝臨時/非合約費用)
        ensure_column(conn, "payments", "payment_schedule_id", "INTEGER")
        # §8 合約付款方式：驅動預計付款排程自動產生（fixed 固定金額 / installment 固定期數 /
        # periodic 週期 / milestone 里程碑%）。installments 供固定期數用。
        ensure_column(conn, "contracts", "payment_method", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "contracts", "installments", "INTEGER NOT NULL DEFAULT 0")
        # 簽呈/請購串接（方案A：只存關聯，不重做簽核系統）：請購可關聯核准它的簽呈，合約可關聯源自的請購
        ensure_column(conn, "purchases", "signoff_id", "INTEGER")
        ensure_column(conn, "contracts", "purchase_id", "INTEGER")
        # 合約模型補齊：原本只有到期日，無起始日／類型／續約增購關係／保固維護期限，
        # 導致「這份約什麼時候開始、是續哪一份、保固到哪天」都得翻紙本合約。
        ensure_column(conn, "contracts", "start_date", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "contracts", "contract_type", "TEXT NOT NULL DEFAULT ''")
        # 續約/增購/整併：三者都是「本約源自哪一份舊約」，用同一組欄位表達，不另開關聯表。
        # parent_contract_id 空＝全新合約；relation_type 見 STATUS_VALUES.contracts.relation_type。
        ensure_column(conn, "contracts", "parent_contract_id", "INTEGER")
        ensure_column(conn, "contracts", "relation_type", "TEXT NOT NULL DEFAULT ''")
        # 保固／維護到期日：與合約到期日不同（合約結束後保固/維護常還在跑），到期提醒要分開看
        ensure_column(conn, "contracts", "warranty_end_date", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "contracts", "maintenance_end_date", "TEXT NOT NULL DEFAULT ''")
        # 助理 2026-08-03 欄位規格：合約主檔要能獨立回答「誰的約、跟誰簽、在哪個機房、
        # 公司合約系統怎麼查、快到期了處理到哪」，這些原本得翻案件或問人。
        ensure_column(conn, "contracts", "system_code", "TEXT NOT NULL DEFAULT ''")      # CT+年+流水，增購掛 A01
        ensure_column(conn, "contracts", "system_seq", "INTEGER NOT NULL DEFAULT 0")     # 發號流水（只計自動發的）
        ensure_column(conn, "contracts", "vendor_tax_id", "TEXT NOT NULL DEFAULT ''")    # 廠商統編，8 碼數字
        ensure_column(conn, "contracts", "owner", "TEXT NOT NULL DEFAULT ''")            # 合約負責人
        ensure_column(conn, "contracts", "group_name", "TEXT NOT NULL DEFAULT ''")       # 組別
        ensure_column(conn, "contracts", "locations", "TEXT NOT NULL DEFAULT ''")        # 地點/機房，可複選（逗號分隔）
        ensure_column(conn, "contracts", "external_code", "TEXT NOT NULL DEFAULT ''")    # 公司內部合約系統編號
        ensure_column(conn, "contracts", "progress_note", "TEXT NOT NULL DEFAULT ''")    # 合約進度說明（黃/紅燈必填）
        ensure_column(conn, "contracts", "end_reason", "TEXT NOT NULL DEFAULT ''")       # 已整併/不續約 → 燈號轉灰
        ensure_column(conn, "contracts", "project_id", "INTEGER")                        # 對應專案
        # 備註是助理 0803 附件一就列的欄位（合約主檔最後一項），先前漏加；
        # source_file/source_row 讓匯入進來的合約指得回盤點表的哪一列（比照案件匯入）
        ensure_column(conn, "contracts", "note", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "contracts", "source_file", "TEXT NOT NULL DEFAULT ''")
        ensure_column(conn, "contracts", "source_row", "INTEGER NOT NULL DEFAULT 0")
        _fill_missing_contract_system_codes(conn)
        _refresh_wbs_progress_from_exec_status(conn)


def _fill_missing_contract_system_codes(conn: sqlite3.Connection) -> int:
    """開機補號：既有合約沒有系統識別碼就配一個。

    system_code 是 2026-08-03 才加的欄位，之前建的合約全是空的，畫面上「系統識別碼」
    那一欄就整排空白。原本要人到後台按一次補號才會有——沒人知道要按，等於預設壞的。
    這裡在建表／補欄位之後直接補掉，重開就是好的。

    冪等且便宜：先數一次還有幾筆沒號，是 0 就直接回。先主約後增購，因為增購的子號
    要接在原合約的識別碼後面，父的還沒號就配不出來。
    """
    if conn.execute("SELECT COUNT(*) n FROM contracts WHERE COALESCE(system_code,'') = ''"
                    ).fetchone()["n"] == 0:
        return 0
    filled = 0
    for is_addon_pass in (False, True):
        rows = conn.execute(
            "SELECT id, relation_type, parent_contract_id, start_date FROM contracts "
            "WHERE COALESCE(system_code, '') = '' ORDER BY id").fetchall()
        for r in rows:
            row_is_addon = str(r["relation_type"] or "") == "addon" and bool(r["parent_contract_id"])
            if row_is_addon is not is_addon_pass:
                continue
            code = _next_contract_system_code(conn, {
                "relation_type": r["relation_type"],
                "parent_contract_id": r["parent_contract_id"],
                "start_date": r["start_date"],
            })
            conn.execute("UPDATE contracts SET system_code = ?, system_seq = ? WHERE id = ?",
                         (code["system_code"], code["system_seq"], r["id"]))
            filled += 1
    return filled


def _refresh_wbs_progress_from_exec_status(conn: sqlite3.Connection) -> int:
    """開機修正：沒拆子項但「執行進度」寫了已完成的工作項，進度補成 100%、燈號重判。

    這些列原本進度是 0%，過了結束日就被自動判成紅燈「已延遲」，畫面上跟同一列的
    「執行進度：已完成」自相矛盾（實際資料裡 26 個專案都中）。修法在 wbs_item_progress，
    但既有資料的 progress／rag 是存在資料庫裡的，要一起刷過才看得到效果。

    只動系統自動判的（rag_manual=0）。人工指定過的燈號不碰——那是使用者的決定。
    """
    rows = conn.execute(
        "SELECT id, project_id, sub_total, sub_done, progress, rag, exec_status, start_date, end_date "
        "FROM project_items WHERE COALESCE(sub_total,0) <= 0 AND COALESCE(progress,0) <= 0 "
        "AND COALESCE(exec_status,'') <> ''").fetchall()
    touched, projects = 0, set()
    for r in rows:
        if not wbs_exec_done(r["exec_status"]):
            continue
        rag = wbs_auto_rag(100.0, r["start_date"], r["end_date"])
        conn.execute(
            "UPDATE project_items SET progress = 100, rag = CASE WHEN COALESCE(rag_manual,0) = 1 "
            "THEN rag ELSE ? END WHERE id = ?", (rag, r["id"]))
        projects.add(r["project_id"])
        touched += 1
    for pid in projects:
        _recompute_project_rollup(conn, pid)     # 專案層的完成度／燈號跟著重算
    return touched


def ensure_column(conn: sqlite3.Connection, table: str, column: str, definition: str) -> None:
    columns = {row["name"] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}
    if column not in columns:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")


_FK_REFS = {
    "case_id": ("cases", "案件"),
    "contract_id": ("contracts", "合約"),
    "signoff_id": ("signoffs", "簽呈"),
    "purchase_id": ("purchases", "請購"),
    "parent_contract_id": ("contracts", "來源合約"),  # 續約/增購/整併指向的舊約
    "project_id": ("projects", "專案"),               # 合約掛的專案（助理 0803）
    "expense_id": ("expense_masters", "費用主檔"),     # 費用區段掛的主檔
    "section_id": ("expense_sections", "費用區段"),    # 排程明細掛的區段
    "schedule_id": ("expense_schedules", "費用排程"),  # 第三層請款／實際費用對應的那一期
    "actual_id": ("expense_actuals", "實際費用明細"),
}


def _validate_fks(conn: sqlite3.Connection, fields: dict[str, Any]) -> None:
    """關聯 ID（case_id / contract_id）若有填，必須指向存在的資料，否則擋下。"""
    for fk, (ref_table, label) in _FK_REFS.items():
        val = fields.get(fk)
        if val is None:
            continue
        if conn.execute(f"SELECT 1 FROM {ref_table} WHERE id = ?", (val,)).fetchone() is None:
            raise ValueError(f"關聯的{label} ID {val} 不存在，請確認後再填。")


_LINEAGE_MAX_DEPTH = 20  # 續約鏈往上追的層數上限（正常合約不會超過；純粹防資料異常繞不完）


def _validate_contract_parent(conn: sqlite3.Connection, contract_id: int, parent_id: int) -> None:
    """續約/增購/整併指的『來源合約』不能是自己，也不能繞回自己（A續B、B又續A）——
    否則續約鏈會追不完，畫面會卡死。"""
    if int(parent_id) == int(contract_id):
        raise ValueError("來源合約不能是自己。")
    seen = {int(contract_id)}
    cur = int(parent_id)
    for _ in range(_LINEAGE_MAX_DEPTH):
        if cur in seen:
            raise ValueError("這樣設定會讓合約續約關係繞成一個圈，請改指到正確的舊約。")
        seen.add(cur)
        row = conn.execute("SELECT parent_contract_id FROM contracts WHERE id = ?", (cur,)).fetchone()
        if row is None or row["parent_contract_id"] is None:
            return
        cur = int(row["parent_contract_id"])


_TAX_ID_OK = re.compile(r"^\d{8}$")


def _validate_contract(conn: sqlite3.Connection, fields: dict[str, Any],
                       before: dict[str, Any] | None = None) -> None:
    """合約主檔的商業檢核（助理 0803 規格）。更新時只送部分欄位，所以拿舊值補齊再判——
    不然「只改到期日」會因為看不到 progress_note 而誤判必填。"""
    def val(key: str) -> str:
        if key in fields:
            return str(fields.get(key) or "").strip()
        return str((before or {}).get(key) or "").strip()

    tax = val("vendor_tax_id")
    if tax and not _TAX_ID_OK.match(tax):
        raise ValueError("廠商統一編號要是 8 碼數字。")

    start, end = val("start_date"), val("end_date")
    if start and end and end[:10] < start[:10]:
        raise ValueError("合約迄日不能早於合約起日。")

    # 「黃燈/紅燈要填進度說明」刻意不在這裡擋存檔——助理規格原文是
    # 「到期警示為黃燈且進度說明未填寫時，不得將合約到期追蹤標示為完成」，
    # 要的是追蹤不能算結案，不是不准建檔。擋存檔會連帶擋掉匯入既有合約、
    # 示範資料與所有「先把約建進來、說明晚點補」的正常流程（那些約本來就快到期了）。
    # 改成標記出來：contract_needs_progress_note() 供清單、待辦與提醒顯示。

    # 增購／附屬一定要指到原合約，而且只能指同一個案件底下的——
    # 跨案件掛增購，金額與追蹤都會算到別人的案子上
    if val("relation_type") == "addon":
        parent_id = fields.get("parent_contract_id", (before or {}).get("parent_contract_id"))
        if not parent_id:
            raise ValueError("增購／附屬合約必須指定原合約。")
        parent = conn.execute(
            "SELECT case_id FROM contracts WHERE id = ?", (int(parent_id),)).fetchone()
        case_id = fields.get("case_id", (before or {}).get("case_id"))
        if parent is not None and case_id and parent["case_id"] and int(parent["case_id"]) != int(case_id):
            raise ValueError("原合約必須是同一個案件底下的合約。")


def _next_contract_system_code(conn: sqlite3.Connection, fields: dict[str, Any]) -> dict[str, Any]:
    """配合約系統識別碼：新購/續約/整併＝CT＋年＋四位流水；增購附屬＝原識別碼＋A＋兩位流水。

    年份取合約起日，沒填就用作業年度——識別碼要能一眼看出是哪一年的約，
    用建檔當下的日期會讓補建的舊約掛到錯的年份。
    """
    if str(fields.get("relation_type") or "") == "addon" and fields.get("parent_contract_id"):
        parent = conn.execute("SELECT system_code FROM contracts WHERE id = ?",
                              (int(fields["parent_contract_id"]),)).fetchone()
        base = str((parent or {})["system_code"] if parent else "").strip()
        if base:
            n = conn.execute(
                "SELECT COUNT(*) c FROM contracts WHERE system_code LIKE ?", (base + "A%",)).fetchone()["c"]
            for i in range(n + 1, n + 100):          # 撞號往後找，重跑不會蓋掉既有子號
                code = f"{base}A{i:02d}"
                if conn.execute("SELECT 1 FROM contracts WHERE system_code = ?", (code,)).fetchone() is None:
                    return {"system_code": code, "system_seq": 0}
    start = str(fields.get("start_date") or "").strip()
    year = start[:4] if len(start) >= 4 and start[:4].isdigit() else get_working_year()
    seq = conn.execute(
        "SELECT COALESCE(MAX(system_seq), 0) + 1 AS n FROM contracts WHERE substr(system_code, 3, 4) = ?",
        (year,)).fetchone()["n"]
    return {"system_code": f"{CONTRACT_PREFIX}{year}{seq:04d}", "system_seq": seq}


def contract_addon_options(case_id: int) -> dict[str, Any]:
    """同一個案件底下可以當「原合約」的既有合約。

    助理 0803 規格：0 筆 → 增購／附屬要停用（不給點）；1 筆 → 自動帶入且不給改；
    2 筆以上 → 給使用者選，且必填。前端照 mode 決定怎麼顯示，判斷邏輯只留這一份。
    """
    with connect() as conn:
        rows = conn.execute(
            "SELECT id, system_code, contract_code, contract_name FROM contracts "
            "WHERE case_id = ? AND status <> 'disabled' ORDER BY id", (case_id,)).fetchall()
    items = [dict(r) for r in rows]
    mode = "disabled" if not items else ("auto" if len(items) == 1 else "choose")
    return {"case_id": case_id, "mode": mode, "count": len(items), "contracts": items,
            "hint": {"disabled": "此案件尚無既有合約，無法建立增購／附屬合約。",
                     "auto": "此案件只有一份合約，系統自動帶入為原合約。",
                     "choose": "此案件有多份合約，請選擇要掛在哪一份底下。"}[mode]}


def contract_lineage(contract_id: int) -> list[dict[str, Any]]:
    """本約的來源鏈（續約/增購/整併一路往上追的舊約），由近而遠。
    主管問「這份維護約續幾年了」時，看這條鏈就知道，不用翻紙本。"""
    out: list[dict[str, Any]] = []
    seen: set[int] = {int(contract_id)}
    with connect() as conn:
        row = conn.execute(
            "SELECT parent_contract_id, relation_type FROM contracts WHERE id = ?", (contract_id,)).fetchone()
        relation = row["relation_type"] if row else ""
        cur = row["parent_contract_id"] if row else None
        while cur is not None and len(out) < _LINEAGE_MAX_DEPTH:
            if int(cur) in seen:
                break  # 資料異常繞圈：停在這裡，不讓畫面追不完
            seen.add(int(cur))
            parent = conn.execute(
                "SELECT id, contract_code, contract_name, amount, start_date, end_date, "
                "parent_contract_id, relation_type FROM contracts WHERE id = ?", (cur,)).fetchone()
            if parent is None:
                break
            item = dict(parent)
            item["relation_to_child"] = relation  # 下一份約是用什麼關係接上來的（續約/增購/整併）
            out.append(item)
            relation = item["relation_type"]
            cur = item["parent_contract_id"]
    return out


def get_working_year() -> str:
    """目前作業年度：新案件『所屬年度』的預設值（例如今年 8 月就開始編明年預算）。
    設定沒填則退回今年。"""
    v = read_settings(["working_year"]).get("working_year", "")
    if str(v).strip():
        return str(v).strip()
    import datetime
    return str(datetime.date.today().year)


# 核銷編號：12 碼無連字號＝功能碼(4)＋西元年(4)＋流水號(4)，例 Sett20260012（主管指定格式）
SETTLE_PREFIX = "Sett"

# 合約系統識別碼：CT＋西元年＋四位流水（例 CT20260001）。增購／附屬掛在原合約底下，
# 用「原識別碼＋A＋兩位流水」（例 CT20260001A01），一眼看得出誰是誰的增購。
# 助理 0803 文件寫成 CT-2026-0001，但主管交代編號不得含連字號，故拿掉分隔符（見 _CODE_OK）。
CONTRACT_PREFIX = "CT"
CONTRACT_LIGHT_LABEL = {"red": "已到期", "yellow": "3 個月內到期", "green": "尚未接近到期",
                        "gray": "已整併／不續約", "none": "未設到期日"}


def contract_expiry_light(end_date: Any, end_reason: Any = "", today: date | None = None) -> str:
    """合約到期警示（助理 0803 規格）：
    紅＝已到期／黃＝距到期日 3 個月內／綠＝超過 3 個月／灰＝已整併或不續約。

    灰燈優先於日期：已經整併或確定不續約的約，再催到期沒有意義。
    沒填到期日回 none——「沒有日期」跟「日期還很遠」是兩回事，混成綠燈會讓人以為查過了。
    即時計算不落地：燈號每天都在變，存進資料庫就得靠排程去刷，漏跑一天就是錯的。
    """
    if str(end_reason or "").strip() in ("merged", "not_renew"):
        return "gray"
    raw = str(end_date or "").strip()
    if not raw:
        return "none"
    try:
        due = datetime.strptime(raw[:10], "%Y-%m-%d").date()
    except ValueError:
        return "none"
    now = today or date.today()
    if due < now:
        return "red"
    return "yellow" if due <= _add_months_date(now, 3) else "green"


def contract_needs_progress_note(row: Any) -> bool:
    """這份合約「到期追蹤還不能算完成」——快到期或已到期，卻沒寫處理到哪。

    助理 0803 規格：黃燈且進度說明未填時，不得把到期追蹤標示為完成。
    做成旗標而不是存檔檢核，既有合約與匯入資料才進得來（它們一進來常常就是黃燈）。
    """
    row = dict(row)   # sqlite3.Row 沒有 .get()，先轉成 dict 統一取值
    light = contract_expiry_light(row.get("end_date"), row.get("end_reason"))
    return light in ("yellow", "red") and not str(row.get("progress_note") or "").strip()


def _add_months_date(d: date, months: int) -> date:
    """d 往後 n 個月；當月沒有那一天就取月底（例：11/30 + 3 個月 → 2/28）。"""
    y, m = divmod(d.year * 12 + (d.month - 1) + months, 12)
    m += 1
    last = [31, 29 if (y % 4 == 0 and y % 100 != 0) or y % 400 == 0 else 28,
            31, 30, 31, 30, 31, 31, 30, 31, 30, 31][m - 1]
    return date(y, m, min(d.day, last))

# 主管交代（2026-08-03）：系統自動產生的編號一律不得含連字號、底線與中文，
# 只能是英數。理由是這些號碼要能貼進其他系統、當檔名、當搜尋關鍵字，
# 分隔符號與全形字在那些地方常出事。
# 範圍限「系統自己配的號」；來源帶進來的編號（公司合約系統編號、發票號碼、
# Excel 匯入原本就有的案件編號）照原樣保留，那是別人的號、不是我們配的。
_CODE_OK = re.compile(r"^[A-Za-z0-9]+$")


def is_system_code_valid(code: Any) -> bool:
    """系統自動產生的編號是否合規（純英數，無 - _ 與中文）。"""
    return bool(_CODE_OK.match(str(code or "")))


def _match_owner_username(conn: sqlite3.Connection, owner_display_name: str | None) -> str | None:
    """把專案的「負責人」欄位（可能是「令狐沖/黃蓉」這種"/"分隔共同負責人）比對到登入帳號的顯示名稱，
    抓到剛好一個相符帳號才回傳其 username，用來讓自動生成的案件自動掛對承辦人。
    抓不到、或抓到不只一個相符帳號，保守回 None——寧可留白讓人工指派，也不要瞎猜指派錯人。"""
    name = str(owner_display_name or "").strip()
    if not name:
        return None
    parts = [p.strip() for p in name.split("/") if p.strip()]
    if not parts:
        return None
    placeholders = ",".join("?" for _ in parts)
    rows = conn.execute(
        f"SELECT username FROM users WHERE disabled = 0 AND display_name IN ({placeholders})", parts
    ).fetchall()
    usernames = {r["username"] for r in rows}
    if len(usernames) == 1:
        return next(iter(usernames))
    return None


def _ensure_case_for(
    conn: sqlite3.Connection, name: str | None, code_hint: str | None, fiscal_year: str | None,
    owner_display_name: str | None = None, established: bool = False,
) -> int | None:
    """使用者的心智模型裡沒有「先建一個叫案件的空殼」這一步——「案子」就是那筆預算/專案本身。
    建預算/專案沒給 case_id 時，用這個名稱找或建一個同名案件、自動掛上，讓使用者感覺不到「案件」這層存在。
    標題完全相同才視為同一案（不做模糊比對，避免系統瞎猜合併不相干的東西——命名不一致要靠既有「＋歸戶」人工改掛）。
    沒有名稱就回 None，呼叫端維持 case_id 為空，不強迫。
    owner_display_name（僅專案有）：若能唯一比對到一個登入帳號，新案件直接掛該帳號為負責人
    （若觸發者本身是承辦，_insert_row 既有規則「承辦建案自動歸自己」會再覆蓋一次，維持原本行為）。
    established：匯入／回填舊資料時傳 True，配出來的案件直接算已成立（使用者拍板 A2）；
    一般在系統裡新建預算/專案時維持 False，那條路配出來的案件仍是申請中（走複核）。"""
    name = str(name or "").strip()
    if not name:
        return None
    existing = conn.execute("SELECT id FROM cases WHERE title = ?", (name,)).fetchone()
    if existing:
        return existing["id"]
    # 案件編號只收「合規的英數代碼」當提示（如專案代碼 PRJ20260001）。
    # 預算/專案匯入常把中文名稱當代碼傳進來，那種一律不用——留空讓 _insert_row 自己配號，
    # 否則案件編號會變成一串中文，違反主管交代的編號規則（見 is_system_code_valid）。
    hint = str(code_hint or "").strip()
    code = hint if is_system_code_valid(hint) else ""
    n = 1
    while code and conn.execute("SELECT 1 FROM cases WHERE case_code = ?", (code,)).fetchone() is not None:
        n += 1
        code = f"{hint}A{n:02d}"   # 撞號往後掛 A02、A03（不用連字號）
    payload: dict[str, Any] = {"case_code": code, "title": name, "fiscal_year": fiscal_year or ""}
    matched = _match_owner_username(conn, owner_display_name)
    if matched:
        payload["owner"] = matched
    if established:
        with import_mode():
            new_case = _insert_row(conn, "cases", payload)
    else:
        new_case = _insert_row(conn, "cases", payload)
    return new_case["id"]


def _established_case_fields(conn: sqlite3.Connection, fiscal_year: Any) -> dict[str, Any]:
    """匯入既有案件時要補的欄位：當年度正式流水號＋已成立狀態。

    approved_by 特意標「（匯入）」而不是只寫操作者帳號——這件案子沒有真的經過雙人複核，
    是匯入時認定它本來就在跑；不標的話事後看不出差別，會誤以為有人複核過（稽核要看得出依據）。
    """
    fy = str(fiscal_year or "").strip() or get_working_year()
    seq = conn.execute(
        "SELECT COALESCE(MAX(seq), 0) + 1 AS n FROM cases WHERE fiscal_year = ?", (fy,)
    ).fetchone()["n"]
    return {
        "fiscal_year": fy,
        "seq": seq,
        "temp_seq": 0,          # 不發暫時號：它不是申請中的案子
        "status": "approved",
        "approved_by": f"{_current_actor.get()}（匯入）",
        "approved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def _insert_row(conn, table: str, payload: dict[str, Any]) -> dict[str, Any]:
    """insert_row 的核心邏輯，吃外部傳入的連線——供需要跨表單一交易的呼叫方使用（如 create_case_wizard）。"""
    scope = _owner_scope.get()
    if table == "cases":
        payload = {**payload, "created_by": _current_actor.get()}  # 記錄建立者，供雙人複核擋自己核自己
        if scope is not None and _scope_kind.get() == "owner":
            payload = {**payload, "owner": scope}  # 承辦建案自動歸自己
        elif scope is not None and _scope_kind.get() == "group":
            # 組長建案（他自己當承辦時）：組別自動帶自己的組，否則他建完會看不到這件案子。
            # 負責人不強制改成自己——組長常代組員開案，負責人由他自己在表單指定。
            if not str(payload.get("group_name") or "").strip():
                payload = {**payload, "group_name": scope}
        # 所屬年度預設＝作業年度；流水號在交易內配發（同年遞增）
        payload = {**payload, "fiscal_year": str(payload.get("fiscal_year") or "").strip() or get_working_year()}
    allowed = allowed_fields()
    fields = {key: value for key, value in payload.items() if key in allowed[table]}
    if not fields:
        raise ValueError("No valid fields supplied.")
    validate_status_fields(table, fields)
    if table in ("budgets", "projects") and not fields.get("case_id"):
        # 案件自動生成：使用者拍板「不需要案件，是系統要幫我建出一個案件」——
        # 建預算/專案時沒指定案件，就用這筆自己的名稱/代碼幫它配一個同名案件。
        name = fields.get("project_name") if table == "projects" else fields.get("budget_code")
        code_hint = fields.get("project_code") if table == "projects" else fields.get("budget_code")
        owner_hint = fields.get("owner") if table == "projects" else None
        cid = _ensure_case_for(conn, name, code_hint, fields.get("fiscal_year"), owner_hint)
        if cid:
            fields["case_id"] = cid
    if table == "cases" and _import_mode.get():
        # 匯入的是已經在跑的舊案子，不是新申請（使用者拍板 A2）：直接配正式流水號、狀態算已成立，
        # 不落草稿也不發 TMP- 暫時號，省掉匯入完還要人工補號那一步。
        fields = {**fields, **_established_case_fields(conn, fields.get("fiscal_year", ""))}
        if not str(fields.get("case_code") or "").strip():
            fields["case_code"] = f"{fields.get('fiscal_year', '')}{fields['seq']:04d}"
    elif table == "cases":
        # 需求書 §4＋使用者拍板(A案)：申請階段只配「暫時號」，核准才配正式流水號。
        # 這樣被駁回／被併走的申請不會吃掉正式號，年度編號不跳號。正式號在 approve_case 配。
        tmp = conn.execute(
            "SELECT COALESCE(MAX(temp_seq), 0) + 1 AS n FROM cases WHERE fiscal_year = ?",
            (fields.get("fiscal_year", ""),)).fetchone()["n"]
        fields = {**fields, "temp_seq": tmp, "seq": 0}
        # 案件編號改由系統產生（助理回饋：承辦不必填）。沒給就用暫時號當編號，
        # 核准後系統編號（Cont/Case+年+流水）另外算，這裡不改它，避免外部已引用的號碼變動。
        # 主管 2026-08-03 交代編號不得含連字號，所以是 TMP20260001 而不是 TMP-2026-0001。
        if not str(fields.get("case_code") or "").strip():
            fields["case_code"] = f"TMP{fields.get('fiscal_year', '')}{tmp:04d}"
    if table == "contracts":
        # 助理 0803 規格：合約主檔要有自己的系統識別碼（跟公司合約系統編號分開），
        # 增購／附屬掛在原合約底下用子號。檢核也在這裡做，走 API 或程式進來都擋得到。
        _validate_contract(conn, fields)
        if not str(fields.get("system_code") or "").strip():
            fields = {**fields, **_next_contract_system_code(conn, fields)}
        # 對應專案是系統關聯（助理註明「無須顯示給使用者」）：合約掛在案件上，
        # 案件底下有專案就自動接起來，不要求人再選一次。
        if not fields.get("project_id") and fields.get("case_id"):
            prj = conn.execute(
                "SELECT id FROM projects WHERE case_id = ? AND status <> 'disabled' ORDER BY id LIMIT 1",
                (int(fields["case_id"]),)).fetchone()
            if prj is not None:
                fields["project_id"] = prj["id"]
    if table == "expense_masters":
        fields = _prepare_expense_master(conn, fields)
    if table == "projects" and not str(fields.get("project_code") or "").strip():
        # 需求書 §6「專案不另設代號」＋助理回饋的新案申請沒有代號欄；系統自己配一個唯一碼，
        # 讓內部關聯與匯出仍有穩定識別（用 id 當流水，必唯一）。
        nxt = conn.execute("SELECT COALESCE(MAX(id), 0) + 1 AS n FROM projects").fetchone()["n"]
        fields["project_code"] = f"PRJ{get_working_year()}{nxt:04d}"
    if table == "payments" and not str(fields.get("settle_no") or "").strip():
        # 核銷編號自動發號：年度取核銷月份，流水號只計自動發的(settle_seq>0)，避免匯入真號污染
        pm = str(fields.get("payment_month") or "").strip()
        year = pm[:4] if len(pm) >= 4 and pm[:4].isdigit() else get_working_year()
        nseq = conn.execute(
            "SELECT COALESCE(MAX(settle_seq), 0) + 1 AS n FROM payments "
            "WHERE substr(payment_month, 1, 4) = ? AND settle_seq > 0",
            (year,)).fetchone()["n"]
        fields = {**fields, "settle_seq": nseq,
                  "settle_no": f"{SETTLE_PREFIX}{year}{nseq:04d}"}
    columns = ", ".join(fields)
    placeholders = ", ".join("?" for _ in fields)
    _validate_fks(conn, fields)
    cursor = conn.execute(
        f"INSERT INTO {table} ({columns}) VALUES ({placeholders})",
        list(fields.values()),
    )
    row_id = cursor.lastrowid
    row = get_row(conn, table, row_id)
    write_audit_log(conn, table, row_id, "create", None, row)
    return row


def insert_row(table: str, payload: dict[str, Any]) -> dict[str, Any]:
    with connect() as conn:
        return _insert_row(conn, table, payload)


def create_case_wizard(
    case: dict[str, Any],
    budget: dict[str, Any] | None,
    signoff: dict[str, Any] | None,
    purchase: dict[str, Any] | None,
    contract: dict[str, Any] | None,
    payment: dict[str, Any] | None,
    project: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """新案申請：一次送出，依序建案件→(可選)專案/合約/費用，全部自動帶上新案的 case_id。
    單一交易：任一步驟失敗，前面已建的一併回滾、什麼都不留下，使用者修正後可整批重送，
    不會卡在「半成功」的狀態。budget/signoff/payment 參數保留給既有呼叫方與匯入流程用。"""
    with connect() as conn:
        case_row = _insert_row(conn, "cases", case)
        case_id = case_row["id"]

        project_row = None
        if project is not None:
            # 專案負責人預設沿用案件負責人（助理回饋：仍可人工改）
            payload = {**project, "case_id": case_id}
            if not str(payload.get("owner") or "").strip():
                payload["owner"] = case.get("owner", "")
            project_row = _insert_row(conn, "projects", payload)

        budget_row = None
        if budget is not None:
            budget_row = _insert_row(conn, "budgets", {**budget, "case_id": case_id})

        signoff_row = None
        if signoff is not None:
            signoff_row = _insert_row(conn, "signoffs", {**signoff, "case_id": case_id})

        purchase_row = None
        if purchase is not None:
            purchase_row = _insert_row(conn, "purchases", {**purchase, "case_id": case_id})

        contract_row = None
        if contract is not None:
            contract_row = _insert_row(conn, "contracts", {**contract, "case_id": case_id})

        payment_row = None
        if payment is not None:
            payment_row = _insert_row(conn, "payments", {**payment, "contract_id": contract_row["id"]})

    return {
        "case": case_row,
        "project": project_row,
        "budget": budget_row,
        "signoff": signoff_row,
        "purchase": purchase_row,
        "contract": contract_row,
        "payment": payment_row,
    }


# ── §10 合約費用調整：同一份合約中途改金額（機櫃增減、電費調價）留下歷史，不蓋掉原值 ──
def list_contract_adjustments(contract_id: int) -> list[dict[str, Any]]:
    """某合約的費用調整歷史，新的在前。"""
    with connect() as conn:
        rows = conn.execute(
            "SELECT * FROM contract_adjustments WHERE contract_id = ? ORDER BY effective_date DESC, id DESC",
            (contract_id,),
        ).fetchall()
        return [dict(r) for r in rows]


def add_contract_adjustment(contract_id: int, new_amount: float, effective_date: str = "",
                            reason: str = "", note: str = "") -> dict[str, Any]:
    """記一筆費用調整：舊值自動取合約現值，算出差額，並把 contracts.amount 更新成調整後的值。
    合約金額永遠是「現在多少錢」，「什麼時候為什麼從多少調到多少」查這張表。
    調整紀錄不提供刪除（稽核用）——填錯就再調一次回去，兩筆都留著才看得出經過。"""
    new_amount = float(new_amount)
    with connect() as conn:
        contract = conn.execute("SELECT * FROM contracts WHERE id = ?", (contract_id,)).fetchone()
        if contract is None:
            raise ValueError(f"合約 ID {contract_id} 不存在。")
        old_amount = float(contract["amount"] or 0)
        if round(new_amount, 2) == round(old_amount, 2):
            raise ValueError("調整後金額與現值相同，沒有東西要記錄。")
        row = _insert_row(conn, "contract_adjustments", {
            "contract_id": contract_id,
            "effective_date": str(effective_date or "").strip(),
            "old_amount": old_amount,
            "new_amount": new_amount,
            "delta": round(new_amount - old_amount, 2),
            "reason": str(reason or "").strip(),
            "note": str(note or "").strip(),
            "created_by": _current_actor.get(),
        })
        before = dict(contract)
        conn.execute("UPDATE contracts SET amount = ? WHERE id = ?", (new_amount, contract_id))
        after = get_row(conn, "contracts", contract_id)
        write_audit_log(conn, "contracts", contract_id, "update", before, after)
        return row


def contract_adjustment_summary(contract_id: int) -> dict[str, Any]:
    """調整摘要：調過幾次、最初金額、累計增減。沒調過就回 0 筆。"""
    rows = list_contract_adjustments(contract_id)
    if not rows:
        return {"count": 0, "original_amount": None, "total_delta": 0.0, "items": []}
    oldest = rows[-1]  # 依 effective_date DESC 排序，最後一筆＝最早那次調整
    return {
        "count": len(rows),
        "original_amount": float(oldest["old_amount"]),
        "total_delta": round(sum(float(r["delta"]) for r in rows), 2),
        "items": rows,
    }


# ── §8 預計付款排程（Payment Schedule）：與實際費用(payments)分離的「預計付款」層 ──
def list_payment_schedules(contract_id: int) -> list[dict[str, Any]]:
    """某合約的所有預計付款排程，依期別序排。"""
    with connect() as conn:
        rows = conn.execute(
            "SELECT * FROM payment_schedules WHERE contract_id = ? ORDER BY seq, id",
            (contract_id,),
        ).fetchall()
        return [dict(r) for r in rows]


def _check_milestone_total(conn: sqlite3.Connection, contract_id: int) -> None:
    """里程碑付款：同一合約 method='milestone' 的 percent 合計必須 = 100（需求書 §8）。
    完全沒有里程碑排程時不檢核。"""
    has = conn.execute(
        "SELECT 1 FROM payment_schedules WHERE contract_id = ? AND method='milestone' LIMIT 1",
        (contract_id,),
    ).fetchone()
    if not has:
        return
    total = conn.execute(
        "SELECT COALESCE(SUM(percent),0) AS s FROM payment_schedules "
        "WHERE contract_id = ? AND method='milestone'",
        (contract_id,),
    ).fetchone()["s"]
    if round(float(total), 4) != 100:
        raise ValueError(f"里程碑付款比例合計為 {round(float(total),2)}%，必須等於 100% 才能送出。")


def validate_milestone_total(contract_id: int) -> None:
    with connect() as conn:
        _check_milestone_total(conn, contract_id)


def contract_payment_summary(contract_id: int) -> dict[str, float]:
    """一份合約的『預計 vs 實際』——需求書 §8 拆分後才算得準：
      planned        = 預計付款排程總額（未取消）
      paid           = 實際費用已付（payments.status='closed'）
      unpaid_planned = 還沒付的預計（planned − 已付），供『還欠多少』用
    預計走 payment_schedules、實際走 payments，兩者不重複計算。"""
    with connect() as conn:
        planned = float(conn.execute(
            "SELECT COALESCE(SUM(planned_amount),0) AS s FROM payment_schedules "
            "WHERE contract_id = ? AND status != 'cancelled'",
            (contract_id,),
        ).fetchone()["s"])
        paid = float(conn.execute(
            "SELECT COALESCE(SUM(CASE WHEN status='closed' THEN payment_amount ELSE 0 END),0) AS s "
            "FROM payments WHERE contract_id = ?",
            (contract_id,),
        ).fetchone()["s"])
    return {"planned": planned, "paid": paid, "unpaid_planned": max(0.0, planned - paid)}


def _split_even(amount: float, n: int, remainder_on: str = "last") -> list[float]:
    """把金額平均分 n 份，除不盡的零頭全歸某一期（不硬性規定——由 remainder_on 決定，因為
    每個合約寫法不同：有的第一期多、有的最後一期多）。各份加總必定 == amount，不會出現
    999,999.99 這種對不起來的數字。產生後每期金額仍可手動改，以符合合約實際分法。"""
    import math
    n = max(1, int(n))
    base = math.floor(amount / n * 100) / 100  # 無條件捨去到分，避免每期進位使總額超過
    parts = [base] * n
    remainder = round(amount - base * n, 2)     # 零頭
    idx = 0 if remainder_on == "first" else n - 1
    parts[idx] = round(parts[idx] + remainder, 2)
    return parts


_FREQ_STEP = {"monthly": 1, "quarterly": 3, "yearly": 12}  # 每月/每季/每年 → 遞增幾個月


def _add_months(ym: str, k: int) -> str:
    """'YYYY-MM' 加 k 個月，回 'YYYY-MM'（跨年自動進位）。給不出合法起始月就回空字串。"""
    try:
        y, m = (int(x) for x in str(ym).split("-")[:2])
    except (ValueError, TypeError):
        return ""
    total = (y * 12 + (m - 1)) + k
    return f"{total // 12:04d}-{total % 12 + 1:02d}"


def contract_has_linked_payment(contract_id: int) -> bool:
    """該合約是否已有任何『實際核銷回指某期排程』——有的話就不准整個重產排程，
    否則會把已付的連結弄丟。"""
    with connect() as conn:
        row = conn.execute(
            "SELECT 1 FROM payments WHERE contract_id = ? AND payment_schedule_id IS NOT NULL LIMIT 1",
            (contract_id,),
        ).fetchone()
        return row is not None


def clear_payment_schedules(contract_id: int) -> int:
    """清掉某合約所有預計排程（重產前用）。若已有實際核銷回指排程則擋下，保護已付紀錄。"""
    if contract_has_linked_payment(contract_id):
        raise ValueError("這份合約已有核銷回填某期排程，不能整個重產；請改用加列/改列。")
    with connect() as conn:
        cur = conn.execute("DELETE FROM payment_schedules WHERE contract_id = ?", (contract_id,))
        return cur.rowcount


def generate_payment_schedules(contract_id: int, method: str, spec: Any = None,
                               remainder_on: str = "last", start_month: str = "",
                               frequency: str = "monthly", base_amount: float | None = None) -> list[dict[str, Any]]:
    """依付款方式自動產生預計付款排程（需求書 §8）。金額＝合約含稅總額（不拆稅、不換匯——那是會計/出納的事）。
      installment（固定期數）: spec = 期數 n → 平均分，零頭歸 remainder_on 指定的那期
      periodic  （週期月租）:  spec = 期數 n，frequency 每月/每季/每年 → 平均分＋日期依週期遞增
      milestone （里程碑）:    spec = [pct, ...]，合計須=100 → amount*pct/100，零頭同上
      fixed     （固定金額）:  單筆 = 合約金額
    remainder_on='first'|'last'：零頭放第一/最後期（合約寫哪期多就選哪期）。
    start_month='YYYY-MM'：有給就自動帶每期預計付款日（下月預計付款、跨年度預算歸屬都靠它）。
    base_amount：只分這個金額（給『把剩餘分N期』用；不給＝分合約總額）。產生後每期仍可手動改。"""
    with connect() as conn:
        c = conn.execute("SELECT amount, case_id FROM contracts WHERE id = ?", (contract_id,)).fetchone()
        if c is None:
            raise ValueError(f"合約 ID {contract_id} 不存在。")
        amount = float(base_amount if base_amount is not None else (c["amount"] or 0))
        case_id = c["case_id"]
        step = _FREQ_STEP.get(frequency, 1)
        base_seq = conn.execute(
            "SELECT COALESCE(MAX(seq),0) AS m FROM payment_schedules WHERE contract_id = ?",
            (contract_id,),
        ).fetchone()["m"]

        rows: list[dict[str, Any]] = []
        if method in ("installment", "periodic"):
            n = int(spec or 1)
            amounts = _split_even(amount, n, remainder_on)
            # 期別編號接續既有排程（base_seq），不從 1 重數——調價後補的期別若也叫「第1期」，
            # 清單與待辦上會出現兩個「第1期」，看的人分不出哪個是哪個。
            for i, a in enumerate(amounts):
                rows.append({"seq": base_seq + i + 1, "method": method, "label": f"第{base_seq + i + 1}期",
                             "planned_amount": a, "due_date": _add_months(start_month, i * step)})
        elif method == "milestone":
            pcts = list(spec or [])
            if round(sum(pcts), 4) != 100:
                raise ValueError(f"里程碑比例合計 {sum(pcts)}%，必須等於 100%。")
            amounts = [math.floor(amount * p / 100 * 100) / 100 for p in pcts]
            rem = round(amount - sum(amounts), 2)
            amounts[0 if remainder_on == "first" else -1] = round(
                amounts[0 if remainder_on == "first" else -1] + rem, 2)
            for i, (p, a) in enumerate(zip(pcts, amounts)):
                rows.append({"seq": base_seq + i + 1, "method": "milestone", "label": f"里程碑{base_seq + i + 1}",
                             "percent": p, "planned_amount": a, "due_date": _add_months(start_month, i)})
        else:  # fixed
            rows = [{"seq": base_seq + 1, "method": "fixed", "label": "全額",
                     "planned_amount": amount, "due_date": start_month}]

        out = [_insert_row(conn, "payment_schedules",
                           {**r, "contract_id": contract_id, "case_id": case_id, "status": "planned"})
               for r in rows]
    return out


def allowed_fields() -> dict[str, set[str]]:
    return {
        "cases": {"case_code", "title", "owner", "status", "amount", "risk_level", "note", "next_step", "due_date", "created_by", "fiscal_year", "seq", "source_file", "source_row",
                  "temp_seq", "review_note", "merged_into_case_id",
                  "group_name", "budget_type", "expense_kind", "budget_item", "source", "description",
                  "reopened_by", "reopened_at", "reopen_reason", "status_note"},
        "contracts": {"contract_code", "contract_name", "vendor_name", "amount", "status", "case_id", "purchase_id", "end_date",
                      "payment_method", "installments",
                      "start_date", "contract_type", "parent_contract_id", "relation_type",
                      "warranty_end_date", "maintenance_end_date",
                      "vendor_tax_id", "owner", "group_name", "locations", "external_code",
                      "progress_note", "end_reason", "project_id",
                      "note", "source_file", "source_row"},
        "payments": {"contract_id", "payment_month", "payment_amount", "invoice_status", "status",
                     "item", "settle_no", "ref_no", "period", "billing_period", "settled_by",
                     "vendor", "approval_level", "owner", "owner_email", "net_amount", "tax_amount",
                     "settle_seq", "payment_schedule_id"},
        "payment_schedules": {"contract_id", "case_id", "seq", "label", "method", "planned_amount",
                              "percent", "due_date", "status", "note"},
        "contract_adjustments": {"contract_id", "effective_date", "old_amount", "new_amount",
                                 "delta", "reason", "note", "created_by"},
        "project_item_extensions": {"item_id", "old_end_date", "new_end_date", "reason", "note", "created_by"},
        "documents": {"file_name", "document_type", "source_note", "status", "case_id", "contract_id"},
        "budgets": {"budget_code", "category", "unit_name", "fiscal_year", "amount", "status", "case_id", "note",
                    "remainder_unit_code", "alloc_method", "alloc_category_kind", "alloc_category",
                    "expense_detail", "fill_dept", "estimator"},
        "unit_headcounts": {"unit_code", "unit_name", "headcount", "source_file"},
        "category_shares": {"category", "unit_code", "unit_name", "share_pct", "source_file"},
        "projects": {"project_code", "project_name", "source", "necessity", "progress", "owner", "status", "case_id", "due_date", "note",
                     "level", "progress_planned", "rag_status", "start_date", "end_date",
                     "vendor_name", "cross_company", "involves_procurement"},
        "signoffs": {"signoff_code", "subject", "applicant", "amount", "status", "sign_date", "case_id", "note", "attachment_ref"},
        "purchases": {"purchase_code", "item_name", "vendor_name", "quantity", "amount", "status", "case_id", "signoff_id", "note"},
        "project_items": {"project_id", "seq", "item_name", "owner", "start_date", "end_date", "exec_status",
                          "sub_total", "sub_done", "progress", "rag", "risk_note", "decision_needed",
                          "support_needed", "duration_days", "status", "rag_manual"},
        "project_subitems": {"item_id", "seq", "name", "owner", "start_date", "end_date",
                             "done", "note", "status"},
        "budget_allocations": {"budget_id", "seq", "unit_code", "unit_name", "share_pct", "amount", "source_file"},
        # 費用模組三層（助理 0803 附件一）
        "expense_masters": {"contract_id", "case_id", "expense_name", "vendor_name", "vendor_tax_id",
                            "start_date", "end_date", "total_amount", "modes", "signoff_ref",
                            "signoff_none_reason", "owner", "note", "status"},
        "expense_sections": {"expense_id", "mode", "section_name", "section_amount", "price_method",
                             "periods", "frequency", "period_start", "period_end", "first_amount",
                             "first_month", "first_due_date", "note", "status", "version",
                             "commit_span_months", "next_amount_rule", "growth_pct", "carry_over",
                             "achievement_basis", "shortfall_action"},
        "expense_schedules": {"section_id", "seq", "milestone_name", "custom_name", "percent",
                              "planned_amount", "expense_month", "billing_start", "billing_end",
                              "due_date", "note", "manual_adjusted", "commit_period"},
        "expense_actuals": {"section_id", "schedule_id", "commit_period", "usage_amount",
                            "billing_start", "billing_end", "description", "adjust_amount",
                            "adjust_reason", "recognized_amount"},
        "expense_settlements": {"expense_id", "section_id", "schedule_id", "actual_id", "settle_month",
                                "billing_start", "billing_end", "vendor_name", "vendor_tax_id",
                                "invoice_date", "invoice_no", "claim_amount", "progress", "confirmed",
                                "settler", "signoff_no", "doc_ref", "diff_reason", "note"},
    }


def validate_status_fields(table: str, fields: dict[str, Any]) -> None:
    for field, valid_values in STATUS_VALUES.get(table, {}).items():
        if field not in fields or fields[field] is None:
            continue
        value = str(fields[field])
        if value not in valid_values:
            allowed = ", ".join(sorted(valid_values))
            raise ValueError(f"Invalid {table}.{field}: {value}. Allowed values: {allowed}.")


def get_row(conn: sqlite3.Connection, table: str, row_id: int) -> dict[str, Any]:
    row = conn.execute(f"SELECT * FROM {table} WHERE id = ?", (row_id,)).fetchone()
    if row is None:
        raise LookupError(f"{table} row {row_id} not found")
    return row


def fetch_one(table: str, row_id: int) -> dict[str, Any] | None:
    """單筆查詢（自開連線），找不到回 None。"""
    with connect() as conn:
        try:
            return dict(get_row(conn, table, row_id))
        except LookupError:
            return None


NULLABLE_FIELDS: dict[str, set[str]] = {
    "contracts": {"case_id", "purchase_id", "parent_contract_id"},
    "documents": {"case_id", "contract_id"},
    "budgets": {"case_id"},
    "projects": {"case_id"},
    "signoffs": {"case_id"},
    "purchases": {"case_id", "signoff_id"},
}


def update_row(table: str, row_id: int, payload: dict[str, Any]) -> dict[str, Any]:
    scope = _owner_scope.get()
    allowed = allowed_fields()
    nullable = NULLABLE_FIELDS.get(table, set())
    # 允許把可為空的外鍵顯式清成 NULL（解除關聯）；其餘欄位仍略過 None。
    fields = {
        key: value
        for key, value in payload.items()
        if key in allowed[table] and (value is not None or key in nullable)
    }
    if scope is not None and table == "cases":
        fields.pop("owner", None)  # 承辦不得竄改案件歸屬（避免竊佔/送人）
        # 限 cases：合約也有 owner（合約負責人），那是合約主檔自己的欄位，
        # 跟「這件案子歸誰」不是同一回事，不能一起擋掉，否則承辦連自己的合約負責人都填不了。
    if not fields:
        raise ValueError("No valid fields supplied.")
    validate_status_fields(table, fields)
    assignments = ", ".join(f"{key} = ?" for key in fields)
    with connect() as conn:
        before = get_row(conn, table, row_id)
        if scope is not None and not _row_in_scope(conn, table, row_id, scope):
            raise LookupError(f"{table} row {row_id} not found")  # 非本人範圍，視同不存在
        _validate_fks(conn, fields)
        if table == "contracts":
            if fields.get("parent_contract_id") is not None:
                _validate_contract_parent(conn, row_id, int(fields["parent_contract_id"]))
            _validate_contract(conn, fields, dict(before))   # 帶舊值：只改一欄也要判得出必填
        cursor = conn.execute(
            f"UPDATE {table} SET {assignments} WHERE id = ?",
            [*fields.values(), row_id],
        )
        if cursor.rowcount == 0:
            raise LookupError(f"{table} row {row_id} not found")
        after = get_row(conn, table, row_id)
        write_audit_log(conn, table, row_id, "update", before, after)
        return after


def disable_row(table: str, row_id: int) -> dict[str, Any]:
    scope = _owner_scope.get()
    if "status" not in allowed_fields()[table]:
        raise ValueError(f"{table} does not support disable.")
    validate_status_fields(table, {"status": "disabled"})
    with connect() as conn:
        before = get_row(conn, table, row_id)
        if scope is not None and not _row_in_scope(conn, table, row_id, scope):
            raise LookupError(f"{table} row {row_id} not found")  # 非本人範圍，視同不存在
        cursor = conn.execute(f"UPDATE {table} SET status = ? WHERE id = ?", ("disabled", row_id))
        if cursor.rowcount == 0:
            raise LookupError(f"{table} row {row_id} not found")
        after = get_row(conn, table, row_id)
        write_audit_log(conn, table, row_id, "disable", before, after)
        return after


def submit_case(case_id: int) -> dict[str, Any]:
    """送出複核：draft/reviewing/returned -> pending_review。承辦可送自己的（套 owner 範圍）。
    退回補件(returned)的案子補完可以直接再送，沿用原暫時號、不用重開一件（需求書 §4）。"""
    scope = _owner_scope.get()
    with connect() as conn:
        before = get_row(conn, "cases", case_id)
        if scope is not None and not _row_in_scope(conn, "cases", case_id, scope):
            raise LookupError(f"cases row {case_id} not found")  # 非本人範圍，視同不存在
        if before["status"] not in ("draft", "reviewing", "returned"):
            raise RuntimeError(f"案件目前狀態為 {before['status']}，無法送出複核。")
        conn.execute("UPDATE cases SET status = 'pending_review' WHERE id = ?", (case_id,))
        after = get_row(conn, "cases", case_id)
        write_audit_log(conn, "cases", case_id, "submit", before, after)
        return after


def cancel_case_review(case_id: int, actor: str, actor_role: str) -> dict[str, Any]:
    """取消複核：pending_review -> draft。原提交者本人，或主管/助理角色，都可以取消
    （跟核准不同，取消沒有「球員兼裁判」風險，不用排除提交者本人）。"""
    with connect() as conn:
        before = get_row(conn, "cases", case_id)
        if before["status"] != "pending_review":
            raise RuntimeError(f"案件目前狀態為 {before['status']}，只有『待複核』能取消複核。")
        is_submitter = (before.get("created_by") or "") == actor
        if not is_submitter and actor_role != "manager_assistant":
            raise PermissionError("只有原提交者或主管/助理能取消複核。")
        conn.execute("UPDATE cases SET status = 'draft' WHERE id = ?", (case_id,))
        after = get_row(conn, "cases", case_id)
        write_audit_log(conn, "cases", case_id, "cancel_review", before, after)
        return after


def approve_case(case_id: int, approver: str) -> dict[str, Any]:
    """核准：pending_review -> approved。雙人複核鐵則——建立者不得核准自己的案件。
    （角色限制「只有助理/主管可核」在 API 層擋；此處核准者看全部，不套 owner 範圍。）"""
    with connect() as conn:
        before = get_row(conn, "cases", case_id)
        if before["status"] != "pending_review":
            raise RuntimeError(f"案件目前狀態為 {before['status']}，只有『待複核』能核准。")
        if (before.get("created_by") or "") == approver:
            raise PermissionError("不能核准自己建立的案件，需由另一人複核。")
        # 正式流水號在這裡才配（使用者拍板 A 案）：申請階段用暫時號，核准通過才佔正式號，
        # 被駁回／被併走的申請不吃號，年度編號不跳號。已配過就不重配（重複核准不會換號）。
        seq = int(before.get("seq") or 0)
        if seq <= 0:
            seq = conn.execute(
                "SELECT COALESCE(MAX(seq), 0) + 1 AS n FROM cases WHERE fiscal_year = ?",
                (before.get("fiscal_year") or "",)).fetchone()["n"]
        conn.execute(
            "UPDATE cases SET status = 'approved', seq = ?, approved_by = ?, approved_at = datetime('now') "
            "WHERE id = ?",
            (seq, approver, case_id),
        )
        after = get_row(conn, "cases", case_id)
        write_audit_log(conn, "cases", case_id, "approve", before, after)
        return after


# ── 需求書 §4 審核關卡：核准以外的三條路（退回補件／併入既有案／拒絕建立）──
# 三者都不刪資料：退件要能補完再送、併案要留得住「這兩件本來是同一件」、拒絕也要留申請紀錄。
def _reviewable(conn: sqlite3.Connection, case_id: int, action_label: str) -> dict[str, Any]:
    before = get_row(conn, "cases", case_id)
    if before["status"] not in ("pending_review", "draft", "returned"):
        raise RuntimeError(f"案件目前狀態為 {before['status']}，無法{action_label}。")
    return before


def return_case(case_id: int, actor: str, reason: str) -> dict[str, Any]:
    """退回補件：pending_review -> returned，必須寫退件原因（否則申請人不知道要補什麼）。
    沿用原暫時號，申請人補完直接再送，不用重開一件。"""
    reason = str(reason or "").strip()
    if not reason:
        raise ValueError("請填退件原因，讓申請人知道要補什麼。")
    with connect() as conn:
        before = _reviewable(conn, case_id, "退回補件")
        conn.execute("UPDATE cases SET status = 'returned', review_note = ? WHERE id = ?", (reason, case_id))
        after = get_row(conn, "cases", case_id)
        write_audit_log(conn, "cases", case_id, "return", before, after, actor=actor)
        return after


def reject_case(case_id: int, actor: str, reason: str) -> dict[str, Any]:
    """拒絕建立：不產生正式 Case（不配正式號），但申請與審核紀錄留著——
    停用會讓它看起來像資料被藏起來，查不到「這件曾經被申請過、被誰以什麼理由駁回」。"""
    reason = str(reason or "").strip()
    if not reason:
        raise ValueError("請填駁回原因，這是審核紀錄的一部分。")
    with connect() as conn:
        before = _reviewable(conn, case_id, "拒絕建立")
        conn.execute("UPDATE cases SET status = 'rejected', review_note = ? WHERE id = ?", (reason, case_id))
        after = get_row(conn, "cases", case_id)
        write_audit_log(conn, "cases", case_id, "reject", before, after, actor=actor)
        return after


# ── 需求書 §4 核准之後的生命週期：已核准 → 進行中 → (暫停) → 已結案；隨時可取消。
# 已結案可重新開啟，但必須記錄重開人/時間/原因（需求書明列）。
# 以「動作」為主軸列出合法來源與目標，不要只列目標——start 與 resume 的目標都是「進行中」，
# 只看目標的話，剛核准的案會被允許「復工」（測試抓到過這個洞）。
CASE_ACTIONS: dict[str, dict[str, Any]] = {
    "start":  {"from": ("approved",), "to": "in_progress", "label": "開始執行"},
    "pause":  {"from": ("in_progress",), "to": "paused", "label": "暫停", "need_reason": True},
    "resume": {"from": ("paused",), "to": "in_progress", "label": "復工"},
    "close":  {"from": ("in_progress",), "to": "closed", "label": "結案"},
    # 取消：核准後、還沒結案前隨時可撤
    "cancel": {"from": ("approved", "in_progress", "paused"), "to": "cancelled",
               "label": "取消案件", "need_reason": True},
    # 重開：需求書 §4 明列必須記錄重開人、時間與原因
    "reopen": {"from": ("closed",), "to": "in_progress", "label": "重新開啟", "need_reason": True},
}
CASE_STATUS_LABEL = {
    "draft": "暫存", "pending_review": "待審核", "reviewing": "審核中", "returned": "退回補件",
    "approved": "已核准", "in_progress": "進行中", "paused": "暫停", "closed": "已結案",
    "cancelled": "已取消", "rejected": "已駁回", "merged": "已併入他案", "disabled": "已停用",
}


def change_case_status(case_id: int, action: str, actor: str, reason: str = "") -> dict[str, Any]:
    """推進案件狀態（核准之後的生命週期）。只允許 CASE_ACTIONS 列出的動作與來源狀態。

    暫停與取消要寫原因（否則事後沒人知道為什麼停在那裡）；
    重開（已結案 → 進行中）依需求書 §4 必填原因，並記下重開人與時間。
    """
    spec = CASE_ACTIONS.get(action)
    if spec is None:
        raise ValueError(f"不支援的狀態動作：{action}")
    target = spec["to"]
    reason = str(reason or "").strip()
    scope = _owner_scope.get()
    with connect() as conn:
        before = get_row(conn, "cases", case_id)
        if scope is not None and not _row_in_scope(conn, "cases", case_id, scope):
            raise LookupError(f"cases row {case_id} not found")  # 不在可視範圍，視同不存在
        current = before["status"]
        if current not in spec["from"]:
            raise RuntimeError(
                f"案件目前是「{CASE_STATUS_LABEL.get(current, current)}」，不能{spec['label']}。")
        if spec.get("need_reason") and not reason:
            raise ValueError(
                "重新開啟已結案的案件必須填原因（需求書 §4）。" if action == "reopen"
                else f"請填{spec['label']}原因。")
        reopening = action == "reopen"

        sets = {"status": target}
        if reopening:
            sets.update({"reopened_by": actor, "reopened_at": "", "reopen_reason": reason})
        if target in ("paused", "cancelled"):
            sets["status_note"] = reason
        assignments = ", ".join(f"{k} = ?" for k in sets)
        extra = ", reopened_at = datetime('now')" if reopening else ""
        conn.execute(f"UPDATE cases SET {assignments}{extra} WHERE id = ?", [*sets.values(), case_id])
        after = get_row(conn, "cases", case_id)
        action = "reopen" if reopening else f"status_{target}"
        write_audit_log(conn, "cases", case_id, action, before, after, actor=actor)
        return after


# 併案要一起搬走的關聯資料：這些都是掛 case_id 的。付款掛合約、合約搬過去付款自然跟著走。
_MERGE_TABLES = ("budgets", "projects", "signoffs", "purchases", "contracts", "documents")


def merge_case_into(case_id: int, target_case_id: int, actor: str, reason: str = "") -> dict[str, Any]:
    """併入既有 Case：把這件申請底下的資料轉到目標案，來源標 merged 並記「併到哪一件」。
    重點是關聯要留得住——只把新申請停用、人工把資料補到舊案，事後沒人看得出這兩件是同一件事。"""
    if int(case_id) == int(target_case_id):
        raise ValueError("不能併入自己。")
    with connect() as conn:
        before = _reviewable(conn, case_id, "併入既有案")
        target = conn.execute("SELECT id, status FROM cases WHERE id = ?", (target_case_id,)).fetchone()
        if target is None:
            raise ValueError(f"要併入的案件 ID {target_case_id} 不存在。")
        if target["status"] in ("merged", "rejected"):
            raise ValueError("目標案件本身已被併入或已駁回，不能當併入對象。")
        moved: dict[str, int] = {}
        for table in _MERGE_TABLES:
            cur = conn.execute(f"UPDATE {table} SET case_id = ? WHERE case_id = ?", (target_case_id, case_id))
            if cur.rowcount:
                moved[table] = cur.rowcount
        conn.execute(
            "UPDATE cases SET status = 'merged', merged_into_case_id = ?, review_note = ? WHERE id = ?",
            (target_case_id, str(reason or "").strip(), case_id))
        after = get_row(conn, "cases", case_id)
        write_audit_log(conn, "cases", case_id, "merge", before, after, actor=actor)
        return {**after, "moved": moved}


CHILD_REFS: dict[str, list[tuple[str, str]]] = {
    "cases": [("contracts", "case_id"), ("documents", "case_id")],
    "contracts": [("payments", "contract_id"), ("documents", "contract_id")],
}


def delete_row(table: str, row_id: int) -> None:
    scope = _owner_scope.get()
    with connect() as conn:
        before = get_row(conn, table, row_id)
        if scope is not None and not _row_in_scope(conn, table, row_id, scope):
            raise LookupError(f"{table} row {row_id} not found")  # 非本人範圍，視同不存在
        # 有子列關聯時不得硬刪（避免靜默孤立子列、金額短少）；請先處理或改用作廢。
        for child_table, fk in CHILD_REFS.get(table, []):
            count = conn.execute(
                f"SELECT COUNT(*) AS c FROM {child_table} WHERE {fk} = ?", (row_id,)
            ).fetchone()["c"]
            if count:
                raise RuntimeError(
                    f"仍有 {count} 筆 {child_table} 關聯此 {table}（id={row_id}），"
                    "請先處理關聯資料或改用作廢。"
                )
        cursor = conn.execute(f"DELETE FROM {table} WHERE id = ?", (row_id,))
        if cursor.rowcount == 0:
            raise LookupError(f"{table} row {row_id} not found")
        write_audit_log(conn, table, row_id, "delete", before, None)


def list_rows(table: str, limit: int = 100) -> list[dict[str, Any]]:
    scope = _owner_scope.get()
    where, params = _scope_where(table, scope) if scope is not None else ("", [])
    sql = f"SELECT * FROM {table}"
    if where:
        sql += f" WHERE {where}"
    sql += " ORDER BY id DESC LIMIT ?"
    with connect() as conn:
        rows = conn.execute(sql, [*params, max(1, min(limit, 500))]).fetchall()
    if table == "contracts":
        # 到期警示即時算（助理 0803）：燈號每天都在變，存進資料庫就得靠排程刷，
        # 漏跑一天畫面就是錯的。讀出來時算，永遠是今天的答案。
        out = []
        for r in rows:
            d = dict(r)
            d["expiry_light"] = contract_expiry_light(d.get("end_date"), d.get("end_reason"))
            # 到期追蹤還不能算完成：快到期／已到期卻沒寫處理到哪（助理 0803）
            d["needs_progress_note"] = contract_needs_progress_note(d)
            out.append(d)
        rows = out
    return rows


def list_projects(limit: int = 100) -> list[dict[str, Any]]:
    """專案清單：沿用 list_rows（保留承辦 owner scope），再用一次 grouped COUNT 補上
    每個專案的工作項數（item_count＝子項目總數、item_done＝完成度 100% 的數），
    讓清單一眼看得出每案有幾個子項目、完成幾個，不必逐案點進去。"""
    rows = list_rows("projects", limit)
    if not rows:
        return rows
    ids = [r["id"] for r in rows]
    placeholders = ",".join("?" * len(ids))
    with connect() as conn:
        counts = conn.execute(
            f"SELECT project_id, COUNT(*) AS total, "
            f"SUM(CASE WHEN progress >= 100 THEN 1 ELSE 0 END) AS done "
            f"FROM project_items WHERE project_id IN ({placeholders}) AND status != 'disabled' "
            f"GROUP BY project_id",
            ids,
        ).fetchall()
    cmap = {c["project_id"]: c for c in counts}
    for r in rows:
        c = cmap.get(r["id"])
        r["item_count"] = int(c["total"]) if c else 0
        r["item_done"] = int(c["done"] or 0) if c else 0
    return rows


def create_import_batch(source_name: str, status: str = "created") -> dict[str, Any]:
    source = source_name.strip()
    if not source:
        raise ValueError("source_name is required.")
    with connect() as conn:
        cursor = conn.execute(
            "INSERT INTO import_batches (source_name, status) VALUES (?, ?)",
            (source, status),
        )
        batch_id = cursor.lastrowid
        batch = get_row(conn, "import_batches", batch_id)
        write_audit_log(conn, "import_batches", batch_id, "create", None, batch)
        return batch


def stage_import_rows(batch_id: int, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not rows:
        raise ValueError("rows must contain at least one row.")
    with connect() as conn:
        batch = get_row(conn, "import_batches", batch_id)
        staged: list[dict[str, Any]] = []
        for index, raw in enumerate(rows, start=1):
            cursor = conn.execute(
                "INSERT INTO import_rows (batch_id, row_number, raw_json, status, error_message) "
                "VALUES (?, ?, ?, ?, ?)",
                (
                    batch_id,
                    index,
                    json.dumps(raw, ensure_ascii=False, sort_keys=True),
                    "staged",
                    None,
                ),
            )
            staged.append(get_row(conn, "import_rows", cursor.lastrowid))
        after = {"batch": batch, "staged_row_count": len(staged)}
        write_audit_log(conn, "import_batches", batch_id, "stage_rows", None, after)
        return staged


def list_import_batches(limit: int = 100) -> list[dict[str, Any]]:
    with connect() as conn:
        return conn.execute(
            "SELECT * FROM import_batches ORDER BY id DESC LIMIT ?",
            (max(1, min(limit, 500)),),
        ).fetchall()


def get_import_batch(batch_id: int) -> dict[str, Any]:
    with connect() as conn:
        batch = get_row(conn, "import_batches", batch_id)
        rows = conn.execute(
            "SELECT * FROM import_rows WHERE batch_id = ? ORDER BY row_number ASC, id ASC",
            (batch_id,),
        ).fetchall()
        return {"batch": batch, "rows": rows}


def list_import_rows(batch_id: int, limit: int = 500) -> list[dict[str, Any]]:
    with connect() as conn:
        get_row(conn, "import_batches", batch_id)
        return conn.execute(
            "SELECT * FROM import_rows WHERE batch_id = ? ORDER BY row_number ASC, id ASC LIMIT ?",
            (batch_id, max(1, min(limit, 500))),
        ).fetchall()


def preview_import_mapping(batch_id: int) -> dict[str, Any]:
    with connect() as conn:
        batch = get_row(conn, "import_batches", batch_id)
        rows = conn.execute(
            "SELECT * FROM import_rows WHERE batch_id = ? ORDER BY row_number ASC, id ASC",
            (batch_id,),
        ).fetchall()
        return mapping_preview(batch, rows)


def confirm_import_batch_cases_dry_run(
    batch_id: int,
    confirmed_fields: list[dict[str, Any]],
) -> dict[str, Any]:
    with connect() as conn:
        batch = get_row(conn, "import_batches", batch_id)
        rows = conn.execute(
            "SELECT * FROM import_rows WHERE batch_id = ? ORDER BY row_number ASC, id ASC",
            (batch_id,),
        ).fetchall()
        preview = mapping_preview(batch, rows)
        existing_case_codes = {
            str(row["case_code"]).strip()
            for row in conn.execute("SELECT case_code FROM cases").fetchall()
        }
        return confirm_cases_dry_run_plan(preview, confirmed_fields, existing_case_codes)


def confirm_import_batch_cases_write(
    batch_id: int,
    confirmed_fields: list[dict[str, Any]],
) -> dict[str, Any]:
    """正式寫入案件。安全閘門：
    - 先跑與 dry-run 相同的驗證（零錯誤、確認齊、批內無重複）→ 任一不過即 raise，完全不寫。
    - 單一交易：全部成功才 commit，中途出錯整批回滾。
    - 冪等：案件編號已存在則跳過（不覆蓋），可安全重跑。
    - 來源舉證：逐列寫稽核，記 batch_id / row_number / source_row_id / actor。
    """
    actor = _current_actor.get()
    with connect() as conn:
        batch = get_row(conn, "import_batches", batch_id)
        rows = conn.execute(
            "SELECT * FROM import_rows WHERE batch_id = ? ORDER BY row_number ASC, id ASC",
            (batch_id,),
        ).fetchall()
        preview = mapping_preview(batch, rows)
        # 用空的 existing 驗證：批內錯誤/確認/重複要擋，但既有編號改為逐列冪等跳過而非整批拒絕。
        plan = confirm_cases_dry_run_plan(preview, confirmed_fields, set())
        existing = {
            str(r["case_code"]).strip() for r in conn.execute("SELECT case_code FROM cases").fetchall()
        }
        created: list[str] = []
        skipped: list[str] = []
        for item in plan["plan"]["cases"]:
            record = item["record"]
            code = str(record.get("case_code", "")).strip()
            if not code or code in existing:
                skipped.append(code)
                continue
            fields = {k: record[k] for k in ("case_code", "title", "owner", "amount") if k in record}
            fields["created_by"] = actor
            # Excel 來源勾稽：把來源檔名＋原始列號寫在案件上，供清單 📎 指回 Excel
            fields["source_file"] = str(batch.get("source_name") or "").strip()
            fields["source_row"] = int(item.get("row_number") or 0)
            # 使用者拍板 A2：匯入的是已經在跑的案子，直接算已成立並配當年度正式號
            # （不落草稿／不發 TMP-，省掉匯入完還要按一次補號）。
            fields.update(_established_case_fields(conn, record.get("fiscal_year")))
            columns = ", ".join(fields)
            placeholders = ", ".join("?" for _ in fields)
            cursor = conn.execute(
                f"INSERT INTO cases ({columns}) VALUES ({placeholders})", list(fields.values())
            )
            row_id = cursor.lastrowid
            after = get_row(conn, "cases", row_id)
            write_audit_log(conn, "cases", row_id, "import", None, {
                **after,
                "import_batch_id": batch_id,
                "import_row_number": item["row_number"],
                "import_source_row_id": item["source_row_id"],
            })
            existing.add(code)
            created.append(code)
        conn.execute("UPDATE import_batches SET status = 'committed' WHERE id = ?", (batch_id,))
        return {
            "dry_run": False,
            "committed": True,
            "batch_id": batch_id,
            "created": created,
            "skipped": skipped,
            "created_count": len(created),
            "skipped_count": len(skipped),
        }


def _norm_date(v: Any) -> str:
    """Excel 日期（datetime 或字串）→ 'YYYY-MM-DD'；空值回空字串。"""
    if v is None or v == "":
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip().replace("/", "-")
    return s[:10]


def _clean_owner(v: Any) -> str:
    """負責人欄防呆：若被填成長句／備註（去識別化或誤填一整段），不當人名，回空字串。"""
    s = " ".join(str(v).split()) if v is not None else ""
    if len(s) > 16 or any(p in s for p in "。？！?!，,；;"):
        return ""
    return s


def _xls_pct(v: Any) -> float:
    """比例欄：<=1 視為小數（0.294→29.4），>1 視為已是百分比。"""
    if v is None or v == "":
        return 0.0
    try:
        f = float(v)
    except (TypeError, ValueError):
        return 0.0
    return round(f * 100, 1) if f <= 1 else round(f, 1)


# ── 合約盤點表匯入（黃助理 2026-08-13 給的「合約盤點_主機組.xlsx」）───────────
# 表頭不在第一列（前面有填寫說明），所以用欄名定位，不寫死列號——各組交回來的檔案
# 說明列數可能不一樣，寫死列號第二個組就爆。
_CONTRACT_COL_MAP = {
    "詩芸備註": "inventory_note",
    "已確認完成": "confirmed",
    "合約編號": "external_code",       # EF-20190726-046：公司的號，有連字號，照原樣留
    "合約名稱": "contract_name",
    "合約系統之內容說明": "content_note",
    "合約狀態": "inventory_status",
    "合約狀態詳細說明": "progress_note",
    "組別": "group_name",
    "合約維護人": "owner",
    "廠商名稱": "vendor_name",
    "廠商統編或ID": "vendor_tax_id",
    "合約開始日": "start_date",
    "合約到期日": "end_date",
}
# 組別正規化：H 欄有「AS400」，那是系統名不是組別（使用者 2026-08-13 裁決：歸主機組）
_GROUP_ALIAS = {"AS400": "主機組"}
# A、G 欄的文字裡藏著關聯合約編號（「已由新合約取代EF-20240416-005」）
# 主機組是 EF-20240416-005；其他組的前綴與位數未必一樣，稍微放寬但仍要有「英文-數字-數字」
# 的骨架，才不會把說明文字裡的日期或金額當成合約編號抓進來。
_CONTRACT_CODE_RE = re.compile(r"[A-Za-z]{2,4}-\d{4,8}-\d{2,4}")
_PREV_OWNER_RE = re.compile(r"原合約維護人[：:]\s*([^\s，,、;；]+)")


def parse_contract_inventory_xlsx(data: bytes) -> dict[str, Any]:
    """解析合約盤點表 → 合約清單＋盤點發現的事（關聯、原維護人）。

    回傳的每一筆都帶 source_sheet 與 source_row，匯錯了查得回原始那一列。
    """
    import io
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out: list[dict[str, Any]] = []
    skipped_sheets: list[str] = []
    try:
        for sheet in wb.sheetnames:
            rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
            header_idx = next(
                (i for i, r in enumerate(rows[:20])
                 if r and any(str(c).strip() == "合約編號" for c in r if c is not None)), None)
            if header_idx is None:
                skipped_sheets.append(sheet)      # 說明頁之類的，沒有合約編號欄
                continue
            header = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
            idx = {_CONTRACT_COL_MAP[h]: i for i, h in enumerate(header) if h in _CONTRACT_COL_MAP}
            for rno, r in enumerate(rows[header_idx + 1:], start=header_idx + 2):
                def val(key: str) -> str:
                    i = idx.get(key)
                    if i is None or i >= len(r) or r[i] is None:
                        return ""
                    return " ".join(str(r[i]).split())
                name = val("contract_name")
                code = val("external_code")
                if not name and not code:
                    continue                       # 整列空白
                item: dict[str, Any] = {
                    "source_sheet": sheet, "source_row": rno,
                    "external_code": code,
                    "contract_name": name,
                    "vendor_name": val("vendor_name"),
                    "vendor_tax_id": val("vendor_tax_id"),
                    "owner": val("owner"),
                    "group_name": _GROUP_ALIAS.get(val("group_name"), val("group_name")),
                    "start_date": _norm_date(_raw(r, idx.get("start_date"))),
                    "end_date": _norm_date(_raw(r, idx.get("end_date"))),
                    "progress_note": val("progress_note"),
                    "inventory_status": val("inventory_status"),
                    "inventory_note": val("inventory_note"),
                    "content_note": val("content_note"),
                    "confirmed": str(val("confirmed")).lower() in ("true", "v", "y", "yes", "是"),
                }
                # 盤點表寫的「原合約維護人：張根榮」＝離職或職務異動，接得上交接功能
                prev = _PREV_OWNER_RE.search(item["inventory_note"])
                item["previous_owner"] = prev.group(1) if prev else ""
                # A、G 欄提到的其他合約編號＝續約/取代/整併關係的線索
                related = set(_CONTRACT_CODE_RE.findall(item["inventory_note"]))
                related |= set(_CONTRACT_CODE_RE.findall(item["progress_note"]))
                related.discard(code)
                item["related_codes"] = sorted(related)
                item["relation_hint"] = _relation_hint(item["progress_note"])
                out.append(item)
    finally:
        wb.close()
    return {"contracts": out, "count": len(out), "skipped_sheets": skipped_sheets}


def _raw(row: list, i: int | None) -> Any:
    return row[i] if (i is not None and i < len(row)) else None


def commit_contract_inventory(rows: list[dict[str, Any]]) -> dict[str, Any]:
    """把盤點表寫進合約主檔。以「公司內部合約編號」為識別鍵：同編號更新、沒見過的新增。

    冪等——助理 8/19 盤點完會再給一次，同一份檔重匯不該長出第二套。
    分兩輪：先把合約都建好，再回頭接關聯（被提到的那份合約可能排在後面，先接會找不到）。
    盤點表的原始欄位（合約狀態、詩芸備註、內容說明）一律留在備註裡，不丟資訊——
    「查無後續合約」是盤點當下的標記，不是合約本身的狀態，硬套成系統狀態會失真。
    """
    actor = _current_actor.get()
    created, updated, skipped = [], [], []
    with connect() as conn:
        # 識別鍵優先用公司的合約編號；盤點表裡有 1 筆沒填編號，那種退回用合約名稱比對——
        # 不然重匯時它每次都被當成新的，第二次就撞 contract_code 的唯一鍵直接 500。
        existing: dict[str, dict[str, Any]] = {}
        for r in conn.execute("SELECT * FROM contracts").fetchall():
            row = dict(r)
            key = str(row.get("external_code") or "").strip() or str(row.get("contract_name") or "").strip()
            if key:
                existing.setdefault(key, row)
        for item in rows:
            code = str(item.get("external_code") or "").strip()
            name = str(item.get("contract_name") or "").strip()
            if not name:
                skipped.append({"row": item.get("source_row"), "reason": "沒有合約名稱"})
                continue
            key = code or name
            note_bits = [b for b in (
                f"盤點狀態：{item['inventory_status']}" if item.get("inventory_status") else "",
                f"內容：{item['content_note']}" if item.get("content_note") else "",
                f"盤點備註：{item['inventory_note']}" if item.get("inventory_note") else "",
            ) if b]
            fields = {
                "contract_name": name,
                "external_code": code,
                "vendor_name": item.get("vendor_name", ""),
                "vendor_tax_id": item.get("vendor_tax_id", ""),
                "owner": item.get("owner", ""),
                "group_name": item.get("group_name", ""),
                "start_date": item.get("start_date", ""),
                "end_date": item.get("end_date", ""),
                "progress_note": item.get("progress_note", ""),
                "note": "；".join(note_bits),
                "source_file": "合約盤點表",
                "source_row": int(item.get("source_row") or 0),
            }
            if key in existing:
                row_id = existing[key]["id"]
                before = existing[key]
                sets = ", ".join(f"{k} = ?" for k in fields)
                conn.execute(f"UPDATE contracts SET {sets} WHERE id = ?", [*fields.values(), row_id])
                after = get_row(conn, "contracts", row_id)
                write_audit_log(conn, "contracts", row_id, "import-update", before, after)
                existing[key] = dict(after)
                updated.append(key)
                continue
            # 盤點表沒有本系統的合約編號，用公司的號當 contract_code；沒有編號就用名稱
            fields["contract_code"] = code or name
            fields.update(_next_contract_system_code(conn, fields))
            columns = ", ".join(fields)
            cur = conn.execute(
                f"INSERT INTO contracts ({columns}) VALUES ({', '.join('?' * len(fields))})",
                list(fields.values()))
            after = get_row(conn, "contracts", cur.lastrowid)
            write_audit_log(conn, "contracts", cur.lastrowid, "import",
                            None, {**dict(after), "import_source": "合約盤點表", "actor": actor})
            existing[key] = dict(after)
            created.append(key)

        # 第二輪：接續約／取代／整併關係（前面建好了才找得到被提到的那一份）
        linked = 0
        for item in rows:
            code = str(item.get("external_code") or "").strip()
            hint = item.get("relation_hint") or ""
            targets = [c for c in (item.get("related_codes") or []) if c in existing]
            if not code or code not in existing or not targets or not hint:
                continue
            me = existing[code]
            if me.get("parent_contract_id"):
                continue                          # 已經接過就不覆蓋（人工改過的優先）
            parent = existing[targets[0]]
            if int(parent["id"]) == int(me["id"]):
                continue
            # 防循環：兩份合約的說明互相提到對方時（真實資料就有），照接會繞成一個圈，
            # 續約鏈往上追會追不完、畫面直接卡死。用既有的循環檢查擋掉，擋到就跳過不接。
            try:
                _validate_contract_parent(conn, int(me["id"]), int(parent["id"]))
            except ValueError:
                skipped.append({"row": item.get("source_row"),
                                "reason": f"{code} 與 {targets[0]} 的說明互相指向對方，"
                                          f"接起來會繞成循環，關聯留白請人工判斷"})
                continue
            conn.execute(
                "UPDATE contracts SET parent_contract_id = ?, relation_type = ? WHERE id = ?",
                (parent["id"], hint, me["id"]))
            write_audit_log(conn, "contracts", me["id"], "link-relation", me, {
                "parent_contract_id": parent["id"], "relation_type": hint,
                "from_note": item.get("progress_note") or item.get("inventory_note")})
            linked += 1

    # 盤點表寫「原合約維護人：某某」＝那個人離職或換職務，接得上離職交接功能
    handover = [{"code": i.get("external_code"), "contract_name": i.get("contract_name"),
                 "previous_owner": i["previous_owner"], "current_owner": i.get("owner")}
                for i in rows if i.get("previous_owner")]
    return {"created": created, "updated": updated, "skipped": skipped,
            "created_count": len(created), "updated_count": len(updated),
            "skipped_count": len(skipped), "linked_count": linked,
            "handover_hints": handover}


def _relation_hint(note: str) -> str:
    """從說明文字判斷它跟被提到那份合約是什麼關係。判斷不出來就留空，不瞎猜。

    「一起追蹤」「參考」這種只是提到對方，不是從屬關係——真實資料裡有
    「115年配合集團會啟動續約 (與EF-20240222-002一起追蹤)」，看到「續約」就接
    會把兩份平行的合約硬掛成父子，而且兩邊互相指就繞成一個圈。
    """
    s = str(note or "")
    if any(k in s for k in ("一起追蹤", "一併追蹤", "併同追蹤", "參考", "相關合約")):
        return ""
    if "整併" in s:
        return "merge"
    if "取代" in s or "續約" in s or "新合約" in s:
        return "renew"
    if "增購" in s or "增補" in s:
        return "addon"
    return ""


def parse_projects_xlsx(data: bytes) -> list[dict[str, Any]]:
    """解析『處級專案進度追蹤總表』.xlsx → 專案清單（依欄名對應，不靠欄位位置）。

    版面：每張工作表＝一個組別；表頭含「專案名稱」。專案為多列一組——
    第一列帶專案層級欄（名稱/必要性/總進度預計%/實際%/總進度燈號，AI 表多一欄「分類」）；
    其後每列是「工作主項目」，各自帶開始日期/結束日期。專案起訖＝各工作項的 min(開始)→max(結束)。
    因 AI 表欄位右移一格，一律用欄名比對，避免位置錯位（先前燈號抓成小數即此故）。"""
    import io
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out: list[dict[str, Any]] = []
    try:
        for sheet_idx, sheet in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            header_idx = None
            for i, r in enumerate(rows[:8]):
                if r and any(str(c).replace("\n", "") == "專案名稱" for c in r if c is not None):
                    header_idx = i
                    break
            if header_idx is None:
                continue
            headers = [str(c).replace("\n", "").strip() if c is not None else "" for c in rows[header_idx]]

            def find(pred, lo=0, hi=None):
                hi = len(headers) if hi is None else hi
                for j in range(lo, hi):
                    if headers[j] and pred(headers[j]):
                        return j
                return None

            work_i = find(lambda h: h == "工作主項目")
            split = work_i if work_i is not None else len(headers)
            name_i = find(lambda h: h == "專案名稱", 0, split)
            if name_i is None:
                continue
            nec_i = find(lambda h: "必要性" in h, 0, split)
            plan_i = find(lambda h: "預計" in h, 0, split)
            act_i = find(lambda h: "實際" in h, 0, split)
            rag_i = find(lambda h: "燈號" in h, 0, split)          # 總進度燈號（專案層級那個）
            lvl_i = find(lambda h: h == "分類", 0, split)          # 只有 AI 表有
            start_i = find(lambda h: "開始" in h, split)
            end_i = find(lambda h: "結束" in h, split)
            owner_i = find(lambda h: "負責人" in h, split)
            # 工作項層級欄位（split＝工作主項目那欄之後）
            item_i = work_i  # 工作主項目
            exec_i = find(lambda h: "執行進度" in h, split)
            subtot_i = find(lambda h: "子項目總數" in h, split)
            subdone_i = find(lambda h: "子項目完成數" in h, split)
            prog_i = find(lambda h: "完成度" in h, split)
            wrag_i = find(lambda h: "燈號" in h, split)            # 工作項燈號（split 後第一個）
            risk_i = find(lambda h: ("風險" in h or "備註" in h), split)
            dec_i = find(lambda h: "需決策" in h, split)
            sup_i = find(lambda h: "需支援" in h, split)
            dur_i = find(lambda h: "持續天數" in h, split)

            def cell(r, i):
                return r[i] if (i is not None and i < len(r)) else None

            def txt(r, i):
                v = cell(r, i)
                return str(v).strip() if v is not None else ""

            def as_int(r, i):
                try:
                    return int(float(cell(r, i)))
                except (TypeError, ValueError):
                    return 0

            def build_item(r, seq_no):
                name = " ".join(txt(r, item_i).split())
                if not name:
                    return None
                return {
                    "seq": seq_no,
                    "item_name": name,
                    "owner": _clean_owner(cell(r, owner_i)),
                    "start_date": _norm_date(cell(r, start_i)),
                    "end_date": _norm_date(cell(r, end_i)),
                    "exec_status": txt(r, exec_i),
                    "sub_total": as_int(r, subtot_i),
                    "sub_done": as_int(r, subdone_i),
                    "progress": _xls_pct(cell(r, prog_i)),
                    "rag": txt(r, wrag_i),
                    "risk_note": txt(r, risk_i),
                    "decision_needed": txt(r, dec_i),
                    "support_needed": txt(r, sup_i),
                    "duration_days": txt(r, dur_i),
                }

            seq = 0
            cur: dict[str, Any] | None = None
            for r in rows[header_idx + 1:]:
                if not r:
                    continue
                nm = " ".join(txt(r, name_i).split())  # 收斂內部換行/多空白，避免分頁標籤爆版
                if nm:  # 新專案起始列
                    if cur is not None:
                        out.append(cur)
                    seq += 1
                    cur = {
                        # 編號一律純英數（主管交代：不用連字號、底線、中文）。原本是「工作表名-流水」，
                        # 工作表名是中文，產出來的代號同時踩到中文與連字號兩條，改成 PRJ＋年＋表序＋流水。
                        "project_code": f"PRJ{get_working_year()}{sheet_idx:02d}{seq:03d}",
                        "project_name": nm,
                        "source": sheet,
                        "necessity": txt(r, nec_i),
                        "progress_planned": _xls_pct(cell(r, plan_i)),
                        "progress": _xls_pct(cell(r, act_i)),
                        "rag_status": txt(r, rag_i),
                        "level": txt(r, lvl_i),
                        "owner": _clean_owner(cell(r, owner_i)),
                        "start_date": _norm_date(cell(r, start_i)),
                        "end_date": _norm_date(cell(r, end_i)),
                        "items": [],
                    }
                    first = build_item(r, 1)
                    if first:
                        cur["items"].append(first)
                elif cur is not None:  # 工作項續列：擴張起訖、補負責人、加一筆工作項
                    sd, ed = _norm_date(cell(r, start_i)), _norm_date(cell(r, end_i))
                    if sd and (not cur["start_date"] or sd < cur["start_date"]):
                        cur["start_date"] = sd
                    if ed and (not cur["end_date"] or ed > cur["end_date"]):
                        cur["end_date"] = ed
                    if not cur["owner"]:
                        cur["owner"] = _clean_owner(cell(r, owner_i))
                    item = build_item(r, len(cur["items"]) + 1)
                    if item:
                        cur["items"].append(item)
            if cur is not None:
                out.append(cur)
    finally:
        wb.close()
    return out


def commit_projects_import(records: list[dict[str, Any]]) -> dict[str, Any]:
    """寫入 projects：單一交易、逐列稽核。以（組別＋專案名稱）為識別鍵——
    同名專案改『更新』（讓更新版總表的起訖日/進度灌進既有資料），沒見過的才『新增』。"""
    _identity = {"project_code", "source", "project_name"}
    fields_allowed = allowed_fields()["projects"]
    with connect() as conn:
        existing: dict[tuple[str, str], int] = {}
        for row in conn.execute("SELECT id, source, project_name FROM projects").fetchall():
            existing[(row["source"] or "", row["project_name"] or "")] = row["id"]
        item_fields = allowed_fields()["project_items"]
        created: list[str] = []
        updated: list[str] = []
        items_written = 0
        rollup_ids: list[int] = []
        for rec in records:
            name = str(rec.get("project_name", "")).strip()
            if not name:
                continue
            key = (rec.get("source", "") or "", name)
            fields = {k: v for k, v in rec.items() if k in fields_allowed}
            if key in existing:
                rid = existing[key]
                before = get_row(conn, "projects", rid)
                upd = {k: v for k, v in fields.items() if k not in _identity}
                if upd:
                    sets = ", ".join(f"{k} = ?" for k in upd)
                    conn.execute(f"UPDATE projects SET {sets} WHERE id = ?", [*upd.values(), rid])
                after = get_row(conn, "projects", rid)
                write_audit_log(conn, "projects", rid, "import-update", before, after)
                updated.append(name)
            else:
                if not fields.get("case_id"):
                    cid = _ensure_case_for(
                        conn, fields.get("project_name"), fields.get("project_code"), fields.get("fiscal_year"),
                        fields.get("owner"), established=True,
                    )
                    if cid:
                        fields["case_id"] = cid
                columns = ", ".join(fields)
                placeholders = ", ".join("?" for _ in fields)
                cur = conn.execute(f"INSERT INTO projects ({columns}) VALUES ({placeholders})", list(fields.values()))
                after = get_row(conn, "projects", cur.lastrowid)
                write_audit_log(conn, "projects", cur.lastrowid, "import", None, {**after, "import_source": "xlsx"})
                rid = cur.lastrowid
                existing[key] = rid
                created.append(name)
            # 工作項：以（project_id＋item_name）為鍵，同名更新、沒見過新增（不刪系統內新增的）
            seen = {r["item_name"]: r["id"] for r in conn.execute(
                "SELECT id, item_name FROM project_items WHERE project_id = ?", (rid,)).fetchall()}
            for it in rec.get("items", []):
                ifields = {k: v for k, v in it.items() if k in item_fields}
                ifields["project_id"] = rid
                # 匯入也走同一套自動計算：進度由子項目數算、燈號正規化（Excel 是中文）後
                # 沒指定就自動判。匯入不強制「紅/黃要填風險點」——舊總表沒填的很多，
                # 擋下來會讓整批匯入失敗；那是人在系統裡編輯時才強制。
                ifields["progress"] = wbs_item_progress(
                    ifields.get("sub_total"), ifields.get("sub_done"), ifields.get("exec_status"))
                rag = normalize_wbs_rag(ifields.get("rag"))
                ifields["rag_manual"] = 1 if rag else 0   # Excel 有填燈號＝視為人工指定
                ifields["rag"] = rag or wbs_auto_rag(
                    ifields["progress"], ifields.get("start_date"), ifields.get("end_date"))
                iname = ifields.get("item_name", "")
                if iname in seen:
                    upd = {k: v for k, v in ifields.items() if k not in ("project_id", "item_name")}
                    if upd:
                        sets = ", ".join(f"{k} = ?" for k in upd)
                        conn.execute(f"UPDATE project_items SET {sets} WHERE id = ?", [*upd.values(), seen[iname]])
                else:
                    cols = ", ".join(ifields)
                    ph = ", ".join("?" for _ in ifields)
                    c2 = conn.execute(f"INSERT INTO project_items ({cols}) VALUES ({ph})", list(ifields.values()))
                    seen[iname] = c2.lastrowid
                items_written += 1
            rollup_ids.append(rid)
        result = {"created_count": len(created), "updated_count": len(updated), "skipped_count": 0,
                  "items_count": items_written, "created": created, "updated": updated}
    # 交易結束後才彙總（recompute 自己開連線），把每個匯入專案的完成%/起訖日/總燈號算出來
    for rid in dict.fromkeys(rollup_ids):
        recompute_project_rollup(rid)
    return result


def list_project_items(project_id: int) -> list[dict[str, Any]]:
    """某專案的工作項清單（排除已停用），依序號排序。

    subitem_count：這個工作項底下有沒有真的子項目清單。畫面要靠它分辨
    「3 是子項算出來的（可以點開）」還是「3 只是 Excel 帶進來的數字（要先拆）」。
    """
    with connect() as conn:
        rows = conn.execute(
            "SELECT i.*, (SELECT COUNT(*) FROM project_subitems s "
            "             WHERE s.item_id = i.id AND s.status != 'disabled') AS subitem_count "
            "FROM project_items i WHERE i.project_id = ? AND i.status != 'disabled' "
            "ORDER BY i.seq ASC, i.id ASC",
            (project_id,),
        ).fetchall()
        return [dict(r) for r in rows]


# ── WBS 自動彙總（助理文件 2026-07-29）──
# 助理定義的算法：
#   WBS 進度% ＝ 子項目完成數 ÷ 子項目總數（系統算，人只填兩個數字）
#   專案完成% ＝ 所有 WBS 的子項目完成數 ÷ 子項目總數（不是各 WBS 進度的平均——
#                20 個子項的 WBS 跟 1 個子項的 WBS 權重本來就不同）
#   專案起訖日 ＝ 第一個 WBS 的開始日、最後一個 WBS 的完成日
#   專案燈號   ＝ 所有 WBS 燈號取最嚴重，紅>黃>綠>白>灰
# 燈號五色（灰＝已完成，跟本系統流程圖的「灰＝不適用」是不同語意，別混用）
WBS_RAG_ORDER = ("red", "yellow", "green", "white", "gray")   # 越前面越嚴重
WBS_RAG_LABEL = {"red": "已延遲", "yellow": "有延遲風險", "green": "如期執行",
                 "white": "未開始", "gray": "已完成"}


def normalize_wbs_rag(value: Any) -> str:
    """人工填的燈號轉成內部代碼。舊資料與 Excel 匯入常是中文（「如期執行」「有延遲」…），
    這裡一併吃下來；認不出來就回空字串＝交給系統自動判定。"""
    v = str(value or "").strip()
    if not v:
        return ""
    if v in WBS_RAG_ORDER:
        return v
    zh = {"紅": "red", "紅燈": "red", "已延遲": "red", "有延遲且可能影響": "red",
          "黃": "yellow", "黃燈": "yellow", "有延遲風險": "yellow", "有延遲但不影響": "yellow",
          "綠": "green", "綠燈": "green", "如期執行": "green", "如期執行中": "green",
          "白": "white", "白燈": "white", "未開始": "white",
          "灰": "gray", "灰燈": "gray", "已完成": "gray"}
    return zh.get(v, "")


def wbs_exec_done(exec_status: Any) -> bool:
    """執行進度這欄（自由文字，多半從 Excel 帶進來）是不是在說「這項做完了」。

    「未完成」「尚未完成」要排除掉——只比對「有沒有完成兩個字」會把它們也算成完成。
    """
    v = " ".join(str(exec_status or "").split())
    if not v:
        return False
    if any(neg in v for neg in ("未完成", "尚未", "未結", "不完成")):
        return False
    return ("完成" in v) or ("結案" in v) or (v in ("100%", "100％", "done", "Done", "DONE"))


def wbs_item_progress(sub_total: Any, sub_done: Any, exec_status: Any = "") -> float:
    """子項目完成數 ÷ 總數 → 進度%。

    沒拆子項（總數 0）時看「執行進度」那欄：承辦寫了「已完成」就算 100%。
    原本一律回 0，於是「執行進度＝已完成、子項數 0」的工作項進度是 0%，
    過了結束日就被自動判成紅燈「已延遲」——畫面上同一列自己打自己，26 個專案都這樣。
    有拆子項時仍以子項為準（那是更精確的資訊，也是助理規格寫的算法）。
    """
    try:
        total = float(sub_total or 0)
        done = float(sub_done or 0)
    except (TypeError, ValueError):
        return 100.0 if wbs_exec_done(exec_status) else 0.0
    if total <= 0:
        return 100.0 if wbs_exec_done(exec_status) else 0.0
    return round(max(0.0, min(100.0, done / total * 100)), 1)


def wbs_auto_rag(progress: float, start_date: Any, end_date: Any) -> str:
    """依進度與起訖日自動判燈（助理定義）：
      灰＝已完成（100%）／白＝未開始（0% 且還沒到開始日，或沒排日期）
      紅＝已延遲（過了完成日還沒做完，會影響整體完成日）
      黃＝有延遲風險（進度落後時間軸推算的預期，或完成日在兩週內還沒做完）
      綠＝如期執行中
    """
    if progress >= 100:
        return "gray"
    today = date.today()
    start = _pdate(start_date)
    end = _pdate(end_date)
    if progress <= 0 and ((start and today < start) or (not start and not end)):
        return "white"
    if end and today > end:
        return "red"
    behind = False
    if start and end and end > start:
        span = (end - start).days
        elapsed = (today - start).days
        if span > 0:
            expected = max(0.0, min(100.0, elapsed / span * 100))
            behind = (expected - progress) > 10       # 落後預期一成以上＝有風險
    near_due = bool(end) and 0 <= (end - today).days <= 14
    return "yellow" if (behind or near_due) else "green"


def list_project_subitems(item_id: int) -> list[dict[str, Any]]:
    with connect() as conn:
        return [dict(r) for r in conn.execute(
            "SELECT * FROM project_subitems WHERE item_id = ? AND status != 'disabled' "
            "ORDER BY seq, id", (item_id,)).fetchall()]


def _rollup_item_from_subitems(conn: sqlite3.Connection, item_id: int) -> dict[str, Any] | None:
    """子項目異動後，把父工作項的「子項總數／已完成／完成度／燈號」重算，再往上滾到專案。

    拆了子項就以子項為準——原本那兩個數字是人工填的，兩邊並存一定會有一天對不起來。
    子項全刪光時把數字歸零並退回用「執行進度」判斷（跟沒拆過的工作項一樣）。
    """
    item = conn.execute("SELECT * FROM project_items WHERE id = ?", (item_id,)).fetchone()
    if item is None:
        return None
    rows = conn.execute(
        "SELECT done FROM project_subitems WHERE item_id = ? AND status != 'disabled'",
        (item_id,)).fetchall()
    total = len(rows)
    done = sum(1 for r in rows if int(r["done"] or 0) == 1)
    progress = wbs_item_progress(total, done, item["exec_status"])
    fields: dict[str, Any] = {"sub_total": total, "sub_done": done, "progress": progress}
    if int(item["rag_manual"] or 0) != 1:
        fields["rag"] = wbs_auto_rag(progress, item["start_date"], item["end_date"])
    assignments = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE project_items SET {assignments} WHERE id = ?", [*fields.values(), item_id])
    _recompute_project_rollup(conn, item["project_id"])
    return dict(conn.execute("SELECT * FROM project_items WHERE id = ?", (item_id,)).fetchone())


def save_project_subitem(item_id: int, payload: dict[str, Any],
                         subitem_id: int | None = None) -> dict[str, Any]:
    """新增或修改一筆子項目，並立刻把父工作項與專案的數字重算。"""
    with connect() as conn:
        if conn.execute("SELECT 1 FROM project_items WHERE id = ?", (item_id,)).fetchone() is None:
            raise LookupError(f"project_items row {item_id} not found")
        fields = {k: v for k, v in payload.items()
                  if k in allowed_fields()["project_subitems"] and v is not None}
        if subitem_id is None:
            if not str(fields.get("name") or "").strip():
                raise ValueError("子項目名稱必填。")
            fields["item_id"] = item_id
            if not fields.get("seq"):
                fields["seq"] = conn.execute(
                    "SELECT COALESCE(MAX(seq), 0) + 1 AS n FROM project_subitems WHERE item_id = ?",
                    (item_id,)).fetchone()["n"]
            row = _insert_row(conn, "project_subitems", fields)
        else:
            before = get_row(conn, "project_subitems", subitem_id)
            if not fields:
                raise ValueError("沒有可更新的欄位。")
            assignments = ", ".join(f"{k} = ?" for k in fields)
            conn.execute(f"UPDATE project_subitems SET {assignments} WHERE id = ?",
                         [*fields.values(), subitem_id])
            row = get_row(conn, "project_subitems", subitem_id)
            write_audit_log(conn, "project_subitems", subitem_id, "update", before, row)
        parent = _rollup_item_from_subitems(conn, item_id)
    return {"subitem": dict(row), "item": parent}


def delete_project_subitem(subitem_id: int) -> dict[str, Any]:
    with connect() as conn:
        before = get_row(conn, "project_subitems", subitem_id)
        item_id = before["item_id"]
        conn.execute("DELETE FROM project_subitems WHERE id = ?", (subitem_id,))
        write_audit_log(conn, "project_subitems", subitem_id, "delete", before, None)
        parent = _rollup_item_from_subitems(conn, item_id)
    return {"deleted": subitem_id, "item": parent}


def split_item_into_subitems(item_id: int) -> dict[str, Any]:
    """把「只有數字沒有清單」的舊工作項拆成子項目。

    既有資料多半是 Excel 帶進來的：sub_total=3、sub_done=3，但那三項是什麼沒人知道。
    這裡照數字產出對應筆數的空白子項（前 N 筆先勾完成，對齊原本的已完成數），
    名稱留「子項目 1、2…」讓承辦自己改——**不猜內容**，猜出來的名字看起來像真的最危險。
    已經有子項就不做，避免重複拆。
    """
    with connect() as conn:
        item = get_row(conn, "project_items", item_id)
        exist = conn.execute("SELECT COUNT(*) n FROM project_subitems WHERE item_id = ?",
                             (item_id,)).fetchone()["n"]
        if exist:
            raise RuntimeError("這個工作項已經有子項目了，不用再拆一次。")
        total = int(item["sub_total"] or 0)
        done = int(item["sub_done"] or 0)
        if total <= 0:
            raise ValueError("這個工作項的子項總數是 0，沒有東西可以拆——請直接新增子項目。")
        for i in range(1, total + 1):
            _insert_row(conn, "project_subitems", {
                "item_id": item_id, "seq": i, "name": f"子項目 {i}",
                "owner": item["owner"] or "", "done": 1 if i <= done else 0,
            })
        parent = _rollup_item_from_subitems(conn, item_id)
    return {"created": total, "done": done, "item": parent}


def recompute_project_rollup(project_id: int) -> dict[str, Any] | None:
    """把某專案底下所有 WBS 彙總回專案主檔（完成%、起訖日、總燈號）。
    專案層這三個欄位改為唯讀衍生——人只維護 WBS，專案數字自動長出來，
    避免「WBS 說落後、專案卻寫如期」這種自己打自己的畫面。沒有 WBS 就不動專案原值。"""
    with connect() as conn:
        return _recompute_project_rollup(conn, project_id)


def _recompute_project_rollup(conn: sqlite3.Connection, project_id: int) -> dict[str, Any] | None:
    """recompute_project_rollup 的核心，吃現成連線——開機遷移時已經在交易裡，
    再開一條連線會跟自己搶鎖。"""
    items = conn.execute(
        "SELECT sub_total, sub_done, progress, rag, start_date, end_date FROM project_items "
        "WHERE project_id = ? AND status != 'disabled'", (project_id,)).fetchall()
    if not items:
        return None
    total = sum(float(i["sub_total"] or 0) for i in items)
    done = sum(float(i["sub_done"] or 0) for i in items)
    # 有拆子項就按子項比例；完全沒拆（總數都是 0）就退回各 WBS 進度的平均
    if total > 0:
        progress = round(max(0.0, min(100.0, done / total * 100)), 1)
    else:
        progress = round(sum(float(i["progress"] or 0) for i in items) / len(items), 1)
    starts = [d for d in (_pdate(i["start_date"]) for i in items) if d]
    ends = [d for d in (_pdate(i["end_date"]) for i in items) if d]
    rags = [str(i["rag"] or "").strip() for i in items]
    rags = [r for r in rags if r in WBS_RAG_ORDER]
    worst = next((r for r in WBS_RAG_ORDER if r in rags), "") if rags else ""
    fields = {
        "progress": progress,
        "start_date": min(starts).isoformat() if starts else "",
        "end_date": max(ends).isoformat() if ends else "",
    }
    if worst:
        fields["rag_status"] = WBS_RAG_LABEL[worst]
    before = get_row(conn, "projects", project_id)
    assignments = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE projects SET {assignments} WHERE id = ?", [*fields.values(), project_id])
    after = get_row(conn, "projects", project_id)
    if any(str(before.get(k)) != str(after.get(k)) for k in fields):
        write_audit_log(conn, "projects", project_id, "update", before, after)
    return after


# ── WBS 展延（第三次回饋 8.4）：工作項逾期需展延結束日時留下歷史，不直接覆蓋（比照 §10 合約調整）──
def list_project_item_extensions(item_id: int) -> list[dict[str, Any]]:
    """某工作項的展延歷史，新的在前。"""
    with connect() as conn:
        rows = conn.execute(
            "SELECT * FROM project_item_extensions WHERE item_id = ? ORDER BY id DESC",
            (item_id,),
        ).fetchall()
        return [dict(r) for r in rows]


def add_project_item_extension(item_id: int, new_end_date: str, reason: str = "", note: str = "") -> dict[str, Any]:
    """記一筆展延：舊結束日自動取工作項現值，更新成展延後的新結束日並重判燈號。
    工作項 end_date 永遠是「現在的結束日」，「什麼時候、為什麼、從哪天展延到哪天」查這張表。
    展延紀錄不提供刪除（稽核用）——填錯就再展延一次回去，兩筆都留著才看得出經過。"""
    new_end_date = str(new_end_date or "").strip()
    if not new_end_date:
        raise ValueError("請填展延後的結束日。")
    with connect() as conn:
        item = conn.execute("SELECT * FROM project_items WHERE id = ?", (item_id,)).fetchone()
        if item is None:
            raise ValueError(f"工作項 ID {item_id} 不存在。")
        old_end_date = str(item["end_date"] or "").strip()
        if new_end_date == old_end_date:
            raise ValueError("展延後的結束日與現值相同，沒有東西要記錄。")
        progress = wbs_item_progress(item["sub_total"], item["sub_done"], item["exec_status"])
        rag_manual = int(item["rag_manual"] or 0) == 1
        rag = normalize_wbs_rag(item["rag"]) if rag_manual else wbs_auto_rag(progress, item["start_date"], new_end_date)
        if rag in ("red", "yellow") and not str(item["risk_note"] or "").strip():
            raise ValueError(f"燈號是「{WBS_RAG_LABEL[rag]}」時，關鍵風險點必填（要先在工作項填清楚卡在哪，再展延）。")
        row = _insert_row(conn, "project_item_extensions", {
            "item_id": item_id,
            "old_end_date": old_end_date,
            "new_end_date": new_end_date,
            "reason": str(reason or "").strip(),
            "note": str(note or "").strip(),
            "created_by": _current_actor.get(),
        })
        before = dict(item)
        conn.execute("UPDATE project_items SET end_date = ?, rag = ? WHERE id = ?", (new_end_date, rag, item_id))
        after = get_row(conn, "project_items", item_id)
        write_audit_log(conn, "project_items", item_id, "update", before, after)
    recompute_project_rollup(item["project_id"])
    return row


def next_project_item_seq(project_id: int) -> int:
    """該專案下一個工作項標號＝目前最大標號＋1（手動新增不用自己想編號，系統自動排）。"""
    with connect() as conn:
        row = conn.execute(
            "SELECT COALESCE(MAX(seq), 0) AS n FROM project_items WHERE project_id = ?", (project_id,)
        ).fetchone()
        return int(row["n"]) + 1


def parse_budget_xlsx(data: bytes) -> list[dict[str, Any]]:
    """解析『預算』.xlsx → 預算清單。此檔為『表單型』：一張工作表＝一筆預算，
    內容是「標籤：值」（預算項目／費用內容／填寫部門／預估人員…）＋右側各年度費用表。
    故用『認標籤』抓值（不是認欄位位置），金額取『全年度費用』欄中最大的一年。"""
    import io
    import openpyxl

    def norm(v: Any) -> str:
        return " ".join(str(v).split()) if v is not None else ""

    label_map = {"預算項目": "category", "填寫部門": "unit_name"}
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out: list[dict[str, Any]] = []
    try:
        for sheet in wb.sheetnames:
            rows = [list(r) if r else [] for r in wb[sheet].iter_rows(values_only=True)]

            def value_right(r: list, i: int) -> str:
                for j in range(i + 1, len(r)):
                    if r[j] is not None and str(r[j]).strip():
                        return norm(r[j])
                return ""

            fields: dict[str, str] = {}
            content = person = ""
            amount_col: int | None = None
            header_row_idx: int | None = None
            for ridx, r in enumerate(rows):
                for i, cell in enumerate(r):
                    key = norm(cell).rstrip("：:")
                    if key in label_map and label_map[key] not in fields:
                        fields[label_map[key]] = value_right(r, i)
                    elif key == "費用內容" and not content:
                        content = value_right(r, i)
                    elif key == "預估人員" and not person:
                        person = value_right(r, i)
                    if norm(cell) == "全年度費用":
                        amount_col = i
                        header_row_idx = ridx

            amount = 0.0
            fiscal_year = ""
            if amount_col is not None:
                for r in rows:
                    if amount_col < len(r):
                        try:
                            fv = float(r[amount_col])
                        except (TypeError, ValueError):
                            fv = None
                        if fv and fv > amount:
                            yr = next((norm(c) for c in r if norm(c).endswith("年度")
                                       and any(ch.isdigit() for ch in norm(c))), "")
                            amount, fiscal_year = round(fv, 2), yr

            if not (fields.get("category") or amount):
                continue  # 空白/非預算表跳過

            # 62 單位共同費用分攤表：找「部門代號」表頭列，往下讀到「合計」；col「合計」＝各單位年度分攤額
            allocations: list[dict[str, Any]] = []
            code_col = name_col = total_col = None
            alloc_hdr = None
            for i, r in enumerate(rows):
                if any(norm(c) == "部門代號" for c in r):
                    alloc_hdr = i
                    for j, c in enumerate(r):
                        k = norm(c)
                        if k == "部門代號":
                            code_col = j
                        elif k == "部門別":
                            name_col = j
                        elif k == "合計":
                            total_col = j
                    break
            if alloc_hdr is not None and name_col is not None and total_col is not None:
                seq = 0
                for r in rows[alloc_hdr + 1:]:
                    nm = norm(r[name_col]) if name_col < len(r) else ""
                    if nm in ("合計", "EOF"):
                        if nm == "合計":
                            break  # 遇到合計列就停
                        continue
                    if not nm:
                        continue
                    try:
                        amt = round(float(r[total_col]), 2) if total_col < len(r) else 0.0
                    except (TypeError, ValueError):
                        amt = 0.0
                    cd = norm(r[code_col]) if (code_col is not None and code_col < len(r)) else ""
                    seq += 1
                    allocations.append({
                        "seq": seq,
                        "unit_code": cd,
                        "unit_name": nm,
                        "amount": amt,
                        "share_pct": round(amt / amount * 100, 2) if amount else 0.0,
                    })

            # 年度×期間明細（budget_periods）：從表頭列找期間欄（含「月」且是範圍），逐年抓金額
            periods_out: list[dict[str, Any]] = []
            if header_row_idx is not None:
                hdr = rows[header_row_idx]
                period_cols = [(norm(c).replace("份", "").strip(), j)
                               for j, c in enumerate(hdr)
                               if "月" in norm(c) and "-" in norm(c) and (amount_col is None or j < amount_col)]
                year_col = (period_cols[0][1] - 1) if period_cols else None
                if period_cols and year_col is not None and year_col >= 0:
                    for r in rows[header_row_idx + 1:]:
                        yr = norm(r[year_col]) if year_col < len(r) else ""
                        if not (yr.endswith("年度") and any(ch.isdigit() for ch in yr)):
                            continue
                        yr_clean = yr.replace("年度", "").strip()
                        for lab, col in period_cols:
                            try:
                                amt = round(float(r[col]), 2) if (col < len(r) and r[col] is not None) else 0.0
                            except (TypeError, ValueError):
                                amt = 0.0
                            periods_out.append({"fiscal_year": yr_clean, "period": lab, "amount": amt})

            out.append({
                "budget_code": norm(sheet)[:60] or f"預算-{len(out) + 1}",
                "category": fields.get("category", ""),
                "unit_name": fields.get("unit_name", ""),
                "expense_detail": content,
                "estimator": person,
                "fiscal_year": fiscal_year,
                "amount": amount,
                "periods": periods_out,
                "allocations": allocations,
            })
    finally:
        wb.close()
    return out


def commit_budgets_import(records: list[dict[str, Any]], source_file: str = "") -> dict[str, Any]:
    """寫入 budgets：單一交易、逐列稽核。以 budget_code（工作表名）為鍵——同名更新、沒見過新增。
    每筆預算的 62 單位分攤明細一併寫入 budget_allocations（以 budget_id+unit_code 為鍵、同碼更新）。
    source_file：這批資料的來源 Excel 檔名，寫進每筆分攤，供單位撞名清單指回來源。"""
    source_file = (source_file or "").strip()
    fields_allowed = allowed_fields()["budgets"]
    alloc_fields = allowed_fields()["budget_allocations"]
    with connect() as conn:
        existing = {r["budget_code"]: r["id"] for r in conn.execute("SELECT id, budget_code FROM budgets").fetchall()}
        created: list[str] = []
        updated: list[str] = []
        alloc_written = 0
        periods_written = 0
        for rec in records:
            code = str(rec.get("budget_code", "")).strip()
            if not code:
                continue
            fields = {k: v for k, v in rec.items() if k in fields_allowed}
            if code in existing:
                rid = existing[code]
                before = get_row(conn, "budgets", rid)
                upd = {k: v for k, v in fields.items() if k != "budget_code"}
                if upd:
                    sets = ", ".join(f"{k} = ?" for k in upd)
                    conn.execute(f"UPDATE budgets SET {sets} WHERE id = ?", [*upd.values(), rid])
                write_audit_log(conn, "budgets", rid, "import-update", before, get_row(conn, "budgets", rid))
                updated.append(code)
            else:
                if not fields.get("case_id"):
                    cid = _ensure_case_for(conn, fields.get("budget_code"), fields.get("budget_code"),
                                           fields.get("fiscal_year"), established=True)
                    if cid:
                        fields["case_id"] = cid
                columns = ", ".join(fields)
                placeholders = ", ".join("?" for _ in fields)
                cur = conn.execute(f"INSERT INTO budgets ({columns}) VALUES ({placeholders})", list(fields.values()))
                after = get_row(conn, "budgets", cur.lastrowid)
                write_audit_log(conn, "budgets", cur.lastrowid, "import", None, {**after, "import_source": "xlsx"})
                rid = cur.lastrowid
                existing[code] = rid
                created.append(code)
            # 分攤明細：以（budget_id, unit_code）為鍵 upsert
            seen = {r["unit_code"]: r["id"] for r in conn.execute(
                "SELECT id, unit_code FROM budget_allocations WHERE budget_id = ?", (rid,)).fetchall()}
            for al in rec.get("allocations", []):
                afields = {k: v for k, v in al.items() if k in alloc_fields}
                afields["budget_id"] = rid
                if source_file:
                    afields["source_file"] = source_file
                ucode = afields.get("unit_code", "")
                if ucode and ucode in seen:
                    upd = {k: v for k, v in afields.items() if k not in ("budget_id", "unit_code")}
                    if upd:
                        sets = ", ".join(f"{k} = ?" for k in upd)
                        conn.execute(f"UPDATE budget_allocations SET {sets} WHERE id = ?", [*upd.values(), seen[ucode]])
                else:
                    cols = ", ".join(afields)
                    ph = ", ".join("?" for _ in afields)
                    c2 = conn.execute(f"INSERT INTO budget_allocations ({cols}) VALUES ({ph})", list(afields.values()))
                    if ucode:
                        seen[ucode] = c2.lastrowid
                alloc_written += 1
            # 年度費用明細（budget_periods）：整批取代這個預算的
            if "periods" in rec:
                conn.execute("DELETE FROM budget_periods WHERE budget_id = ?", (rid,))
                for pr in rec.get("periods", []):
                    fy = str(pr.get("fiscal_year") or "").strip()
                    period = str(pr.get("period") or "").strip()
                    if not fy or not period:
                        continue
                    try:
                        amt = float(pr.get("amount") or 0)
                    except (TypeError, ValueError):
                        amt = 0.0
                    conn.execute(
                        "INSERT INTO budget_periods (budget_id, fiscal_year, period, amount) VALUES (?, ?, ?, ?)",
                        (rid, fy, period, amt))
                    periods_written += 1
        return {"created_count": len(created), "updated_count": len(updated), "skipped_count": 0,
                "allocations_count": alloc_written, "periods_count": periods_written,
                "created": created, "updated": updated}


def list_budget_allocations(budget_id: int) -> list[dict[str, Any]]:
    """某費用項目的單位分攤明細（依分攤額大到小），並算出『整數分攤』：
    各單位四捨五入到元，湊不齊的尾數歸給『尾數承擔單位』（預設＝填寫部門，可用 remainder_unit_code 覆寫），
    使整數欄合計＝項目總額。回傳每列含 amount(精確)、amount_int(整數)、is_remainder_unit、remainder。"""
    with connect() as conn:
        budget = get_row(conn, "budgets", budget_id)
        rows = [dict(r) for r in conn.execute(
            "SELECT * FROM budget_allocations WHERE budget_id = ? ORDER BY amount DESC, seq ASC",
            (budget_id,)).fetchall()]
    if not rows:
        return rows
    total = int(round(float(budget.get("amount") or 0)))
    for r in rows:
        r["amount_int"] = int(round(float(r.get("amount") or 0)))
        r["is_remainder_unit"] = False
        r["remainder"] = 0
    remainder = total - sum(r["amount_int"] for r in rows)
    # 決定承擔單位：明指 remainder_unit_code > 對到填寫部門(unit_name) > 分攤額最大者
    rem_code = str(budget.get("remainder_unit_code") or "").strip()
    bunit = str(budget.get("unit_name") or "").strip()
    absorber = None
    if rem_code:
        absorber = next((r for r in rows if r["unit_code"] == rem_code), None)
    if absorber is None and bunit:
        absorber = next((r for r in rows if r["unit_name"] == bunit), None)
    if absorber is None:
        absorber = rows[0]  # 分攤額最大者
    absorber["amount_int"] += remainder
    absorber["is_remainder_unit"] = True
    absorber["remainder"] = remainder
    return rows


def update_budget_allocation(alloc_id: int, amount: Any = None, share_pct: Any = None) -> dict[str, Any]:
    """人工改單一單位的分攤金額或比例。

    使用者實際卡住的地方：分攤表只能整批重算（換分攤方法），改不了個別單位——
    但實務上談定的分攤常常是「大致按比例、某一兩個單位另議」，沒有人工微調就只能改 Excel 重匯。

    給金額就以金額為準、比例跟著算；給比例就以比例為準、金額跟著算（兩者一致，不會各說各話）。
    改完把預算的分攤方法降回 fixed——不然下次任何人按「重算」都會把這次人工談好的結果洗掉。
    """
    with connect() as conn:
        row = get_row(conn, "budget_allocations", alloc_id)
        budget = get_row(conn, "budgets", row["budget_id"])
        total = float(budget.get("amount") or 0)
        if amount is not None:
            new_amount = round(float(amount), 2)
            new_pct = round(new_amount / total * 100, 4) if total > 0 else 0.0
        elif share_pct is not None:
            new_pct = round(float(share_pct), 4)
            new_amount = round(total * new_pct / 100, 2)
        else:
            raise ValueError("請給金額或比例其中一個。")
        if new_amount < 0:
            raise ValueError("分攤金額不能是負數。")
        conn.execute("UPDATE budget_allocations SET amount = ?, share_pct = ? WHERE id = ?",
                     (new_amount, new_pct, alloc_id))
        after = get_row(conn, "budget_allocations", alloc_id)
        write_audit_log(conn, "budget_allocations", alloc_id, "manual-adjust", row, after)
        if str(budget.get("alloc_method") or "") != "fixed":
            before_b = dict(budget)
            conn.execute("UPDATE budgets SET alloc_method = 'fixed' WHERE id = ?", (row["budget_id"],))
            write_audit_log(conn, "budgets", row["budget_id"], "alloc-method-locked",
                            before_b, get_row(conn, "budgets", row["budget_id"]))
    return budget_allocation_check(row["budget_id"])


def budget_allocation_check(budget_id: int) -> dict[str, Any]:
    """分攤合計 vs 費用項目金額。改了一個單位就會對不上，要當場說差多少，
    不然人得自己拿計算機加 20 幾列。"""
    rows = list_budget_allocations(budget_id)
    with connect() as conn:
        budget = get_row(conn, "budgets", budget_id)
    total = float(budget.get("amount") or 0)
    allocated = round(sum(float(r.get("amount") or 0) for r in rows), 2)
    diff = round(total - allocated, 2)
    return {"budget_id": budget_id, "total": total, "allocated": allocated, "diff": diff,
            "balanced": abs(diff) < 0.01, "alloc_method": budget.get("alloc_method"),
            "allocations": rows}


def budget_unit_rollup(unit_code: str | None = None) -> dict[str, Any]:
    """以單位為主的彙總：每個單位在所有費用項目的分攤合計（部門負擔表）。
    經單位主檔別名解析：合併過的撞名（如同碼異名）會認到同一單位、合併加總；
    未裁決的維持原 (代號,名稱)。帶 unit_code 則另回該單位被攤的每一筆明細。"""
    with connect() as conn:
        alias_map, _masters = _load_alias_map(conn)
        rows = [dict(r) for r in conn.execute(
            "SELECT a.unit_code, a.unit_name, a.amount, a.share_pct, b.budget_code, b.category, b.fiscal_year "
            "FROM budget_allocations a JOIN budgets b ON b.id = a.budget_id").fetchall()]

    def resolve(code: str, name: str) -> tuple[str, str, str]:
        m = alias_map.get((code, name))
        if m:
            return (f"m{m['master_id']}", m["canonical_code"], m["canonical_name"])
        return (f"r::{code}::{name}", code, name)

    groups: dict[str, dict[str, Any]] = {}
    detail: list[dict[str, Any]] = []
    for r in rows:
        code = str(r.get("unit_code") or "").strip()
        name = str(r.get("unit_name") or "").strip()
        key, disp_code, disp_name = resolve(code, name)
        g = groups.setdefault(key, {"unit_code": disp_code, "unit_name": disp_name,
                                    "item_count": 0, "total_amount": 0.0})
        g["item_count"] += 1
        g["total_amount"] += float(r.get("amount") or 0)
        if unit_code is not None and disp_code == str(unit_code):
            detail.append({"amount": r.get("amount"), "share_pct": r.get("share_pct"),
                           "budget_code": r.get("budget_code"), "category": r.get("category"),
                           "fiscal_year": r.get("fiscal_year")})

    units = sorted(groups.values(), key=lambda g: -g["total_amount"])
    for u in units:
        u["total_amount"] = round(u["total_amount"])
    result: dict[str, Any] = {"units": units}
    if unit_code is not None:
        detail.sort(key=lambda d: -float(d.get("amount") or 0))
        result["detail"] = detail
    return result


def _load_alias_map(conn) -> tuple[dict[tuple[str, str], dict[str, Any]], dict[int, dict[str, Any]]]:
    """讀單位主檔＋別名，回傳：
    - alias_map：{(代號, 名稱) → {master_id, canonical_code, canonical_name}}
    - masters：{master_id → {id, canonical_code, canonical_name, ...}}"""
    masters = {m["id"]: dict(m) for m in conn.execute(
        "SELECT id, canonical_code, canonical_name, status, note FROM unit_master").fetchall()}
    alias_map: dict[tuple[str, str], dict[str, Any]] = {}
    for a in conn.execute("SELECT master_id, alias_code, alias_name FROM unit_aliases").fetchall():
        m = masters.get(a["master_id"])
        if not m:
            continue
        alias_map[(str(a["alias_code"] or "").strip(), str(a["alias_name"] or "").strip())] = {
            "master_id": m["id"], "canonical_code": m["canonical_code"], "canonical_name": m["canonical_name"]}
    return alias_map, masters


def unit_conflicts() -> dict[str, Any]:
    """單位主檔：偵測跨資料的『撞名』——同一代號對到多個名稱、或同一名稱對到多個代號。
    掃描 budget_allocations、unit_headcounts；已在單位主檔裁決過（每個變體都有別名）的組別視為『已處理』，
    不再列入待確認。每個變體附上它目前對到的主檔（canonical），供前端顯示與合併操作。"""
    # 每個資料表的中文分類（沒記到來源檔名時的退路標籤）
    sources = [
        ("budget_allocations", "預算分攤"),
        ("unit_headcounts", "人數基準"),
    ]
    # 收集所有 (代號, 名稱) 出現次數，並記「來源」＝實際 Excel 檔名（記不到才退回分類標籤）
    pairs: dict[tuple[str, str], dict[str, Any]] = {}
    with connect() as conn:
        alias_map, _masters = _load_alias_map(conn)
        for table, label in sources:
            rows = conn.execute(
                f"SELECT COALESCE(unit_code,'') AS c, COALESCE(unit_name,'') AS n, "
                f"COALESCE(source_file,'') AS f, COUNT(*) AS cnt "
                f"FROM {table} GROUP BY c, n, f"
            ).fetchall()
            for r in rows:
                code = str(r["c"]).strip()
                name = str(r["n"]).strip()
                if not name and not code:
                    continue
                fname = str(r["f"]).strip()
                # 有檔名就顯示檔名；舊資料沒記檔名，退回「分類（未記檔名）」
                src = fname if fname else f"{label}（舊資料·未記檔名）"
                key = (code, name)
                slot = pairs.setdefault(key, {"unit_code": code, "unit_name": name, "count": 0, "sources": set()})
                slot["count"] += int(r["cnt"])
                slot["sources"].add(src)

    def _entry(code: str, name: str, slot: dict[str, Any]) -> dict[str, Any]:
        m = alias_map.get((code, name))
        return {"unit_code": code, "unit_name": name, "count": slot["count"],
                "sources": sorted(slot["sources"]),
                "master": m,  # None＝尚未裁決；否則為它目前對到的主檔
                "resolved": bool(m)}

    by_code: dict[str, list[dict[str, Any]]] = {}
    by_name: dict[str, list[dict[str, Any]]] = {}
    for (code, name), slot in pairs.items():
        entry = _entry(code, name, slot)
        if code:
            by_code.setdefault(code, []).append(entry)
        if name:
            by_name.setdefault(name, []).append(entry)

    # 一碼多名；已全部裁決（每個變體都有別名）者視為已處理，不再列入待確認
    code_conflicts = []
    resolved_codes = 0
    for code, entries in by_code.items():
        names = {e["unit_name"] for e in entries if e["unit_name"]}
        if len(names) <= 1:
            continue
        if all(e["resolved"] for e in entries):
            resolved_codes += 1
            continue
        code_conflicts.append({"unit_code": code, "variants": sorted(entries, key=lambda e: -e["count"])})
    # 一名多碼（含「有代號 vs 空代號」）
    name_conflicts = []
    resolved_names = 0
    for name, entries in by_name.items():
        codes = {e["unit_code"] for e in entries}
        if len(codes) <= 1:
            continue
        if all(e["resolved"] for e in entries):
            resolved_names += 1
            continue
        name_conflicts.append({"unit_name": name, "variants": sorted(entries, key=lambda e: -e["count"])})

    code_conflicts.sort(key=lambda x: x["unit_code"])
    name_conflicts.sort(key=lambda x: x["unit_name"])
    return {
        "code_conflicts": code_conflicts,
        "name_conflicts": name_conflicts,
        "summary": {
            "code_conflicts": len(code_conflicts),
            "name_conflicts": len(name_conflicts),
            "resolved_groups": resolved_codes + resolved_names,
            "distinct_pairs": len(pairs),
        },
    }


def list_unit_master() -> dict[str, Any]:
    """單位主檔清單：每個主檔 + 它底下的別名（代號/名稱），供檢視與解除合併。"""
    with connect() as conn:
        masters = [dict(m) for m in conn.execute(
            "SELECT id, canonical_code, canonical_name, status, note FROM unit_master ORDER BY canonical_code, id").fetchall()]
        aliases = [dict(a) for a in conn.execute(
            "SELECT id, master_id, alias_code, alias_name FROM unit_aliases ORDER BY id").fetchall()]
    by_master: dict[int, list[dict[str, Any]]] = {}
    for a in aliases:
        by_master.setdefault(a["master_id"], []).append(a)
    for m in masters:
        m["aliases"] = by_master.get(m["id"], [])
    return {"masters": masters, "count": len(masters)}


def create_unit_master(canonical_code: str, canonical_name: str, note: str = "") -> dict[str, Any]:
    """主動新增一個乾淨單位（跟合併機制不同——合併是被動處理既有撞名資料，這是事前登記，
    給表單下拉選單用）。建立前擋撞名：跟現有主檔的代號/名稱、或現有別名撞到就拒絕並指出撞到誰，
    避免髒資料從源頭就重複，而不是等資料進來後才靠合併機制事後補救。"""
    code = (canonical_code or "").strip()
    name = (canonical_name or "").strip()
    if not name:
        raise ValueError("請填單位名稱。")
    with connect() as conn:
        if code:
            dup_code = conn.execute(
                "SELECT canonical_name FROM unit_master WHERE canonical_code = ?", (code,)
            ).fetchone()
            if dup_code:
                raise ValueError(f"代號「{code}」已存在於單位主檔（對應「{dup_code['canonical_name']}」），不能重複。")
        dup_name = conn.execute(
            "SELECT canonical_code FROM unit_master WHERE canonical_name = ?", (name,)
        ).fetchone()
        if dup_name:
            raise ValueError(f"「{name}」已存在於單位主檔（代號 {dup_name['canonical_code'] or '—'}），不能重複新增。")
        dup_alias = conn.execute(
            "SELECT um.canonical_name FROM unit_aliases ua "
            "JOIN unit_master um ON um.id = ua.master_id WHERE ua.alias_name = ?", (name,)
        ).fetchone()
        if dup_alias:
            raise ValueError(f"「{name}」已被登記為「{dup_alias['canonical_name']}」的別名，不能重複新增；如需調整請到「撞名待確認」處理。")
        cur = conn.execute(
            "INSERT INTO unit_master (canonical_code, canonical_name, note) VALUES (?, ?, ?)",
            (code, name, note),
        )
        row_id = cur.lastrowid
        row = get_row(conn, "unit_master", row_id)
        write_audit_log(conn, "unit_master", row_id, "create", None, row)
    return row


def list_personnel_master(include_disabled: bool = False) -> dict[str, Any]:
    """人員主檔清單：給案件/簽呈/預算/付款/專案表單的人員欄位下拉選單用。
    後台管理要看得到已停用的（才能重新啟用），所以用 include_disabled 切換。"""
    where = "" if include_disabled else "WHERE status <> 'disabled'"
    with connect() as conn:
        masters = [dict(m) for m in conn.execute(
            f"SELECT id, name, group_name, email, status, note FROM personnel_master {where} "
            "ORDER BY group_name, name").fetchall()]
    groups = sorted({m["group_name"] for m in masters if m["group_name"]})
    return {"masters": masters, "count": len(masters), "groups": groups,
            "missing_email": sum(1 for m in masters
                                 if m["status"] != "disabled" and not str(m["email"] or "").strip()),
            "missing_group": sum(1 for m in masters
                                 if m["status"] != "disabled" and not str(m["group_name"] or "").strip())}


def personnel_email(name: str) -> str:
    """人名 → email。通知的收件人欄位（核銷者、負責人）存的是人名不是帳號，
    沒有這層對照就寄不出去（助理 2026-08-13 卡在這裡）。"""
    n = str(name or "").strip()
    if not n:
        return ""
    with connect() as conn:
        row = conn.execute(
            "SELECT email FROM personnel_master WHERE name = ? AND status <> 'disabled'",
            (n,)).fetchone()
    return str(row["email"]).strip() if row and row["email"] else ""


def create_personnel_master(name: str, note: str = "", group_name: str = "",
                            email: str = "") -> dict[str, Any]:
    """主動新增一個人員（給表單下拉選單用）。建立前擋撞名，避免同一人被打成兩種寫法。"""
    n = (name or "").strip()
    if not n:
        raise ValueError("請填人員姓名。")
    with connect() as conn:
        dup = conn.execute("SELECT id FROM personnel_master WHERE name = ?", (n,)).fetchone()
        if dup:
            raise ValueError(f"「{n}」已存在於人員主檔，不能重複新增。")
        cur = conn.execute(
            "INSERT INTO personnel_master (name, group_name, note, email) VALUES (?, ?, ?, ?)",
            (n, (group_name or "").strip(), note, (email or "").strip()))
        row_id = cur.lastrowid
        row = get_row(conn, "personnel_master", row_id)
        write_audit_log(conn, "personnel_master", row_id, "create", None, row)
    return row


def update_personnel_master(person_id: int, fields: dict[str, Any]) -> dict[str, Any]:
    """改人員資料（換組、改名、停用/啟用、備註）。人會轉組、會離職，這些都要能改。"""
    allowed = {k: v for k, v in fields.items()
               if k in ("name", "group_name", "note", "status", "email") and v is not None}
    if not allowed:
        raise ValueError("沒有可更新的欄位。")
    if "name" in allowed:
        allowed["name"] = str(allowed["name"]).strip()
        if not allowed["name"]:
            raise ValueError("人員姓名不能空白。")
    if "status" in allowed and allowed["status"] not in ("active", "disabled"):
        raise ValueError("人員狀態只能是 active 或 disabled。")
    with connect() as conn:
        before = get_row(conn, "personnel_master", person_id)
        if "name" in allowed:
            dup = conn.execute("SELECT id FROM personnel_master WHERE name = ? AND id <> ?",
                               (allowed["name"], person_id)).fetchone()
            if dup:
                raise ValueError(f"「{allowed['name']}」已存在於人員主檔。")
        assignments = ", ".join(f"{k} = ?" for k in allowed)
        conn.execute(f"UPDATE personnel_master SET {assignments} WHERE id = ?",
                     [*allowed.values(), person_id])
        after = get_row(conn, "personnel_master", person_id)
        write_audit_log(conn, "personnel_master", person_id, "update", before, after)
        return after


def delete_personnel_master(person_id: int) -> None:
    """真的刪掉一筆人員。已經被表單引用過的名字不會跟著消失（那些欄位存的是文字），
    所以刪除只影響「以後還選不選得到」，不會弄壞歷史資料。"""
    with connect() as conn:
        before = get_row(conn, "personnel_master", person_id)
        conn.execute("DELETE FROM personnel_master WHERE id = ?", (person_id,))
        write_audit_log(conn, "personnel_master", person_id, "delete", before, None)


# 示範名單：四組各三人，讓下拉選單一開始就有東西可選（真名單由後台自行維護／匯入）。
# note 一律標「示範資料」，之後要清掉一眼就分得出來。
DEMO_PERSONNEL = {
    "資料庫組": ["張淑芬", "李冠廷", "吳佩珊"],
    "網路組": ["黃俊傑", "劉美玲", "蔡承翰"],
    "主機組": ["王志明", "陳怡君", "林建宏"],
    "專案及流程管理組": ["鄭雅婷", "許家豪", "楊書瑋"],
}


def seed_demo_personnel() -> dict[str, Any]:
    """載入示範人員名單（四組各三人）。已存在的同名人員跳過不覆蓋，可重複執行。"""
    created, skipped = [], []
    with connect() as conn:
        for group, names in DEMO_PERSONNEL.items():
            for name in names:
                exists = conn.execute("SELECT id FROM personnel_master WHERE name = ?", (name,)).fetchone()
                if exists:
                    skipped.append(name)
                    continue
                cur = conn.execute(
                    "INSERT INTO personnel_master (name, group_name, note) VALUES (?, ?, '示範資料')",
                    (name, group))
                row = get_row(conn, "personnel_master", cur.lastrowid)
                write_audit_log(conn, "personnel_master", cur.lastrowid, "create", None, row)
                created.append(row)
    return {"created": created, "created_count": len(created),
            "skipped": skipped, "skipped_count": len(skipped)}


def _find_or_create_master(conn, canonical_code: str, canonical_name: str, note: str = "") -> int:
    """以 (canonical_code, canonical_name) 找主檔，沒有就建。回傳 master_id。"""
    code = (canonical_code or "").strip()
    name = (canonical_name or "").strip()
    row = conn.execute(
        "SELECT id FROM unit_master WHERE canonical_code = ? AND canonical_name = ?", (code, name)).fetchone()
    if row:
        return row["id"]
    cur = conn.execute(
        "INSERT INTO unit_master (canonical_code, canonical_name, note) VALUES (?, ?, ?)", (code, name, note))
    return cur.lastrowid


def _attach_alias(conn, master_id: int, alias_code: str, alias_name: str) -> int | None:
    """把 (代號, 名稱) 掛到某主檔；若別名已存在則改指到此主檔（重新裁決可覆蓋）。
    回傳『前一個 master_id』（原本沒別名則 None），供決策紀錄記 undo。"""
    code = (alias_code or "").strip()
    name = (alias_name or "").strip()
    existing = conn.execute(
        "SELECT id, master_id FROM unit_aliases WHERE alias_code = ? AND alias_name = ?", (code, name)).fetchone()
    if existing:
        prev = existing["master_id"]
        conn.execute("UPDATE unit_aliases SET master_id = ? WHERE id = ?", (master_id, existing["id"]))
        return prev
    conn.execute(
        "INSERT INTO unit_aliases (master_id, alias_code, alias_name) VALUES (?, ?, ?)", (master_id, code, name))
    return None


def _record_decision(conn, action: str, reason: str, detail: dict[str, Any], undo_ops: list[dict[str, Any]]) -> int:
    cur = conn.execute(
        "INSERT INTO unit_decisions (action, reason, actor, detail_json, undo_ops_json) VALUES (?, ?, ?, ?, ?)",
        (action, (reason or "").strip(), _current_actor.get(),
         json.dumps(detail, ensure_ascii=False), json.dumps(undo_ops, ensure_ascii=False)))
    return cur.lastrowid


def _cleanup_empty_masters(conn) -> int:
    rows = conn.execute(
        "SELECT id FROM unit_master WHERE id NOT IN (SELECT DISTINCT master_id FROM unit_aliases)").fetchall()
    for r in rows:
        conn.execute("DELETE FROM unit_master WHERE id = ?", (r["id"],))
    return len(rows)


def unit_merge_impact(variants: list[dict[str, Any]]) -> dict[str, Any]:
    """影響預覽：這些變體目前在『預算分攤』佔幾筆、金額多少（讓使用者按下前看清後果）。"""
    rows = 0
    amount = 0.0
    per: list[dict[str, Any]] = []
    with connect() as conn:
        for v in variants:
            code = str(v.get("unit_code", "")).strip()
            name = str(v.get("unit_name", "")).strip()
            r = conn.execute(
                "SELECT COUNT(*) AS n, COALESCE(SUM(amount),0) AS amt FROM budget_allocations "
                "WHERE COALESCE(unit_code,'')=? AND COALESCE(unit_name,'')=?", (code, name)).fetchone()
            rows += int(r["n"])
            amount += float(r["amt"])
            per.append({"unit_code": code, "unit_name": name, "rows": int(r["n"]), "amount": round(float(r["amt"]))})
    return {"rows": rows, "amount": round(amount), "per_variant": per}


def merge_units(variants: list[dict[str, Any]], canonical_code: str, canonical_name: str, reason: str = "") -> dict[str, Any]:
    """合併：這些變體是同一單位，以 (canonical_code, canonical_name) 為準。
    建/取主檔，把每個變體掛成別名。非破壞式：原始資料不動，讀取時經別名認到同一主檔。
    reason 必填（防呆＋留依據）；記入 unit_decisions，可逐筆復原。"""
    if not variants:
        raise ValueError("沒有要合併的單位變體。")
    if not (reason or "").strip():
        raise ValueError("請填『為什麼這樣判斷』的理由，才能裁決。")
    cname = (canonical_name or "").strip()
    if not cname and not (canonical_code or "").strip():
        raise ValueError("請指定要以哪個為準（代號或名稱至少一個）。")
    with connect() as conn:
        master_id = _find_or_create_master(conn, canonical_code, canonical_name)
        undo_ops = []
        for v in variants:
            code = str(v.get("unit_code", ""))
            name = str(v.get("unit_name", ""))
            prev = _attach_alias(conn, master_id, code, name)
            undo_ops.append({"alias_code": code.strip(), "alias_name": name.strip(), "prev_master_id": prev})
        _cleanup_empty_masters(conn)
        detail = {"canonical_code": (canonical_code or "").strip(), "canonical_name": cname, "variants": variants}
        did = _record_decision(conn, "merge", reason, detail, undo_ops)
    return {"master_id": master_id, "merged": len(variants), "decision_id": did,
            "canonical_code": (canonical_code or "").strip(), "canonical_name": cname}


def reassign_unit(variant: dict[str, Any], canonical_code: str, canonical_name: str, reason: str = "") -> dict[str, Any]:
    """逐筆改派：某一筆撞名變體其實屬於別的單位（常見於代號打錯），
    把它單獨掛到指定的主單位（現有的、或用正確代號/名稱新建），不影響同組其他筆。非破壞式、可復原。"""
    if not variant or (not str(variant.get("unit_code", "")).strip() and not str(variant.get("unit_name", "")).strip()):
        raise ValueError("沒有要改派的單位。")
    if not (reason or "").strip():
        raise ValueError("請填『為什麼這樣判斷』的理由，才能改派。")
    cname = (canonical_name or "").strip()
    if not cname and not (canonical_code or "").strip():
        raise ValueError("請指定要改派到哪個單位（代號或名稱至少一個）。")
    code = str(variant.get("unit_code", ""))
    name = str(variant.get("unit_name", ""))
    with connect() as conn:
        master_id = _find_or_create_master(conn, canonical_code, canonical_name)
        prev = _attach_alias(conn, master_id, code, name)
        _cleanup_empty_masters(conn)
        undo_ops = [{"alias_code": code.strip(), "alias_name": name.strip(), "prev_master_id": prev}]
        detail = {"canonical_code": (canonical_code or "").strip(), "canonical_name": cname, "variants": [variant]}
        did = _record_decision(conn, "reassign", reason, detail, undo_ops)
    return {"master_id": master_id, "decision_id": did,
            "canonical_code": (canonical_code or "").strip(), "canonical_name": cname}


def split_units(variants: list[dict[str, Any]], reason: str = "") -> dict[str, Any]:
    """分開保留：這些變體是不同單位，各自成為一個主檔（別名＝自己）。裁決後不再列為待確認。"""
    if not variants:
        raise ValueError("沒有要分開的單位變體。")
    if not (reason or "").strip():
        raise ValueError("請填『為什麼這樣判斷』的理由，才能裁決。")
    made = 0
    with connect() as conn:
        undo_ops = []
        for v in variants:
            code = str(v.get("unit_code", ""))
            name = str(v.get("unit_name", ""))
            mid = _find_or_create_master(conn, code, name)
            prev = _attach_alias(conn, mid, code, name)
            undo_ops.append({"alias_code": code.strip(), "alias_name": name.strip(), "prev_master_id": prev})
            made += 1
        _cleanup_empty_masters(conn)
        did = _record_decision(conn, "split", reason, {"variants": variants}, undo_ops)
    return {"split": made, "decision_id": did}


def list_unit_decisions(limit: int = 100) -> dict[str, Any]:
    """決策紀錄：誰、何時、把什麼合併/分開、為什麼，供檢視與逐筆復原。"""
    with connect() as conn:
        rows = [dict(r) for r in conn.execute(
            "SELECT id, action, reason, actor, detail_json, undone, created_at "
            "FROM unit_decisions ORDER BY id DESC LIMIT ?", (limit,)).fetchall()]
    out = []
    for r in rows:
        try:
            detail = json.loads(r.get("detail_json") or "{}")
        except (ValueError, TypeError):
            detail = {}
        variants = detail.get("variants", [])
        out.append({
            "id": r["id"], "action": r["action"], "reason": r["reason"], "actor": r["actor"],
            "undone": bool(r["undone"]), "created_at": r["created_at"],
            "canonical_code": detail.get("canonical_code", ""), "canonical_name": detail.get("canonical_name", ""),
            "variants": [{"unit_code": v.get("unit_code", ""), "unit_name": v.get("unit_name", "")} for v in variants],
        })
    return {"decisions": out, "count": len(out)}


def undo_decision(decision_id: int) -> dict[str, Any]:
    """復原某次裁決：依 undo_ops 把每個別名還原到前一個歸屬（原本沒有就刪掉），並清掉空主檔。"""
    with connect() as conn:
        row = conn.execute(
            "SELECT undo_ops_json, undone FROM unit_decisions WHERE id = ?", (decision_id,)).fetchone()
        if not row:
            raise ValueError("找不到該筆裁決紀錄。")
        if row["undone"]:
            raise ValueError("這筆裁決已經復原過了。")
        try:
            ops = json.loads(row["undo_ops_json"] or "[]")
        except (ValueError, TypeError):
            ops = []
        for op in ops:
            code = str(op.get("alias_code", "")).strip()
            name = str(op.get("alias_name", "")).strip()
            prev = op.get("prev_master_id")
            if prev is None:
                conn.execute("DELETE FROM unit_aliases WHERE alias_code = ? AND alias_name = ?", (code, name))
            else:
                conn.execute("UPDATE unit_aliases SET master_id = ? WHERE alias_code = ? AND alias_name = ?",
                             (prev, code, name))
        removed = _cleanup_empty_masters(conn)
        conn.execute("UPDATE unit_decisions SET undone = 1 WHERE id = ?", (decision_id,))
    return {"undone": decision_id, "removed_masters": removed}


def reset_unit_decisions() -> dict[str, Any]:
    """一鍵還原：清掉所有單位裁決（別名＋主檔），回到剛匯入的原始狀態。
    原始 budget_allocations / unit_headcounts 本就沒被動過，所以這是保證級的後悔藥。"""
    with connect() as conn:
        n_alias = conn.execute("SELECT COUNT(*) AS n FROM unit_aliases").fetchone()["n"]
        n_master = conn.execute("SELECT COUNT(*) AS n FROM unit_master").fetchone()["n"]
        conn.execute("DELETE FROM unit_aliases")
        conn.execute("DELETE FROM unit_master")
        conn.execute("UPDATE unit_decisions SET undone = 1 WHERE undone = 0")
        write_audit_log(conn, "unit_master", 0, "reset-all",
                        {"aliases": n_alias, "masters": n_master}, {"aliases": 0, "masters": 0})
    return {"removed_aliases": n_alias, "removed_masters": n_master}


def unlink_alias(alias_id: int) -> dict[str, Any]:
    """解除某別名（還原裁決）；若主檔已無任何別名，一併刪除該空主檔。"""
    with connect() as conn:
        row = conn.execute("SELECT master_id FROM unit_aliases WHERE id = ?", (alias_id,)).fetchone()
        if not row:
            raise ValueError("找不到該別名。")
        master_id = row["master_id"]
        conn.execute("DELETE FROM unit_aliases WHERE id = ?", (alias_id,))
        left = conn.execute("SELECT COUNT(*) AS n FROM unit_aliases WHERE master_id = ?", (master_id,)).fetchone()["n"]
        removed_master = False
        if left == 0:
            conn.execute("DELETE FROM unit_master WHERE id = ?", (master_id,))
            removed_master = True
        write_audit_log(conn, "unit_master", master_id, "unlink-alias", None,
                        {"alias_id": alias_id, "removed_master": removed_master})
    return {"unlinked": alias_id, "removed_master": removed_master}


# ==== 名稱歸納（案件名/專案名/廠商名）：比照單位主檔，把同一實體的不同寫法歸成一個 ====
NAME_SOURCES: dict[str, list[tuple[str, str]]] = {
    "case": [("cases", "title")],
    "project": [("projects", "project_name")],
    "vendor": [("contracts", "vendor_name"), ("payments", "vendor"), ("purchases", "vendor_name")],
    "budget": [("budgets", "budget_code")],   # 費用項目名：同一實體不同寫法（端點APT防護 vs …授權暨維護）先歸一
}
NAME_KIND_LABEL = {"case": "案件名稱", "project": "專案名稱", "vendor": "廠商名稱", "budget": "預算項目"}


# ── 人員盤點與離職交接（使用者 2026-08-12）───────────────────────────────
# 「一個人名下有什麼」不是一個查詢就有答案，因為系統裡有兩種存法：
#   案件的負責人存**登入帳號**（ap03），其他模組存**人名**（林信成）。
# 而且人名欄位允許共同負責人（實際資料裡有「陳昱杉/洪似妮」），所以不能只用 = 比對。
# 盤點盲了，交接就是盲的——先看得到他有什麼，才談得上轉給誰。
PERSON_FIELDS = [
    # (表, 欄位, 顯示名稱, 比對方式, 結案狀態)
    ("cases", "owner", "案件", "account", ("closed", "cancelled", "merged", "rejected", "disabled")),
    ("projects", "owner", "專案", "name", ("completed", "disabled")),
    ("contracts", "owner", "合約", "name", ("closed", "disabled")),
    ("project_items", "owner", "工作項", "name", ("disabled",)),
    ("project_subitems", "owner", "子項目", "name", ("disabled",)),
    ("expense_masters", "owner", "費用主檔", "name", ("closed", "disabled")),
    ("payments", "owner", "付款", "name", ("closed", "disabled")),
    ("signoffs", "applicant", "簽呈", "name", ("approved", "rejected", "disabled")),
]
_NAME_SEPARATORS = ("/", "、", "&", "／", ",", "，")


def _name_matches(cell: Any, person: str) -> bool:
    """欄位值是不是「這個人」。共同負責人（陳昱杉/洪似妮）也要算他一份。"""
    raw = str(cell or "").strip()
    if not raw or not person:
        return False
    if raw == person:
        return True
    parts = [raw]
    for sep in _NAME_SEPARATORS:
        parts = [p for chunk in parts for p in chunk.split(sep)]
    return person in [p.strip() for p in parts if p.strip()]


def personnel_workload(person_name: str = "", username: str = "") -> dict[str, Any]:
    """某個人名下有哪些資料，按模組分「進行中／已結案」。

    兩個 key 都要給才盤得完整：person_name 比人名欄位、username 比案件的帳號欄位。
    只給其中一個也能跑，但另一半會是 0——回傳裡標出來，不要讓人以為真的沒有。
    """
    person_name = str(person_name or "").strip()
    username = str(username or "").strip()
    blocks: list[dict[str, Any]] = []
    with connect() as conn:
        existing = {r["name"] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'")}
        for table, col, label, kind, closed_states in PERSON_FIELDS:
            if table not in existing:
                continue
            key = username if kind == "account" else person_name
            if not key:
                continue
            rows = [dict(r) for r in conn.execute(
                f"SELECT * FROM {table} WHERE COALESCE({col},'') <> ''").fetchall()]
            mine = [r for r in rows
                    if (str(r[col]).strip() == key if kind == "account" else _name_matches(r[col], key))]
            if not mine:
                continue
            active = [r for r in mine if str(r.get("status") or "") not in closed_states]
            blocks.append({
                "table": table, "field": col, "label": label, "match_by": kind,
                "total": len(mine), "active": len(active), "closed": len(mine) - len(active),
                "sample": [_workload_title(table, r) for r in mine[:5]],
            })
    return {
        "person_name": person_name, "username": username,
        "blocks": blocks,
        "total": sum(b["total"] for b in blocks),
        "active": sum(b["active"] for b in blocks),
        "closed": sum(b["closed"] for b in blocks),
        "note": "案件比對登入帳號，其他模組比對人名；共同負責人（A/B）也算他一份。",
    }


def _workload_title(table: str, row: dict[str, Any]) -> str:
    for key in ("title", "project_name", "contract_name", "item_name", "name",
                "expense_name", "subject", "settle_no", "payment_month"):
        if str(row.get(key) or "").strip():
            return str(row[key]).strip()
    return f"#{row.get('id')}"


def personnel_workload_overview() -> dict[str, Any]:
    """全部人員的負擔一覽：誰身上有多少東西。交接前先看這張，才知道要找誰接。

    名單來源是「人員主檔 ∪ 資料裡實際出現過的名字」，不是只有主檔——
    實測這台的主檔只登記 1 個人，但專案負責人有 8 個以上（林義昌、黎世偉/游穗宗…）。
    只看主檔的話，真正有工作要交接的人全部盤不到，等於這功能白做。
    共同負責人（A/B）會拆開各算一份。

    一次把資料撈進記憶體分組，不對每個人各掃一次全表（人一多會慢得誇張）。
    """
    people: dict[str, dict[str, Any]] = {}

    def slot(name: str) -> dict[str, Any]:
        return people.setdefault(name, {
            "name": name, "group_name": "", "status": "", "username": "",
            "in_master": False, "total": 0, "active": 0, "closed": 0, "blocks": {},
        })

    with connect() as conn:
        existing = {r["name"] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'")}
        for m in conn.execute(
                "SELECT name, group_name, status FROM personnel_master").fetchall():
            s = slot(str(m["name"]).strip())
            s.update({"group_name": m["group_name"] or "", "status": m["status"] or "",
                      "in_master": True})
        by_account = {}
        for u in conn.execute(
                "SELECT username, display_name FROM users WHERE COALESCE(display_name,'') <> ''").fetchall():
            by_account[str(u["username"])] = str(u["display_name"])
            slot(str(u["display_name"]))["username"] = str(u["username"])

        for table, col, label, kind, closed_states in PERSON_FIELDS:
            if table not in existing:
                continue
            for r in conn.execute(
                    f"SELECT {col} AS person, status FROM {table} "
                    f"WHERE COALESCE({col},'') <> ''").fetchall():
                raw = str(r["person"]).strip()
                # 案件存的是帳號，換算回人名再併帳；查不到對照就用帳號本身當顯示名
                names = [by_account.get(raw, raw)] if kind == "account" else _split_persons(raw)
                closed = str(r["status"] or "") in closed_states
                for name in names:
                    s = slot(name)
                    s["total"] += 1
                    if closed:
                        s["closed"] += 1
                    else:
                        s["active"] += 1
                        s["blocks"][label] = s["blocks"].get(label, 0) + 1

    out = sorted(people.values(), key=lambda x: (-x["active"], -x["total"], x["name"]))
    return {"people": out, "count": len(out),
            "not_in_master": sum(1 for p in out if not p["in_master"] and p["total"]),
            "unassigned_hint": "名單同時來自人員主檔與實際資料；標「未登記」的人有資料卻不在主檔，"
                               "建議補登記。沒有登入帳號的人，案件那一塊會是 0（案件比對的是帳號）。"}


def _split_persons(cell: Any) -> list[str]:
    """把「陳昱杉/洪似妮」拆成兩個人。"""
    parts = [str(cell or "").strip()]
    for sep in _NAME_SEPARATORS:
        parts = [p for chunk in parts for p in chunk.split(sep)]
    return [p.strip() for p in parts if p.strip()]


# 從資料裡自動補登記人員（使用者 2026-08-12：「如果系統有抓到人 可以自動幫我建立人嗎」）。
# 不能照單全收——實際資料裡就有「蔡維庭 黎世偉 吳季凌 游穗宗」這種四個人塞一格用空白分隔的，
# 直接建會生出一個名字很長的假人。所以：拆得開的拆開、可疑的標出來、一律先預覽再建。
_GROUP_HINT = re.compile(r"(主機組|網路組|資料庫組|專案及流程管理組|[一-鿿]{1,6}組)")


# 不是人名的字眼：這些是欄位被拿來寫備註留下的（「由網路組協助，待確認負責人」）
_NOT_A_NAME = ("待確認", "未指派", "未定", "待補", "待定", "協助", "負責人", "組別",
               "TBD", "tbd", "N/A", "n/a", "無")


def _looks_like_person(name: str) -> bool:
    """像不像一個人的名字。中文姓名通常 2–5 字，超過多半是把欄位當備註在寫。

    只用長度不夠：「由網路組協助，待確認負責人」會被逗號拆成兩段 6 字的片段矇混過關，
    所以再擋一組明顯不是名字的字眼。判斷錯了也不會出事——只是不預設勾選，人還是能自己勾。
    """
    n = str(name or "").strip()
    if not (2 <= len(n) <= 5):
        return False
    return not any(k in n for k in _NOT_A_NAME)


def _guess_group(sources: list[str]) -> str:
    """從專案來源（匯入的工作表名，如「網路組處級專案」）推組別。推不出來就留空，不瞎猜。"""
    for s in sources:
        m = _GROUP_HINT.search(str(s or ""))
        if m:
            return m.group(1)
    return ""


def suggest_personnel_from_data() -> dict[str, Any]:
    """掃出資料裡出現過、但還沒登記在人員主檔的人，附推測組別與可疑標記。

    可疑的定義（借用 _clean_owner 同一套判斷）：太長、含標點、或看起來是一格塞多個人。
    這些不預設勾選，讓人自己看過再決定。
    """
    with connect() as conn:
        existing = {str(r["name"]).strip() for r in conn.execute(
            "SELECT name FROM personnel_master").fetchall()}
        tables = {r["name"] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'")}
        accounts = {str(r["username"]) for r in conn.execute("SELECT username FROM users").fetchall()}
        found: dict[str, dict[str, Any]] = {}
        for table, col, label, kind, _closed in PERSON_FIELDS:
            if table not in tables or kind == "account":
                continue                      # 帳號那塊不是人名，不拿來建人員
            src_col = "source" if table == "projects" else None
            sql = f"SELECT {col} AS person" + (", source" if src_col else "") + f" FROM {table} " \
                  f"WHERE COALESCE({col},'') <> ''"
            for r in conn.execute(sql).fetchall():
                raw = str(r["person"]).strip()
                source = str(r["source"]) if src_col else ""
                # 空白分隔的多人（蔡維庭 黎世偉 吳季凌 游穗宗）拆開，但標記讓人確認
                spaced = len(raw.split()) > 1 and all(2 <= len(p) <= 4 for p in raw.split())
                names = raw.split() if spaced else _split_persons(raw)
                for name in names:
                    if not name or name in existing or name in accounts:
                        continue
                    slot = found.setdefault(name, {
                        "name": name, "count": 0, "sources": [], "from": set(),
                        "suspect": "", "raw_samples": set(), "clean_hit": False})
                    slot["count"] += 1
                    slot["from"].add(label)
                    slot["raw_samples"].add(raw)
                    if source:
                        slot["sources"].append(source)
                    if spaced:
                        slot["suspect"] = "原本一格塞了多個人（用空白分隔），系統拆開了，請確認拆得對不對"
                    elif not _looks_like_person(name):
                        slot["suspect"] = "看起來不像人名（太長或含「待確認」這類字眼），可能是誤填"
                    else:
                        # 這個人也單獨出現在別的地方＝名字本身沒問題，不該因為某一筆髒資料被連坐
                        slot["clean_hit"] = True
    out = []
    for slot in found.values():
        # 只要這個人有一次是乾淨地單獨出現，就不算可疑——不然一筆髒資料會連坐到
        # 好幾個正常的人（實例：「蔡維庭 黎世偉 吳季凌 游穗宗」害游穗宗也被標）
        suspect = "" if slot["clean_hit"] else slot["suspect"]
        out.append({
            "name": slot["name"], "count": slot["count"],
            "group_name": _guess_group(slot["sources"]),
            "from": sorted(slot["from"]),
            "suspect": suspect,
            "raw_sample": sorted(slot["raw_samples"], key=len)[-1],   # 秀最長的那個，看得出問題在哪
            "recommend": not suspect,       # 可疑的不預設勾選
        })
    out.sort(key=lambda x: (bool(x["suspect"]), -x["count"], x["name"]))
    return {"candidates": out, "count": len(out),
            "recommended": sum(1 for x in out if x["recommend"]),
            "note": "組別是從專案來源工作表（例：網路組處級專案）推的，推不出來就留空，可事後補。"}


def create_personnel_from_data(names: list[str], group_overrides: dict[str, str] | None = None,
                               email_overrides: dict[str, str] | None = None) -> dict[str, Any]:
    """把選定的名字建進人員主檔。已存在的跳過（冪等），組別與 email 可個別填。

    email 一併收：助理 2026-08-13 反映人員＋組別＋EMAIL 沒填好就沒辦法繼續測，
    建完還要再一個個補 email 等於白做一半。
    """
    picked = [str(n).strip() for n in (names or []) if str(n).strip()]
    if not picked:
        raise ValueError("沒有選到任何人。")
    suggested = {c["name"]: c for c in suggest_personnel_from_data()["candidates"]}
    overrides = group_overrides or {}
    created, skipped = [], []
    with connect() as conn:
        existing = {str(r["name"]).strip() for r in conn.execute(
            "SELECT name FROM personnel_master").fetchall()}
        for name in picked:
            if name in existing:
                skipped.append(name)
                continue
            group = overrides.get(name) or (suggested.get(name, {}) or {}).get("group_name", "")
            email = (email_overrides or {}).get(name, "").strip()
            # personnel_master 不走 allowed_fields，比照 create_personnel_master 直接寫入
            cur = conn.execute(
                "INSERT INTO personnel_master (name, group_name, note, email) VALUES (?, ?, ?, ?)",
                (name, group, "由既有資料自動補登記", email))
            row = get_row(conn, "personnel_master", cur.lastrowid)
            write_audit_log(conn, "personnel_master", cur.lastrowid, "create", None, row)
            existing.add(name)
            created.append({"name": name, "group_name": group, "email": email, "id": cur.lastrowid})
    return {"created": created, "skipped": skipped,
            "created_count": len(created), "skipped_count": len(skipped)}


def handover_preview(from_name: str, from_username: str = "", include_closed: bool = False) -> dict[str, Any]:
    """交接前先看會動到哪幾筆、哪幾筆不動。按下去才知道動到誰，那是最糟的設計。"""
    w = personnel_workload(from_name, from_username)
    will, keep = [], []
    for b in w["blocks"]:
        n = b["total"] if include_closed else b["active"]
        if n:
            will.append({"label": b["label"], "count": n, "table": b["table"]})
        skipped = 0 if include_closed else b["closed"]
        if skipped:
            keep.append({"label": b["label"], "count": skipped, "table": b["table"]})
    return {
        "from_name": from_name, "from_username": from_username,
        "include_closed": include_closed,
        "will_transfer": will, "will_keep": keep,
        "transfer_count": sum(x["count"] for x in will),
        "keep_count": sum(x["count"] for x in keep),
        "keep_reason": "已結案／已停用的維持原承辦：那是歷史事實（這案子當初誰做的），"
                       "而且結案的也不會再產生待辦。要一起轉請勾「連已結案的一起轉」。",
    }


def handover_apply(from_name: str, to_name: str, from_username: str = "", to_username: str = "",
                   include_closed: bool = False, reason: str = "") -> dict[str, Any]:
    """把某人名下的資料整批轉給接手人。單一交易：全部成功才算數，中途失敗整批回滾。

    共同負責人只換自己那一份（「陳昱杉/洪似妮」轉走陳昱杉 → 變成「接手人/洪似妮」），
    不會把另一個人也一起換掉。
    """
    from_name, to_name = str(from_name or "").strip(), str(to_name or "").strip()
    if not from_name or not to_name:
        raise ValueError("請指定離職者與接手人。")
    if from_name == to_name:
        raise ValueError("離職者與接手人是同一個人。")
    actor = _current_actor.get()
    moved: list[dict[str, Any]] = []
    with connect() as conn:
        existing = {r["name"] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'")}
        for table, col, label, kind, closed_states in PERSON_FIELDS:
            if table not in existing:
                continue
            key = from_username if kind == "account" else from_name
            new_key = to_username if kind == "account" else to_name
            if not key or not new_key:
                continue
            rows = [dict(r) for r in conn.execute(
                f"SELECT * FROM {table} WHERE COALESCE({col},'') <> ''").fetchall()]
            count = 0
            for r in rows:
                if kind == "account":
                    if str(r[col]).strip() != key:
                        continue
                    new_value = new_key
                else:
                    if not _name_matches(r[col], key):
                        continue
                    new_value = _replace_person(str(r[col]), key, new_key)
                if not include_closed and str(r.get("status") or "") in closed_states:
                    continue
                before = dict(r)
                conn.execute(f"UPDATE {table} SET {col} = ? WHERE id = ?", (new_value, r["id"]))
                after = get_row(conn, table, r["id"])
                write_audit_log(conn, table, r["id"], "handover", before, {
                    **dict(after), "handover_from": key, "handover_to": new_key,
                    "handover_by": actor, "handover_reason": reason})
                count += 1
            if count:
                moved.append({"label": label, "table": table, "count": count})
    return {"from_name": from_name, "to_name": to_name, "include_closed": include_closed,
            "moved": moved, "moved_count": sum(m["count"] for m in moved), "actor": actor}


def _replace_person(cell: str, old: str, new: str) -> str:
    """把共同負責人字串裡的某一個人換掉，其他人與原本的分隔符維持不變。"""
    out, buf, result = [], "", []
    for ch in cell:
        if ch in _NAME_SEPARATORS:
            result.append(buf)
            result.append(ch)
            buf = ""
        else:
            buf += ch
    result.append(buf)
    for token in result:
        out.append(new if token.strip() == old else token)
    return "".join(out)


# ── 跨模組串接：案件／專案／預算其實在講同一件事，但名字都不一樣 ──────────────
# 使用者 2026-08-12 舉的實例：
#   案件「青浦機房搬遷」／專案「青浦機房搬遷專案」／預算「桃園青浦機房」
# 既有的名稱歸納是「同類別內」合併別名（中華電信＝中華電信有限公司），
# 這裡是「跨類別」找出它們指的是同一件事，把專案／預算歸戶到同一個案件底下。
#
# 純字串演算法，離線可跑（使用者明講不能依賴 AI，怕以後功能失效）：
# 取最長共同片段，長度與佔比都要過門檻才算候選，而且**一律由人裁決**——
# 「桃園機房搬遷」與「青浦機房搬遷」共用「機房搬遷」，自動併就把兩個案子併成一個了。
_NAME_NOISE = ("專案", "計畫", "案件", "作業", "系統", "費用", "採購", "建置", "案")


def _name_core(name: Any) -> str:
    """比對前先把常見尾綴與空白拿掉，讓「青浦機房搬遷」和「青浦機房搬遷專案」對得上。"""
    s = "".join(str(name or "").split())
    for _ in range(3):                      # 「…專案計畫」這種疊字尾綴，剝幾層
        for tail in _NAME_NOISE:
            if len(s) > len(tail) + 2 and s.endswith(tail):
                s = s[: -len(tail)]
                break
        else:
            break
    return s


def longest_common_part(a: Any, b: Any) -> str:
    """兩個名稱的最長共同片段（連續字元）。中文沒有空白斷詞，用這個最直觀，
    而且算出來的東西可以直接顯示給人看——「因為都有『青浦機房』」比「相似度 0.72」好懂。"""
    x, y = _name_core(a), _name_core(b)
    if not x or not y:
        return ""
    best_len, best_end = 0, 0
    prev = [0] * (len(y) + 1)
    for i in range(1, len(x) + 1):
        cur = [0] * (len(y) + 1)
        for j in range(1, len(y) + 1):
            if x[i - 1] == y[j - 1]:
                cur[j] = prev[j - 1] + 1
                if cur[j] > best_len:
                    best_len, best_end = cur[j], i
        prev = cur
    return x[best_end - best_len: best_end]


def names_look_related(a: Any, b: Any, min_len: int = 3, min_ratio: float = 0.5) -> tuple[bool, str]:
    """兩個名稱像不像同一件事 →（像不像, 共同片段）。

    門檻刻意保守：共同片段至少 3 個字，而且要佔短名稱一半以上。
    寧可漏掉幾個讓人手動歸戶，也不要把「桃園機房搬遷」跟「青浦機房搬遷」配在一起——
    併錯比沒併更難救（資料已經掛過去了）。
    """
    part = longest_common_part(a, b)
    if len(part) < min_len:
        return False, part
    shorter = min(len(_name_core(a)), len(_name_core(b))) or 1
    return (len(part) / shorter) >= min_ratio, part


def cross_kind_link_candidates(limit: int = 200) -> dict[str, Any]:
    """找出「名字不同但可能是同一件事」的專案／預算，建議歸戶到哪個案件。

    只看還沒掛案件、或掛到別的案件的資料；已經在同一個案件底下的不重複提示。
    """
    with connect() as conn:
        cases = [dict(r) for r in conn.execute(
            "SELECT id, case_code, title FROM cases WHERE status != 'disabled'").fetchall()]
        projects = [dict(r) for r in conn.execute(
            "SELECT id, project_code AS code, project_name AS name, case_id FROM projects "
            "WHERE status != 'disabled'").fetchall()]
        budgets = [dict(r) for r in conn.execute(
            "SELECT id, budget_code AS code, budget_code AS name, case_id FROM budgets "
            "WHERE status != 'disabled'").fetchall()]
    case_title = {c["id"]: c["title"] for c in cases}
    out: list[dict[str, Any]] = []
    for kind, rows in (("project", projects), ("budget", budgets)):
        for r in rows:
            for c in cases:
                if r["case_id"] and int(r["case_id"]) == int(c["id"]):
                    continue                       # 已經掛在這個案件下了
                related, part = names_look_related(r["name"], c["title"])
                if not related:
                    continue
                out.append({
                    "kind": kind, "id": r["id"], "code": r["code"], "name": r["name"],
                    "current_case_id": r["case_id"],
                    "current_case_title": case_title.get(r["case_id"]) if r["case_id"] else None,
                    "suggest_case_id": c["id"], "suggest_case_code": c["case_code"],
                    "suggest_case_title": c["title"], "common_part": part,
                })
    # 同一筆資料可能對到多個案件：共同片段長的排前面，讓人先看最像的那個
    out.sort(key=lambda x: (-len(x["common_part"]), x["kind"], str(x["name"])))
    return {"candidates": out[:limit], "total": len(out),
            "note": "共同片段是純字串比對算出來的（不連網、不用 AI）；要不要歸戶由你決定。"}


def apply_cross_kind_link(kind: str, row_id: int, case_id: int) -> dict[str, Any]:
    """把某筆專案／預算歸戶到指定案件（使用者裁決後才會走到這裡）。"""
    table = {"project": "projects", "budget": "budgets"}.get(kind)
    if not table:
        raise ValueError(f"只支援專案與預算的跨模組歸戶，收到：{kind}")
    with connect() as conn:
        before = get_row(conn, table, row_id)
        get_row(conn, "cases", case_id)              # 案件不存在會 raise LookupError
        conn.execute(f"UPDATE {table} SET case_id = ? WHERE id = ?", (case_id, row_id))
        after = get_row(conn, table, row_id)
        write_audit_log(conn, table, row_id, "cross-link", before, after)
    return {"kind": kind, "id": row_id, "case_id": case_id, "row": dict(after)}


def _name_alias_map(conn, kind: str) -> dict[str, str]:
    """{別名 → 主名(canonical)}（限某 kind）。"""
    rows = conn.execute(
        "SELECT a.alias_name, m.canonical_name FROM name_aliases a JOIN name_master m ON m.id = a.master_id "
        "WHERE a.kind = ?", (kind,)).fetchall()
    return {str(r["alias_name"]): str(r["canonical_name"]) for r in rows}


def list_name_values(kind: str) -> dict[str, Any]:
    """回某 kind 目前所有不同名稱（跨來源表去重、計數），並附它目前歸到的主名（若已裁決）。
    供前端做相似度分群、裁決合併。"""
    if kind not in NAME_SOURCES:
        raise ValueError(f"未知的名稱種類：{kind}")
    counts: dict[str, int] = {}
    with connect() as conn:
        for table, col in NAME_SOURCES[kind]:
            for r in conn.execute(
                f"SELECT COALESCE({col},'') AS v, COUNT(*) AS n FROM {table} GROUP BY v").fetchall():
                name = str(r["v"]).strip()
                if name:
                    counts[name] = counts.get(name, 0) + int(r["n"])
        amap = _name_alias_map(conn, kind)
    values = [{"name": n, "count": c, "canonical": amap.get(n)} for n, c in counts.items()]
    values.sort(key=lambda x: (-x["count"], x["name"]))
    return {"kind": kind, "values": values, "resolved": sum(1 for v in values if v["canonical"])}


def _record_name_decision(conn, kind: str, action: str, reason: str, detail: dict, undo_ops: list) -> int:
    cur = conn.execute(
        "INSERT INTO name_decisions (kind, action, reason, actor, detail_json, undo_ops_json) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        (kind, action, (reason or "").strip(), _current_actor.get(),
         json.dumps(detail, ensure_ascii=False), json.dumps(undo_ops, ensure_ascii=False)))
    return cur.lastrowid


def _attach_name_alias(conn, master_id: int, kind: str, alias_name: str) -> int | None:
    name = (alias_name or "").strip()
    existing = conn.execute(
        "SELECT id, master_id FROM name_aliases WHERE kind = ? AND alias_name = ?", (kind, name)).fetchone()
    if existing:
        prev = existing["master_id"]
        conn.execute("UPDATE name_aliases SET master_id = ? WHERE id = ?", (master_id, existing["id"]))
        return prev
    conn.execute("INSERT INTO name_aliases (master_id, kind, alias_name) VALUES (?, ?, ?)", (master_id, kind, name))
    return None


def _find_or_create_name_master(conn, kind: str, canonical_name: str) -> int:
    name = (canonical_name or "").strip()
    row = conn.execute(
        "SELECT id FROM name_master WHERE kind = ? AND canonical_name = ?", (kind, name)).fetchone()
    if row:
        return row["id"]
    cur = conn.execute("INSERT INTO name_master (kind, canonical_name) VALUES (?, ?)", (kind, name))
    return cur.lastrowid


def _cleanup_empty_name_masters(conn) -> int:
    rows = conn.execute(
        "SELECT id FROM name_master WHERE id NOT IN (SELECT DISTINCT master_id FROM name_aliases)").fetchall()
    for r in rows:
        conn.execute("DELETE FROM name_master WHERE id = ?", (r["id"],))
    return len(rows)


def merge_names(kind: str, names: list[str], canonical_name: str, reason: str = "") -> dict[str, Any]:
    """把這些名稱視為同一實體，以 canonical_name 為準（其餘掛成別名）。非破壞式、可復原。"""
    if kind not in NAME_SOURCES:
        raise ValueError(f"未知的名稱種類：{kind}")
    names = [str(n).strip() for n in (names or []) if str(n).strip()]
    if not names:
        raise ValueError("沒有要合併的名稱。")
    if not (reason or "").strip():
        raise ValueError("請填『為什麼這樣判斷』的理由，才能裁決。")
    cname = (canonical_name or "").strip()
    if not cname:
        raise ValueError("請指定要以哪個名稱為準。")
    with connect() as conn:
        master_id = _find_or_create_name_master(conn, kind, cname)
        undo_ops = []
        for n in names:
            prev = _attach_name_alias(conn, master_id, kind, n)
            undo_ops.append({"alias_name": n, "prev_master_id": prev})
        _cleanup_empty_name_masters(conn)
        did = _record_name_decision(conn, kind, "merge", reason,
                                    {"canonical_name": cname, "names": names}, undo_ops)
    return {"master_id": master_id, "merged": len(names), "decision_id": did, "canonical_name": cname}


def split_names(kind: str, names: list[str], reason: str = "") -> dict[str, Any]:
    """這些名稱各自是不同實體（各自成主名）。裁決後不再列為待確認。"""
    if kind not in NAME_SOURCES:
        raise ValueError(f"未知的名稱種類：{kind}")
    names = [str(n).strip() for n in (names or []) if str(n).strip()]
    if not names:
        raise ValueError("沒有要分開的名稱。")
    if not (reason or "").strip():
        raise ValueError("請填『為什麼這樣判斷』的理由，才能裁決。")
    with connect() as conn:
        undo_ops = []
        for n in names:
            mid = _find_or_create_name_master(conn, kind, n)
            prev = _attach_name_alias(conn, mid, kind, n)
            undo_ops.append({"alias_name": n, "prev_master_id": prev})
        _cleanup_empty_name_masters(conn)
        did = _record_name_decision(conn, kind, "split", reason, {"names": names}, undo_ops)
    return {"split": len(names), "decision_id": did}


def list_name_decisions(kind: str | None = None, limit: int = 100) -> dict[str, Any]:
    where = "WHERE kind = ?" if kind else ""
    params = ([kind] if kind else []) + [limit]
    with connect() as conn:
        rows = [dict(r) for r in conn.execute(
            f"SELECT id, kind, action, reason, actor, detail_json, undone, created_at "
            f"FROM name_decisions {where} ORDER BY id DESC LIMIT ?", params).fetchall()]
    out = []
    for r in rows:
        try:
            detail = json.loads(r.get("detail_json") or "{}")
        except (ValueError, TypeError):
            detail = {}
        out.append({"id": r["id"], "kind": r["kind"], "action": r["action"], "reason": r["reason"],
                    "actor": r["actor"], "undone": bool(r["undone"]), "created_at": r["created_at"],
                    "canonical_name": detail.get("canonical_name", ""), "names": detail.get("names", [])})
    return {"decisions": out, "count": len(out)}


def undo_name_decision(decision_id: int) -> dict[str, Any]:
    with connect() as conn:
        row = conn.execute("SELECT undo_ops_json, undone FROM name_decisions WHERE id = ?", (decision_id,)).fetchone()
        if not row:
            raise ValueError("找不到該筆裁決紀錄。")
        if row["undone"]:
            raise ValueError("這筆裁決已經復原過了。")
        try:
            ops = json.loads(row["undo_ops_json"] or "[]")
        except (ValueError, TypeError):
            ops = []
        d = conn.execute("SELECT kind FROM name_decisions WHERE id = ?", (decision_id,)).fetchone()
        kind = d["kind"] if d else ""
        for op in ops:
            name = str(op.get("alias_name", "")).strip()
            prev = op.get("prev_master_id")
            if prev is None:
                conn.execute("DELETE FROM name_aliases WHERE kind = ? AND alias_name = ?", (kind, name))
            else:
                conn.execute("UPDATE name_aliases SET master_id = ? WHERE kind = ? AND alias_name = ?",
                             (prev, kind, name))
        removed = _cleanup_empty_name_masters(conn)
        conn.execute("UPDATE name_decisions SET undone = 1 WHERE id = ?", (decision_id,))
    return {"undone": decision_id, "removed_masters": removed}


def reset_name_decisions(kind: str | None = None) -> dict[str, Any]:
    """一鍵還原某 kind（或全部）的名稱裁決。原始資料本就沒被動過。"""
    with connect() as conn:
        if kind:
            mids = [r["id"] for r in conn.execute("SELECT id FROM name_master WHERE kind = ?", (kind,)).fetchall()]
            n_alias = conn.execute("SELECT COUNT(*) AS n FROM name_aliases WHERE kind = ?", (kind,)).fetchone()["n"]
            conn.execute("DELETE FROM name_aliases WHERE kind = ?", (kind,))
            for mid in mids:
                conn.execute("DELETE FROM name_master WHERE id = ?", (mid,))
            conn.execute("UPDATE name_decisions SET undone = 1 WHERE kind = ? AND undone = 0", (kind,))
            return {"kind": kind, "removed_aliases": n_alias, "removed_masters": len(mids)}
        n_alias = conn.execute("SELECT COUNT(*) AS n FROM name_aliases").fetchone()["n"]
        n_master = conn.execute("SELECT COUNT(*) AS n FROM name_master").fetchone()["n"]
        conn.execute("DELETE FROM name_aliases")
        conn.execute("DELETE FROM name_master")
        conn.execute("UPDATE name_decisions SET undone = 1 WHERE undone = 0")
        return {"removed_aliases": n_alias, "removed_masters": n_master}


def resolve_name(kind: str, name: str) -> str:
    """把一個名稱解析成它的主名（未裁決則原樣回傳）。供之後報表/彙總用。"""
    with connect() as conn:
        amap = _name_alias_map(conn, kind)
    return amap.get(str(name).strip(), name)


def parse_headcount_xlsx(data: bytes) -> list[dict[str, Any]]:
    """解析『費用分攤表（人數）』.xlsx → 人數基準。認表頭『人數』欄那一列，讀 代號/部門/人數。"""
    import io
    import openpyxl

    def norm(v: Any) -> str:
        return " ".join(str(v).split()) if v is not None else ""

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out: list[dict[str, Any]] = []
    try:
        for sheet in wb.sheetnames:
            rows = [list(r) if r else [] for r in wb[sheet].iter_rows(values_only=True)]
            hdr = code_c = name_c = hc_c = None
            for i, r in enumerate(rows[:8]):
                if any(norm(c) == "人數" for c in r):
                    hdr = i
                    for j, c in enumerate(r):
                        k = norm(c)
                        if k in ("代號", "部門代號", "單位代碼"):
                            code_c = j
                        elif k in ("部門", "部門別", "單位", "單位名稱"):
                            name_c = j
                        elif k == "人數":
                            hc_c = j
                    break
            if hdr is None or hc_c is None:
                continue
            for r in rows[hdr + 1:]:
                name = norm(r[name_c]) if (name_c is not None and name_c < len(r)) else ""
                if name in ("合計", "小計"):
                    break
                if not name:
                    continue
                code = norm(r[code_c]) if (code_c is not None and code_c < len(r)) else ""
                try:
                    hc = int(float(r[hc_c])) if (hc_c < len(r) and r[hc_c] not in (None, "")) else 0
                except (TypeError, ValueError):
                    hc = 0
                out.append({"unit_code": code, "unit_name": name, "headcount": hc})
    finally:
        wb.close()
    return out


def commit_headcounts_import(records: list[dict[str, Any]], source_file: str = "") -> dict[str, Any]:
    """寫入人數基準：以 unit_code 為鍵 upsert（無代號者以 unit_name 為鍵）。
    source_file：來源 Excel 檔名，寫進每筆，供單位撞名清單指回來源。"""
    allowed = allowed_fields()["unit_headcounts"]
    source_file = (source_file or "").strip()
    with connect() as conn:
        existing: dict[str, int] = {}
        for r in conn.execute("SELECT id, unit_code, unit_name FROM unit_headcounts").fetchall():
            existing[(r["unit_code"] or "").strip() or ("＃" + (r["unit_name"] or ""))] = r["id"]
        created = updated = 0
        for rec in records:
            fields = {k: v for k, v in rec.items() if k in allowed}
            if source_file:
                fields["source_file"] = source_file
            key = str(rec.get("unit_code", "")).strip() or ("＃" + str(rec.get("unit_name", "")))
            if key in existing:
                rid = existing[key]
                upd = {k: v for k, v in fields.items() if k != "unit_code"}
                if upd:
                    conn.execute(f"UPDATE unit_headcounts SET {', '.join(f'{k} = ?' for k in upd)} WHERE id = ?",
                                 [*upd.values(), rid])
                updated += 1
            else:
                cols = ", ".join(fields)
                ph = ", ".join("?" for _ in fields)
                cur = conn.execute(f"INSERT INTO unit_headcounts ({cols}) VALUES ({ph})", list(fields.values()))
                existing[key] = cur.lastrowid
                created += 1
        return {"created_count": created, "updated_count": updated}


def list_headcounts() -> list[dict[str, Any]]:
    with connect() as conn:
        return [dict(r) for r in conn.execute(
            "SELECT * FROM unit_headcounts ORDER BY headcount DESC, id ASC").fetchall()]


def parse_category_shares_xlsx(data: bytes) -> list[dict[str, Any]]:
    """解析『資訊架構部費用分攤表』的『對照』表 → 類別基準。
    表頭第一列是類別（台股功能/複委託功能/台、複共用功能…），第二列標(現行)/(NEW)；
    只取 NEW 欄。每個有百分比的儲存格 → (類別, 代號, 名稱, 百分比)。"""
    import io
    import openpyxl

    def norm(v: Any) -> str:
        return " ".join(str(v).split()) if v is not None else ""

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out: list[dict[str, Any]] = []
    try:
        # 找「對照」表；找不到就退回最後一張
        sheet = next((s for s in wb.sheetnames if "對照" in s), wb.sheetnames[-1])
        rows = [list(r) if r else [] for r in wb[sheet].iter_rows(values_only=True)]
        if len(rows) < 3:
            return out
        hdr_cat, hdr_ver = rows[0], rows[1]
        width = max(len(hdr_cat), len(hdr_ver))
        # 類別名稱往右填滿（合併儲存格只有左上有值）
        cats, last = [], ""
        for j in range(width):
            v = norm(hdr_cat[j]) if j < len(hdr_cat) else ""
            if v:
                last = v
            cats.append(last)
        # 取 NEW 欄
        new_cols = [j for j in range(2, width)
                    if j < len(hdr_ver) and "NEW" in norm(hdr_ver[j]).upper()]
        for r in rows[2:]:
            code = norm(r[0]) if len(r) > 0 else ""
            name = norm(r[1]) if len(r) > 1 else ""
            if not code and not name:
                continue
            if code in ("合計", "小計") or name in ("合計", "小計"):
                break
            for j in new_cols:
                if j >= len(r) or r[j] in (None, ""):
                    continue
                try:
                    pct = float(r[j])
                except (TypeError, ValueError):
                    continue
                if pct <= 0:
                    continue
                out.append({"category": cats[j], "unit_code": code, "unit_name": name,
                            "share_pct": round(pct * 100, 4)})
    finally:
        wb.close()
    return out


def commit_category_shares_import(records: list[dict[str, Any]], source_file: str = "") -> dict[str, Any]:
    """寫入類別基準：以 (類別, 代號, 名稱) 為鍵 upsert。整批重匯前先清掉同來源舊資料避免殘留。"""
    allowed = allowed_fields()["category_shares"]
    source_file = (source_file or "").strip()
    with connect() as conn:
        existing = {(r["category"], r["unit_code"], r["unit_name"]): r["id"] for r in conn.execute(
            "SELECT id, category, unit_code, unit_name FROM category_shares").fetchall()}
        written = 0
        cats: set[str] = set()
        for rec in records:
            fields = {k: v for k, v in rec.items() if k in allowed}
            if source_file:
                fields["source_file"] = source_file
            key = (rec.get("category", ""), rec.get("unit_code", ""), rec.get("unit_name", ""))
            cats.add(rec.get("category", ""))
            if key in existing:
                upd = {k: v for k, v in fields.items() if k not in ("category", "unit_code", "unit_name")}
                if upd:
                    conn.execute(f"UPDATE category_shares SET {', '.join(f'{k} = ?' for k in upd)} WHERE id = ?",
                                 [*upd.values(), existing[key]])
            else:
                cols = ", ".join(fields)
                ph = ", ".join("?" for _ in fields)
                conn.execute(f"INSERT INTO category_shares ({cols}) VALUES ({ph})", list(fields.values()))
            written += 1
        return {"written": written, "categories": sorted(c for c in cats if c)}


def list_category_shares(category: str | None = None) -> dict[str, Any]:
    """類別基準：回各類別清單（含單位數、百分比合計），帶 category 則回該類別各單位明細。"""
    with connect() as conn:
        cats = [dict(r) for r in conn.execute(
            "SELECT category, COUNT(*) AS units, ROUND(SUM(share_pct), 2) AS pct_sum "
            "FROM category_shares GROUP BY category ORDER BY category").fetchall()]
        result: dict[str, Any] = {"categories": cats}
        if category is not None:
            result["shares"] = [dict(r) for r in conn.execute(
                "SELECT unit_code, unit_name, share_pct FROM category_shares "
                "WHERE category = ? ORDER BY share_pct DESC", (category,)).fetchall()]
        return result


def compute_budget_allocations(budget_id: int) -> dict[str, Any]:
    """依預算的 alloc_method 重算分攤並寫入 budget_allocations。
    headcount：amount = 費用 × 該單位人數 ÷ 總人數。
    category：amount = 費用 × 該類別下該單位%（整數化＋尾數承擔在 list_budget_allocations 處理）。"""
    with connect() as conn:
        budget = get_row(conn, "budgets", budget_id)
        method = str(budget.get("alloc_method") or "fixed")
        total = float(budget.get("amount") or 0)
        if method == "fixed":
            return {"method": "fixed", "written": 0, "note": "固定金額：沿用現有分攤，未重算"}
        if method == "category":
            cat = str(budget.get("alloc_category") or "").strip()
            if not cat:
                raise ValueError("請先選一個分攤類別（台股功能／複委託功能／台、複共用功能…）。")
            shares = conn.execute(
                "SELECT unit_code, unit_name, share_pct FROM category_shares "
                "WHERE category = ? AND share_pct > 0 ORDER BY share_pct DESC", (cat,)).fetchall()
            if not shares:
                raise ValueError(f"類別「{cat}」在基準表裡查不到資料，請先匯入類別基準（對照表）。")
            conn.execute("DELETE FROM budget_allocations WHERE budget_id = ?", (budget_id,))
            written = 0
            for seq, s in enumerate(shares, start=1):
                pct = float(s["share_pct"])
                conn.execute(
                    "INSERT INTO budget_allocations (budget_id, seq, unit_code, unit_name, share_pct, amount) "
                    "VALUES (?, ?, ?, ?, ?, ?)",
                    (budget_id, seq, s["unit_code"], s["unit_name"], round(pct, 2), round(total * pct / 100, 2)))
                written += 1
            write_audit_log(conn, "budgets", budget_id, "recompute-category", None,
                            {"category": cat, "units": written, "total": total})
            return {"method": "category", "written": written, "category": cat}
        if method != "headcount":
            raise ValueError(f"未知分攤方法：{method}")
        hcs = conn.execute(
            "SELECT unit_code, unit_name, headcount FROM unit_headcounts WHERE headcount > 0 "
            "ORDER BY headcount DESC, id ASC").fetchall()
        total_hc = sum(int(h["headcount"]) for h in hcs)
        if total_hc <= 0:
            raise ValueError("人數基準表是空的或總人數為 0，請先匯入人數表。")
        conn.execute("DELETE FROM budget_allocations WHERE budget_id = ?", (budget_id,))
        written = 0
        for seq, h in enumerate(hcs, start=1):
            hc = int(h["headcount"])
            conn.execute(
                "INSERT INTO budget_allocations (budget_id, seq, unit_code, unit_name, share_pct, amount) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (budget_id, seq, h["unit_code"], h["unit_name"],
                 round(hc / total_hc * 100, 2), round(total * hc / total_hc, 2)))
            written += 1
        write_audit_log(conn, "budgets", budget_id, "recompute-headcount", None,
                        {"units": written, "total": total, "total_headcount": total_hc})
        return {"method": "headcount", "written": written, "total_headcount": total_hc}


def preflight_import_batch_confirm(
    batch_id: int,
    confirmed_fields: list[dict[str, Any]],
    accepted_warning_codes: list[str],
) -> dict[str, Any]:
    with connect() as conn:
        batch = get_row(conn, "import_batches", batch_id)
        rows = conn.execute(
            "SELECT * FROM import_rows WHERE batch_id = ? ORDER BY row_number ASC, id ASC",
            (batch_id,),
        ).fetchall()
        preview = mapping_preview(batch, rows)
        existing_case_codes = {
            str(row["case_code"]).strip()
            for row in conn.execute("SELECT case_code FROM cases").fetchall()
        }
        return confirm_preflight_report(
            preview,
            confirmed_fields,
            accepted_warning_codes,
            existing_case_codes,
        )


def write_audit_log(
    conn: sqlite3.Connection,
    table: str,
    row_id: int,
    action: str,
    before: dict[str, Any] | None,
    after: dict[str, Any] | None,
    actor: str | None = None,
) -> None:
    if actor is None:
        actor = _current_actor.get()
    conn.execute(
        "INSERT INTO audit_logs (table_name, row_id, action, before_json, after_json, actor) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        (
            table,
            row_id,
            action,
            json.dumps(before, ensure_ascii=False, sort_keys=True) if before is not None else None,
            json.dumps(after, ensure_ascii=False, sort_keys=True) if after is not None else None,
            actor,
        ),
    )


def list_audit_logs(
    limit: int = 100,
    table_name: str | None = None,
    row_id: int | None = None,
    action: str | None = None,
) -> list[dict[str, Any]]:
    clauses: list[str] = []
    params: list[Any] = []
    if table_name:
        clauses.append("table_name = ?")
        params.append(table_name)
    if row_id is not None:
        clauses.append("row_id = ?")
        params.append(row_id)
    if action:
        clauses.append("action = ?")
        params.append(action)
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    params.append(max(1, min(limit, 500)))
    with connect() as conn:
        return conn.execute(
            f"SELECT * FROM audit_logs {where} ORDER BY id DESC LIMIT ?",
            params,
        ).fetchall()


def dashboard_summary() -> dict[str, Any]:
    scope = _owner_scope.get()

    def _clause(table: str) -> tuple[str, list[Any]]:
        if scope is None:
            return "", []
        where, params = _scope_where(table, scope)
        return (f" WHERE {where}" if where else ""), params

    with connect() as conn:
        counts = {}
        for table in ("cases", "contracts", "payments", "documents"):
            clause, params = _clause(table)
            counts[table] = conn.execute(
                f"SELECT COUNT(*) AS count FROM {table}{clause}", params
            ).fetchone()["count"]
        c_clause, c_params = _clause("contracts")
        money = conn.execute(
            f"SELECT COALESCE(SUM(amount), 0) AS contract_amount FROM contracts{c_clause}", c_params
        ).fetchone()
        p_where, p_params = _scope_where("payments", scope) if scope is not None else ("", [])
        due_sql = (
            "SELECT COALESCE(SUM(payment_amount), 0) AS pending_payment_amount "
            "FROM payments WHERE status <> 'closed'"
        )
        if p_where:
            due_sql += f" AND {p_where}"
        due = conn.execute(due_sql, p_params).fetchone()
        return {
            "counts": counts,
            "contract_amount": money["contract_amount"],
            "pending_payment_amount": due["pending_payment_amount"],
        }


def manager_focus() -> dict[str, Any]:
    """主管每天真正要看的三件事（助理 2026-08-03 回饋：儀表板不要放各模組的統計報表）：
    本月新成立的案子、已經出事的（合約逾期＋專案延遲）、下個月要付多少錢。
    每一項都同時回數字與明細，點卡片就能往下看，不用再切到別的模組去找。
    依 owner 範圍過濾：承辦看自己的、組長看本組、部長看全部。
    """
    scope = _owner_scope.get()
    today = date.today()
    this_month = today.strftime("%Y-%m")
    next_month = f"{today.year + 1}-01" if today.month == 12 else f"{today.year}-{today.month + 1:02d}"
    iso_today = today.isoformat()

    def _tail(table: str) -> tuple[str, list[Any]]:
        if scope is None:
            return "", []
        where, params = _scope_where(table, scope)
        return (f" AND {where}" if where else ""), params

    with connect() as conn:
        # 一、本月新成立：核准當下才算成立，所以看 approved_at；沒有 approved_at 的舊資料退回看建立時間
        ct, cp = _tail("cases")
        new_cases = conn.execute(
            "SELECT id, case_code, title, owner, amount, fiscal_year, seq, "
            "COALESCE(NULLIF(approved_at,''), created_at) AS established_at FROM cases "
            "WHERE status NOT IN ('draft','pending_review','returned','rejected','merged','disabled') "
            f"AND substr(COALESCE(NULLIF(approved_at,''), created_at), 1, 7) = ?{ct} "
            "ORDER BY established_at DESC", [this_month, *cp]).fetchall()

        # 二、已經出事的：合約過了到期日還在生效、專案過了結束日還沒完成
        kt, kp = _tail("contracts")
        overdue_contracts = conn.execute(
            "SELECT id, contract_code, contract_name, vendor_name, end_date, amount, case_id "
            "FROM contracts WHERE status <> 'disabled' AND COALESCE(end_date,'') <> '' "
            f"AND end_date < ?{kt} ORDER BY end_date", [iso_today, *kp]).fetchall()
        jt, jp = _tail("projects")
        delayed_projects = conn.execute(
            "SELECT id, project_code, project_name, owner, end_date, progress, rag_status, case_id "
            "FROM projects WHERE status <> 'disabled' AND COALESCE(end_date,'') <> '' "
            f"AND end_date < ? AND COALESCE(progress,0) < 100{jt} ORDER BY end_date", [iso_today, *jp]).fetchall()

        # 三、下個月要付的錢：只算已核准案件的付款（未複核的錢不讓主管當真，比照決策總覽）
        pt, pp = _tail("payments")
        due_rows = conn.execute(
            "SELECT p.id, p.payment_month, p.payment_amount, p.item, p.vendor, p.status, "
            "k.contract_code, k.contract_name FROM payments p "
            "LEFT JOIN contracts k ON k.id = p.contract_id "
            f"WHERE p.payment_month = ? AND {_APPROVED_PAYMENT_CLAUSE}{pt} "
            "ORDER BY p.payment_amount DESC", [next_month, *pp]).fetchall()

    to_dict = lambda rows: [dict(r) for r in rows]
    return {
        "this_month": this_month,
        "next_month": next_month,
        "new_cases": {"count": len(new_cases), "items": to_dict(new_cases)},
        "at_risk": {
            "count": len(overdue_contracts) + len(delayed_projects),
            "overdue_contracts": to_dict(overdue_contracts),
            "delayed_projects": to_dict(delayed_projects),
        },
        "next_month_payment": {
            "total": round(sum(float(r["payment_amount"] or 0) for r in due_rows), 2),
            "items": to_dict(due_rows),
        },
    }


def todo_cards() -> dict[str, Any]:
    """待辦事項的四個區塊（助理 2026-08-03 回饋：依角色顯示自己該關注的，不要每個角色看到同一份）。

    範圍靠既有的 owner scope 自動收斂：承辦只看自己負責的案件、組長看本組、部長看全部。
    這裡一律把四塊都算出來，「哪幾塊要顯示」由前端依角色決定（承辦沒有審核權，就不給他看待審核）。
    期限採助理指定的口徑：合約三個月內、WBS 兩週內、核銷看當月。
    """
    scope = _owner_scope.get()
    today = date.today()
    this_month = today.strftime("%Y-%m")
    in_3m = (today + timedelta(days=90)).isoformat()
    in_2w = (today + timedelta(days=14)).isoformat()
    iso_today = today.isoformat()

    def _tail(table: str, alias: str = "") -> tuple[str, list[Any]]:
        if scope is None:
            return "", []
        where, params = _scope_where(table, scope, alias)
        return (f" AND {where}" if where else ""), params

    with connect() as conn:
        ct, cp = _tail("cases")
        pending = conn.execute(
            "SELECT id, case_code, title, owner, amount, created_by, fiscal_year, temp_seq "
            f"FROM cases WHERE status = 'pending_review'{ct} ORDER BY created_at", cp).fetchall()
        new_approved = conn.execute(
            "SELECT id, case_code, title, owner, amount, fiscal_year, seq, "
            "COALESCE(NULLIF(approved_at,''), created_at) AS established_at FROM cases "
            "WHERE status NOT IN ('draft','pending_review','returned','rejected','merged','disabled') "
            f"AND substr(COALESCE(NULLIF(approved_at,''), created_at), 1, 7) = ?{ct} "
            "ORDER BY established_at DESC", [this_month, *cp]).fetchall()

        kt, kp = _tail("contracts")
        contracts = conn.execute(
            "SELECT id, contract_code, contract_name, vendor_name, end_date, amount, case_id "
            "FROM contracts WHERE status <> 'disabled' AND COALESCE(end_date,'') <> '' "
            f"AND end_date <= ?{kt} ORDER BY end_date", [in_3m, *kp]).fetchall()

        # WBS 工作項沒有自己的 case_id，經 projects 掛回案件，所以 scope 條件套在 projects 上
        jt, jp = _tail("projects", "p")
        wbs = conn.execute(
            "SELECT i.id, i.item_name, i.owner, i.end_date, i.progress, i.rag, i.sub_total, i.sub_done, "
            "p.id AS project_id, p.project_name, p.case_id FROM project_items i "
            "JOIN projects p ON p.id = i.project_id "
            "WHERE i.status <> 'disabled' AND p.status <> 'disabled' AND COALESCE(i.end_date,'') <> '' "
            f"AND i.end_date <= ? AND COALESCE(i.progress,0) < 100{jt} ORDER BY i.end_date", [in_2w, *jp]).fetchall()

        pt, pp = _tail("payments")
        settle = conn.execute(
            "SELECT p.id, p.payment_month, p.payment_amount, p.item, p.vendor, p.status, p.settle_no, "
            "k.contract_code FROM payments p LEFT JOIN contracts k ON k.id = p.contract_id "
            f"WHERE p.payment_month = ? AND p.status <> 'closed'{pt} "
            "ORDER BY p.payment_amount DESC", [this_month, *pp]).fetchall()

    d = lambda rows: [dict(r) for r in rows]
    overdue = [r for r in wbs if str(r["end_date"]) < iso_today]
    return {
        "this_month": this_month,
        "pending_review": {"count": len(pending), "items": d(pending)},
        "new_approved": {"count": len(new_approved), "items": d(new_approved)},
        "contracts_expiring": {"count": len(contracts), "items": d(contracts), "window": "三個月內"},
        "wbs_due": {"count": len(wbs), "overdue": len(overdue), "items": d(wbs), "window": "兩週內"},
        "settlements": {"count": len(settle), "items": d(settle),
                        "total": round(sum(float(r["payment_amount"] or 0) for r in settle), 2)},
    }


def monthly_spending_summary() -> list[dict[str, Any]]:
    """依月份彙總付款：每月總額、已付(closed)、待付(其餘)、筆數。依 owner 範圍過濾。"""
    scope = _owner_scope.get()
    where, params = _scope_where("payments", scope) if scope is not None else ("", [])
    sql = (
        "SELECT payment_month AS month, COUNT(*) AS count, "
        "COALESCE(SUM(payment_amount), 0) AS total, "
        "COALESCE(SUM(CASE WHEN status = 'closed' THEN payment_amount ELSE 0 END), 0) AS paid, "
        "COALESCE(SUM(CASE WHEN status <> 'closed' THEN payment_amount ELSE 0 END), 0) AS pending "
        "FROM payments"
    )
    if where:
        sql += f" WHERE {where}"
    sql += " GROUP BY payment_month ORDER BY payment_month DESC"
    with connect() as conn:
        return conn.execute(sql, params).fetchall()


def monthly_spending_status(months_back: int = 6, months_ahead: int = 6,
                            group_name: str = "") -> dict[str, Any]:
    """處長要的「每月支出狀態」（使用者 2026-08-12：不是核決門檻，是要看每個月支出）。

    既有的月度支出只算 payments（實際核銷），等於只看得到已經發生的錢。
    處長要掌握的是「這個月還要付多少、下個月要準備多少」，所以三個數字一起給：
      預計應付＝費用排程（第二層）＋合約付款排程；實際已付＝已結案的核銷；
      待付＝已登錄但還沒付掉的核銷。
    過去月份看「預估準不準」，未來月份看「要準備多少錢」。

    group_name：處長看全部，但要拆組別時用得到（經案件的組別過濾）。
    """
    today = date.today()
    base = today.year * 12 + (today.month - 1)
    span = [f"{(base + d) // 12:04d}-{(base + d) % 12 + 1:02d}"
            for d in range(-months_back, months_ahead + 1)]
    rows = {m: {"month": m, "planned": 0.0, "paid": 0.0, "unpaid": 0.0,
                "planned_count": 0, "paid_count": 0} for m in span}
    gfilter = str(group_name or "").strip()

    with connect() as conn:
        tables = {r["name"] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'")}

        def in_group(case_id: Any) -> bool:
            if not gfilter:
                return True
            if not case_id:
                return False
            r = conn.execute("SELECT group_name FROM cases WHERE id = ?", (case_id,)).fetchone()
            return bool(r) and str(r["group_name"] or "") == gfilter

        # 預計：費用模組第二層的排程（已確認的才算——草稿還在喬，不能拿來當資金預估）
        if "expense_schedules" in tables:
            for r in conn.execute(
                    "SELECT s.expense_month AS m, s.planned_amount AS amt, e.case_id AS case_id "
                    "FROM expense_schedules s "
                    "JOIN expense_sections sec ON sec.id = s.section_id "
                    "JOIN expense_masters e ON e.id = sec.expense_id "
                    "WHERE sec.status = 'confirmed' AND COALESCE(sec.archived,0) = 0").fetchall():
                m = str(r["m"] or "")[:7]
                if m in rows and in_group(r["case_id"]):
                    rows[m]["planned"] += float(r["amt"] or 0)
                    rows[m]["planned_count"] += 1

        # 預計：既有的合約付款排程（§8 那套，還沒轉到費用模組的合約走這裡）
        if "payment_schedules" in tables:
            for r in conn.execute(
                    "SELECT due_date, planned_amount AS amt, case_id FROM payment_schedules "
                    "WHERE COALESCE(status,'') <> 'paid'").fetchall():
                m = str(r["due_date"] or "")[:7]
                if m in rows and in_group(r["case_id"]):
                    rows[m]["planned"] += float(r["amt"] or 0)
                    rows[m]["planned_count"] += 1

        # 實際：核銷。closed＝已付，其餘＝登錄了還沒付
        for r in conn.execute(
                "SELECT p.payment_month AS m, p.payment_amount AS amt, p.status AS st, "
                "       c.case_id AS case_id "
                "FROM payments p LEFT JOIN contracts c ON c.id = p.contract_id").fetchall():
            m = str(r["m"] or "")[:7]
            if m not in rows or not in_group(r["case_id"]):
                continue
            amt = float(r["amt"] or 0)
            if str(r["st"] or "") == "closed":
                rows[m]["paid"] += amt
                rows[m]["paid_count"] += 1
            else:
                rows[m]["unpaid"] += amt

    this_month = today.strftime("%Y-%m")
    out = []
    for m in span:
        d = rows[m]
        d["diff"] = round(d["paid"] + d["unpaid"] - d["planned"], 2)   # 實際 vs 預計
        d["is_past"] = m < this_month
        d["is_current"] = m == this_month
        for k in ("planned", "paid", "unpaid"):
            d[k] = round(d[k], 2)
        out.append(d)
    return {
        "months": out, "this_month": this_month, "group_name": gfilter,
        "ahead_total": round(sum(d["planned"] for d in out if not d["is_past"] and not d["is_current"]), 2),
        "current": next((d for d in out if d["is_current"]), None),
        "note": "預計＝已確認的費用排程＋未付的合約付款排程；實際＝核銷（已付／待付）。"
                "草稿狀態的排程不列入，那還在喬，不能當資金預估。",
    }


def unit_budget_vs_actual(fiscal_year: int | None = None) -> dict[str, Any]:
    """單位別「預算 vs 實付」彙總（給主管看整體錢花在哪、超支在哪）。

    - 預算＝budgets.amount 依 unit_name 加總（排除停用）。
    - 實付＝付款經合約掛到案件（payment→contract.case_id），再經該案預算的 unit_name 歸到單位。
      已付＝status='closed'；待付＝其餘。案件沒有任何預算可歸單位的付款進「未歸單位」。
    - fiscal_year 有給就只算該年度的預算，以及該年度（payment_month 前四碼）的付款；不給＝全部年度。
    回傳 rows（依預算大→小）＋totals＋unattributed（未歸單位付款）＋years（可選年度清單）。
    """
    UNFILLED = "（未填單位）"
    with connect() as conn:
        # 可選年度（給前端下拉）：預算年度 ∪ 付款年度。過濾明顯異常年（如民國年/髒資料），只留合理西元範圍。
        years: set[int] = set()

        def _add_year(raw: Any) -> None:
            try:
                y = int(raw)
            except (TypeError, ValueError):
                return
            if 2000 <= y <= 2100:
                years.add(y)

        for r in conn.execute("SELECT DISTINCT fiscal_year FROM budgets WHERE fiscal_year IS NOT NULL"):
            _add_year(r["fiscal_year"])
        for r in conn.execute("SELECT DISTINCT substr(payment_month,1,4) AS y FROM payments WHERE payment_month <> ''"):
            _add_year(r["y"])

        # 預算：依單位加總
        bsql = "SELECT unit_name, COALESCE(SUM(amount),0) AS s FROM budgets WHERE status <> 'disabled'"
        bp: list[Any] = []
        if fiscal_year is not None:
            bsql += " AND fiscal_year = ?"
            bp.append(fiscal_year)
        bsql += " GROUP BY unit_name"
        budget_by_unit: dict[str, float] = {}
        for r in conn.execute(bsql, bp):
            unit = (r["unit_name"] or "").strip() or UNFILLED
            budget_by_unit[unit] = budget_by_unit.get(unit, 0.0) + (r["s"] or 0)

        # 案 → 單位：取該案（不限年度）第一筆非停用預算的 unit_name（空白也算，落到「未填單位」）
        case_unit: dict[int, str] = {}
        for r in conn.execute(
            "SELECT case_id, unit_name FROM budgets "
            "WHERE status <> 'disabled' AND case_id IS NOT NULL ORDER BY id"
        ):
            cid = r["case_id"]
            if cid not in case_unit:
                case_unit[cid] = (r["unit_name"] or "").strip() or UNFILLED

        # 付款 → 案（經合約）→ 單位
        psql = (
            "SELECT k.case_id AS case_id, p.payment_amount AS amt, p.status AS st "
            "FROM payments p JOIN contracts k ON k.id = p.contract_id WHERE 1=1"
        )
        pp: list[Any] = []
        if fiscal_year is not None:
            psql += " AND substr(p.payment_month,1,4) = ?"
            pp.append(str(fiscal_year))
        paid_by_unit: dict[str, float] = {}
        pending_by_unit: dict[str, float] = {}
        unattributed_paid = 0.0
        unattributed_pending = 0.0
        for r in conn.execute(psql, pp):
            unit = case_unit.get(r["case_id"])
            amt = r["amt"] or 0
            closed = (r["st"] == "closed")
            if unit is None:  # 該案沒有任何預算 → 無法歸單位
                if closed:
                    unattributed_paid += amt
                else:
                    unattributed_pending += amt
                continue
            target = paid_by_unit if closed else pending_by_unit
            target[unit] = target.get(unit, 0.0) + amt

    units = set(budget_by_unit) | set(paid_by_unit) | set(pending_by_unit)
    rows: list[dict[str, Any]] = []
    for u in units:
        budget = budget_by_unit.get(u, 0.0)
        paid = paid_by_unit.get(u, 0.0)
        pending = pending_by_unit.get(u, 0.0)
        rows.append({
            "unit": u,
            "budget": budget,
            "paid": paid,
            "pending": pending,
            "remaining": budget - paid,
            "usage_pct": round(paid / budget * 100, 1) if budget else None,
            "over": budget > 0 and paid > budget,
        })
    rows.sort(key=lambda x: (-(x["budget"] or 0), x["unit"]))
    totals = {
        "budget": sum(budget_by_unit.values()),
        "paid": sum(paid_by_unit.values()) + unattributed_paid,
        "pending": sum(pending_by_unit.values()) + unattributed_pending,
    }
    totals["remaining"] = totals["budget"] - totals["paid"]
    return {
        "fiscal_year": fiscal_year,
        "years": sorted(years, reverse=True),
        "rows": rows,
        "totals": totals,
        "unattributed": {"paid": unattributed_paid, "pending": unattributed_pending},
    }


def vendor_amount_summary() -> dict[str, Any]:
    """廠商別金額彙總（給主管看跟哪家廠商往來金額最大、有沒有實付超過合約金額）。

    - 合約金額＝contracts.amount 依 vendor_name 加總（排除停用）。
    - 實付＝該廠商名下所有合約的付款加總；已付=status='closed'、待付=其餘。
    - 合約沒填廠商的併入「（未填廠商）」，不像單位別報表那樣需要跨表歸戶，
      這裡合約本身就帶 vendor_name，故不設「未歸戶」桶。
    - 不分年度：合約本身無所屬年度欄位，金額是合約存續期間的總額，不是逐年概念。
    回傳 rows（依合約金額大→小）＋totals。
    """
    UNFILLED = "（未填廠商）"
    with connect() as conn:
        contract_rows = conn.execute(
            "SELECT id, vendor_name, amount FROM contracts WHERE status <> 'disabled'"
        ).fetchall()
        contract_vendor: dict[int, str] = {}
        amount_by_vendor: dict[str, float] = {}
        for r in contract_rows:
            vendor = (r["vendor_name"] or "").strip() or UNFILLED
            contract_vendor[r["id"]] = vendor
            amount_by_vendor[vendor] = amount_by_vendor.get(vendor, 0.0) + (r["amount"] or 0)

        paid_by_vendor: dict[str, float] = {}
        pending_by_vendor: dict[str, float] = {}
        for r in conn.execute(
            "SELECT p.contract_id AS contract_id, p.payment_amount AS amt, p.status AS st "
            "FROM payments p JOIN contracts k ON k.id = p.contract_id WHERE k.status <> 'disabled'"
        ):
            vendor = contract_vendor.get(r["contract_id"])
            if vendor is None:
                continue
            amt = r["amt"] or 0
            target = paid_by_vendor if r["st"] == "closed" else pending_by_vendor
            target[vendor] = target.get(vendor, 0.0) + amt

    vendors = set(amount_by_vendor) | set(paid_by_vendor) | set(pending_by_vendor)
    rows: list[dict[str, Any]] = []
    for v in vendors:
        contract_amount = amount_by_vendor.get(v, 0.0)
        paid = paid_by_vendor.get(v, 0.0)
        pending = pending_by_vendor.get(v, 0.0)
        rows.append({
            "vendor": v,
            "contract_amount": contract_amount,
            "paid": paid,
            "pending": pending,
            "remaining": contract_amount - paid,
            "usage_pct": round(paid / contract_amount * 100, 1) if contract_amount else None,
            "over": contract_amount > 0 and paid > contract_amount,
        })
    rows.sort(key=lambda x: (-(x["contract_amount"] or 0), x["vendor"]))
    totals = {
        "contract_amount": sum(amount_by_vendor.values()),
        "paid": sum(paid_by_vendor.values()),
        "pending": sum(pending_by_vendor.values()),
    }
    totals["remaining"] = totals["contract_amount"] - totals["paid"]
    return {"rows": rows, "totals": totals}


def expense_category_summary(dimension: str = "budget") -> dict[str, Any]:
    """費用類別分析：錢花在哪一類。

    「類別」有兩種合理讀法，這裡兩種都給、由使用者切換，不預先幫他決定：
      dimension='budget'  ：走預算類別（基礎建設/工具/資訊安全…）。付款→合約→案件→該案預算的
                            category。一個案子底下若有多個不同類別的預算，歸屬有歧義，
                            不硬猜，統一進「多類別（需人工歸戶）」讓人自己看。
      dimension='contract'：走合約類型（採購/維護/租賃/軟體授權/服務）。直接掛在合約上，
                            沒有歧義，但顆粒度是合約而非預算科目。
    金額口徑：已付＝payments.status='closed'（真的花掉的），待付＝其餘（已排未付）。
    停用的合約不列入。承辦只看自己案件下的（沿用 owner scope）。
    """
    UNCLASSIFIED = "（未分類）"
    MIXED = "（多類別，需人工歸戶）"
    scope = _owner_scope.get()
    with connect() as conn:
        where = "k.status <> 'disabled'"
        params: list[Any] = []
        if scope is not None:
            sw, sp = _scope_where("contracts", scope, alias="k")
            if sw:
                where = f"{where} AND {sw}"
                params += sp
        contracts = conn.execute(
            f"SELECT k.id, k.case_id, k.contract_type, k.amount FROM contracts k WHERE {where}", params).fetchall()
        if not contracts:
            return {"dimension": dimension, "rows": [], "totals": {"paid": 0.0, "pending": 0.0, "contract_amount": 0.0}}

        # 案件 → 該案預算的類別集合（一個案子多個不同類別＝歸屬有歧義，不硬猜）
        case_category: dict[int, str] = {}
        if dimension == "budget":
            for r in conn.execute(
                "SELECT case_id, category FROM budgets WHERE case_id IS NOT NULL AND status <> 'disabled'"):
                cat = (r["category"] or "").strip()
                if not cat:
                    continue
                cur = case_category.get(r["case_id"])
                case_category[r["case_id"]] = cat if cur is None else (cur if cur == cat else MIXED)

        def _key_of(row) -> str:
            if dimension == "contract":
                return (row["contract_type"] or "").strip() or UNCLASSIFIED
            return case_category.get(row["case_id"]) or UNCLASSIFIED

        key_by_contract = {r["id"]: _key_of(r) for r in contracts}
        contract_amount: dict[str, float] = {}
        for r in contracts:
            key = key_by_contract[r["id"]]
            contract_amount[key] = contract_amount.get(key, 0.0) + float(r["amount"] or 0)

        paid: dict[str, float] = {}
        pending: dict[str, float] = {}
        counts: dict[str, int] = {}
        marks = ",".join("?" * len(key_by_contract))
        for r in conn.execute(
            f"SELECT contract_id, payment_amount, status FROM payments WHERE contract_id IN ({marks})",
            tuple(key_by_contract)):
            key = key_by_contract.get(r["contract_id"])
            if key is None:
                continue
            bucket = paid if r["status"] == "closed" else pending
            bucket[key] = bucket.get(key, 0.0) + float(r["payment_amount"] or 0)
            counts[key] = counts.get(key, 0) + 1

    rows = []
    for key in set(contract_amount) | set(paid) | set(pending):
        rows.append({
            "category": key,
            "contract_amount": contract_amount.get(key, 0.0),
            "paid": paid.get(key, 0.0),
            "pending": pending.get(key, 0.0),
            "payment_count": counts.get(key, 0),
            "needs_attention": key in (UNCLASSIFIED, MIXED),  # 前端標出來提示要人工歸戶
        })
    rows.sort(key=lambda x: (-(x["paid"] or 0), -(x["contract_amount"] or 0), x["category"]))
    return {
        "dimension": dimension,
        "rows": rows,
        "totals": {
            "contract_amount": sum(contract_amount.values()),
            "paid": sum(paid.values()),
            "pending": sum(pending.values()),
        },
    }


_TODO_HORIZON_DAYS = 30   # 未來 30 天內要處理的才進待辦，再遠的還不用煩


def cases_needing_attention() -> list[dict[str, Any]]:
    """待辦事項：卡在流程上的案子，加上快到日子的事。

    使用者拍板（2026-07-29）改為由日期自動生成，不再靠人工填「下一步」：
      - 卡在流程：待複核、審核中、退回補件（要有人動手才會前進）
      - 合約／保固／維護到期：已過期或 30 天內
      - 付款排程：預計付款日已過或 30 天內、且尚未付款
    承辦只看自己的（合約經案件歸屬過濾）。
    """
    scope = _owner_scope.get()
    today = date.today()
    horizon = (today + timedelta(days=_TODO_HORIZON_DAYS)).isoformat()
    today_s = today.isoformat()
    items: list[dict[str, Any]] = []

    with connect() as conn:
        # ① 卡在審核流程的案件
        where = "status IN ('pending_review', 'reviewing', 'returned')"
        params: list[Any] = []
        if scope is not None:
            where = f"({where}) AND owner = ?"
            params.append(scope)
        for r in conn.execute(
            f"SELECT id, case_code, title, status, owner, amount, review_note FROM cases WHERE {where} "
            "ORDER BY id DESC LIMIT 100", params):
            items.append({
                "kind": "case", "id": r["id"], "case_code": r["case_code"], "title": r["title"],
                "status": r["status"], "owner": r["owner"], "amount": r["amount"],
                "detail": r["review_note"] or "", "due_date": "", "days_left": None,
            })

        # ②③ 合約相關的到期日：合約/保固/維護到期、預計付款日
        cw, cp = ("", [])
        if scope is not None:
            cw, cp = _scope_where("contracts", scope, alias="k")
        tail = f" AND {cw}" if cw else ""

        for col, kind, label in EXPIRY_KINDS:
            for r in conn.execute(
                f"SELECT k.id, k.case_id, k.contract_code, k.contract_name, k.{col} AS due "
                f"FROM contracts k WHERE k.{col} <> '' AND k.{col} <= ? AND k.status <> 'disabled'{tail} "
                "ORDER BY due LIMIT 100", [horizon, *cp]):
                due = _pdate(r["due"])
                items.append({
                    "kind": kind, "id": r["case_id"], "case_code": r["contract_code"],
                    "title": r["contract_name"], "status": "expiring", "owner": "", "amount": 0,
                    "detail": label, "due_date": r["due"],
                    "days_left": (due - today).days if due else None,
                })

        for r in conn.execute(
            "SELECT ps.id, k.case_id, k.contract_code, ps.label, ps.due_date, ps.planned_amount "
            "FROM payment_schedules ps JOIN contracts k ON k.id = ps.contract_id "
            f"WHERE ps.status = 'planned' AND ps.due_date <> '' AND ps.due_date <= ? "
            f"AND k.status <> 'disabled'{tail} ORDER BY ps.due_date LIMIT 100", [horizon[:7], *cp]):
            due = _pdate(r["due_date"] + "-01" if len(r["due_date"]) == 7 else r["due_date"])
            items.append({
                "kind": "payment_due", "id": r["case_id"], "case_code": r["contract_code"],
                "title": f"{r['label']} 預計付款", "status": "expiring", "owner": "",
                "amount": r["planned_amount"], "detail": "預計付款日", "due_date": r["due_date"],
                "days_left": (due - today).days if due else None,
            })

    # 最急的排前面：已過期 → 快到期 → 卡流程（沒有日期的排最後）
    items.sort(key=lambda x: (x["days_left"] is None, x["days_left"] if x["days_left"] is not None else 0))
    return items


# 到期提醒分階段：合約沒人管就是自動續約或斷保，所以不是「快到期」一句話帶過，
# 而是按剩餘天數分四階段，讓主管知道哪些要現在處理、哪些還能排。
# 合約到期／保固到期／維護到期是三種不同的事（保固常晚於合約），三種一起看才不會漏。
EXPIRY_STAGES = ["overdue", "d7", "d30", "d60", "d90"]
EXPIRY_KINDS = [("end_date", "contract", "合約到期"),
                ("warranty_end_date", "warranty", "保固到期"),
                ("maintenance_end_date", "maintenance", "維護到期")]


def _expiry_stage(days_left: int, within_days: int) -> str | None:
    """剩餘天數 → 階段。已過期最急；超過 within_days 不列入（回 None）。"""
    if days_left < 0:
        return "overdue"
    if days_left <= 7:
        return "d7"
    if days_left <= 30:
        return "d30"
    if days_left <= 60:
        return "d60"
    if days_left <= within_days:
        return "d90"
    return None


def expiring_contracts(within_days: int = 90) -> dict[str, Any]:
    """到期雷達：合約／保固／維護三種到期日，按 90/60/30/7 天分階段（已過期另計）。
    一份合約三種日期各算一筆，因為要處理的事不一樣（續約 vs 續保 vs 續維護）。
    未作廢才列入；承辦只看自己案件下的合約。"""
    scope = _owner_scope.get()
    today = date.today()
    threshold = (today + timedelta(days=within_days)).isoformat()
    date_cols = " OR ".join(f"({col} <> '' AND {col} <= ?)" for col, _, _ in EXPIRY_KINDS)
    where = f"({date_cols}) AND status <> 'disabled'"
    params: list[Any] = [threshold] * len(EXPIRY_KINDS)
    if scope is not None:
        sw, sp = _scope_where("contracts", scope)
        if sw:
            where = f"({where}) AND {sw}"
            params += sp
    with connect() as conn:
        rows = conn.execute(
            f"SELECT * FROM contracts WHERE {where} ORDER BY end_date ASC LIMIT 200", params).fetchall()

    items: list[dict[str, Any]] = []
    for row in rows:
        for col, kind, kind_label in EXPIRY_KINDS:
            raw = str(row[col] or "").strip()
            if not raw:
                continue
            due = _pdate(raw)
            if due is None:
                continue
            days_left = (due - today).days
            stage = _expiry_stage(days_left, within_days)
            if stage is None:
                continue
            items.append({
                "contract_id": row["id"], "contract_code": row["contract_code"],
                "contract_name": row["contract_name"], "vendor_name": row["vendor_name"],
                "amount": row["amount"], "case_id": row["case_id"],
                "kind": kind, "kind_label": kind_label,
                "due_date": raw, "days_left": days_left, "stage": stage,
            })
    items.sort(key=lambda x: (x["due_date"], x["contract_code"]))
    counts = {stage: sum(1 for x in items if x["stage"] == stage) for stage in EXPIRY_STAGES}
    return {"items": items, "counts": counts, "total": len(items), "within_days": within_days}


def overdue_reminders(within_days: int = 14) -> list[dict[str, Any]]:
    """催辦清單：逾期或即將到期、但尚未完成的『案件』與『合約』，供主動提醒。
    - 案件：有預計完成日(due_date) 且未核准/未作廢。
    - 合約：有到期日(end_date) 且未作廢。
    依 owner 範圍過濾（承辦只看自己的）。逾期(date<今天)標 overdue，其餘標 soon。"""
    scope = _owner_scope.get()
    today = date.today()
    today_s = today.isoformat()
    horizon = (today + timedelta(days=within_days)).isoformat()

    def _days(date_str: str) -> int:
        try:
            return (date.fromisoformat(date_str) - today).days
        except ValueError:
            return 0

    items: list[dict[str, Any]] = []
    with connect() as conn:
        # 案件：未完成且有預計完成日
        c_where = "due_date <> '' AND due_date <= ? AND status NOT IN ('approved', 'disabled')"
        c_params: list[Any] = [horizon]
        if scope is not None:
            c_where = f"({c_where}) AND owner = ?"
            c_params.append(scope)
        for r in conn.execute(
            f"SELECT id, case_code, title, owner, due_date, status FROM cases WHERE {c_where} ORDER BY due_date ASC LIMIT 100",
            c_params,
        ).fetchall():
            items.append({
                "type": "case", "id": r["id"], "code": r["case_code"], "title": r["title"],
                "owner": r["owner"], "date": r["due_date"], "status": r["status"],
                "days": _days(r["due_date"]), "severity": "overdue" if r["due_date"] < today_s else "soon",
            })

        # 專案：未完成且有預計完成日（比照案件納入催辦）
        pj_where = "due_date <> '' AND due_date <= ? AND status NOT IN ('completed', 'disabled')"
        pj_params: list[Any] = [horizon]
        if scope is not None:
            sw, sp = _scope_where("projects", scope)
            if sw:
                pj_where = f"({pj_where}) AND {sw}"
                pj_params += sp
        for r in conn.execute(
            f"SELECT id, project_code, project_name, owner, due_date, status FROM projects WHERE {pj_where} ORDER BY due_date ASC LIMIT 100",
            pj_params,
        ).fetchall():
            items.append({
                "type": "project", "id": r["id"], "code": r["project_code"], "title": r["project_name"],
                "owner": r["owner"], "date": r["due_date"], "status": r["status"],
                "days": _days(r["due_date"]), "severity": "overdue" if r["due_date"] < today_s else "soon",
            })

        # 合約：未作廢且有到期日
        k_where = "k.end_date <> '' AND k.end_date <= ? AND k.status <> 'disabled'"
        k_params: list[Any] = [horizon]
        if scope is not None:
            sw, sp = _scope_where("contracts", scope)
            if sw:
                k_where = f"({k_where}) AND {sw.replace('case_id', 'k.case_id')}"
                k_params += sp
        for r in conn.execute(
            "SELECT k.id, k.contract_code, k.contract_name, c.owner, k.end_date, k.status "
            f"FROM contracts k LEFT JOIN cases c ON c.id = k.case_id WHERE {k_where} ORDER BY k.end_date ASC LIMIT 100",
            k_params,
        ).fetchall():
            items.append({
                "type": "contract", "id": r["id"], "code": r["contract_code"], "title": r["contract_name"],
                "owner": r["owner"], "date": r["end_date"], "status": r["status"],
                "days": _days(r["end_date"]), "severity": "overdue" if r["end_date"] < today_s else "soon",
            })

        # 專案工作項：真正在跑的最小單位。原本催辦只看案件/專案/合約，工作項逾期
        # 只能逐案點進專案才看得到，等於最容易卡住的那一層反而沒有主動提醒。
        # progress < 100 才算未完成——做完的不追，才不會讓催辦清單被完成項灌滿。
        i_where = ("i.end_date <> '' AND i.end_date <= ? AND i.progress < 100 "
                   "AND i.status <> 'disabled' AND p.status <> 'disabled'")
        i_params: list[Any] = [horizon]
        if scope is not None:
            i_where = f"({i_where}) AND i.owner = ?"
            i_params.append(scope)
        for r in conn.execute(
            "SELECT i.id, i.item_name, i.owner, i.end_date, i.exec_status, i.progress, "
            "p.project_code, p.project_name "
            f"FROM project_items i JOIN projects p ON p.id = i.project_id WHERE {i_where} "
            "ORDER BY i.end_date ASC LIMIT 100",
            i_params,
        ).fetchall():
            items.append({
                "type": "project_item", "id": r["id"],
                "code": r["project_code"] or r["project_name"],
                # 帶上專案名，否則清單只看到「安裝與上線」這種工作項名稱，認不出是哪一案
                "title": f'{r["item_name"]}（{r["project_name"]}）',
                "owner": r["owner"], "date": r["end_date"], "status": r["exec_status"],
                "days": _days(r["end_date"]), "severity": "overdue" if r["end_date"] < today_s else "soon",
            })

    items.sort(key=lambda x: x["date"])
    return items


def orphan_payments() -> list[dict[str, Any]]:
    """未歸戶付款：所屬合約沒有掛案件（case_id 為空）→ 沒人追、CIO 也看不到。給主管檢視。"""
    with connect() as conn:
        return conn.execute(
            "SELECT p.id, p.payment_month, p.payment_amount, p.status, k.contract_code "
            "FROM payments p JOIN contracts k ON k.id = p.contract_id "
            "WHERE k.case_id IS NULL AND p.status <> 'disabled' ORDER BY p.id DESC LIMIT 100"
        ).fetchall()


def pending_approvals() -> list[dict[str, Any]]:
    """待我複核：狀態為 pending_review 且非我建立的案件（雙人複核，不能核自己的）。"""
    actor = _current_actor.get()
    with connect() as conn:
        return conn.execute(
            "SELECT id, case_code, title, owner, amount, created_by FROM cases "
            "WHERE status = 'pending_review' AND created_by <> ? ORDER BY id DESC LIMIT 100",
            (actor,),
        ).fetchall()


# 雙人複核規則 (b)：核准前不算數 —— CIO 畫面的金額只計「已核准（含之後）」案件下的付款。
# 需求書 §4 的狀態機在核准之後還會往 進行中／暫停／已結案 走，這些都是「核准過的錢」，
# 一律要算進去；只有 已取消 例外（案子撤了，錢不該再算）。少列任何一個，案件一開始執行
# CIO 的數字就會憑空掉一塊。
LIVE_CASE_STATUSES = ("approved", "in_progress", "paused", "closed")
_LIVE_CASE_LIST = ", ".join(f"'{s}'" for s in LIVE_CASE_STATUSES)
_APPROVED_PAYMENT_CLAUSE = (
    "contract_id IN (SELECT id FROM contracts WHERE case_id IN "
    f"(SELECT id FROM cases WHERE status IN ({_LIVE_CASE_LIST})))"
)


def cio_overview() -> dict[str, Any]:
    """CIO 決策總覽：大方向資金（下月應付 / 要準備的資金）+ 下月要出的款（可下探至案件）。
    金額只算『已核准』案件（未複核的錢不讓 CIO 看到當真）；並依 owner 範圍過濾。"""
    scope = _owner_scope.get()
    today = date.today()
    this_month = today.strftime("%Y-%m")
    if today.month == 12:
        next_month = f"{today.year + 1}-01"
    else:
        next_month = f"{today.year}-{today.month + 1:02d}"

    pw, pp = _scope_where("payments", scope) if scope is not None else ("", [])
    tail = f" AND {pw}" if pw else ""
    approved = f" AND {_APPROVED_PAYMENT_CLAUSE}"  # 只算已核准案件的付款

    with connect() as conn:
        def _sum(cond: str, params: list[Any]) -> float:
            return conn.execute(
                f"SELECT COALESCE(SUM(payment_amount), 0) AS s FROM payments WHERE {cond}",
                params,
            ).fetchone()["s"]

        next_month_total = _sum(f"payment_month = ?{tail}{approved}", [next_month, *pp])
        this_month_total = _sum(f"payment_month = ?{tail}{approved}", [this_month, *pp])
        funds_to_prepare = _sum(f"status <> 'closed'{tail}{approved}", [*pp])  # 尚未結案 = 要準備的資金

        # D：未來 6 個月現金流預測（含本月），只算已核准案件的付款
        forecast = []
        y, m = today.year, today.month
        for _ in range(6):
            mon = f"{y}-{m:02d}"
            forecast.append({"month": mon, "total": _sum(f"payment_month = ?{tail}{approved}", [mon, *pp])})
            m = 1 if m == 12 else m + 1
            y = y + 1 if m == 1 else y

        # 下月要出的每一筆款，連到所屬案件（供 CIO 逐層下探）；只列已核准案件。
        # budget_links=0 代表案件沒關聯任何預算 → 視為「預算外/計畫外」支出。
        # E：case_budget_total>0 且 案件付款合計>預算合計 → 超支。
        detail_sql = (
            "SELECT c.id AS case_id, c.case_code, c.title AS case_title, c.owner, "
            "k.contract_code, p.payment_month, p.payment_amount, p.status, "
            "(SELECT COUNT(*) FROM budgets b WHERE b.case_id = c.id AND b.status <> 'disabled') AS budget_links, "
            "(SELECT COALESCE(SUM(b.amount),0) FROM budgets b WHERE b.case_id = c.id AND b.status <> 'disabled') AS case_budget_total, "
            "(SELECT COALESCE(SUM(pp.payment_amount),0) FROM payments pp JOIN contracts kk ON kk.id = pp.contract_id WHERE kk.case_id = c.id) AS case_payment_total "
            "FROM payments p JOIN contracts k ON k.id = p.contract_id "
            f"JOIN cases c ON c.id = k.case_id WHERE p.payment_month = ? AND c.status IN ({_LIVE_CASE_LIST})"
        )
        detail_params: list[Any] = [next_month]
        if scope is not None:
            detail_sql += " AND c.owner = ?"
            detail_params.append(scope)
        detail_sql += " ORDER BY p.payment_amount DESC LIMIT 100"
        upcoming = []
        unplanned_total = 0.0
        overspent_count = 0
        for r in conn.execute(detail_sql, detail_params).fetchall():
            item = dict(r)
            item["unplanned"] = (r["budget_links"] or 0) == 0  # 無對應預算＝計畫外
            item["overspent"] = (r["case_budget_total"] or 0) > 0 and (r["case_payment_total"] or 0) > r["case_budget_total"]
            if item["unplanned"]:
                unplanned_total += r["payment_amount"]
            if item["overspent"]:
                overspent_count += 1
            upcoming.append(item)

        # §8 預計付款排程（未付）＝真正的「待付款 / 下月預計付款 / 現金流」。實際核銷(payments)只
        # 涵蓋已經核銷的，還沒核銷的預計得靠排程。跨年度預算歸屬也靠 due_date 的年份。
        # 只算已核准案件、依 owner 範圍過濾（與上方口徑一致）。
        sched_scope = " AND c.owner = ?" if scope is not None else ""

        def _sched_sum(cond: str, extra: list[Any]) -> float:
            params = ([scope] if scope is not None else []) + extra
            return conn.execute(
                "SELECT COALESCE(SUM(ps.planned_amount),0) AS s FROM payment_schedules ps "
                "JOIN contracts k ON k.id = ps.contract_id JOIN cases c ON c.id = k.case_id "
                f"WHERE ps.status <> 'paid' AND c.status IN ({_LIVE_CASE_LIST}){sched_scope} AND {cond}",
                params,
            ).fetchone()["s"]

        payable_planned = _sched_sum("1=1", [])                          # 待付款＝所有預計未付
        next_month_planned = _sched_sum("ps.due_date = ?", [next_month]) # 下月預計付款（來自排程）
        this_month_planned = _sched_sum("ps.due_date = ?", [this_month])
        planned_forecast = []
        _y, _m = today.year, today.month
        for _ in range(6):
            _mon = f"{_y}-{_m:02d}"
            planned_forecast.append({"month": _mon, "total": _sched_sum("ps.due_date = ?", [_mon])})
            _m = 1 if _m == 12 else _m + 1
            _y = _y + 1 if _m == 1 else _y

        # 跨年度：待付款按「預計付款日的年份」拆（100萬季付跨年→75萬算今年、25萬算明年，
        # 各由該年度預算支付）。前端可據此顯示各年度要準備的資金。
        _year_rows = conn.execute(
            "SELECT substr(ps.due_date,1,4) AS yr, COALESCE(SUM(ps.planned_amount),0) AS s "
            "FROM payment_schedules ps JOIN contracts k ON k.id=ps.contract_id "
            "JOIN cases c ON c.id=k.case_id "
            f"WHERE ps.status <> 'paid' AND c.status IN ({_LIVE_CASE_LIST}){sched_scope} AND ps.due_date <> '' "
            "GROUP BY yr ORDER BY yr",
            ([scope] if scope is not None else []),
        ).fetchall()
        payable_by_year = {r["yr"]: float(r["s"]) for r in _year_rows}

    return {
        "this_month": this_month,
        "next_month": next_month,
        "next_month_total": next_month_total,
        "this_month_total": this_month_total,
        "funds_to_prepare": funds_to_prepare,
        # §8 排程口徑（預計未付）——待付款、下月預計、現金流用這個才完整
        "payable_planned": payable_planned,
        "next_month_planned": next_month_planned,
        "this_month_planned": this_month_planned,
        "planned_forecast": planned_forecast,
        "payable_by_year": payable_by_year,
        "unplanned_next_month": unplanned_total,  # 下月「預算外/計畫外」金額
        "overspent_count": overspent_count,       # 下月清單中超支案件數
        "forecast": forecast,                     # 未來 6 個月現金流
        "upcoming_next_month": upcoming,
        # 到期雷達摘要：合約/保固/維護沒人管就是自動續約或斷保，CIO 一頁要看得到還有幾件沒處理
        "expiry_counts": expiring_contracts()["counts"],
    }


CHANGE_TABLE_LABEL = {
    "cases": "案件", "contracts": "合約", "payments": "付款", "budgets": "預算",
    "projects": "專案", "signoffs": "簽呈", "purchases": "請購", "documents": "文件",
}
CHANGE_ACTION_LABEL = {
    "create": "新增", "update": "更新", "disable": "停用", "delete": "刪除",
    "submit": "送出複核", "approve": "核准", "cancel_review": "取消複核", "import": "匯入新增", "import-update": "匯入更新",
}


def cio_changes_since_last_view() -> dict[str, Any]:
    """CIO「自上次查看以來」變動摘要：查看即視為已讀，下次再顯示這之後的變動。

    以 audit_logs 為準（涵蓋所有模組的 create/update/disable/delete/submit/approve/匯入），
    比逐表比對 created_at 準確——能抓到「既有資料被改動」，不只新增的筆數。
    游標用 audit_logs.id（嚴格遞增）而非時間字串：時間字串只到秒，兩次查看之間若
    有動作跟「標記已讀」落在同一秒會被 `created_at > since` 漏掉；id 不會有這問題。
    每個帳號各自記自己的上次查看進度（settings 表 cio_last_seen_id/cio_last_seen_at:{actor} 鍵）。
    """
    actor = _current_actor.get()
    id_key = f"cio_last_seen_id:{actor}"
    at_key = f"cio_last_seen_at:{actor}"
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with connect() as conn:
        current_max_id = conn.execute("SELECT COALESCE(MAX(id), 0) AS m FROM audit_logs").fetchone()["m"]
        prev = conn.execute("SELECT value FROM settings WHERE key = ?", (id_key,)).fetchone()
        prev_at = conn.execute("SELECT value FROM settings WHERE key = ?", (at_key,)).fetchone()

        def _mark_seen() -> None:
            for k, v in ((id_key, str(current_max_id)), (at_key, now)):
                conn.execute(
                    "INSERT INTO settings(key, value) VALUES(?, ?) "
                    "ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=CURRENT_TIMESTAMP",
                    (k, v),
                )

        if not prev or not str(prev["value"]).strip():
            _mark_seen()
            return {"first_visit": True, "since": None, "changes": [], "total_count": 0}

        since_id = int(prev["value"])
        since = prev_at["value"] if prev_at else None
        rows = conn.execute(
            "SELECT table_name, action, COUNT(*) AS c FROM audit_logs "
            "WHERE id > ? GROUP BY table_name, action ORDER BY c DESC",
            (since_id,),
        ).fetchall()
        _mark_seen()

    changes = [
        {
            "table": r["table_name"],
            "table_label": CHANGE_TABLE_LABEL.get(r["table_name"], r["table_name"]),
            "action": r["action"],
            "action_label": CHANGE_ACTION_LABEL.get(r["action"], r["action"]),
            "count": r["c"],
        }
        for r in rows
    ]
    return {
        "first_visit": False,
        "since": since,
        "changes": changes,
        "total_count": sum(c["count"] for c in changes),
    }


def get_db_user(username: str) -> dict[str, Any] | None:
    with connect() as conn:
        return conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()


def list_db_users() -> list[dict[str, Any]]:
    with connect() as conn:
        rows = conn.execute(
            "SELECT username, role_code, display_name, email, disabled, group_name "
            "FROM users ORDER BY username").fetchall()
    return rows


def create_db_user(username: str, role_code: str, display_name: str, email: str, password_hash: str,
                   group_name: str = "") -> None:
    with connect() as conn:
        if conn.execute("SELECT 1 FROM users WHERE username = ?", (username,)).fetchone():
            raise ValueError(f"帳號 {username} 已存在。")
        conn.execute(
            "INSERT INTO users(username, role_code, display_name, email, password_hash, group_name) "
            "VALUES(?,?,?,?,?,?)",
            (username, role_code, display_name, email, password_hash, (group_name or "").strip()),
        )


def update_db_user(username: str, fields: dict[str, Any]) -> None:
    allowed = {"role_code", "display_name", "email", "disabled", "password_hash", "group_name"}
    sets = {k: v for k, v in fields.items() if k in allowed and v is not None}
    if not sets:
        return
    assignments = ", ".join(f"{k} = ?" for k in sets)
    with connect() as conn:
        cur = conn.execute(f"UPDATE users SET {assignments} WHERE username = ?", [*sets.values(), username])
        if cur.rowcount == 0:
            raise LookupError(f"帳號 {username} 不存在。")


def delete_db_user(username: str) -> None:
    with connect() as conn:
        cur = conn.execute("DELETE FROM users WHERE username = ?", (username,))
        if cur.rowcount == 0:
            raise LookupError(f"帳號 {username} 不存在。")


def backup_database(dest_path: str) -> None:
    """用 SQLite 線上備份 API 把整個資料庫複製到 dest_path（即使有連線也一致）。"""
    with connect() as src:
        dst = sqlite3.connect(dest_path)
        try:
            src.backup(dst)
        finally:
            dst.close()


def reset_database() -> dict[str, Any]:
    """整個資料庫重置（測試用危險操作）：先自動備份現有 .db，再清空所有資料表、流水號歸零。
    只清「資料」，不動 schema 本身——重置後結構還在，不用重開服務。呼叫端（main.py）要先擋
    settings.allow_db_reset，這裡不重複檢查。"""
    settings = get_settings()
    db_path = Path(settings.database_path)
    backup_path = ""
    if db_path.exists():
        backup_dir = db_path.parent / "reset_backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        dest = backup_dir / f"{db_path.stem}_before_reset_{stamp}.db"
        backup_database(str(dest))
        backup_path = str(dest)

    with connect() as conn:
        tables = [r["name"] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
        ).fetchall()]
        for t in tables:
            conn.execute(f"DELETE FROM {t}")
        has_seq = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='sqlite_sequence'"
        ).fetchone()
        if has_seq:
            conn.execute("DELETE FROM sqlite_sequence")  # 自增流水號歸零，重置後 id 從 1 開始
    return {"reset": True, "tables_cleared": len(tables), "backup_path": backup_path}


def read_setting(key: str, default: str = "") -> str:
    with connect() as conn:
        row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
        return row["value"] if row else default


def read_settings(keys: list[str]) -> dict[str, str]:
    with connect() as conn:
        rows = conn.execute("SELECT key, value FROM settings").fetchall()
    have = {r["key"]: r["value"] for r in rows}
    return {k: have.get(k, "") for k in keys}


def write_settings(values: dict[str, str]) -> None:
    """upsert 一批設定。空字串代表清空該鍵；未出現的鍵不動。"""
    with connect() as conn:
        for key, value in values.items():
            conn.execute(
                "INSERT INTO settings(key, value) VALUES(?, ?) "
                "ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=CURRENT_TIMESTAMP",
                (key, "" if value is None else str(value)),
            )


def search_records(query: str) -> list[dict[str, Any]]:
    pattern = f"%{query}%"
    scope = _owner_scope.get()
    results: list[dict[str, Any]] = []
    # 每列：(型別, 表, 代號欄, 標題欄, 明細欄, [額外可搜欄], 回案件的 JOIN, 系統編號前綴)
    # JOIN 回案件後，把「年度-流水號 尾碼／年-流水號／前綴-年-流水號」也納入比對，
    # 才搜得到前端即時組出來、DB 沒存的系統編號（搜 0003 能一次撈齊同案各階段）。
    entities = [
        ("case", "cases", "case_code", "title", "owner", ["note", "next_step"], "", "Case"),
        ("contract", "contracts", "contract_code", "contract_name", "vendor_name", [],
         "LEFT JOIN cases c ON c.id = t.case_id", "Cont"),
        ("document", "documents", "file_name", "document_type", "source_note", [],
         "LEFT JOIN cases c ON c.id = t.case_id", None),
        ("budget", "budgets", "budget_code", "category", "unit_name", ["note"],
         "LEFT JOIN cases c ON c.id = t.case_id", "Budg"),
        ("project", "projects", "project_code", "project_name", "source", ["owner", "necessity", "note"],
         "LEFT JOIN cases c ON c.id = t.case_id", "Proj"),
        ("signoff", "signoffs", "signoff_code", "subject", "applicant", ["note"],
         "LEFT JOIN cases c ON c.id = t.case_id", "Sign"),
        ("purchase", "purchases", "purchase_code", "item_name", "vendor_name", ["note"],
         "LEFT JOIN cases c ON c.id = t.case_id", "Purc"),
        ("payment", "payments", "settle_no", "item", "vendor", ["ref_no", "period"],
         "LEFT JOIN contracts k ON k.id = t.contract_id LEFT JOIN cases c ON c.id = k.case_id", "Paym"),
    ]
    with connect() as conn:
        for typ, source, code_field, title_field, extra_field, more_fields, join, prefix in entities:
            cref = "t" if source == "cases" else "c"  # 案件本身的年/流水號在自己身上
            search_fields = [code_field, title_field, extra_field, *more_fields]
            ors = [f"t.{f} LIKE ?" for f in search_fields]
            params: list[Any] = [pattern] * len(search_fields)
            # 系統編號比對（12 碼無連字號＝功能碼＋西元年＋流水號）：四位尾碼、年+尾碼、前綴+年+尾碼
            ors.append(f"printf('%04d', {cref}.seq) LIKE ?")
            params.append(pattern)
            ors.append(f"({cref}.fiscal_year || printf('%04d', {cref}.seq)) LIKE ?")
            params.append(pattern)
            if prefix:
                ors.append(f"('{prefix}' || {cref}.fiscal_year || printf('%04d', {cref}.seq)) LIKE ?")
                params.append(pattern)
            sql = (
                f"SELECT t.id AS id, t.{code_field} AS code, t.{title_field} AS title, "
                f"t.{extra_field} AS detail FROM {source} t {join} WHERE ({' OR '.join(ors)})"
            )
            if scope is not None:
                sw, sp = _scope_where(source, scope, alias="t")
                if sw:
                    sql += f" AND {sw}"
                    params += sp
            sql += " ORDER BY t.id DESC LIMIT 50"
            rows = conn.execute(sql, params).fetchall()
            results.extend({"type": typ, **row} for row in rows)

        # 專案「工作主項目」子項目：使用者反饋 Excel 表格裡的細項（如「集團聯合議價，合約」）搜不到。
        # 子項目沒有獨立頁面可編輯，比對到就導回所屬專案（type 沿用 project，id 用專案的 id）。
        item_fields = ["item_name", "owner", "risk_note", "decision_needed", "support_needed"]
        item_ors = " OR ".join(f"i.{f} LIKE ?" for f in item_fields)
        item_sql = (
            "SELECT pr.id AS id, pr.project_code AS code, i.item_name AS title, "
            "('屬於「' || pr.project_name || '」的工作項目') AS detail "
            f"FROM project_items i JOIN projects pr ON pr.id = i.project_id WHERE ({item_ors})"
        )
        item_params: list[Any] = [pattern] * len(item_fields)
        if scope is not None:
            sw, sp = _scope_where("projects", scope)
            if sw:
                item_sql += f" AND pr.id IN (SELECT id FROM projects WHERE {sw})"
                item_params += sp
        item_sql += " ORDER BY i.id DESC LIMIT 50"
        results.extend({"type": "project_item", **row} for row in conn.execute(item_sql, item_params).fetchall())
    return results


def case_360(case_id: int) -> dict[str, Any]:
    scope = _owner_scope.get()
    with connect() as conn:
        case = get_row(conn, "cases", case_id)
        if scope is not None and case.get("owner") != scope:
            raise LookupError(f"cases row {case_id} not found")  # 非本人案件，視同不存在
        contracts = conn.execute("SELECT * FROM contracts WHERE case_id = ? ORDER BY id DESC", (case_id,)).fetchall()
        payments = conn.execute(
            "SELECT p.* FROM payments p JOIN contracts c ON c.id = p.contract_id "
            "WHERE c.case_id = ? ORDER BY p.payment_month DESC",
            (case_id,),
        ).fetchall()
        documents = conn.execute("SELECT * FROM documents WHERE case_id = ? ORDER BY id DESC", (case_id,)).fetchall()
        # 追查「這筆費用對應的預算/專案/簽呈/請購」——CIO 下探時要看得到整條控管鏈
        budgets = conn.execute("SELECT * FROM budgets WHERE case_id = ? ORDER BY id DESC", (case_id,)).fetchall()
        projects = conn.execute("SELECT * FROM projects WHERE case_id = ? ORDER BY id DESC", (case_id,)).fetchall()
        signoffs = conn.execute("SELECT * FROM signoffs WHERE case_id = ? ORDER BY id DESC", (case_id,)).fetchall()
        purchases = conn.execute("SELECT * FROM purchases WHERE case_id = ? ORDER BY id DESC", (case_id,)).fetchall()
        # §8：全案付款彙總＝這案子「花多少、欠多少」的核心數字（主管一頁看完的重點）。
        # 預計走付款排程、已付走實際核銷(closed)，兩者不重複計算（見需求書 §8）。
        planned_total = float(conn.execute(
            "SELECT COALESCE(SUM(ps.planned_amount),0) AS s FROM payment_schedules ps "
            "JOIN contracts c ON c.id = ps.contract_id "
            "WHERE c.case_id = ? AND ps.status <> 'cancelled'", (case_id,)).fetchone()["s"])
        paid_total = float(conn.execute(
            "SELECT COALESCE(SUM(CASE WHEN p.status='closed' THEN p.payment_amount ELSE 0 END),0) AS s "
            "FROM payments p JOIN contracts c ON c.id = p.contract_id WHERE c.case_id = ?",
            (case_id,)).fetchone()["s"])
        # 向下鑽取：案件層的「花多少、欠多少」再拆到每一份合約，主管才看得出是哪一份還欠。
        # 用 GROUP BY 一次撈完（不逐份合約發查詢，避免 N+1）。
        planned_by_ct = {r["cid"]: float(r["s"]) for r in conn.execute(
            "SELECT ps.contract_id AS cid, COALESCE(SUM(ps.planned_amount),0) AS s FROM payment_schedules ps "
            "JOIN contracts c ON c.id = ps.contract_id "
            "WHERE c.case_id = ? AND ps.status <> 'cancelled' GROUP BY ps.contract_id", (case_id,)).fetchall()}
        paid_by_ct = {r["cid"]: float(r["s"]) for r in conn.execute(
            "SELECT p.contract_id AS cid, "
            "COALESCE(SUM(CASE WHEN p.status='closed' THEN p.payment_amount ELSE 0 END),0) AS s "
            "FROM payments p JOIN contracts c ON c.id = p.contract_id "
            "WHERE c.case_id = ? GROUP BY p.contract_id", (case_id,)).fetchall()}
        # 續約/增購/整併：帶出「本約源自哪一份舊約」的編號，追溯鏈才看得出這不是全新的約
        parent_ids = {r["parent_contract_id"] for r in contracts if r["parent_contract_id"] is not None}
        parent_codes = {}
        if parent_ids:
            marks = ",".join("?" * len(parent_ids))
            parent_codes = {r["id"]: r["contract_code"] for r in conn.execute(
                f"SELECT id, contract_code FROM contracts WHERE id IN ({marks})", tuple(parent_ids)).fetchall()}
        contract_rows = []
        for row in contracts:
            item = dict(row)
            planned = planned_by_ct.get(item["id"], 0.0)
            paid = paid_by_ct.get(item["id"], 0.0)
            item["planned_total"] = planned
            item["paid_total"] = paid
            item["unpaid_planned"] = max(0.0, planned - paid)
            item["parent_contract_code"] = parent_codes.get(item.get("parent_contract_id"), "")
            contract_rows.append(item)
        return {
            "case": case,
            "contracts": contract_rows,
            "payments": payments,
            "documents": documents,
            "budgets": budgets,
            "projects": projects,
            "signoffs": signoffs,
            "purchases": purchases,
            "totals": {
                "contract_amount": sum(row["amount"] for row in contracts),
                "payment_amount": sum(row["payment_amount"] for row in payments),
                "document_count": len(documents),
                "budget_amount": sum(row["amount"] for row in budgets),
                "signoff_amount": sum(row["amount"] for row in signoffs),
                "purchase_amount": sum(row["amount"] for row in purchases),
                # §8 核心：預計付款總額 / 已付 / 還欠
                "planned_total": planned_total,
                "paid_total": paid_total,
                "unpaid_planned": max(0.0, planned_total - paid_total),
            },
        }


# ── 案件線性進度圖／處理優先矩陣：由系統自動推導，唯讀（不手改、不匯入） ──
# 五色燈：green=完成 white=還沒輪到 orange=東西到了/快逾期待處理 red=已逾期 na=不適用(灰)
# 一律畫 7 個階段；灰燈＝這案用不到（在「第一個有資料階段」之前又沒資料）
_STAGE_ORDER = ["budget", "project", "signoff", "contract", "purchase", "payment", "invoice"]
_STAGE_LABEL = {"budget": "預算", "project": "專案", "signoff": "簽呈",
                "contract": "合約", "purchase": "請購", "payment": "付款", "invoice": "發票"}
_TONE_RANK = {"na": -1, "green": 0, "white": 1, "orange": 2, "red": 3}
_ORANGE_WINDOW_DAYS = 7          # 距期限 7 天內＝橘燈（快逾期）
_AMOUNT_HIGH = 10_000_000        # 金額 ≥ 1000 萬＝矩陣「金額/影響高」


def _pdate(value: Any):
    """寬鬆解析日期字串（YYYY-MM-DD / YYYY/MM/DD），失敗回 None。"""
    s = str(value or "").strip().replace("/", "-")
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        return None


def _tone_by_deadline(deadline, *, overdue="red", near="orange", ok="white"):
    """依距今天數給燈：已過→overdue，7 天內→near，其餘→ok；無日期→ok。回 (tone, days)。"""
    if deadline is None:
        return ok, None
    days = (deadline - date.today()).days
    if days < 0:
        return overdue, days
    if days <= _ORANGE_WINDOW_DAYS:
        return near, days
    return ok, days


def _worst(tones: list[str]) -> str | None:
    """取最嚴重的燈（red>orange>white>green）；空清單回 None（該階段不適用）。"""
    tones = [t for t in tones if t]
    if not tones:
        return None
    return max(tones, key=lambda t: _TONE_RANK.get(t, 0))


def _month_tone(payment_month: str, *, done: bool):
    """付款/發票期程比對當月：已完成→green；期程已過未完→red；當月→orange；未來→white。回 (tone, days)。"""
    if done:
        return "green", None
    cur = date.today().strftime("%Y-%m")
    pm = str(payment_month or "").strip()[:7]
    if not pm:
        return "white", None
    # 以每月 1 日估算距今天數，供矩陣急迫度排序
    d = _pdate(pm + "-01")
    days = (d - date.today()).days if d else None
    if pm < cur:
        return "red", days
    if pm == cur:
        return "orange", days
    return "white", days


def _case_stage_lights(case, budgets, projects, signoffs, contracts, purchases, payments):
    """回傳 (stages, urgency_days)：一律 7 個階段 [{key,label,tone,days}]。
    有資料的階段照 綠/白/橘/紅 算；沒資料者：在「第一個有資料階段」之前＝灰(na,不適用)，之後＝白(還沒輪到)。
    natural[key] = (自然燈 or None, 急迫天數)；None 代表這階段沒有任何關聯資料。"""
    natural: dict[str, tuple[Any, Any]] = {}

    # 預算：有非停用預算＝已編列（綠）
    natural["budget"] = ("green", None) if budgets else (None, None)

    # 專案：completed/進度100＝綠；否則看 due_date/end_date
    p_tones, p_days = [], []
    for p in projects:
        try:
            prog = float(p["progress"] or 0)
        except (TypeError, ValueError):
            prog = 0.0
        if p["status"] == "completed" or prog >= 100:
            p_tones.append("green")
        else:
            t, d = _tone_by_deadline(_pdate(p["due_date"] or p["end_date"]))
            p_tones.append(t)
            if d is not None:
                p_days.append(d)
    pt = _worst(p_tones)
    natural["project"] = (pt, min(p_days) if (p_days and pt in ("orange", "red")) else None)

    # 簽呈：approved＝綠、rejected＝紅、其餘（草稿/送審）＝白
    s_tones = ["green" if s["status"] == "approved" else "red" if s["status"] == "rejected" else "white"
               for s in signoffs]
    natural["signoff"] = (_worst(s_tones), None)

    # 合約：closed＝用完（綠）；否則看到期日
    k_tones, k_days = [], []
    for k in contracts:
        if k["status"] == "closed":
            k_tones.append("green")
        else:
            t, d = _tone_by_deadline(_pdate(k["end_date"]))
            k_tones.append(t)
            if d is not None:
                k_days.append(d)
    kt = _worst(k_tones)
    natural["contract"] = (kt, min(k_days) if (k_days and kt in ("orange", "red")) else None)

    # 請購：closed＝用完（綠）；其餘（pending/ordered/arrived）＝白
    natural["purchase"] = (_worst(["green" if q["status"] == "closed" else "white" for q in purchases]), None)

    # 付款：closed＝付畢（綠）；期程已過未付＝紅、當月＝橘、未來＝白
    pay_tones, pay_days = [], []
    for p in payments:
        t, d = _month_tone(p["payment_month"], done=(p["status"] == "closed"))
        pay_tones.append(t)
        if d is not None and t in ("orange", "red"):
            pay_days.append(d)
    payt = _worst(pay_tones)
    natural["payment"] = (payt, min(pay_days) if pay_days else None)

    # 發票：verified＝核銷畢（綠）；否則依期程：已過＝紅、當月＝橘、未來＝白
    inv_tones, inv_days = [], []
    for p in payments:
        t, d = _month_tone(p["payment_month"], done=(p["invoice_status"] == "verified"))
        inv_tones.append(t)
        if d is not None and t in ("orange", "red"):
            inv_days.append(d)
    invt = _worst(inv_tones)
    natural["invoice"] = (invt, min(inv_days) if inv_days else None)

    # 第一個有資料的階段；之前的空階段＝灰(不適用)、之後的空階段＝白(還沒輪到)
    first_active = next((i for i, k in enumerate(_STAGE_ORDER) if natural[k][0] is not None), None)
    stages: list[dict[str, Any]] = []
    days_pool: list[int] = []
    for i, key in enumerate(_STAGE_ORDER):
        tone, days = natural[key]
        if tone is None:
            tone = "na" if (first_active is not None and i < first_active) else "white"
            days = None
        stages.append({"key": key, "label": _STAGE_LABEL[key], "tone": tone, "days": days})
        if days is not None and tone in ("orange", "red"):
            days_pool.append(days)
    urgency = min(days_pool) if days_pool else None
    return stages, urgency


def _case_block(stages: list[dict[str, Any]]) -> dict[str, str]:
    """目前卡點 pill：忽略灰燈(不適用)，取最嚴重的非綠階段。全（適用階段）綠＝完成。"""
    applicable = [s for s in stages if s["tone"] != "na"]
    if not applicable:
        return {"text": "尚未建立流程", "tone": "white"}
    active = [s for s in applicable if s["tone"] != "green"]
    if not active:
        return {"text": "完成", "tone": "green"}
    worst = max(active, key=lambda s: _TONE_RANK[s["tone"]])
    suffix = {"red": "已逾期", "orange": "待處理·近期限", "white": "處理中"}[worst["tone"]]
    return {"text": f"{worst['label']}{suffix}", "tone": worst["tone"]}


def case_progress_overview() -> dict[str, Any]:
    """案件線性進度圖＋處理優先矩陣的資料：每案八階段燈號、卡點、金額、急迫度、矩陣落點。"""
    scope = _owner_scope.get()
    items: list[dict[str, Any]] = []
    with connect() as conn:
        if scope is not None:
            cases = conn.execute(
                "SELECT * FROM cases WHERE owner = ? AND status <> 'disabled' ORDER BY id DESC",
                (scope,)).fetchall()
        else:
            cases = conn.execute(
                "SELECT * FROM cases WHERE status <> 'disabled' ORDER BY id DESC").fetchall()
        for c in cases:
            cid = c["id"]
            budgets = conn.execute("SELECT * FROM budgets WHERE case_id=? AND status<>'disabled'", (cid,)).fetchall()
            projects = conn.execute("SELECT * FROM projects WHERE case_id=? AND status<>'disabled'", (cid,)).fetchall()
            signoffs = conn.execute("SELECT * FROM signoffs WHERE case_id=? AND status<>'disabled'", (cid,)).fetchall()
            contracts = conn.execute("SELECT * FROM contracts WHERE case_id=? AND status<>'disabled'", (cid,)).fetchall()
            purchases = conn.execute("SELECT * FROM purchases WHERE case_id=? AND status<>'disabled'", (cid,)).fetchall()
            payments = conn.execute(
                "SELECT p.* FROM payments p JOIN contracts k ON k.id=p.contract_id "
                "WHERE k.case_id=? AND p.status<>'disabled'", (cid,)).fetchall()

            stages, urgency = _case_stage_lights(c, budgets, projects, signoffs, contracts, purchases, payments)
            block = _case_block(stages)
            # 階段別：done=全完成、not_started=還沒動(無綠也無急)、active=進行中/有風險（矩陣過濾用）
            tones = [s["tone"] for s in stages]
            if block["tone"] == "green":
                phase = "done"
            elif "green" not in tones and not any(t in ("orange", "red") for t in tones):
                phase = "not_started"
            else:
                phase = "active"
            # 金額：優先用案件金額，退回合約總額、預算總額
            amount = float(c["amount"] or 0) or sum(k["amount"] for k in contracts) or sum(b["amount"] for b in budgets)
            worst_tone = block["tone"]
            urgent = worst_tone in ("orange", "red")
            # 使用者拍板（2026-07-29）：金額不代表優先，只用時間當急迫度，不再切四象限。
            # reason 保留一句人看得懂的說明，回答「為什麼排這個位置」。
            if urgency is None:
                reason = "無期限可判斷"
            elif urgency < 0:
                reason = f"已逾期 {-urgency} 天"
            elif urgency <= _ORANGE_WINDOW_DAYS:
                reason = f"剩 {urgency} 天"
            else:
                reason = f"還有 {urgency} 天"
            # 落點：單一時間軸——x 依急迫度，逾期在右、還很久在左
            # X：逾期→最右、近期限→右、未來越久→左、無期限→偏左
            if urgency is None:
                x = 24.0
            elif urgency < 0:
                x = 68 + min(-urgency / 30.0, 1.0) * 24   # 逾期越久越右 68..92
            elif urgency <= _ORANGE_WINDOW_DAYS:
                x = 54 + (_ORANGE_WINDOW_DAYS - urgency) / _ORANGE_WINDOW_DAYS * 12  # 近 54..66
            else:
                x = 48 - min((urgency - _ORANGE_WINDOW_DAYS) / 90.0, 1.0) * 34       # 遠 48..14
            # 穩定微抖動（用 case_id，避免同值完全疊住、但重整不亂跳）
            x = max(6, min(94, round(x + (cid % 7) - 3)))
            # 只剩時間一軸，垂直位置純粹用來把重疊的點錯開，不代表任何意義
            y = 18 + (cid % 9) * 8
            items.append({
                "case_id": cid,
                "case_code": c["case_code"],
                "title": c["title"],
                "owner": c["owner"] or "",
                "fiscal_year": c["fiscal_year"] if "fiscal_year" in c.keys() else "",
                "seq": c["seq"] if "seq" in c.keys() else 0,
                "amount": amount,
                "stages": stages,
                "block": block,
                "phase": phase,
                "urgency_days": urgency,
                "matrix": {"reason": reason, "x": x, "y": y, "tone": worst_tone},
            })
    return {"items": items}


# ── Step 3：舊資料補號（系統編號要件 fiscal_year+seq、核銷編號 settle_no）──
# 冪等：只補「缺號」的列，已有號的不動；由管理員手動觸發（先 preview 後正式）。
# 補號時要跳過的狀態：這些案件本來就不該佔用正式號。
# rejected（駁回）與 merged（併入他案）尤其不能補——它們佔住一個號就是永久跳號，
# 那個號永遠不會有有效案件對應，正好違反「沒過的申請不吃正式號」（需求書 §4 ＋ A 案）。
NO_NUMBER_STATUSES = ("disabled", "rejected", "merged")
_NO_NUMBER_LIST = ", ".join(f"'{s}'" for s in NO_NUMBER_STATUSES)


def backfill_status() -> dict[str, Any]:
    """回報還有多少舊資料缺號，並按狀態分組——按下補號之前要看得到會動到哪些案件
    （匯入來的草稿要補，但如果裡面混著真的還在等審核的新申請，那是不同的事）。"""
    with connect() as conn:
        cases_missing = conn.execute(
            f"SELECT COUNT(*) n FROM cases WHERE status NOT IN ({_NO_NUMBER_LIST}) "
            "AND (COALESCE(fiscal_year,'')='' OR COALESCE(seq,0)=0)").fetchone()["n"]
        by_status = {r["status"]: r["n"] for r in conn.execute(
            f"SELECT status, COUNT(*) n FROM cases WHERE status NOT IN ({_NO_NUMBER_LIST}) "
            "AND (COALESCE(fiscal_year,'')='' OR COALESCE(seq,0)=0) GROUP BY status")}
        skipped = {r["status"]: r["n"] for r in conn.execute(
            f"SELECT status, COUNT(*) n FROM cases WHERE status IN ({_NO_NUMBER_LIST}) "
            "AND COALESCE(seq,0)=0 GROUP BY status")}
        settle_missing = conn.execute(
            "SELECT COUNT(*) n FROM payments WHERE status <> 'disabled' "
            "AND COALESCE(settle_no,'')=''").fetchone()["n"]
        case_link_missing = conn.execute(
            "SELECT (SELECT COUNT(*) FROM budgets WHERE status <> 'disabled' AND (case_id IS NULL OR case_id = 0)) + "
            "(SELECT COUNT(*) FROM projects WHERE status <> 'disabled' AND (case_id IS NULL OR case_id = 0)) n"
        ).fetchone()["n"]
    return {"cases_missing": cases_missing, "cases_by_status": by_status, "skipped_by_status": skipped,
            "settle_missing": settle_missing, "case_link_missing": case_link_missing}


def case_code_cleanup_plan() -> dict[str, Any]:
    """列出「系統配的案件編號」有哪些不合新規則（主管 2026-08-03：不得含 - _ 中文），
    以及會被改成什麼。先看再按，不要讓人按下去才知道動到誰。

    只動系統自己配的號：`source_file` 有值的是 Excel 帶進來的真編號（別人的號），
    改掉就對不回原始檔，一律跳過並列在 kept 裡。
    """
    with connect() as conn:
        rows = conn.execute(
            "SELECT id, case_code, source_file, fiscal_year, seq, temp_seq FROM cases ORDER BY id"
        ).fetchall()
        taken = {str(r["case_code"] or "") for r in rows}
        changes: list[dict[str, Any]] = []
        kept: list[dict[str, Any]] = []
        for r in rows:
            code = str(r["case_code"] or "")
            if is_system_code_valid(code):
                continue
            if str(r["source_file"] or "").strip():
                kept.append({"id": r["id"], "case_code": code, "reason": "Excel 匯入帶進來的原始編號"})
                continue
            fy = str(r["fiscal_year"] or "").strip() or get_working_year()
            if code.upper().startswith("TMP-"):
                new = "TMP" + code[4:].replace("-", "")      # 只是把連字號拿掉，號碼本身不變
            elif int(r["seq"] or 0) > 0:
                new = f"{fy}{int(r['seq']):04d}"             # 已配正式號 → 用正式號當編號
            else:
                new = f"TMP{fy}{int(r['temp_seq'] or 0):04d}"
            base, n = new, 1
            while new in taken:                              # 撞號往後掛 A02、A03（不用連字號）
                n += 1
                new = f"{base}A{n:02d}"
            taken.discard(code)
            taken.add(new)
            changes.append({"id": r["id"], "from": code, "to": new})
    return {"changes": changes, "kept": kept,
            "change_count": len(changes), "kept_count": len(kept)}


def backfill_case_codes() -> dict[str, Any]:
    """依 case_code_cleanup_plan 實際換號，逐筆寫稽核（換號是對外可見的事，要留紀錄）。
    冪等：已經合規的不會再動，重跑不會一直換號。"""
    plan = case_code_cleanup_plan()
    actor = _current_actor.get()
    with connect() as conn:
        for item in plan["changes"]:
            before = get_row(conn, "cases", item["id"])
            conn.execute("UPDATE cases SET case_code = ? WHERE id = ?", (item["to"], item["id"]))
            after = get_row(conn, "cases", item["id"])
            write_audit_log(conn, "cases", item["id"], "recode", before,
                            {**after, "recode_by": actor, "recode_reason": "編號規則：只能英數"})
    return {"changed": plan["change_count"], "kept": plan["kept_count"], "details": plan["changes"]}


def backfill_case_numbers() -> int:
    """回填舊案件的系統編號要件：缺 fiscal_year 用 created_at 年度、缺 seq 於該年度續號。

    跳過 NO_NUMBER_STATUSES：被駁回／已併入他案的不能補號，它們佔住一個正式號就是
    永久跳號（那個號永遠沒有有效案件對應）。這支是 v0.46 時代寫的，當時所有案件在建立
    當下就配號、沒有「暫時號」概念，所以原本只排除 disabled——2026-07-30 實測發現它會
    把剛被駁回的案件也配上正式號，才補上這道過濾。
    """
    filled = 0
    with connect() as conn:
        rows = conn.execute(
            f"SELECT id, fiscal_year, seq, created_at FROM cases "
            f"WHERE status NOT IN ({_NO_NUMBER_LIST}) AND (COALESCE(fiscal_year,'')='' OR COALESCE(seq,0)=0) "
            "ORDER BY created_at, id").fetchall()
        for r in rows:
            fy = str(r["fiscal_year"] or "").strip()
            if not fy:
                created = str(r["created_at"] or "")
                fy = created[:4] if created[:4].isdigit() else get_working_year()
            nxt = conn.execute(
                "SELECT COALESCE(MAX(seq),0)+1 n FROM cases WHERE fiscal_year=?", (fy,)).fetchone()["n"]
            conn.execute("UPDATE cases SET fiscal_year=?, seq=? WHERE id=?", (fy, nxt, r["id"]))
            filled += 1
    return filled


def _prepare_expense_master(conn: sqlite3.Connection, fields: dict[str, Any]) -> dict[str, Any]:
    """第一層費用主檔的帶入與檢核（助理 0803 附件一第四節）。

    有合約：廠商、統編、期間、總費用、承辦人一律由合約主檔帶入。總費用是**例外欄位**——
      助理寫明有合約時唯讀反灰、不得人工修改，所以這裡直接以合約為準覆蓋送進來的值，
      切換關聯合約時金額也會跟著換。
    無合約：合約起迄日清空（助理寫「停用、不得輸入」），總費用改由人工填。
    兩種情形總費用都必填且要大於 0——它是第二層所有區段加總的檢核基準，0 的話整個檢核失去意義。
    """
    fields = dict(fields)
    contract_id = fields.get("contract_id")
    if contract_id:
        k = conn.execute(
            "SELECT contract_name, vendor_name, vendor_tax_id, start_date, end_date, amount, owner, case_id "
            "FROM contracts WHERE id = ?", (int(contract_id),)).fetchone()
        if k is None:
            raise ValueError(f"關聯的合約 ID {contract_id} 不存在，請確認後再填。")
        fields["total_amount"] = float(k["amount"] or 0)          # 唯讀：以合約為準
        for src, dst in (("contract_name", "expense_name"), ("vendor_name", "vendor_name"),
                         ("vendor_tax_id", "vendor_tax_id"), ("start_date", "start_date"),
                         ("end_date", "end_date"), ("owner", "owner"), ("case_id", "case_id")):
            if not str(fields.get(dst) or "").strip():
                fields[dst] = k[src]
    else:
        fields["start_date"] = ""
        fields["end_date"] = ""
    if float(fields.get("total_amount") or 0) <= 0:
        raise ValueError("合約總費用（含稅）必填，且要大於 0——第二層所有費用區段都用它做加總檢核。")
    modes = [m.strip() for m in str(fields.get("modes") or "").split(",") if m.strip()]
    if not modes:
        raise ValueError("請至少選一種費用排程模式（里程碑／定期費用／最低承諾金額，可複選）。")
    bad = [m for m in modes if m not in EXPENSE_MODE_LABEL]
    if bad:
        raise ValueError(f"不認得的費用排程模式：{'、'.join(bad)}。")
    fields["modes"] = ",".join(modes)
    if not str(fields.get("signoff_ref") or "").strip() and not str(fields.get("signoff_none_reason") or "").strip():
        raise ValueError("請填簽呈／請購編號；沒有編號的話，請在「無編號原因」說明為什麼。")
    return fields


# ── 費用模組第二層：排程產生、金額檢核、預覽與確認（助理 0803 附件一第五節）──────
EXPENSE_MODE_LABEL = {"milestone": "里程碑", "periodic": "定期費用", "commitment": "最低承諾金額"}
MILESTONE_NAMES = ["簽約款", "交付款", "驗收款", "自訂"]
_FREQ_MONTHS = {"monthly": 1, "quarterly": 3, "semi": 6, "yearly": 12}


def _month_add(ym: str, months: int) -> str:
    """'2026-01' 往後 n 個月 → '2026-04'。給定期費用順延費用年月用。"""
    raw = str(ym or "").strip().replace("/", "-")[:7]
    if len(raw) < 7 or not raw[:4].isdigit():
        return ""
    y, m = int(raw[:4]), int(raw[5:7])
    total = y * 12 + (m - 1) + months
    return f"{total // 12:04d}-{total % 12 + 1:02d}"


def _date_add_months(ds: str, months: int) -> str:
    raw = str(ds or "").strip()[:10]
    try:
        d = datetime.strptime(raw, "%Y-%m-%d").date()
    except ValueError:
        return ""
    return _add_months_date(d, months).isoformat()


def generate_section_schedules(section_id: int) -> dict[str, Any]:
    """依費用區段的設定產生排程明細。

    里程碑：依總期數產生 N 筆「可編輯的空白列」——助理明確寫過不得只存第一期後由系統推測，
      各期名稱/比例/金額/日期都要能各自輸入。比例計價時金額由系統算（區段金額×該期比例）。
    定期費用：依第一期資料＋頻率推算後續各期（費用年月與預計應付日一起順延），
      金額先全部帶第一期，之後可在預覽畫面逐期改。

    重跑會重建：只清掉這個區段自己的明細再重產，人工調過的值會不見——所以呼叫端要先擋住
    已確認的區段（confirmed 要走 reopen 建新版本）。
    """
    with connect() as conn:
        sec = get_row(conn, "expense_sections", section_id)
        if sec["status"] == "confirmed":
            raise RuntimeError("這個費用區段已確認，要改請先『重新編輯』（會建立新版本並保留原版）。")
        conn.execute("DELETE FROM expense_schedules WHERE section_id = ?", (section_id,))
        mode = sec["mode"]
        rows: list[dict[str, Any]] = []
        if mode == "milestone":
            n = int(sec["periods"] or 0)
            if n < 1:
                raise ValueError("里程碑總期數要填 1 以上的整數。")
            for i in range(1, n + 1):
                rows.append({"seq": i, "milestone_name": "", "percent": 0, "planned_amount": 0})
        elif mode == "periodic":
            n = int(sec["periods"] or 0)
            step = _FREQ_MONTHS.get(str(sec["frequency"] or ""), 0)
            if n < 1 or not step:
                raise ValueError("定期費用要填費用頻率與期數（期數為 1 以上的整數）。")
            for i in range(n):
                rows.append({
                    "seq": i + 1,
                    "planned_amount": float(sec["first_amount"] or 0),
                    "expense_month": _month_add(sec["first_month"], i * step),
                    "due_date": _date_add_months(sec["first_due_date"], i * step),
                })
        elif mode == "commitment":
            rows = _commitment_rows(sec)
        else:
            raise ValueError(f"{EXPENSE_MODE_LABEL.get(mode, mode)}模式的排程產生還沒開放。")
        for r in rows:
            _insert_row(conn, "expense_schedules", {**r, "section_id": section_id})
    return section_preview(section_id)


def _commitment_rows(sec: Any) -> list[dict[str, Any]]:
    """最低承諾金額（助理 0803 附件一 5.3）：先算各承諾期間的起訖，再依費用頻率在期間內鋪排程。

    承諾金額是「這段期間至少要用掉多少」的門檻，不是每期一次付清——所以排程列的金額是
    該承諾期的承諾金額攤到各期，實際發生多少要等第三層登錄後才算得出達成率。
    即使承諾金額提前達成，仍保留後續排程到期間屆滿（助理明講）。
    """
    n = int(sec["periods"] or 0)                      # 承諾期數
    span = int(sec["commit_span_months"] or 0)        # 每期期間長度（月）
    step = _FREQ_MONTHS.get(str(sec["frequency"] or ""), 0)
    start = str(sec["period_start"] or "").strip()
    if n < 1 or span < 1 or not step:
        raise ValueError("最低承諾金額要填承諾期數、每期期間長度（月）與費用頻率。")
    if not start:
        raise ValueError("最低承諾金額要填第一期承諾起日。")
    if span % step:
        raise ValueError(
            f"每期期間長度 {span} 個月，除不盡費用頻率（{FREQ_LABEL_ZH.get(sec['frequency'], sec['frequency'])}）；"
            "請調整成整數倍，否則同一期會被切一半。")
    rule = str(sec["next_amount_rule"] or "same")
    growth = float(sec["growth_pct"] or 0)
    rows: list[dict[str, Any]] = []
    seq = 0
    amount = float(sec["first_amount"] or 0)
    for p in range(n):                                 # 每個承諾期
        p_start = _date_add_months(start, p * span)
        per_period = round(amount / (span // step), 2) if span >= step else amount
        for k in range(span // step):
            seq += 1
            offset = p * span + k * step
            rows.append({
                "seq": seq,
                "commit_period": p + 1,
                "planned_amount": per_period,
                "expense_month": _month_add(p_start[:7], k * step),
                "billing_start": _date_add_months(start, offset),
                "billing_end": _date_add_months(start, offset + step),
                "note": f"第 {p + 1} 承諾期",
            })
        if rule == "growth":
            amount = round(amount * (1 + growth / 100), 2)
    return rows


FREQ_LABEL_ZH = {"monthly": "每月", "quarterly": "每季", "semi": "每半年", "yearly": "每年"}


def commitment_achievement(section_id: int) -> dict[str, Any]:
    """各承諾期的達成情形（助理 0803）：承諾金額、實際認列、達成率、未達差額、超額與轉入。

    認列金額＝第三層登錄的實際費用（依設定的認列基礎），沒登錄的期別就是 0——
    這時要顯示「尚未登錄」而不是「達成率 0%」，兩者意思完全不同。
    """
    with connect() as conn:
        sec = dict(get_row(conn, "expense_sections", section_id))
        rows = [dict(r) for r in conn.execute(
            "SELECT * FROM expense_schedules WHERE section_id = ? ORDER BY seq", (section_id,)).fetchall()]
        actuals = {r["schedule_id"]: r for r in conn.execute(
            "SELECT schedule_id, SUM(recognized_amount) AS recognized, COUNT(*) AS n "
            "FROM expense_actuals WHERE section_id = ? GROUP BY schedule_id", (section_id,)).fetchall()}
    periods: dict[int, dict[str, Any]] = {}
    for r in rows:
        p = int(r["commit_period"] or 0)
        d = periods.setdefault(p, {"commit_period": p, "committed": 0.0, "recognized": 0.0,
                                   "logged": 0, "schedule_count": 0})
        d["committed"] += float(r["planned_amount"] or 0)
        d["schedule_count"] += 1
        act = actuals.get(r["id"])
        if act:
            d["recognized"] += float(act["recognized"] or 0)
            d["logged"] += int(act["n"] or 0)
    out = []
    carry = 0.0
    for p in sorted(periods):
        d = periods[p]
        committed = round(d["committed"], 2)
        recognized = round(d["recognized"] + (carry if int(sec["carry_over"] or 0) else 0), 2)
        shortfall = round(max(committed - recognized, 0), 2)
        excess = round(max(recognized - committed, 0), 2)
        carry = excess if int(sec["carry_over"] or 0) else 0.0
        out.append({**d, "committed": committed, "recognized": recognized,
                    "shortfall": shortfall, "excess": excess,
                    # 一筆都沒登錄 → 給 None 表示「還不知道」，不要拿 0% 混充「沒達成」
                    "rate": None if not d["logged"] else (round(recognized / committed * 100, 1) if committed else None),
                    "carry_in_next": carry})
    return {"section_id": section_id, "carry_over": bool(int(sec["carry_over"] or 0)),
            "basis": sec["achievement_basis"] or "usage",
            "shortfall_action": sec["shortfall_action"], "periods": out}


def _money_eq(a: float, b: float) -> bool:
    """金額比對容忍到分（浮點數直接比會因為 0.1+0.2 這種誤差誤報不符）。"""
    return abs(round(float(a or 0) - float(b or 0), 2)) < 0.01


def section_preview(section_id: int) -> dict[str, Any]:
    """放大檢視用的排程預覽＋檢核結果。

    檢核不通過時要講清楚差多少、差在哪一段——只說「檢核失敗」等於要人自己去猜。
    """
    with connect() as conn:
        sec = dict(get_row(conn, "expense_sections", section_id))
        rows = [dict(r) for r in conn.execute(
            "SELECT * FROM expense_schedules WHERE section_id = ? ORDER BY seq, id", (section_id,)).fetchall()]
        master = dict(get_row(conn, "expense_masters", sec["expense_id"]))
    problems: list[str] = []
    total = round(sum(float(r["planned_amount"] or 0) for r in rows), 2)
    section_amount = float(sec["section_amount"] or 0)
    if not rows:
        problems.append("還沒有排程明細，請先產生排程。")
    if sec["mode"] == "milestone":
        if sec["price_method"] == "percent":
            pct = round(sum(float(r["percent"] or 0) for r in rows), 4)
            if not _money_eq(pct, 100):
                problems.append(f"各期比例合計 {pct}%，要等於 100%（差 {round(100 - pct, 4)}%）。")
        missing = [r["seq"] for r in rows if not str(r["milestone_name"] or "").strip()]
        if missing:
            problems.append(f"第 {'、'.join(str(s) for s in missing)} 期還沒選里程碑名稱。")
        custom_missing = [r["seq"] for r in rows
                          if str(r["milestone_name"] or "") == "自訂" and not str(r["custom_name"] or "").strip()]
        if custom_missing:
            problems.append(f"第 {'、'.join(str(s) for s in custom_missing)} 期選了「自訂」，自訂里程碑備註必填。")
    if rows and not _money_eq(total, section_amount):
        diff = round(section_amount - total, 2)
        problems.append(
            f"各期應付費用合計 {total:,.0f} 元，與費用區段金額 {section_amount:,.0f} 元"
            f"差 {diff:,.0f} 元（{'少' if diff > 0 else '多'}了）。")
    return {"section": sec, "schedules": rows, "master": master,
            "scheduled_total": total, "section_amount": section_amount,
            "problems": problems, "can_confirm": not problems}


def expense_master_check(expense_id: int) -> dict[str, Any]:
    """第一層總費用 vs 第二層各費用區段的加總（混合型時每個模式一個區段）。"""
    with connect() as conn:
        master = dict(get_row(conn, "expense_masters", expense_id))
        # 只算現行版本：重新編輯留下來的舊版標了 archived，再算一次會讓總額憑空翻倍
        sections = [dict(r) for r in conn.execute(
            "SELECT * FROM expense_sections WHERE expense_id = ? AND archived = 0 ORDER BY id",
            (expense_id,)).fetchall()]
    section_total = round(sum(float(s["section_amount"] or 0) for s in sections), 2)
    total = float(master["total_amount"] or 0)
    diff = round(total - section_total, 2)
    modes = [m for m in str(master["modes"] or "").split(",") if m]
    missing = [EXPENSE_MODE_LABEL.get(m, m) for m in modes
               if not any(s["mode"] == m for s in sections)]
    return {"expense_id": expense_id, "total_amount": total, "section_total": section_total,
            "diff": diff, "balanced": _money_eq(total, section_total),
            "missing_sections": missing, "sections": sections}


def confirm_section(section_id: int) -> dict[str, Any]:
    """確認排程：檢核全過才准，並記下確認人與時間（助理 0803 要求留存）。"""
    actor = _current_actor.get()
    preview = section_preview(section_id)
    if preview["problems"]:
        raise ValueError("；".join(preview["problems"]))
    with connect() as conn:
        before = get_row(conn, "expense_sections", section_id)
        conn.execute(
            "UPDATE expense_sections SET status = 'confirmed', confirmed_by = ?, confirmed_at = ? WHERE id = ?",
            (actor, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), section_id))
        after = get_row(conn, "expense_sections", section_id)
        write_audit_log(conn, "expense_sections", section_id, "confirm", before, after)
    return dict(after)


def reopen_section(section_id: int) -> dict[str, Any]:
    """已確認的排程要改：建立新版本並保留原版（助理 0803「應保留原排程版本」）。

    原版整段複製成一筆 confirmed 的舊版紀錄（含明細），目前這筆退回草稿讓人改——
    直接就地改的話，之前確認過什麼就查不到了。
    """
    actor = _current_actor.get()
    with connect() as conn:
        sec = dict(get_row(conn, "expense_sections", section_id))
        if sec["status"] != "confirmed":
            raise RuntimeError("只有『已確認』的費用區段需要重新編輯；草稿直接改就好。")
        keep = {k: v for k, v in sec.items() if k in allowed_fields()["expense_sections"]}
        keep.update({"status": "confirmed", "version": int(sec["version"] or 1)})
        archived = _insert_row(conn, "expense_sections", keep)
        # 舊版標成已封存：清單仍看得到（查得到當初確認了什麼），但金額不再進任何加總
        conn.execute(
            "UPDATE expense_sections SET confirmed_by = ?, confirmed_at = ?, archived = 1 WHERE id = ?",
            (sec["confirmed_by"], sec["confirmed_at"], archived["id"]))
        for r in conn.execute("SELECT * FROM expense_schedules WHERE section_id = ?", (section_id,)).fetchall():
            row = {k: v for k, v in dict(r).items() if k in allowed_fields()["expense_schedules"]}
            _insert_row(conn, "expense_schedules", {**row, "section_id": archived["id"]})
        conn.execute(
            "UPDATE expense_sections SET status = 'draft', version = ?, confirmed_by = '', confirmed_at = '' "
            "WHERE id = ?", (int(sec["version"] or 1) + 1, section_id))
        after = get_row(conn, "expense_sections", section_id)
        write_audit_log(conn, "expense_sections", section_id, "reopen", sec,
                        {**dict(after), "archived_section_id": archived["id"]})
    return {"section": dict(after), "archived_section_id": archived["id"], "actor": actor}


def list_expense_sections(expense_id: int) -> list[dict[str, Any]]:
    with connect() as conn:
        rows = conn.execute(
            "SELECT * FROM expense_sections WHERE expense_id = ? ORDER BY id", (expense_id,)).fetchall()
        out = []
        for r in rows:
            d = dict(r)
            d["schedule_count"] = conn.execute(
                "SELECT COUNT(*) n FROM expense_schedules WHERE section_id = ?", (r["id"],)).fetchone()["n"]
            out.append(d)
    return out


# ── 費用模組第三層：實際費用明細與請款／核銷（助理 0803 附件一第六節）────────────
SETTLE_PROGRESS_LABEL = {
    "invoice_pending": "發票尚未收到", "ready_to_sign": "可預備上簽",
    "signing": "款項簽核中", "approved": "款項已核准", "submitted": "提交會計（結案）",
}


def add_expense_actual(schedule_id: int, payload: dict[str, Any]) -> dict[str, Any]:
    """登錄最低承諾金額的當期實際費用。認列金額由系統算（使用金額＋調整金額），不讓人手填——
    手填的話跟兩個來源欄位對不起來，達成率就查不出是怎麼算的。"""
    with connect() as conn:
        sched = get_row(conn, "expense_schedules", schedule_id)
        sec = get_row(conn, "expense_sections", sched["section_id"])
        if sec["mode"] != "commitment":
            raise ValueError("實際費用明細只適用「最低承諾金額」模式；其他模式請直接建立請款／核銷資料。")
        if sec["status"] != "confirmed":
            raise RuntimeError("這個費用區段還沒確認排程，先確認後再登錄實際費用。")
        usage = float(payload.get("usage_amount") or 0)
        adjust = float(payload.get("adjust_amount") or 0)
        if adjust and not str(payload.get("adjust_reason") or "").strip():
            raise ValueError("有填調整金額就要寫調整原因（折讓、退款或前期調整都要說明）。")
        row = _insert_row(conn, "expense_actuals", {
            **{k: v for k, v in payload.items() if k in allowed_fields()["expense_actuals"]},
            "section_id": sched["section_id"],
            "schedule_id": schedule_id,
            "commit_period": sched["commit_period"],
            "billing_start": payload.get("billing_start") or sched["billing_start"],
            "billing_end": payload.get("billing_end") or sched["billing_end"],
            "recognized_amount": round(usage + adjust, 2),
        })
    return dict(row)


def create_settlement(schedule_id: int, payload: dict[str, Any]) -> dict[str, Any]:
    """建立請款／核銷資料。

    助理 0803 的硬規則：一次作業只能對「一筆已確認的排程」＋「一張發票」，不得複選。
    所以這裡用 schedule_id 當入口，其餘關聯（費用主檔、區段、廠商、計費期間）全部由系統帶，
    人只填發票與請款金額——欄位少一個就少一個填錯的機會。
    """
    with connect() as conn:
        sched = get_row(conn, "expense_schedules", schedule_id)
        sec = get_row(conn, "expense_sections", sched["section_id"])
        master = get_row(conn, "expense_masters", sec["expense_id"])
        if sec["status"] != "confirmed":
            raise RuntimeError("這筆排程所屬的費用區段還沒確認，確認後才能建立請款／核銷資料。")
        inv = str(payload.get("invoice_no") or "").strip()
        if inv:
            dup = conn.execute(
                "SELECT id FROM expense_settlements WHERE invoice_no = ? AND schedule_id = ?",
                (inv, schedule_id)).fetchone()
            if dup is not None:
                raise ValueError(f"發票號碼 {inv} 已經對這一期排程建過請款資料了（第 {dup['id']} 筆）。")
        claim = float(payload.get("claim_amount") or 0)
        diff = round(float(sched["planned_amount"] or 0) - claim, 2)
        if abs(diff) >= 0.01 and not str(payload.get("diff_reason") or "").strip():
            raise ValueError(
                f"本次請款 {claim:,.0f} 元與排程應付 {float(sched['planned_amount'] or 0):,.0f} 元"
                f"差 {abs(diff):,.0f} 元，請填差異原因。")
        row = _insert_row(conn, "expense_settlements", {
            **{k: v for k, v in payload.items() if k in allowed_fields()["expense_settlements"]},
            "expense_id": sec["expense_id"],
            "section_id": sched["section_id"],
            "schedule_id": schedule_id,
            "settle_month": payload.get("settle_month") or sched["expense_month"],
            "billing_start": payload.get("billing_start") or sched["billing_start"],
            "billing_end": payload.get("billing_end") or sched["billing_end"],
            "vendor_name": payload.get("vendor_name") or master["vendor_name"],
            "vendor_tax_id": payload.get("vendor_tax_id") or master["vendor_tax_id"],
            "progress": payload.get("progress") or "invoice_pending",
        })
    return settlement_view(row["id"])


def settlement_view(settlement_id: int) -> dict[str, Any]:
    """單筆請款／核銷＋系統算出來的請款差異（排程應付 − 本次請款）。"""
    with connect() as conn:
        row = dict(get_row(conn, "expense_settlements", settlement_id))
        sched = dict(get_row(conn, "expense_schedules", row["schedule_id"]))
    planned = float(sched["planned_amount"] or 0)
    row["planned_amount"] = planned
    row["claim_diff"] = round(planned - float(row["claim_amount"] or 0), 2)
    row["progress_label"] = SETTLE_PROGRESS_LABEL.get(row["progress"], row["progress"])
    # 助理：「確認完成」只在進度為「可預備上簽」時出現
    row["can_confirm"] = row["progress"] == "ready_to_sign"
    return row


def list_settlements(expense_id: int) -> dict[str, Any]:
    """某費用主檔底下的請款／核銷，附累計與待請款（助理 0803 第七節的自動計算）。"""
    with connect() as conn:
        rows = [dict(r) for r in conn.execute(
            "SELECT * FROM expense_settlements WHERE expense_id = ? ORDER BY id DESC", (expense_id,)).fetchall()]
        scheduled = conn.execute(
            "SELECT COALESCE(SUM(sc.planned_amount), 0) AS total FROM expense_schedules sc "
            "JOIN expense_sections se ON se.id = sc.section_id "
            "WHERE se.expense_id = ? AND se.archived = 0",   # 封存的舊版不再計入排程總額
            (expense_id,)).fetchone()["total"]
    claimed = round(sum(float(r["claim_amount"] or 0) for r in rows), 2)
    for r in rows:
        r["progress_label"] = SETTLE_PROGRESS_LABEL.get(r["progress"], r["progress"])
    return {"expense_id": expense_id, "settlements": rows,
            "scheduled_total": round(float(scheduled or 0), 2),
            "claimed_total": claimed,
            "unclaimed_total": round(float(scheduled or 0) - claimed, 2)}


def update_settlement_progress(settlement_id: int, progress: str | None = None,
                               confirmed: bool | None = None) -> dict[str, Any]:
    """推進處理進度／勾確認完成，並回傳「這一步要通知誰」。

    助理 0803 的通知邏輯：承辦自己備齊 → 勾確認完成 → 通知核銷者；
    行政整理好改成「可預備上簽」→ 通知承辦確認 → 承辦勾完成 → 再通知核銷者。
    這裡只回傳該通知誰，實際寄信走既有的通知模組（沒設 SMTP 時本來就只會記錄不寄）。
    """
    with connect() as conn:
        before = dict(get_row(conn, "expense_settlements", settlement_id))
        fields: dict[str, Any] = {}
        if progress is not None:
            if progress not in SETTLE_PROGRESS_LABEL:
                raise ValueError(f"不認得的處理進度：{progress}。")
            fields["progress"] = progress
        if confirmed is not None:
            if confirmed and (fields.get("progress", before["progress"]) != "ready_to_sign"):
                raise ValueError("只有處理進度是「可預備上簽」時才能勾確認完成。")
            fields["confirmed"] = 1 if confirmed else 0
        if not fields:
            raise ValueError("沒有要更新的內容。")
        sets = ", ".join(f"{k} = ?" for k in fields)
        conn.execute(f"UPDATE expense_settlements SET {sets} WHERE id = ?",
                     [*fields.values(), settlement_id])
        after = get_row(conn, "expense_settlements", settlement_id)
        write_audit_log(conn, "expense_settlements", settlement_id, "progress", before, after)
    notify = ""
    if fields.get("confirmed"):
        notify = "settler"      # 承辦確認完成 → 通知核銷者
    elif fields.get("progress") == "ready_to_sign":
        notify = "owner"        # 行政備妥 → 通知承辦確認
    return {**settlement_view(settlement_id), "notify": notify}


# 標準採購流程的工作項（黃助理 0803 附件二第三點，另一位助理 0807 的流程圖也是同一份）。
# 助理原話：「系統不預先限制 WBS 工作項目名稱，上述僅為建議的標準工作項目」——
# 所以做成後台可維護的清單，承辦仍可自己增刪，這裡只是「勾了涉及請購或合約就先幫你排好」。
STANDARD_WBS_ITEMS = ["需求確認", "廠商報價", "上簽申請與核准", "議價", "合約簽訂", "執行／建置", "驗收", "結案"]


def standard_wbs_items() -> list[str]:
    raw = read_settings(["opt_wbs_standard_items"]).get("opt_wbs_standard_items", "")
    items = [x.strip() for x in raw.split(",") if x.strip()]
    return items or list(STANDARD_WBS_ITEMS)


def apply_standard_wbs(project_id: int, owner: str = "") -> dict[str, Any]:
    """把標準採購流程的工作項排進這個專案。

    冪等：同名工作項已經有了就跳過，不重複建立也不覆蓋既有內容（承辦可能已經填了進度）。
    每一項都是完整的 WBS 項目（有負責人、起訖、子項目數、燈號、關鍵風險點），
    不是流程圖上的文字節點——助理特別強調過這點。
    日期留空：這時候還不知道各階段何時做，硬塞日期會讓燈號一建好就亂判。
    """
    with connect() as conn:
        project = get_row(conn, "projects", project_id)
        exist = {str(r["item_name"]).strip()
                 for r in conn.execute("SELECT item_name FROM project_items WHERE project_id = ?",
                                       (project_id,)).fetchall()}
        seq = conn.execute(
            "SELECT COALESCE(MAX(seq), 0) AS n FROM project_items WHERE project_id = ?",
            (project_id,)).fetchone()["n"]
        created, skipped = [], []
        for name in standard_wbs_items():
            if name in exist:
                skipped.append(name)
                continue
            seq += 1
            _insert_row(conn, "project_items", {
                "project_id": project_id,
                "seq": seq,
                "item_name": name,
                "owner": owner or project["owner"] or "",
                "sub_total": 0,
                "sub_done": 0,
                "progress": 0,
                "rag": "",          # 沒起訖日就先不判燈號，等承辦填了日期再自動判
                "status": "active",
            })
            created.append(name)
    recompute_project_rollup(project_id)
    return {"project_id": project_id, "created": created, "skipped": skipped,
            "created_count": len(created), "skipped_count": len(skipped)}


def backfill_contract_system_codes() -> dict[str, int]:
    """手動補合約系統識別碼（後台備援）。

    正常情況這裡會回 0——開機時 `_fill_missing_contract_system_codes()` 已經補完了。
    留這支是為了「補完之後又從別處灌進沒有識別碼的資料」這種情況，不用重開服務。
    補號規則與開機那支同一份，不另外寫一套。
    """
    with connect() as conn:
        filled = _fill_missing_contract_system_codes(conn)
    return {"filled": filled}


def backfill_settle_numbers() -> int:
    """回填舊付款的核銷編號：年度取核銷月份，於該年度續號（沿用自動發號規則）。"""
    filled = 0
    with connect() as conn:
        rows = conn.execute(
            "SELECT id, payment_month FROM payments "
            "WHERE status <> 'disabled' AND COALESCE(settle_no,'')='' "
            "ORDER BY payment_month, id").fetchall()
        for r in rows:
            pm = str(r["payment_month"] or "").strip()
            year = pm[:4] if (len(pm) >= 4 and pm[:4].isdigit()) else get_working_year()
            nxt = conn.execute(
                "SELECT COALESCE(MAX(settle_seq),0)+1 n FROM payments "
                "WHERE substr(payment_month,1,4)=? AND settle_seq>0", (year,)).fetchone()["n"]
            conn.execute("UPDATE payments SET settle_seq=?, settle_no=? WHERE id=?",
                         (nxt, f"{SETTLE_PREFIX}{year}{nxt:04d}", r["id"]))
            filled += 1
    return filled


def backfill_case_links() -> int:
    """回填舊有預算/專案的案件關聯：使用者拍板「匯進來就該自動配一個案件」不只套新資料、
    舊的也要補——比照 v0.9.92 新建時的規則（_ensure_case_for），沒 case_id 的就補一個同名案件掛上。
    專案有「負責人」欄位，能唯一比對到帳號的話案件負責人一併補上（方案A，見 _match_owner_username）；
    budgets 沒有負責人欄位，維持只補案件關聯、不補負責人。
    另外處理「案件已存在但沒負責人」的既有案件（早期回填留下的孤兒）：只要它掛的專案負責人現在能唯一比對到帳號，
    也一併補上——因為 _ensure_case_for 只在『新建案件』那一刻才會嘗試比對，同名案件已存在時不會回頭補。"""
    filled = 0
    with connect() as conn:
        budget_rows = conn.execute(
            "SELECT id, budget_code AS name, budget_code AS code FROM budgets "
            "WHERE status <> 'disabled' AND (case_id IS NULL OR case_id = 0)"
        ).fetchall()
        for r in budget_rows:
            cid = _ensure_case_for(conn, r["name"], r["code"], None, established=True)
            if cid:
                conn.execute("UPDATE budgets SET case_id = ? WHERE id = ?", (cid, r["id"]))
                filled += 1

        project_rows = conn.execute(
            "SELECT id, project_name AS name, project_code AS code, owner FROM projects "
            "WHERE status <> 'disabled' AND (case_id IS NULL OR case_id = 0)"
        ).fetchall()
        for r in project_rows:
            cid = _ensure_case_for(conn, r["name"], r["code"], None, r["owner"], established=True)
            if cid:
                conn.execute("UPDATE projects SET case_id = ? WHERE id = ?", (cid, r["id"]))
                filled += 1

        orphan_owner_cases = conn.execute(
            "SELECT c.id AS case_id, p.owner AS owner FROM cases c "
            "JOIN projects p ON p.case_id = c.id "
            "WHERE COALESCE(c.owner, '') = '' AND COALESCE(p.owner, '') <> '' "
            "AND c.status <> 'disabled' AND p.status <> 'disabled'"
        ).fetchall()
        for r in orphan_owner_cases:
            matched = _match_owner_username(conn, r["owner"])
            if matched:
                conn.execute("UPDATE cases SET owner = ? WHERE id = ? AND COALESCE(owner, '') = ''",
                             (matched, r["case_id"]))
    return filled


def backfill_all_numbers() -> dict[str, int]:
    """一次補齊案件系統編號、付款核銷編號、預算/專案的案件關聯，回報各補幾筆。"""
    cases_filled = backfill_case_numbers()
    settle_filled = backfill_settle_numbers()
    case_links_filled = backfill_case_links()
    return {"cases_filled": cases_filled, "settle_filled": settle_filled, "case_links_filled": case_links_filled}


# ── L3 預算年度費用比較：全年度／年增差異皆讀取時動態算（不存死），#DIV/0! 改語意標示 ──
def budget_annual_comparison(budget_id: int) -> dict[str, Any]:
    """某預算項目的年度費用比較：各年度期間金額、全年度加總、與相鄰前一年的費用差異。"""
    with connect() as conn:
        budget = get_row(conn, "budgets", budget_id)  # 不存在會 raise LookupError
        rows = conn.execute(
            "SELECT fiscal_year, period, amount FROM budget_periods WHERE budget_id = ? "
            "ORDER BY fiscal_year, id", (budget_id,)).fetchall()
        notes = {r["fiscal_year"]: r["note"] for r in conn.execute(
            "SELECT fiscal_year, note FROM budget_year_notes WHERE budget_id = ?", (budget_id,)).fetchall()}
    periods: list[str] = []          # 期間依出現順序（例：1-9月 / 10-12月）
    by_year: dict[str, dict[str, float]] = {}
    for r in rows:
        p = str(r["period"] or "").strip()
        if p and p not in periods:
            periods.append(p)
        y = str(r["fiscal_year"] or "").strip()
        by_year.setdefault(y, {})
        by_year[y][p] = by_year[y].get(p, 0.0) + float(r["amount"] or 0)
    years_out = []
    prev_total = None
    prev_pmap: dict[str, float] | None = None
    for y in sorted(by_year.keys()):
        pmap = by_year[y]
        total = sum(pmap.values())
        if prev_total is None:
            diff, diff_pct, note = None, None, "新增·無前期基準"
        elif prev_total == 0:
            diff, diff_pct, note = total, None, "前期為 0，僅顯示增額"   # 取代 Excel 的 #DIV/0!
        else:
            diff = total - prev_total
            diff_pct = round(diff / prev_total * 100, 1)
            note = ""
        # 各期間 vs 前一年同期的差異%（例：115 的 1-9月 → 116 的 1-9月 差多少%）
        period_diff_pct: dict[str, float | None] = {}
        for p in periods:
            if prev_pmap is None:
                period_diff_pct[p] = None
            else:
                pv = prev_pmap.get(p, 0.0)
                period_diff_pct[p] = round((pmap.get(p, 0.0) - pv) / pv * 100, 1) if pv else None
        years_out.append({
            "fiscal_year": y,
            "periods": {p: pmap.get(p, 0.0) for p in periods},
            "period_diff_pct": period_diff_pct,
            "annual_total": total,
            "diff": diff,
            "diff_pct": diff_pct,
            "diff_note": note,
            "note": notes.get(y, ""),   # 主管/助理可編輯的備註
        })
        prev_total = total
        prev_pmap = pmap
    return {
        "budget": {
            "id": budget_id,
            "budget_code": budget.get("budget_code", ""),
            "category": budget.get("category", ""),          # 預算項目
            "expense_detail": budget.get("expense_detail", ""),
            "unit_name": budget.get("unit_name", ""),
            "fill_dept": budget.get("fill_dept", ""),
            "estimator": budget.get("estimator", ""),
        },
        "periods": periods,
        "years": years_out,
    }


def set_budget_periods(budget_id: int, rows: list[dict[str, Any]]) -> dict[str, Any]:
    """整批取代某預算的年度費用明細（budget_periods）。rows: [{fiscal_year, period, amount}]，空年度/期間略過。"""
    clean = []
    for r in rows:
        fy = str(r.get("fiscal_year") or "").strip()
        period = str(r.get("period") or "").strip()
        if not fy or not period:
            continue
        try:
            amt = float(r.get("amount") or 0)
        except (TypeError, ValueError):
            amt = 0.0
        clean.append((fy, period, amt))
    with connect() as conn:
        get_row(conn, "budgets", budget_id)  # 不存在會 raise LookupError
        conn.execute("DELETE FROM budget_periods WHERE budget_id = ?", (budget_id,))
        for fy, period, amt in clean:
            conn.execute(
                "INSERT INTO budget_periods (budget_id, fiscal_year, period, amount) VALUES (?, ?, ?, ?)",
                (budget_id, fy, period, amt))
    return {"budget_id": budget_id, "count": len(clean)}


def set_budget_year_note(budget_id: int, fiscal_year: str, note: str) -> dict[str, Any]:
    """寫入/更新某預算某年度的備註（一預算一年一筆，upsert）。"""
    with connect() as conn:
        get_row(conn, "budgets", budget_id)  # 不存在會 raise LookupError
        conn.execute(
            "INSERT INTO budget_year_notes (budget_id, fiscal_year, note) VALUES (?, ?, ?) "
            "ON CONFLICT(budget_id, fiscal_year) DO UPDATE SET note = excluded.note",
            (budget_id, str(fiscal_year).strip(), str(note)))
    return {"budget_id": budget_id, "fiscal_year": str(fiscal_year).strip(), "note": str(note)}

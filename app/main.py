from contextlib import asynccontextmanager
import csv
import hashlib
import hmac
import io
import os
from pathlib import Path
import secrets
import sqlite3
import subprocess
import time
from typing import Any

from fastapi import Depends, FastAPI, HTTPException, Query, Request, Response
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field, field_validator, model_validator

from app.dev_console import console_status, run_console_command
from app.notify import send_digests, send_test
from app.seed_data import clear_demo_data, demo_counts, load_demo_data
from app.import_mapping import mapping_draft_catalog
from app.settings import get_settings
from app.store import (
    approve_case,
    backfill_all_numbers,
    backfill_status,
    backup_database,
    budget_annual_comparison,
    set_budget_periods,
    set_budget_year_note,
    case_360,
    case_progress_overview,
    cases_needing_attention,
    cio_overview,
    orphan_payments,
    overdue_reminders,
    pending_approvals,
    submit_case,
    confirm_import_batch_cases_dry_run,
    confirm_import_batch_cases_write,
    parse_projects_xlsx,
    commit_projects_import,
    parse_budget_xlsx,
    commit_budgets_import,
    list_project_items,
    list_budget_allocations,
    budget_unit_rollup,
    unit_conflicts,
    list_unit_master,
    create_unit_master,
    merge_units,
    split_units,
    reassign_unit,
    unlink_alias,
    unit_merge_impact,
    list_unit_decisions,
    undo_decision,
    reset_unit_decisions,
    list_name_values,
    merge_names,
    split_names,
    list_name_decisions,
    undo_name_decision,
    reset_name_decisions,
    get_working_year,
    parse_headcount_xlsx,
    commit_headcounts_import,
    list_headcounts,
    parse_category_shares_xlsx,
    commit_category_shares_import,
    list_category_shares,
    compute_budget_allocations,
    expiring_contracts,
    create_import_batch,
    dashboard_summary,
    delete_row,
    disable_row,
    read_settings as store_get_settings,
    write_settings as store_set_settings,
    get_db_user as store_get_db_user,
    list_db_users as store_list_db_users,
    create_db_user as store_create_db_user,
    update_db_user as store_update_db_user,
    delete_db_user as store_delete_db_user,
    get_import_batch,
    initialize_database,
    insert_row,
    list_import_batches,
    list_import_rows,
    list_audit_logs,
    list_rows,
    cio_changes_since_last_view,
    create_case_wizard,
    monthly_spending_summary,
    unit_budget_vs_actual,
    vendor_amount_summary,
    preflight_import_batch_confirm,
    preview_import_mapping,
    search_records,
    set_current_actor,
    set_owner_scope,
    stage_import_rows,
    update_row,
)


def _load_dotenv(env_path: Path | None = None) -> None:
    """極簡 .env 載入器（無外部依賴）：把 .env 的鍵值放進環境變數。
    用 setdefault，所以已存在的環境變數（測試/正式環境已設定者）優先，不被覆寫。"""
    env_path = env_path or (Path(__file__).resolve().parent.parent / ".env")
    if not env_path.exists():
        return
    for raw in env_path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, val = line.partition("=")
        os.environ.setdefault(key.strip(), val.strip().strip('"').strip("'"))


_load_dotenv()


class CaseIn(BaseModel):
    case_code: str = Field(min_length=1)
    title: str = Field(min_length=1)
    owner: str = ""
    status: str = "draft"
    amount: float = 0
    risk_level: str = "normal"
    note: str = ""
    next_step: str = ""
    due_date: str = ""
    fiscal_year: str = ""  # 所屬年度；空＝用作業年度


class CasePatch(BaseModel):
    case_code: str | None = Field(default=None, min_length=1)
    title: str | None = Field(default=None, min_length=1)
    owner: str | None = None
    status: str | None = None
    amount: float | None = None
    risk_level: str | None = None
    note: str | None = None
    next_step: str | None = None
    due_date: str | None = None
    fiscal_year: str | None = None


class ContractIn(BaseModel):
    contract_code: str = Field(min_length=1)
    contract_name: str = Field(min_length=1)
    vendor_name: str = ""
    amount: float | None = 0
    status: str = "active"
    case_id: int | None = None
    end_date: str = ""

    @field_validator("amount")
    @classmethod
    def _amount_default(cls, v: float | None) -> float:
        # 前端金額留空會送 null；視為 0，避免使用者不填金額就存不進去
        return 0.0 if v is None else v


class ContractPatch(BaseModel):
    contract_code: str | None = Field(default=None, min_length=1)
    contract_name: str | None = Field(default=None, min_length=1)
    vendor_name: str | None = None
    amount: float | None = None
    status: str | None = None
    case_id: int | None = None
    end_date: str | None = None


class PaymentIn(BaseModel):
    contract_id: int
    payment_month: str = Field(pattern=r"^\d{4}-\d{2}$")
    payment_amount: float
    invoice_status: str = "not_received"
    status: str = "pending"
    # 對齊真實費用整合表
    item: str = ""
    settle_no: str = ""
    ref_no: str = ""
    period: str = ""
    billing_period: str = ""
    settled_by: str = ""
    vendor: str = ""
    approval_level: str = ""
    owner: str = ""
    owner_email: str = ""
    net_amount: float | None = 0
    tax_amount: float | None = 0

    @field_validator("net_amount", "tax_amount")
    @classmethod
    def _amount_default(cls, v: float | None) -> float:
        # 前端金額留空會送 null；視為 0，避免使用者不填未稅/稅額就存不進去（同 ContractIn 慣例）
        return 0.0 if v is None else v


class PaymentPatch(BaseModel):
    contract_id: int | None = None
    payment_month: str | None = Field(default=None, pattern=r"^\d{4}-\d{2}$")
    payment_amount: float | None = None
    item: str | None = None
    settle_no: str | None = None
    ref_no: str | None = None
    period: str | None = None
    billing_period: str | None = None
    settled_by: str | None = None
    vendor: str | None = None
    approval_level: str | None = None
    owner: str | None = None
    owner_email: str | None = None
    net_amount: float | None = None
    tax_amount: float | None = None
    invoice_status: str | None = None
    status: str | None = None


class DocumentIn(BaseModel):
    file_name: str = Field(min_length=1)
    document_type: str = "other"
    source_note: str = ""
    status: str = "active"
    case_id: int | None = None
    contract_id: int | None = None


class DocumentPatch(BaseModel):
    file_name: str | None = Field(default=None, min_length=1)
    document_type: str | None = None
    source_note: str | None = None
    status: str | None = None
    case_id: int | None = None
    contract_id: int | None = None


class BudgetIn(BaseModel):
    budget_code: str = Field(min_length=1)
    category: str = ""
    unit_name: str = ""
    expense_detail: str = ""
    fill_dept: str = ""
    estimator: str = ""
    fiscal_year: str = ""
    amount: float = 0
    status: str = "active"
    case_id: int | None = None
    note: str = ""
    alloc_method: str = "fixed"
    alloc_category_kind: str = ""
    alloc_category: str = ""


class BudgetPatch(BaseModel):
    budget_code: str | None = Field(default=None, min_length=1)
    category: str | None = None
    unit_name: str | None = None
    expense_detail: str | None = None
    fill_dept: str | None = None
    estimator: str | None = None
    fiscal_year: str | None = None
    amount: float | None = None
    status: str | None = None
    case_id: int | None = None
    note: str | None = None
    remainder_unit_code: str | None = None
    alloc_method: str | None = None
    alloc_category_kind: str | None = None
    alloc_category: str | None = None


class BudgetYearNoteIn(BaseModel):
    fiscal_year: str = Field(min_length=1)
    note: str = ""


class BudgetPeriodRow(BaseModel):
    fiscal_year: str = ""
    period: str = ""
    amount: float = 0


class BudgetPeriodsIn(BaseModel):
    periods: list[BudgetPeriodRow] = Field(default_factory=list)


class UnitVariant(BaseModel):
    unit_code: str = ""
    unit_name: str = ""


class UnitCreateIn(BaseModel):
    canonical_code: str = ""
    canonical_name: str = Field(min_length=1)
    note: str = ""


class UnitMergeIn(BaseModel):
    variants: list[UnitVariant]
    canonical_code: str = ""
    canonical_name: str = ""
    reason: str = ""


class UnitSplitIn(BaseModel):
    variants: list[UnitVariant]
    reason: str = ""


class UnitReassignIn(BaseModel):
    variant: UnitVariant
    canonical_code: str = ""
    canonical_name: str = ""
    reason: str = ""


class UnitImpactIn(BaseModel):
    variants: list[UnitVariant]


class NameMergeIn(BaseModel):
    kind: str
    names: list[str]
    canonical_name: str = ""
    reason: str = ""


class NameSplitIn(BaseModel):
    kind: str
    names: list[str]
    reason: str = ""


class ProjectIn(BaseModel):
    project_code: str = Field(min_length=1)
    project_name: str = Field(min_length=1)
    source: str = ""
    necessity: str = ""
    progress: float = 0
    owner: str = ""
    status: str = "active"
    case_id: int | None = None
    due_date: str = ""
    note: str = ""
    level: str = ""
    progress_planned: float = 0
    rag_status: str = ""
    start_date: str = ""
    end_date: str = ""


class ProjectPatch(BaseModel):
    project_code: str | None = Field(default=None, min_length=1)
    project_name: str | None = Field(default=None, min_length=1)
    source: str | None = None
    necessity: str | None = None
    progress: float | None = None
    owner: str | None = None
    status: str | None = None
    case_id: int | None = None
    due_date: str | None = None
    note: str | None = None
    level: str | None = None
    progress_planned: float | None = None
    rag_status: str | None = None
    start_date: str | None = None
    end_date: str | None = None


class ProjectItemIn(BaseModel):
    item_name: str = Field(min_length=1)
    owner: str = ""
    start_date: str = ""
    end_date: str = ""
    exec_status: str = ""
    sub_total: int = 0
    sub_done: int = 0
    progress: float = 0
    rag: str = ""
    risk_note: str = ""
    decision_needed: str = ""
    support_needed: str = ""
    duration_days: str = ""
    seq: int = 0


class ProjectItemPatch(BaseModel):
    item_name: str | None = Field(default=None, min_length=1)
    owner: str | None = None
    start_date: str | None = None
    end_date: str | None = None
    exec_status: str | None = None
    sub_total: int | None = None
    sub_done: int | None = None
    progress: float | None = None
    rag: str | None = None
    risk_note: str | None = None
    decision_needed: str | None = None
    support_needed: str | None = None
    duration_days: str | None = None
    seq: int | None = None
    status: str | None = None


class SignoffIn(BaseModel):
    signoff_code: str = Field(min_length=1)
    subject: str = Field(min_length=1)
    applicant: str = ""
    amount: float = 0
    status: str = "draft"
    sign_date: str = ""
    case_id: int | None = None
    note: str = ""
    attachment_ref: str = ""


class SignoffPatch(BaseModel):
    signoff_code: str | None = Field(default=None, min_length=1)
    subject: str | None = Field(default=None, min_length=1)
    applicant: str | None = None
    amount: float | None = None
    status: str | None = None
    sign_date: str | None = None
    case_id: int | None = None
    note: str | None = None
    attachment_ref: str | None = None


class PurchaseIn(BaseModel):
    purchase_code: str = Field(min_length=1)
    item_name: str = Field(min_length=1)
    vendor_name: str = ""
    quantity: float = 0
    amount: float = 0
    status: str = "pending"
    case_id: int | None = None
    note: str = ""


class PurchasePatch(BaseModel):
    purchase_code: str | None = Field(default=None, min_length=1)
    item_name: str | None = Field(default=None, min_length=1)
    vendor_name: str | None = None
    quantity: float | None = None
    amount: float | None = None
    status: str | None = None
    case_id: int | None = None
    note: str | None = None


class CaseWizardCaseIn(BaseModel):
    case_code: str = Field(min_length=1)
    title: str = Field(min_length=1)
    owner: str = ""
    amount: float = 0
    fiscal_year: str = ""
    note: str = ""
    next_step: str = ""
    due_date: str = ""


class CaseWizardBudgetIn(BaseModel):
    budget_code: str = Field(min_length=1)
    category: str = ""
    unit_name: str = ""
    amount: float = 0
    note: str = ""


class CaseWizardSignoffIn(BaseModel):
    signoff_code: str = Field(min_length=1)
    subject: str = Field(min_length=1)
    applicant: str = ""
    amount: float = 0
    sign_date: str = ""
    note: str = ""


class CaseWizardPurchaseIn(BaseModel):
    purchase_code: str = Field(min_length=1)
    item_name: str = Field(min_length=1)
    vendor_name: str = ""
    quantity: float = 0
    amount: float = 0
    note: str = ""


class CaseWizardContractIn(BaseModel):
    contract_code: str = Field(min_length=1)
    contract_name: str = Field(min_length=1)
    vendor_name: str = ""
    amount: float | None = 0
    end_date: str = ""

    @field_validator("amount")
    @classmethod
    def _amount_default(cls, v: float | None) -> float:
        return 0.0 if v is None else v


class CaseWizardPaymentIn(BaseModel):
    payment_month: str = Field(pattern=r"^\d{4}-\d{2}$")
    payment_amount: float
    item: str = ""
    net_amount: float | None = 0
    tax_amount: float | None = 0

    @field_validator("net_amount", "tax_amount")
    @classmethod
    def _amount_default(cls, v: float | None) -> float:
        return 0.0 if v is None else v


class CaseWizardIn(BaseModel):
    """一條龍新案精靈：單頁多步驟表單一次送出。預算/簽呈/請購/合約/付款皆可跳過(None)；
    付款要掛在合約下，若填了付款卻沒填合約，直接擋在驗證層（不用等寫進資料庫才發現）。"""
    case: CaseWizardCaseIn
    budget: CaseWizardBudgetIn | None = None
    signoff: CaseWizardSignoffIn | None = None
    purchase: CaseWizardPurchaseIn | None = None
    contract: CaseWizardContractIn | None = None
    payment: CaseWizardPaymentIn | None = None

    @model_validator(mode="after")
    def _payment_needs_contract(self) -> "CaseWizardIn":
        if self.payment is not None and self.contract is None:
            raise ValueError("付款步驟需要先填合約（付款要掛在合約下）")
        return self


class ImportBatchIn(BaseModel):
    source_name: str = Field(min_length=1)
    status: str = "created"


class ImportRowsIn(BaseModel):
    rows: list[dict[str, Any]] = Field(min_length=1)


class ConfirmedImportField(BaseModel):
    row_number: int = Field(ge=1)
    target_table: str = Field(min_length=1)
    target_field: str = Field(min_length=1)


class ImportConfirmIn(BaseModel):
    dry_run: bool = True
    target_tables: list[str] = Field(default_factory=lambda: ["cases"], min_length=1)
    confirmed_fields: list[ConfirmedImportField] = Field(default_factory=list)
    accepted_warning_codes: list[str] = Field(default_factory=list)


class DevConsoleRunIn(BaseModel):
    command_id: str = Field(min_length=1)
    dry_run: bool = False


class LoginIn(BaseModel):
    username: str = Field(min_length=1)
    password: str = ""  # 試辦免密碼模式下可留空


class SettingsPatch(BaseModel):
    smtp_host: str | None = None
    smtp_port: str | None = None
    smtp_user: str | None = None
    smtp_from: str | None = None
    smtp_password: str | None = None  # write-only；GET 永不回傳
    email_map: str | None = None
    notify_enabled: str | None = None
    opt_budget_categories: str | None = None
    opt_project_necessity: str | None = None
    opt_project_level: str | None = None
    opt_project_rag: str | None = None


class UserCreateIn(BaseModel):
    username: str = Field(min_length=1)
    role_code: str = Field(min_length=1)
    display_name: str = ""
    email: str = ""
    password: str = Field(min_length=1)


class UserPatch(BaseModel):
    role_code: str | None = None
    display_name: str | None = None
    email: str | None = None
    disabled: bool | None = None
    password: str | None = None


SETTINGS_PUBLIC_KEYS = [
    "smtp_host", "smtp_port", "smtp_user", "smtp_from", "email_map", "notify_enabled",
    "opt_budget_categories", "opt_project_necessity", "opt_project_level", "opt_project_rag",
]

# 主檔選項預設（後台未設定時採用）
OPTION_DEFAULTS = {
    "opt_budget_categories": "基礎建設,工具,資訊安全,電子交易平台,其他",
    "opt_project_necessity": "必要,非必要",
    "opt_project_level": "公司級,處級,部級",
    "opt_project_rag": "如期執行,已完成,未開始,有延遲但不影響,有延遲且可能影響",
}


def _option_list(key: str) -> list[str]:
    raw = store_get_settings([key])[key] or OPTION_DEFAULTS.get(key, "")
    return [x.strip() for x in raw.split(",") if x.strip()]


CSV_COLUMNS: dict[str, list[tuple[str, str]]] = {
    "contracts": [("contract_code", "合約編號"), ("contract_name", "合約名稱"), ("vendor_name", "廠商"), ("amount", "金額"), ("status", "狀態"), ("end_date", "到期日"), ("case_id", "案件ID")],
    "payments": [("item", "核銷項目"), ("settle_no", "核銷編號"), ("ref_no", "參照碼"), ("vendor", "廠商"), ("period", "期別"), ("billing_period", "計費期間"), ("payment_month", "核銷月份"), ("net_amount", "未稅金額"), ("tax_amount", "營業稅"), ("payment_amount", "含稅/核銷金額"), ("approval_level", "簽核層級"), ("settled_by", "核銷者"), ("owner", "窗口"), ("owner_email", "窗口Email"), ("invoice_status", "發票狀態"), ("status", "處理進度"), ("contract_id", "合約ID")],
    "documents": [("file_name", "檔案名稱"), ("document_type", "類型"), ("source_note", "來源說明"), ("status", "狀態"), ("case_id", "案件ID"), ("contract_id", "合約ID")],
    "budgets": [("budget_code", "預算編號"), ("category", "類別"), ("unit_name", "單位"), ("fiscal_year", "年度"), ("amount", "金額"), ("status", "狀態"), ("case_id", "案件ID")],
    "projects": [("project_code", "標號"), ("project_name", "專案名稱"), ("level", "專案分類"), ("necessity", "必要性"), ("start_date", "開始日"), ("end_date", "結束日"), ("progress_planned", "進度預計%"), ("progress", "進度實際%"), ("rag_status", "燈號"), ("owner", "負責人"), ("source", "來源"), ("status", "狀態"), ("due_date", "預計完成日"), ("case_id", "案件ID")],
    "signoffs": [("signoff_code", "簽呈編號"), ("subject", "主旨"), ("applicant", "申請人"), ("amount", "金額"), ("status", "狀態"), ("sign_date", "簽核日"), ("case_id", "案件ID")],
    "purchases": [("purchase_code", "請購編號"), ("item_name", "品項"), ("vendor_name", "廠商"), ("quantity", "數量"), ("amount", "金額"), ("status", "狀態"), ("case_id", "案件ID")],
}

# 後端建置日期／標記（單一來源）：由 /health 回傳，前端徽章拿來跟自己的版本比對。
# 每次改後端就 bump；若前端徽章顯示的後端日期不對，代表 uvicorn 沒重啟。
BACKEND_BUILD = "v0.9.69 · 2026-07-11 · 單位主檔可主動新增+預算單位名稱改下拉"

# 試辦免密碼登入：預設關（測試維持嚴格密碼驗證）；上線試辦的伺服器用環境變數 PILOT_PASSWORDLESS=1 打開。
# 打開後，內建帳號（ap01~ap04/admin）從下拉選單選角色即可登入、不需密碼。僅供 localhost 試辦，勿用於正式環境。
def pilot_passwordless() -> bool:
    return os.getenv("PILOT_PASSWORDLESS", "0") == "1"

AUTH_COOKIE_NAME = "ai_fee_user"
HANDLER_FORBIDDEN_PREFIXES = ("/api/audit-logs", "/api/import-batches", "/api/dev-console")  # 承辦不得使用（稽核/匯入/開發者控制台，含 demo-data）
LOCAL_AUTH_USERS: dict[str, dict[str, Any]] = {
    "ap01": {
        "username": "ap01",
        "role_code": "cio",
        "role_name": "CIO",
        "display_name": "CIO",
        # CIO 極簡：只給「決策總覽」，其他模組連卡片都不顯示（不是灰色，是隱藏）
        "default_module": "cio-overview",
        "allowed_modules": ["cio-overview"],
        "allowed_actions": ["read"],
    },
    "ap02": {
        "username": "ap02",
        "role_code": "manager_assistant",
        "role_name": "主管/助理",
        "display_name": "主管/助理",
        "default_module": "cases-module",
        "allowed_modules": [
            "budget",
            "projects",
            "portfolio",
            "signoff",
            "cases-module",
            "data-review",
            "contracts-module",
            "purchases",
            "payments-module",
            "io-center",
            "unit-admin",
            "data-admin",
            "fee-alloc",
            "name-admin",
        ],
        "allowed_actions": ["read", "edit", "import_preview", "preflight"],
    },
    "ap03": {
        "username": "ap03",
        "role_code": "handler",
        "role_name": "承辦",
        "display_name": "承辦",
        "default_module": "cases-module",
        "allowed_modules": ["projects", "portfolio", "cases-module", "purchases", "payments-module", "data-review", "data-admin"],
        "allowed_actions": ["read", "edit"],
    },
    # 第二位助理：雙人複核需要「另一位」助理來核，助理自己建的案件不能自己核
    "ap04": {
        "username": "ap04",
        "role_code": "manager_assistant",
        "role_name": "助理B",
        "display_name": "助理B",
        "default_module": "cases-module",
        "allowed_modules": [
            "budget",
            "projects",
            "portfolio",
            "signoff",
            "cases-module",
            "data-review",
            "contracts-module",
            "purchases",
            "payments-module",
            "io-center",
            "unit-admin",
            "data-admin",
            "fee-alloc",
            "name-admin",
        ],
        "allowed_actions": ["read", "edit", "import_preview", "preflight"],
    },
    # 系統管理員：只進「系統管理」後台（設定/稽核/狀態）。未來改為指定 AD 帳號當 admin。
    "admin": {
        "username": "admin",
        "role_code": "admin",
        "role_name": "系統管理員",
        "display_name": "系統管理員",
        "default_module": "admin-console",
        "allowed_modules": ["admin-console"],
        "allowed_actions": ["read", "edit", "admin"],
    },
}


SESSION_SECRET = os.getenv("SESSION_SECRET") or secrets.token_hex(32)
SESSION_MAX_AGE = int(os.getenv("SESSION_MAX_AGE_SECONDS", "28800"))  # 預設 8 小時後過期


def _sign_session(username: str) -> str:
    payload = f"{username}.{int(time.time())}"
    sig = hmac.new(SESSION_SECRET.encode(), payload.encode(), hashlib.sha256).hexdigest()
    return f"{payload}.{sig}"


def _verify_session(token: str) -> str | None:
    """驗證簽章 session cookie。回傳可信任的 username，否則 None。
    只認得本服務簽出來、且未過期的 token；偽造或過期一律拒絕。"""
    parts = (token or "").rsplit(".", 2)
    if len(parts) != 3:
        return None
    username, ts, sig = parts
    expected = hmac.new(SESSION_SECRET.encode(), f"{username}.{ts}".encode(), hashlib.sha256).hexdigest()
    if not hmac.compare_digest(sig, expected) or username not in LOCAL_AUTH_USERS:
        return None
    try:
        age = time.time() - int(ts)
    except ValueError:
        return None
    if age < 0 or age > SESSION_MAX_AGE:
        return None
    return username


def _hash_password(password: str, salt: bytes | None = None) -> str:
    salt = salt or secrets.token_bytes(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode(), salt, 200_000)
    return f"{salt.hex()}${dk.hex()}"


def _verify_password(password: str, stored: str) -> bool:
    try:
        salt_hex, hash_hex = stored.split("$", 1)
    except ValueError:
        return False
    dk = hashlib.pbkdf2_hmac("sha256", password.encode(), bytes.fromhex(salt_hex), 200_000)
    return hmac.compare_digest(dk.hex(), hash_hex)


def _load_user_password_hashes() -> dict[str, str]:
    """每個帳號密碼由環境變數 {USERNAME}_PASSWORD 提供（例：AP01_PASSWORD），
    啟動時各自加鹽雜湊存記憶體。未設定者無法登入。密碼絕不進原始碼。"""
    hashes: dict[str, str] = {}
    for username in LOCAL_AUTH_USERS:
        pw = os.getenv(f"{username.upper()}_PASSWORD", "")
        if pw:
            hashes[username] = _hash_password(pw)
    return hashes


USER_PASSWORD_HASHES = _load_user_password_hashes()


# 角色範本：由內建帳號自動萃取，供 DB 建立的帳號套用同一套模組/動作權限。
ROLE_TEMPLATES: dict[str, dict[str, Any]] = {}
for _acct in LOCAL_AUTH_USERS.values():
    ROLE_TEMPLATES.setdefault(_acct["role_code"], {
        "role_name": _acct["role_name"],
        "default_module": _acct["default_module"],
        "allowed_modules": list(_acct["allowed_modules"]),
        "allowed_actions": list(_acct["allowed_actions"]),
    })


def get_account(username: str) -> dict[str, Any] | None:
    """解析帳號（內建 or DB 建立）→ 完整權限視圖；找不到回 None。內建帳號永不停用。"""
    if not username:
        return None
    if username in LOCAL_AUTH_USERS:
        return {**LOCAL_AUTH_USERS[username], "disabled": False, "email": "", "builtin": True}
    row = store_get_db_user(username)
    if not row:
        return None
    tmpl = ROLE_TEMPLATES.get(row["role_code"])
    if not tmpl:
        return None
    return {
        "username": row["username"],
        "role_code": row["role_code"],
        "role_name": tmpl["role_name"],
        "display_name": row["display_name"] or row["username"],
        "default_module": tmpl["default_module"],
        "allowed_modules": list(tmpl["allowed_modules"]),
        "allowed_actions": list(tmpl["allowed_actions"]),
        "disabled": bool(row["disabled"]),
        "email": row["email"],
        "builtin": False,
    }


def verify_login(username: str, password: str) -> dict[str, Any] | None:
    """驗證登入。內建帳號用環境變數密碼；DB 帳號用其 password_hash（停用者拒絕）。回帳號視圖或 None。"""
    acct = get_account(username)
    if not acct or acct.get("disabled"):
        return None
    if username in USER_PASSWORD_HASHES:
        return acct if _verify_password(password, USER_PASSWORD_HASHES[username]) else None
    row = store_get_db_user(username)
    if not row or not row["password_hash"] or not _verify_password(password, row["password_hash"]):
        return None
    return acct


def ok(data: Any) -> dict[str, Any]:
    return {"ok": True, "data": data}


def auth_user_payload(username: str) -> dict[str, Any]:
    user = get_account(username)
    if not user or user.get("disabled"):
        raise HTTPException(status_code=401, detail="登入狀態已失效，請重新登入。")
    payload = dict(user)
    payload.pop("builtin", None)
    return payload


@asynccontextmanager
async def lifespan(app: FastAPI):
    initialize_database()
    yield


async def bind_actor(request: Request) -> None:
    """把已驗證的登入者綁到本請求（async 依賴，確保 contextvar 傳到同步端點）。"""
    username = _verify_session(request.cookies.get(AUTH_COOKIE_NAME, ""))
    set_current_actor(username or "anonymous")
    acct = get_account(username or "") or {}
    set_owner_scope(username if acct.get("role_code") == "handler" else None)


def create_app() -> FastAPI:
    settings = get_settings()
    web_dir = Path(__file__).resolve().parent / "web"
    app = FastAPI(
        title="Fee Contract Control",
        version="0.2.0-fresh",
        description="Fresh implementation for fee, contract, payment, document, and case tracking.",
        lifespan=lifespan,
        dependencies=[Depends(bind_actor)],
    )
    app.mount("/static", StaticFiles(directory=web_dir), name="static")

    @app.middleware("http")
    async def _auth_gate(request: Request, call_next):
        # 所有 /api/* 都要登入，例外只有 /api/auth/*（登入/登出/查身分）。
        path = request.url.path
        if path.startswith("/api/") and not path.startswith("/api/auth/"):
            username = _verify_session(request.cookies.get(AUTH_COOKIE_NAME, ""))
            if not username:
                return JSONResponse({"detail": "請先登入。"}, status_code=401)
            user = get_account(username)
            if not user or user.get("disabled"):
                return JSONResponse({"detail": "帳號已停用或不存在，請重新登入。"}, status_code=401)
            if user["role_code"] == "handler" and any(path.startswith(p) for p in HANDLER_FORBIDDEN_PREFIXES):
                return JSONResponse({"detail": "權限不足：承辦無權使用此功能。"}, status_code=403)
            if path.startswith("/api/admin/") and user["role_code"] != "admin":
                return JSONResponse({"detail": "權限不足：僅系統管理員可用。"}, status_code=403)
            if request.method in ("POST", "PATCH", "PUT", "DELETE"):
                if "edit" not in user["allowed_actions"]:
                    return JSONResponse({"detail": "權限不足：此帳號僅供檢視。"}, status_code=403)
        return await call_next(request)

    @app.get("/", include_in_schema=False)
    def home() -> FileResponse:
        # index.html 不快取：瀏覽器每次都拿最新 HTML（連帶抓到最新的 ?v= 靜態檔），
        # 徹底避免「後端已更新、使用者卻卡在舊頁面」。靜態檔本身另用 ?v= 破快取。
        return FileResponse(
            web_dir / "index.html",
            headers={"Cache-Control": "no-cache, no-store, must-revalidate"},
        )

    @app.get("/dev-console", include_in_schema=False)
    def dev_console_home() -> FileResponse:
        return FileResponse(web_dir / "dev-console.html")

    @app.get("/health")
    def health() -> dict[str, Any]:
        return {
            "ok": True,
            "service": settings.app_name,
            "version": "0.2.0-fresh",
            "build": BACKEND_BUILD,
            "database": {"type": "sqlite", "path": settings.database_path},
        }

    @app.get("/api/auth/options")
    def auth_options() -> dict[str, Any]:
        # 公開端點：前端登入頁用來決定「下拉選角色 / 是否要密碼」，並取得可選帳號清單。
        accounts = [
            {"username": u, "label": a["role_name"]}
            for u, a in LOCAL_AUTH_USERS.items()
        ]
        return ok({"passwordless": pilot_passwordless(), "accounts": accounts})

    @app.post("/api/auth/login")
    def login(payload: LoginIn, response: Response) -> dict[str, Any]:
        username = payload.username.strip().lower()
        # 試辦免密碼：內建帳號選了就登入（跳過密碼）；其餘一律走正常密碼驗證。
        passwordless_ok = pilot_passwordless() and username in LOCAL_AUTH_USERS
        if not passwordless_ok and not verify_login(username, payload.password):
            raise HTTPException(status_code=401, detail="帳號或密碼錯誤。")
        response.set_cookie(
            AUTH_COOKIE_NAME,
            _sign_session(username),
            httponly=True,
            samesite="lax",
        )
        return ok(auth_user_payload(username))

    @app.get("/api/auth/me")
    def current_user(request: Request) -> dict[str, Any]:
        username = _verify_session(request.cookies.get(AUTH_COOKIE_NAME, ""))
        if not username:
            raise HTTPException(status_code=401, detail="登入狀態已失效，請重新登入。")
        return ok(auth_user_payload(username))

    @app.post("/api/auth/logout")
    def logout(response: Response) -> dict[str, Any]:
        response.delete_cookie(AUTH_COOKIE_NAME)
        return ok({"logged_out": True})

    @app.get("/api/options")
    def options() -> dict[str, Any]:
        # 給各模組表單的下拉選項（任何登入者可讀），內容由 admin 在後台維護。
        return ok({
            "budget_categories": _option_list("opt_budget_categories"),
            "project_necessity": _option_list("opt_project_necessity"),
            "project_level": _option_list("opt_project_level"),
            "project_rag": _option_list("opt_project_rag"),
        })

    @app.get("/api/dashboard")
    def dashboard() -> dict[str, Any]:
        return ok(dashboard_summary())

    @app.get("/api/reports/monthly-spending")
    def monthly_spending() -> dict[str, Any]:
        return ok(monthly_spending_summary())

    @app.get("/api/todo")
    def todo_cases() -> dict[str, Any]:
        return ok(cases_needing_attention())

    @app.get("/api/reports/expiring-contracts")
    def expiring() -> dict[str, Any]:
        return ok(expiring_contracts())

    @app.get("/api/reports/cio-overview")
    def cio_overview_report() -> dict[str, Any]:
        return ok(cio_overview())

    @app.get("/api/reports/cio-changes-since-last-view")
    def cio_changes_report() -> dict[str, Any]:
        # 讀取本身即視為「已讀」，下次查看只顯示這之後的變動（見 store 函式說明）。
        return ok(cio_changes_since_last_view())

    @app.get("/api/reports/unit-budget-vs-actual")
    def unit_budget_vs_actual_report(year: int | None = None) -> dict[str, Any]:
        return ok(unit_budget_vs_actual(year))

    @app.get("/api/reports/vendor-amount-summary")
    def vendor_amount_summary_report() -> dict[str, Any]:
        return ok(vendor_amount_summary())

    @app.get("/api/reports/cio-overview.xlsx", include_in_schema=False)
    def cio_overview_export_xlsx() -> Response:
        import openpyxl
        from openpyxl.styles import Font

        data = cio_overview()
        wb = openpyxl.Workbook()
        bold = Font(bold=True)

        ws1 = wb.active
        ws1.title = "資金總覽"
        ws1.append(["項目", "數值"])
        ws1["A1"].font = bold
        ws1["B1"].font = bold
        ws1.append(["本月", data["this_month"]])
        ws1.append(["本月應付總額", data["this_month_total"]])
        ws1.append(["下月", data["next_month"]])
        ws1.append(["下月應付總額", data["next_month_total"]])
        ws1.append(["要準備的資金（尚未結案）", data["funds_to_prepare"]])
        ws1.append(["下月預算外金額", data["unplanned_next_month"]])
        ws1.append(["下月超支案件數", data["overspent_count"]])

        ws2 = wb.create_sheet("未來六個月現金流預測")
        ws2.append(["月份", "應付總額"])
        for c in ("A1", "B1"):
            ws2[c].font = bold
        for f in data["forecast"]:
            ws2.append([f["month"], f["total"]])

        ws3 = wb.create_sheet("下月要出的款明細")
        headers = ["案件編號", "案件名稱", "負責人", "合約編號", "核銷月份", "金額", "狀態", "預算外", "超支"]
        ws3.append(headers)
        for c in range(1, len(headers) + 1):
            ws3.cell(row=1, column=c).font = bold
        for r in data["upcoming_next_month"]:
            ws3.append([
                r["case_code"], r["case_title"], r.get("owner", ""), r.get("contract_code", ""),
                r["payment_month"], r["payment_amount"], r["status"],
                "是" if r["unplanned"] else "否", "是" if r["overspent"] else "否",
            ])

        buf = io.BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=cio-overview.xlsx"},
        )

    @app.get("/api/reports/reminders")
    def reminders() -> dict[str, Any]:
        return ok(overdue_reminders())

    @app.get("/api/reports/orphan-payments")
    def orphan_payments_report() -> dict[str, Any]:
        return ok(orphan_payments())

    @app.get("/api/reports/pending-approvals")
    def pending_approvals_report() -> dict[str, Any]:
        return ok(pending_approvals())

    @app.post("/api/reports/reminders/notify")
    def reminders_notify() -> dict[str, Any]:
        return ok(send_digests())

    # ---- 系統管理後台（僅 admin，中介層已擋非 admin）----
    def _settings_view() -> dict[str, Any]:
        vals = store_get_settings(SETTINGS_PUBLIC_KEYS)
        vals["smtp_password_set"] = bool(store_get_settings(["smtp_password"])["smtp_password"])
        return vals

    @app.get("/api/admin/settings")
    def admin_get_settings() -> dict[str, Any]:
        return ok(_settings_view())

    @app.patch("/api/admin/settings")
    def admin_patch_settings(payload: SettingsPatch) -> dict[str, Any]:
        data = payload.model_dump(exclude_unset=True)
        if data.get("smtp_password") == "":
            data.pop("smtp_password")  # 空＝不動密碼（避免誤清）
        store_set_settings(data)
        return ok(_settings_view())

    @app.post("/api/admin/settings/test-email")
    def admin_test_email(payload: dict[str, Any]) -> dict[str, Any]:
        try:
            return ok(send_test(str(payload.get("to", "")).strip()))
        except OSError as exc:
            raise HTTPException(status_code=502, detail=f"寄信失敗：{exc}") from exc

    @app.get("/api/admin/backup", include_in_schema=False)
    def admin_backup() -> FileResponse:
        import tempfile

        dest = str(Path(tempfile.gettempdir()) / "cl_fee_backup.db")
        backup_database(dest)
        return FileResponse(dest, filename="cl_fee_backup.db", media_type="application/octet-stream")

    # ---- 帳號與權限管理（內建帳號唯讀，DB 帳號可增/改/停用/刪）----
    @app.get("/api/admin/users")
    def admin_list_users() -> dict[str, Any]:
        builtin = [
            {"username": u, "role_code": a["role_code"], "role_name": a["role_name"],
             "display_name": a["display_name"], "email": "", "disabled": False, "builtin": True}
            for u, a in LOCAL_AUTH_USERS.items()
        ]
        db = [
            {"username": r["username"], "role_code": r["role_code"],
             "role_name": ROLE_TEMPLATES.get(r["role_code"], {}).get("role_name", r["role_code"]),
             "display_name": r["display_name"], "email": r["email"],
             "disabled": bool(r["disabled"]), "builtin": False}
            for r in store_list_db_users()
        ]
        roles = [{"code": k, "name": v["role_name"]} for k, v in ROLE_TEMPLATES.items()]
        return ok({"users": builtin + db, "roles": roles})

    @app.post("/api/admin/users", status_code=201)
    def admin_create_user(payload: UserCreateIn) -> dict[str, Any]:
        username = payload.username.strip().lower()
        if username in LOCAL_AUTH_USERS:
            raise HTTPException(status_code=409, detail="帳號與內建帳號名稱衝突。")
        if payload.role_code not in ROLE_TEMPLATES:
            raise HTTPException(status_code=422, detail="角色不存在。")
        try:
            store_create_db_user(username, payload.role_code, payload.display_name, payload.email, _hash_password(payload.password))
        except ValueError as exc:
            raise HTTPException(status_code=409, detail=str(exc)) from exc
        return ok({"username": username})

    @app.patch("/api/admin/users/{username}")
    def admin_update_user(username: str, payload: UserPatch) -> dict[str, Any]:
        if username in LOCAL_AUTH_USERS:
            raise HTTPException(status_code=403, detail="內建帳號不可修改。")
        data = payload.model_dump(exclude_unset=True)
        if "role_code" in data and data["role_code"] not in ROLE_TEMPLATES:
            raise HTTPException(status_code=422, detail="角色不存在。")
        if "password" in data:
            pw = data.pop("password")
            if pw:
                data["password_hash"] = _hash_password(pw)
        if "disabled" in data:
            data["disabled"] = 1 if data["disabled"] else 0
        try:
            store_update_db_user(username, data)
        except LookupError as exc:
            raise HTTPException(status_code=404, detail=str(exc)) from exc
        return ok({"username": username})

    @app.delete("/api/admin/users/{username}", status_code=204)
    def admin_delete_user(username: str) -> None:
        if username in LOCAL_AUTH_USERS:
            raise HTTPException(status_code=403, detail="內建帳號不可刪除。")
        try:
            store_delete_db_user(username)
        except LookupError as exc:
            raise HTTPException(status_code=404, detail=str(exc)) from exc

    @app.get("/api/dev-console/status")
    def dev_console_status() -> dict[str, Any]:
        return ok(console_status())

    @app.post("/api/dev-console/run")
    def dev_console_run(payload: DevConsoleRunIn) -> dict[str, Any]:
        try:
            return ok(run_console_command(payload.command_id, payload.dry_run))
        except KeyError as exc:
            raise HTTPException(status_code=404, detail="Unknown control panel command.") from exc
        except subprocess.TimeoutExpired as exc:
            raise HTTPException(status_code=504, detail=f"Command timed out after {exc.timeout} seconds.") from exc

    # 測試種子資料（demo）：掛在 /api/dev-console 底下 → 承辦被既有前綴擋、CIO 唯讀擋，只有主管/助理可用。
    @app.get("/api/dev-console/demo-data/status")
    def demo_data_status() -> dict[str, Any]:
        return ok(demo_counts())

    @app.post("/api/dev-console/demo-data/load")
    def demo_data_load() -> dict[str, Any]:
        return ok(load_demo_data())

    @app.post("/api/dev-console/demo-data/clear")
    def demo_data_clear() -> dict[str, Any]:
        return ok(clear_demo_data())

    # Step 3 舊資料補號：只補缺號、冪等；掛 dev-console（承辦擋、CIO 唯讀擋）。
    @app.get("/api/dev-console/backfill/status")
    def backfill_numbers_status() -> dict[str, Any]:
        return ok(backfill_status())

    @app.post("/api/dev-console/backfill/run")
    def backfill_numbers_run() -> dict[str, Any]:
        return ok(backfill_all_numbers())

    @app.get("/api/audit-logs")
    def audit_logs(
        limit: int = Query(100, ge=1, le=500),
        table_name: str | None = None,
        row_id: int | None = None,
        action: str | None = None,
    ) -> dict[str, Any]:
        return ok(
            list_audit_logs(
                limit=limit,
                table_name=table_name,
                row_id=row_id,
                action=action,
            )
        )

    @app.post("/api/import-batches", status_code=201)
    def create_import_batch_endpoint(payload: ImportBatchIn) -> dict[str, Any]:
        try:
            return ok(create_import_batch(payload.source_name, payload.status))
        except ValueError as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc

    @app.get("/api/import-batches")
    def import_batches(limit: int = Query(100, ge=1, le=500)) -> dict[str, Any]:
        return ok(list_import_batches(limit))

    @app.get("/api/import-batches/{batch_id}")
    def import_batch(batch_id: int) -> dict[str, Any]:
        try:
            return ok(get_import_batch(batch_id))
        except LookupError as exc:
            raise HTTPException(status_code=404, detail=str(exc)) from exc

    @app.post("/api/import-batches/{batch_id}/rows", status_code=201)
    def stage_import_batch_rows(batch_id: int, payload: ImportRowsIn) -> dict[str, Any]:
        try:
            return ok(stage_import_rows(batch_id, payload.rows))
        except ValueError as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc
        except LookupError as exc:
            raise HTTPException(status_code=404, detail=str(exc)) from exc

    @app.get("/api/import-batches/{batch_id}/rows")
    def import_batch_rows(
        batch_id: int,
        limit: int = Query(500, ge=1, le=500),
    ) -> dict[str, Any]:
        try:
            return ok(list_import_rows(batch_id, limit))
        except LookupError as exc:
            raise HTTPException(status_code=404, detail=str(exc)) from exc

    @app.get("/api/import-batches/{batch_id}/mapping-preview")
    def import_batch_mapping_preview(batch_id: int) -> dict[str, Any]:
        try:
            return ok(preview_import_mapping(batch_id))
        except ValueError as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc
        except LookupError as exc:
            raise HTTPException(status_code=404, detail=str(exc)) from exc

    @app.post("/api/import-batches/{batch_id}/confirm-preflight")
    def preflight_import_confirm(batch_id: int, payload: ImportConfirmIn) -> dict[str, Any]:
        if payload.target_tables != ["cases"]:
            raise HTTPException(status_code=400, detail='Only target_tables=["cases"] is supported.')
        try:
            return ok(
                preflight_import_batch_confirm(
                    batch_id,
                    [field.model_dump() for field in payload.confirmed_fields],
                    payload.accepted_warning_codes,
                )
            )
        except LookupError as exc:
            raise HTTPException(status_code=404, detail=str(exc)) from exc

    @app.post("/api/import-batches/{batch_id}/confirm")
    def confirm_import_batch(batch_id: int, payload: ImportConfirmIn) -> dict[str, Any]:
        # dry_run=true → 試算計畫；dry_run=false → 正式寫入（帶交易/冪等/來源舉證/稽核閘門）。
        if payload.target_tables != ["cases"]:
            raise HTTPException(status_code=400, detail='Only target_tables=["cases"] is supported.')
        fields = [field.model_dump() for field in payload.confirmed_fields]
        try:
            if payload.dry_run:
                return ok(confirm_import_batch_cases_dry_run(batch_id, fields))
            return ok(confirm_import_batch_cases_write(batch_id, fields))
        except ValueError as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=409, detail=str(exc)) from exc
        except LookupError as exc:
            raise HTTPException(status_code=404, detail=str(exc)) from exc

    @app.get("/api/import-mapping-draft")
    def import_mapping_draft() -> dict[str, Any]:
        return ok(mapping_draft_catalog())

    @app.post("/api/cases", status_code=201)
    def create_case(payload: CaseIn) -> dict[str, Any]:
        if payload.status == "approved":
            raise HTTPException(status_code=422, detail="案件需經雙人複核核准，不能直接建立為『已核准』。")
        return handle_create("cases", payload.model_dump())

    @app.post("/api/case-wizard", status_code=201)
    def case_wizard(payload: CaseWizardIn) -> dict[str, Any]:
        # 單一交易：任一步驟寫入失敗（含案號/合約號撞號），前面已建的一併回滾，什麼都不留下。
        try:
            result = create_case_wizard(
                case=payload.case.model_dump(),
                budget=payload.budget.model_dump() if payload.budget else None,
                signoff=payload.signoff.model_dump() if payload.signoff else None,
                purchase=payload.purchase.model_dump() if payload.purchase else None,
                contract=payload.contract.model_dump() if payload.contract else None,
                payment=payload.payment.model_dump() if payload.payment else None,
            )
        except ValueError as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc
        except sqlite3.IntegrityError as exc:
            raise HTTPException(status_code=422, detail=f"編號重複或關聯資料不存在：{exc}") from exc
        return ok(result)

    @app.get("/api/cases")
    def cases(limit: int = Query(100, ge=1, le=500)) -> dict[str, Any]:
        return ok(list_rows("cases", limit))

    @app.get("/api/cases/progress")
    def cases_progress() -> dict[str, Any]:
        # 線性進度圖＋處理優先矩陣：系統自動推導、唯讀
        return ok(case_progress_overview())

    @app.get("/api/cases.csv", include_in_schema=False)
    def export_cases_csv() -> Response:
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["案件編號", "案件名稱", "狀態", "金額", "負責人", "風險", "備註", "下一步"])
        for r in list_rows("cases", 500):
            writer.writerow([
                r.get("case_code", ""), r.get("title", ""), r.get("status", ""),
                r.get("amount", 0), r.get("owner", ""), r.get("risk_level", ""),
                r.get("note", ""), r.get("next_step", ""),
            ])
        return Response(
            content="﻿" + buf.getvalue(),  # BOM 讓 Excel 正確辨識 UTF-8
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=cases.csv"},
        )

    @app.patch("/api/cases/{case_id}")
    def update_case(case_id: int, payload: CasePatch) -> dict[str, Any]:
        if payload.status == "approved":
            raise HTTPException(status_code=422, detail="案件需經雙人複核核准，不能直接改為『已核准』。")
        return handle_change("cases", case_id, payload.model_dump(exclude_unset=True))

    @app.post("/api/cases/{case_id}/submit")
    def submit_case_endpoint(case_id: int) -> dict[str, Any]:
        try:
            return ok(submit_case(case_id))
        except LookupError as exc:
            raise HTTPException(status_code=404, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=409, detail=str(exc)) from exc

    @app.post("/api/cases/{case_id}/approve")
    def approve_case_endpoint(case_id: int, request: Request) -> dict[str, Any]:
        approver = _verify_session(request.cookies.get(AUTH_COOKIE_NAME, "")) or ""
        if (get_account(approver) or {}).get("role_code") != "manager_assistant":
            raise HTTPException(status_code=403, detail="只有助理/主管能複核核准案件。")
        try:
            return ok(approve_case(case_id, approver))
        except LookupError as exc:
            raise HTTPException(status_code=404, detail=str(exc)) from exc
        except PermissionError as exc:
            raise HTTPException(status_code=403, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=409, detail=str(exc)) from exc

    @app.post("/api/cases/{case_id}/disable")
    def disable_case(case_id: int) -> dict[str, Any]:
        return handle_disable("cases", case_id)

    @app.delete("/api/cases/{case_id}", status_code=204)
    def delete_case(case_id: int) -> None:
        handle_delete("cases", case_id)

    @app.get("/api/cases/{case_id}/360")
    def case_detail(case_id: int) -> dict[str, Any]:
        try:
            return ok(case_360(case_id))
        except LookupError as exc:
            raise HTTPException(status_code=404, detail=str(exc)) from exc

    @app.post("/api/contracts", status_code=201)
    def create_contract(payload: ContractIn) -> dict[str, Any]:
        return handle_create("contracts", payload.model_dump())

    @app.get("/api/contracts")
    def contracts(limit: int = Query(100, ge=1, le=500)) -> dict[str, Any]:
        return ok(list_rows("contracts", limit))

    @app.patch("/api/contracts/{contract_id}")
    def update_contract(contract_id: int, payload: ContractPatch) -> dict[str, Any]:
        return handle_change("contracts", contract_id, payload.model_dump(exclude_unset=True))

    @app.post("/api/contracts/{contract_id}/disable")
    def disable_contract(contract_id: int) -> dict[str, Any]:
        return handle_disable("contracts", contract_id)

    @app.delete("/api/contracts/{contract_id}", status_code=204)
    def delete_contract(contract_id: int) -> None:
        handle_delete("contracts", contract_id)

    @app.post("/api/payments", status_code=201)
    def create_payment(payload: PaymentIn) -> dict[str, Any]:
        return handle_create("payments", payload.model_dump())

    @app.get("/api/payments")
    def payments(limit: int = Query(100, ge=1, le=500)) -> dict[str, Any]:
        return ok(list_rows("payments", limit))

    @app.patch("/api/payments/{payment_id}")
    def update_payment(payment_id: int, payload: PaymentPatch) -> dict[str, Any]:
        return handle_change("payments", payment_id, payload.model_dump(exclude_unset=True))

    @app.post("/api/payments/{payment_id}/disable")
    def disable_payment(payment_id: int) -> dict[str, Any]:
        return handle_disable("payments", payment_id)

    @app.delete("/api/payments/{payment_id}", status_code=204)
    def delete_payment(payment_id: int) -> None:
        handle_delete("payments", payment_id)

    @app.post("/api/documents", status_code=201)
    def create_document(payload: DocumentIn) -> dict[str, Any]:
        return handle_create("documents", payload.model_dump())

    @app.get("/api/documents")
    def documents(limit: int = Query(100, ge=1, le=500)) -> dict[str, Any]:
        return ok(list_rows("documents", limit))

    @app.patch("/api/documents/{document_id}")
    def update_document(document_id: int, payload: DocumentPatch) -> dict[str, Any]:
        return handle_change("documents", document_id, payload.model_dump(exclude_unset=True))

    @app.post("/api/documents/{document_id}/disable")
    def disable_document(document_id: int) -> dict[str, Any]:
        return handle_disable("documents", document_id)

    @app.delete("/api/documents/{document_id}", status_code=204)
    def delete_document(document_id: int) -> None:
        handle_delete("documents", document_id)

    # ---- 預算 budgets ----
    @app.post("/api/budgets", status_code=201)
    def create_budget(payload: BudgetIn) -> dict[str, Any]:
        return handle_create("budgets", payload.model_dump())

    @app.get("/api/budgets")
    def budgets(limit: int = Query(100, ge=1, le=500)) -> dict[str, Any]:
        return ok(list_rows("budgets", limit))

    @app.get("/api/budgets/{budget_id}/annual")
    def budget_annual(budget_id: int) -> dict[str, Any]:
        # L3 年度費用比較（唯讀衍生）
        try:
            return ok(budget_annual_comparison(budget_id))
        except LookupError:
            raise HTTPException(status_code=404, detail="預算項目不存在")

    @app.put("/api/budgets/{budget_id}/annual-note")
    def budget_annual_note(budget_id: int, payload: BudgetYearNoteIn) -> dict[str, Any]:
        # 主管/助理寫每年備註
        try:
            return ok(set_budget_year_note(budget_id, payload.fiscal_year, payload.note))
        except LookupError:
            raise HTTPException(status_code=404, detail="預算項目不存在")

    @app.put("/api/budgets/{budget_id}/periods")
    def budget_periods_save(budget_id: int, payload: BudgetPeriodsIn) -> dict[str, Any]:
        # 承辦編輯：整批取代年度費用明細
        try:
            return ok(set_budget_periods(budget_id, [r.model_dump() for r in payload.periods]))
        except LookupError:
            raise HTTPException(status_code=404, detail="預算項目不存在")

    @app.patch("/api/budgets/{budget_id}")
    def update_budget(budget_id: int, payload: BudgetPatch) -> dict[str, Any]:
        return handle_change("budgets", budget_id, payload.model_dump(exclude_unset=True))

    @app.post("/api/budgets/{budget_id}/disable")
    def disable_budget(budget_id: int) -> dict[str, Any]:
        return handle_disable("budgets", budget_id)

    @app.delete("/api/budgets/{budget_id}", status_code=204)
    def delete_budget(budget_id: int) -> None:
        handle_delete("budgets", budget_id)

    # ---- 預算共同費用分攤（查詢呈現）----
    @app.get("/api/budgets/{budget_id}/allocations")
    def budget_allocations(budget_id: int) -> dict[str, Any]:
        # 以費用項目看：這筆共用費用攤給哪些單位、各多少
        return ok(list_budget_allocations(budget_id))

    @app.get("/api/budget-units")
    def budget_units(unit_code: str | None = Query(None)) -> dict[str, Any]:
        # 以單位看：各單位在所有項目的分攤合計；帶 unit_code 則回該單位的每筆明細
        return ok(budget_unit_rollup(unit_code))

    # ---- 單位管理 Step 1：撞名偵測（唯讀待確認清單）----
    @app.get("/api/unit-conflicts")
    def get_unit_conflicts() -> dict[str, Any]:
        # 掃描目前資料，挑出「同代號多名 / 同名多代號」給使用者裁決
        return ok(unit_conflicts())

    # ---- 單位管理 Step 2：合併 / 分開 / 主檔（單位主檔＋別名對照，非破壞式）----
    def _require_unit_editor(request: Request) -> dict[str, Any]:
        who = get_account(_verify_session(request.cookies.get(AUTH_COOKIE_NAME, "")) or "") or {}
        if who.get("role_code") not in ("manager_assistant", "admin"):
            raise HTTPException(status_code=403, detail="只有主管/助理可裁決單位合併。")
        return who

    @app.get("/api/unit-master")
    def get_unit_master() -> dict[str, Any]:
        return ok(list_unit_master())

    @app.post("/api/unit-master", status_code=201)
    def post_unit_master(payload: UnitCreateIn, request: Request) -> dict[str, Any]:
        # 主動新增乾淨單位（給表單下拉選單用），跟合併機制分開；建立前擋撞名，避免髒資料從源頭重複。
        _require_unit_editor(request)
        try:
            return ok(create_unit_master(payload.canonical_code, payload.canonical_name, payload.note))
        except ValueError as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc

    @app.post("/api/unit-impact")
    def post_unit_impact(payload: UnitImpactIn) -> dict[str, Any]:
        # 影響預覽：這些變體現在佔幾筆分攤、金額多少（裁決前先看清後果）
        return ok(unit_merge_impact([v.model_dump() for v in payload.variants]))

    @app.post("/api/unit-merge")
    def post_unit_merge(payload: UnitMergeIn, request: Request) -> dict[str, Any]:
        _require_unit_editor(request)
        variants = [v.model_dump() for v in payload.variants]
        try:
            return ok(merge_units(variants, payload.canonical_code, payload.canonical_name, payload.reason))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.post("/api/unit-split")
    def post_unit_split(payload: UnitSplitIn, request: Request) -> dict[str, Any]:
        _require_unit_editor(request)
        variants = [v.model_dump() for v in payload.variants]
        try:
            return ok(split_units(variants, payload.reason))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.post("/api/unit-reassign")
    def post_unit_reassign(payload: UnitReassignIn, request: Request) -> dict[str, Any]:
        _require_unit_editor(request)
        try:
            return ok(reassign_unit(payload.variant.model_dump(), payload.canonical_code, payload.canonical_name, payload.reason))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.post("/api/unit-alias/{alias_id}/unlink")
    def post_unit_alias_unlink(alias_id: int, request: Request) -> dict[str, Any]:
        _require_unit_editor(request)
        try:
            return ok(unlink_alias(alias_id))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.get("/api/unit-decisions")
    def get_unit_decisions() -> dict[str, Any]:
        # 決策紀錄：誰、何時、把什麼合併/分開、為什麼
        return ok(list_unit_decisions())

    @app.post("/api/unit-decisions/{decision_id}/undo")
    def post_unit_decision_undo(decision_id: int, request: Request) -> dict[str, Any]:
        _require_unit_editor(request)
        try:
            return ok(undo_decision(decision_id))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.post("/api/unit-reset")
    def post_unit_reset(request: Request) -> dict[str, Any]:
        # 一鍵還原：清掉所有裁決，回到剛匯入的原始狀態（原始資料本就沒動過）
        _require_unit_editor(request)
        return ok(reset_unit_decisions())

    # ---- 名稱歸納：案件名/專案名/廠商名 撞名清洗（比照單位主檔）----
    @app.get("/api/name-values")
    def get_name_values(kind: str = Query(...)) -> dict[str, Any]:
        try:
            return ok(list_name_values(kind))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.post("/api/name-merge")
    def post_name_merge(payload: NameMergeIn, request: Request) -> dict[str, Any]:
        _require_unit_editor(request)
        try:
            return ok(merge_names(payload.kind, payload.names, payload.canonical_name, payload.reason))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.post("/api/name-split")
    def post_name_split(payload: NameSplitIn, request: Request) -> dict[str, Any]:
        _require_unit_editor(request)
        try:
            return ok(split_names(payload.kind, payload.names, payload.reason))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.get("/api/name-decisions")
    def get_name_decisions(kind: str | None = Query(None)) -> dict[str, Any]:
        return ok(list_name_decisions(kind))

    @app.post("/api/name-decisions/{decision_id}/undo")
    def post_name_undo(decision_id: int, request: Request) -> dict[str, Any]:
        _require_unit_editor(request)
        try:
            return ok(undo_name_decision(decision_id))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.post("/api/name-reset")
    def post_name_reset(request: Request, kind: str | None = Query(None)) -> dict[str, Any]:
        _require_unit_editor(request)
        return ok(reset_name_decisions(kind))

    # ---- 作業年度：新案件「所屬年度」的預設，避免今年寫明年預算時搞混 ----
    @app.get("/api/working-year")
    def working_year() -> dict[str, Any]:
        return ok({"working_year": get_working_year()})

    @app.post("/api/working-year")
    def set_working_year(request: Request, year: str = Query(...)) -> dict[str, Any]:
        _require_unit_editor(request)
        y = (year or "").strip()
        if not (y.isdigit() and len(y) == 4):
            raise HTTPException(status_code=400, detail="作業年度請填四位數字年份，例如 2027。")
        store_set_settings({"working_year": y})
        return ok({"working_year": y})

    # ---- 按人數分攤：人數基準表 + 重算 ----
    @app.get("/api/budget-headcounts")
    def budget_headcounts() -> dict[str, Any]:
        return ok(list_headcounts())

    @app.post("/api/budget-headcounts/import-xlsx")
    async def import_headcounts_xlsx(request: Request, commit: bool = Query(False), filename: str = Query("")) -> dict[str, Any]:
        who = get_account(_verify_session(request.cookies.get(AUTH_COOKIE_NAME, "")) or "") or {}
        if who.get("role_code") not in ("manager_assistant", "admin"):
            raise HTTPException(status_code=403, detail="只有主管/助理可匯入人數表。")
        data = await request.body()
        if not data:
            raise HTTPException(status_code=400, detail="請選擇要上傳的 Excel 檔。")
        try:
            records = parse_headcount_xlsx(data)
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(status_code=400, detail=f"Excel 解析失敗（請確認是人數表格式）：{exc}") from exc
        if not commit:
            return ok({"preview": True, "count": len(records), "sample": records[:10]})
        return ok({"preview": False, **commit_headcounts_import(records, source_file=filename)})

    @app.post("/api/budgets/{budget_id}/recompute")
    def recompute_budget(budget_id: int) -> dict[str, Any]:
        # 依該預算的分攤方法重算分攤（edit 權限；CIO 唯讀由中介層擋）
        try:
            return ok(compute_budget_allocations(budget_id))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    # ---- 按類別分攤：類別基準表（對照表）----
    @app.get("/api/category-shares")
    def category_shares(category: str | None = Query(None)) -> dict[str, Any]:
        # 回類別清單（供分攤方法的類別下拉）；帶 category 則回該類別各單位%
        return ok(list_category_shares(category))

    @app.post("/api/category-shares/import-xlsx")
    async def import_category_shares_xlsx(request: Request, commit: bool = Query(False), filename: str = Query("")) -> dict[str, Any]:
        who = get_account(_verify_session(request.cookies.get(AUTH_COOKIE_NAME, "")) or "") or {}
        if who.get("role_code") not in ("manager_assistant", "admin"):
            raise HTTPException(status_code=403, detail="只有主管/助理可匯入類別基準表。")
        data = await request.body()
        if not data:
            raise HTTPException(status_code=400, detail="請選擇要上傳的 Excel 檔。")
        try:
            records = parse_category_shares_xlsx(data)
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(status_code=400, detail=f"Excel 解析失敗（請確認是含『對照』表的費用分攤表）：{exc}") from exc
        if not commit:
            cats = sorted({r["category"] for r in records if r.get("category")})
            return ok({"preview": True, "count": len(records), "categories": cats, "sample": records[:10]})
        return ok({"preview": False, **commit_category_shares_import(records, source_file=filename)})

    # ---- 專案 projects ----
    @app.post("/api/projects", status_code=201)
    def create_project(payload: ProjectIn) -> dict[str, Any]:
        return handle_create("projects", payload.model_dump())

    @app.get("/api/projects")
    def projects(limit: int = Query(100, ge=1, le=500)) -> dict[str, Any]:
        return ok(list_rows("projects", limit))

    @app.patch("/api/projects/{project_id}")
    def update_project(project_id: int, payload: ProjectPatch) -> dict[str, Any]:
        return handle_change("projects", project_id, payload.model_dump(exclude_unset=True))

    @app.post("/api/projects/{project_id}/disable")
    def disable_project(project_id: int) -> dict[str, Any]:
        return handle_disable("projects", project_id)

    @app.delete("/api/projects/{project_id}", status_code=204)
    def delete_project(project_id: int) -> None:
        handle_delete("projects", project_id)

    # ---- 工作項 project_items（進度總表點進去的細節，可由主管/助理/承辦維護；CIO 唯讀）----
    @app.get("/api/projects/{project_id}/items")
    def project_items(project_id: int) -> dict[str, Any]:
        return ok(list_project_items(project_id))

    @app.post("/api/projects/{project_id}/items", status_code=201)
    def create_project_item(project_id: int, payload: ProjectItemIn) -> dict[str, Any]:
        data = payload.model_dump()
        data["project_id"] = project_id
        return handle_create("project_items", data)

    @app.patch("/api/project-items/{item_id}")
    def update_project_item(item_id: int, payload: ProjectItemPatch) -> dict[str, Any]:
        return handle_change("project_items", item_id, payload.model_dump(exclude_unset=True))

    @app.post("/api/project-items/{item_id}/disable")
    def disable_project_item(item_id: int) -> dict[str, Any]:
        return handle_disable("project_items", item_id)

    @app.delete("/api/project-items/{item_id}", status_code=204)
    def delete_project_item(item_id: int) -> None:
        handle_delete("project_items", item_id)

    @app.post("/api/projects/import-xlsx")
    async def import_projects_xlsx(request: Request, commit: bool = Query(False)) -> dict[str, Any]:
        # 上傳『處級專案進度追蹤總表』.xlsx；commit=false 只預覽、true 正式寫入（交易/冪等/稽核）。
        who = get_account(_verify_session(request.cookies.get(AUTH_COOKIE_NAME, "")) or "") or {}
        if who.get("role_code") not in ("manager_assistant", "admin"):
            raise HTTPException(status_code=403, detail="只有主管/助理可匯入專案。")
        data = await request.body()
        if not data:
            raise HTTPException(status_code=400, detail="請選擇要上傳的 Excel 檔。")
        try:
            records = parse_projects_xlsx(data)
        except Exception as exc:  # noqa: BLE001 — openpyxl 各種解析錯誤都回同一句
            raise HTTPException(status_code=400, detail=f"Excel 解析失敗（請確認是專案總表格式）：{exc}") from exc
        if not commit:
            return ok({"preview": True, "count": len(records), "sample": records[:10]})
        return ok({"preview": False, **commit_projects_import(records)})

    @app.post("/api/budgets/import-xlsx")
    async def import_budgets_xlsx(request: Request, commit: bool = Query(False), filename: str = Query("")) -> dict[str, Any]:
        # 上傳『預算』.xlsx（表單型：一張工作表＝一筆預算）；commit=false 預覽、true 正式寫入。
        # filename：前端帶上原始檔名，寫進分攤資料，供單位撞名清單指回來源。
        who = get_account(_verify_session(request.cookies.get(AUTH_COOKIE_NAME, "")) or "") or {}
        if who.get("role_code") not in ("manager_assistant", "admin"):
            raise HTTPException(status_code=403, detail="只有主管/助理可匯入預算。")
        data = await request.body()
        if not data:
            raise HTTPException(status_code=400, detail="請選擇要上傳的 Excel 檔。")
        try:
            records = parse_budget_xlsx(data)
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(status_code=400, detail=f"Excel 解析失敗（請確認是預算表格式）：{exc}") from exc
        if not commit:
            return ok({"preview": True, "count": len(records), "sample": records[:10]})
        return ok({"preview": False, **commit_budgets_import(records, source_file=filename)})

    # ---- 簽呈 signoffs ----
    @app.post("/api/signoffs", status_code=201)
    def create_signoff(payload: SignoffIn) -> dict[str, Any]:
        return handle_create("signoffs", payload.model_dump())

    @app.get("/api/signoffs")
    def signoffs(limit: int = Query(100, ge=1, le=500)) -> dict[str, Any]:
        return ok(list_rows("signoffs", limit))

    @app.patch("/api/signoffs/{signoff_id}")
    def update_signoff(signoff_id: int, payload: SignoffPatch) -> dict[str, Any]:
        return handle_change("signoffs", signoff_id, payload.model_dump(exclude_unset=True))

    @app.post("/api/signoffs/{signoff_id}/disable")
    def disable_signoff(signoff_id: int) -> dict[str, Any]:
        return handle_disable("signoffs", signoff_id)

    @app.delete("/api/signoffs/{signoff_id}", status_code=204)
    def delete_signoff(signoff_id: int) -> None:
        handle_delete("signoffs", signoff_id)

    # ---- 請購 purchases ----
    @app.post("/api/purchases", status_code=201)
    def create_purchase(payload: PurchaseIn) -> dict[str, Any]:
        return handle_create("purchases", payload.model_dump())

    @app.get("/api/purchases")
    def purchases(limit: int = Query(100, ge=1, le=500)) -> dict[str, Any]:
        return ok(list_rows("purchases", limit))

    @app.patch("/api/purchases/{purchase_id}")
    def update_purchase(purchase_id: int, payload: PurchasePatch) -> dict[str, Any]:
        return handle_change("purchases", purchase_id, payload.model_dump(exclude_unset=True))

    @app.post("/api/purchases/{purchase_id}/disable")
    def disable_purchase(purchase_id: int) -> dict[str, Any]:
        return handle_disable("purchases", purchase_id)

    @app.delete("/api/purchases/{purchase_id}", status_code=204)
    def delete_purchase(purchase_id: int) -> None:
        handle_delete("purchases", purchase_id)

    # ---- 各模組 CSV 匯出（Excel 友善，含 UTF-8 BOM；依 owner 範圍匯出）----
    def _module_csv(table: str) -> Response:
        cols = CSV_COLUMNS[table]
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow([label for _, label in cols])
        for r in list_rows(table, 500):
            writer.writerow([r.get(key, "") for key, _ in cols])
        return Response(
            content="﻿" + buf.getvalue(),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={table}.csv"},
        )

    for _table in CSV_COLUMNS:
        app.add_api_route(f"/api/{_table}.csv", (lambda t: lambda: _module_csv(t))(_table), include_in_schema=False)

    @app.get("/api/search")
    def search(q: str = Query(min_length=1)) -> dict[str, Any]:
        return ok(search_records(q))

    @app.get("/api/cmdb")
    def cmdb_placeholder() -> dict[str, Any]:
        return ok(
            {
                "status": "reserved",
                "reserved_fields": ["cmdb_ci_id", "service_owner", "asset_tag"],
                "next_step": "Connect to enterprise CMDB API after credentials and schema are approved.",
            }
        )

    return app


def handle_create(table: str, payload: dict[str, Any]) -> dict[str, Any]:
    try:
        return ok(insert_row(table, payload))
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


def handle_change(table: str, row_id: int, payload: dict[str, Any]) -> dict[str, Any]:
    try:
        return ok(update_row(table, row_id, payload))
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except LookupError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


def handle_disable(table: str, row_id: int) -> dict[str, Any]:
    try:
        return ok(disable_row(table, row_id))
    except LookupError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


def handle_delete(table: str, row_id: int) -> None:
    try:
        delete_row(table, row_id)
    except RuntimeError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    except LookupError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


app = create_app()

"""週進度報告工具：抓數字、比對流程圖、產出前檢查。

用法（在專案根目錄執行；中文路徑用 py 跑，不要用 PowerShell 的 Get-Content 改檔）：

    py scripts\\weekly_report_tools.py stats  <舊報告.html> <新報告.html>
        兩份報告的整體％、各階段％、四態統計並排，寫信要的數字一次拿到。

    py scripts\\weekly_report_tools.py flow-diff <舊報告.html> <新報告.html>
        流程圖上「這次新完成」的節點編號與行號（用來標橘框）。

    py scripts\\weekly_report_tools.py mark <報告.html> <行號,行號,...>
        把指定行的已完成節點框改成橘色（本次新完成）。行號由 flow-diff 給。

    py scripts\\weekly_report_tools.py verify <報告.html>
        產出前檢查：HTML 標籤閉合、無 BOM、橘框數、舊版號有沒有殘留。

    py scripts\\weekly_report_tools.py xlsx <舊報告.html> <新報告.html> <輸出.xlsx>
        產週報的第二個附件：進度對照＋WBS 全項目清單的 Excel。
        主管要自己篩選、排序、貼進他自己的報表時用這份（HTML 那份是給人看的，
        Excel 這份是給人用的）。差額欄是公式不是算好的數字，改了前後值會自己重算。

舊報告＝上次寄出那份，放在 docs\\報告存檔\\。改報告前一定要先備份一份帶日期的，
否則下次沒有東西可以對比（規則見 AI\\週報告_TEMPLATE.md）。
"""
from __future__ import annotations

from html.parser import HTMLParser
from pathlib import Path
import re
import sys

sys.stdout.reconfigure(encoding="utf-8")

# 自閉合或 SVG 內不需要成對計算的標籤
VOID = {"meta", "link", "br", "hr", "img", "input", "path", "rect", "circle",
        "polygon", "text", "line", "marker", "use", "tspan"}

NODE_LINE = re.compile(r'^<(rect|polygon)\s')
NODE_CODE = re.compile(r'>([A-E]\d)</text>')
ORANGE = "#d9860b"          # 本次新完成
GREEN = "#1b6b4a"           # 先前已完成


def read(path: str) -> str:
    return Path(path).read_text(encoding="utf-8")


def stats(path: str) -> dict:
    s = read(path)
    return {
        "overall": (re.findall(r'class="pct">(\d+)%', s) or ["?"])[0],
        "stages": re.findall(
            r'stage-tag">([^<]+)</span><span class="stage-title">([^<]+)</span>\s*'
            r'<span class="stage-pct"><span class="mini-bar"><i style="width:(\d+)%', s),
        "tiles": re.findall(r'class="num">(\d+)</div><div class="lbl">([^<]+)<', s),
    }


def cmd_stats(old: str, new: str) -> None:
    a, b = stats(old), stats(new)
    print(f"整體：{a['overall']}% → {b['overall']}%")
    print("\n各階段：")
    bmap = {name: pct for _tag, name, pct in b["stages"]}
    for _tag, name, pct in a["stages"]:
        after = bmap.get(name, "?")
        delta = f"+{int(after) - int(pct)}" if after.isdigit() else "?"
        print(f"  {name:　<16} {pct:>3}% → {after:>3}%  ({delta})")
    print("\n四態統計：")
    bt = dict((lbl, n) for n, lbl in b["tiles"])
    for n, lbl in a["tiles"]:
        print(f"  {lbl:　<12} {n:>3} → {bt.get(lbl, '?'):>3}")


def flow_states(path: str) -> dict[str, tuple[str, int]]:
    """{節點編號: (done|todo, 行號)}；節點框後第一個編號文字就是它的編號。"""
    out: dict[str, tuple[str, int]] = {}
    pending = None
    for i, line in enumerate(read(path).split("\n"), 1):
        if NODE_LINE.match(line) and ('rx="10"' in line or line.startswith("<polygon")):
            if "stroke-dasharray" in line:
                pending = ("todo", i)
            elif GREEN in line or ORANGE in line:
                pending = ("done", i)
            else:
                pending = None
        m = NODE_CODE.search(line)
        if m and pending:
            out[m.group(1)] = pending
            pending = None
    return out


def cmd_flow_diff(old: str, new: str) -> None:
    a, b = flow_states(old), flow_states(new)
    newly = sorted(k for k in b if b[k][0] == "done" and a.get(k, ("todo", 0))[0] == "todo")
    print("本次新完成（標橘框）：", ", ".join(newly) or "（無）")
    print("行號：", ",".join(str(b[k][1]) for k in newly))
    print("仍未開發：", ", ".join(sorted(k for k in b if b[k][0] == "todo")) or "（無）")
    print("先前就完成：", ", ".join(sorted(k for k in b if b[k][0] == "done" and k not in newly)))


def cmd_mark(path: str, rows: str) -> None:
    p = Path(path)
    lines = p.read_text(encoding="utf-8").split("\n")
    targets = [int(x) for x in rows.replace(" ", "").split(",") if x]
    for n in targets:
        i = n - 1
        s = lines[i]
        if ORANGE in s:
            continue                      # 已經標過，重跑不會壞
        assert f'stroke="{GREEN}"' in s, f"第 {n} 行不是已完成節點：{s[:90]}"
        lines[i] = s.replace(f'stroke="{GREEN}"', f'stroke="{ORANGE}"').replace(
            'stroke-width="2.5"', 'stroke-width="3"')
    p.write_text("\n".join(lines), encoding="utf-8", newline="")
    print(f"已標記 {len(targets)} 個節點為本次新完成")


class _Balance(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.stack: list[str] = []
        self.bad: list[str] = []

    def handle_starttag(self, tag, attrs):
        if tag not in VOID:
            self.stack.append(tag)

    def handle_endtag(self, tag):
        if tag in VOID:
            return
        if self.stack and self.stack[-1] == tag:
            self.stack.pop()
        else:
            self.bad.append(tag)


def cmd_verify(path: str) -> None:
    s = read(path)
    h = _Balance()
    h.feed(s)
    tabs = re.findall(r'class="tab[^"]*"[^>]*>([^<]+)</button>', s)
    orange = len(re.findall(r'<(?:rect|polygon)[^>]*stroke="' + ORANGE + r'"', s))
    old_versions = sorted(set(re.findall(r'v0\.\d+\.\d+', s)))
    print(f"[{'OK  ' if not h.stack and not h.bad else 'FAIL'}] 標籤閉合 未閉合={h.stack} 錯配={h.bad[:5]}")
    print(f"[{'OK  ' if not s.startswith(chr(0xFEFF)) else 'FAIL'}] 無 BOM")
    print(f"[{'OK  ' if len(tabs) >= 6 else 'FAIL'}] 頁籤 {len(tabs)} 個：{' / '.join(tabs)}")
    print(f"[{'OK  ' if orange else 'WARN'}] 流程圖橘框（本次新完成）{orange} 個")
    head = re.search(r'class="eyebrow">.*?<span>(v0\.\d+\.\d+)</span>.*?<span>(\d{4}-\d{2}-\d{2})</span>', s, re.S)
    foot = re.search(r'產生於 (\d{4}-\d{2}-\d{2}) · (v0\.\d+\.\d+)', s)
    print(f"[INFO] 頁首：{head.group(1) if head else '?'} / {head.group(2) if head else '?'}"
          f"　頁尾：{foot.group(2) if foot else '?'} / {foot.group(1) if foot else '?'}　←這兩處要一致且是最新")
    print(f"[INFO] 內文提到的版號（更新紀錄會有舊的，正常）：{', '.join(old_versions)}")


ITEM_RE = re.compile(
    r'<li class="item"><span class="chip (done|partial|todo)">[^<]+</span>'
    r'<div class="item-body"><div class="item-name">(.*?)</div>'
    r'(?:<div class="item-note">(.*?)</div>)?', re.S)
STAGE_SPLIT = re.compile(
    r'<div class="stage-head"><span class="stage-tag">([^<]+)</span>'
    r'<span class="stage-title">([^<]+)</span>')
CHIP_LABEL = {"done": "已完成", "partial": "進行中", "todo": "待辦"}


def _text(html: str) -> str:
    """把 item 名稱/說明裡的標籤與標記拿掉，只留人看的字。"""
    s = re.sub(r'<[^>]+>', "", html or "")
    return " ".join(s.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">").split())


def wbs_items(path: str) -> list[tuple[str, str, str, str]]:
    """[(階段, 項目, 狀態, 說明)]，依報告上的順序。"""
    s = read(path)
    marks = list(STAGE_SPLIT.finditer(s))
    out: list[tuple[str, str, str, str]] = []
    for i, m in enumerate(marks):
        stage = f"{m.group(1)} {m.group(2)}"
        block = s[m.end(): marks[i + 1].start() if i + 1 < len(marks) else len(s)]
        for chip, name, note in ITEM_RE.findall(block):
            out.append((stage, _text(name), CHIP_LABEL.get(chip, chip), _text(note)))
    return out


def cmd_xlsx(old: str, new: str, out: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    a, b = stats(old), stats(new)
    old_label = Path(old).stem.replace("開發進度報告_", "")
    new_label = Path(new).stem.replace("開發進度報告_", "") or "現在"

    F = "Arial"
    head_font = Font(name=F, bold=True, color="FFFFFF", size=11)
    head_fill = PatternFill("solid", fgColor="1F4E6B")
    title_font = Font(name=F, bold=True, size=14)
    body = Font(name=F, size=11)
    thin = Side(style="thin", color="D0D6DE")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(vertical="top", wrap_text=True)

    def header(ws, row, labels, widths):
        for c, (label, w) in enumerate(zip(labels, widths), start=1):
            cell = ws.cell(row=row, column=c, value=label)
            cell.font, cell.fill, cell.border, cell.alignment = head_font, head_fill, box, center
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = ws.cell(row=row + 1, column=1)

    wb = Workbook()

    # ── 進度對照 ──
    ws = wb.active
    ws.title = "進度對照"
    ws["A1"] = "費用合約控管平台 — 開發進度對照"
    ws["A1"].font = title_font
    ws["A2"] = f"左＝{old_label} 那份報告當下的狀態，右＝{new_label}。差額欄是公式，改了前後數字會自己重算。"
    ws["A2"].font = Font(name=F, size=9, color="7D8894")

    header(ws, 4, ["項目", f"開發前（{old_label}）", f"開發後（{new_label}）", "差額"], [26, 18, 18, 10])
    r = 5
    ws.cell(row=r, column=1, value="整體開發進度").font = Font(name=F, bold=True, size=11)
    ws.cell(row=r, column=2, value=int(a["overall"]) / 100)
    ws.cell(row=r, column=3, value=int(b["overall"]) / 100)
    ws.cell(row=r, column=4, value=f"=C{r}-B{r}")
    r += 1

    bmap = {name: pct for _t, name, pct in b["stages"]}
    for _tag, name, pct in a["stages"]:
        after = bmap.get(name)
        ws.cell(row=r, column=1, value=name).font = body
        ws.cell(row=r, column=2, value=int(pct) / 100)
        ws.cell(row=r, column=3, value=int(after) / 100 if after else None)
        ws.cell(row=r, column=4, value=f"=C{r}-B{r}")
        r += 1

    for row in ws.iter_rows(min_row=5, max_row=r - 1, min_col=1, max_col=4):
        for cell in row:
            cell.border = box
            if cell.column > 1:
                cell.font, cell.number_format, cell.alignment = body, "0%", center

    r += 1
    ws.cell(row=r, column=1, value="功能項目統計").font = title_font
    r += 1
    header(ws, r, ["狀態", f"開發前（{old_label}）", f"開發後（{new_label}）", "差額"], [26, 18, 18, 10])
    ws.freeze_panes = "A5"          # 凍在第一張表的表頭，不被第二段覆寫
    start = r + 1
    bt = {lbl: n for n, lbl in b["tiles"]}
    for n, lbl in a["tiles"]:
        r += 1
        ws.cell(row=r, column=1, value=lbl).font = body
        ws.cell(row=r, column=2, value=int(n))
        ws.cell(row=r, column=3, value=int(bt.get(lbl, 0)))
        ws.cell(row=r, column=4, value=f"=C{r}-B{r}")
    for row in ws.iter_rows(min_row=start, max_row=r, min_col=1, max_col=4):
        for cell in row:
            cell.border = box
            if cell.column > 1:
                cell.font, cell.alignment = body, center

    r += 2
    ws.cell(row=r, column=1, value="資料來源：開發進度報告 HTML（本檔由 scripts\\weekly_report_tools.py xlsx 自動產生，不要手改）")
    ws.cell(row=r, column=1).font = Font(name=F, size=9, color="7D8894")

    # ── WBS 全項目 ──
    ws2 = wb.create_sheet("WBS 全項目")
    ws2["A1"] = f"WBS 功能項目清單（{new_label}）"
    ws2["A1"].font = title_font
    ws2["A2"] = "可用篩選看某個階段、或只看「待辦」。狀態欄：已完成／進行中／待辦。"
    ws2["A2"].font = Font(name=F, size=9, color="7D8894")
    header(ws2, 4, ["階段", "功能項目", "狀態", "說明"], [22, 34, 10, 78])

    fills = {"已完成": PatternFill("solid", fgColor="E7F3EC"),
             "進行中": PatternFill("solid", fgColor="FBF0DD"),
             "待辦": PatternFill("solid", fgColor="EEF0F3")}
    rows = wbs_items(new)
    for i, (stage, name, status, note) in enumerate(rows, start=5):
        for c, v in enumerate((stage, name, status, note), start=1):
            cell = ws2.cell(row=i, column=c, value=v)
            cell.font, cell.border, cell.alignment = body, box, wrap
            if c == 3:
                cell.fill, cell.alignment = fills.get(status, fills["待辦"]), center
    ws2.auto_filter.ref = f"A4:D{4 + len(rows)}"

    Path(out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"已產生 {out}")
    print(f"  進度對照：整體 {a['overall']}% → {b['overall']}%，{len(a['stages'])} 個階段")
    print(f"  WBS 全項目：{len(rows)} 列（" + "、".join(
        f"{s} {sum(1 for x in rows if x[2] == s)}" for s in ("已完成", "進行中", "待辦")) + "）")
    print("  提醒：openpyxl 寫的公式沒有快取值，要用 LibreOffice/Excel 開過一次才有數字")


def main() -> int:
    if len(sys.argv) < 3:
        print(__doc__)
        return 1
    cmd, args = sys.argv[1], sys.argv[2:]
    table = {"stats": cmd_stats, "flow-diff": cmd_flow_diff, "mark": cmd_mark,
             "verify": cmd_verify, "xlsx": cmd_xlsx}
    if cmd not in table:
        print(__doc__)
        return 1
    table[cmd](*args)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

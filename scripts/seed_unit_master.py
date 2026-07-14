#!/usr/bin/env python3
"""一次性：把「docs/一、預算.xlsx」裡的部門代號/部門別，灌進單位主檔（unit_master）。
給承辦這類角色用的預算/專案表單「單位名稱」下拉選單，選項就是來自這裡。

用法：
    python scripts/seed_unit_master.py            # 對 SQLITE_PATH（或預設路徑）灌資料
    python scripts/seed_unit_master.py --dry-run   # 只印出會新增哪些，不寫入

冪等：同代號或同名稱已存在就跳過（不會報錯中斷、不會重複），可重複執行。
"""
from __future__ import annotations

import sys
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl

from app import store

XLSX_PATH = Path(__file__).resolve().parents[1] / "docs" / "一、預算.xlsx"
DRY_RUN = "--dry-run" in sys.argv


def extract_units() -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = next(i for i, r in enumerate(rows) if any(str(c).strip() == "部門代號" for c in r if c is not None))
    units: list[dict[str, str]] = []
    for r in rows[header_idx + 1:]:
        name = r[2] if len(r) > 2 else None
        if name is None:
            continue
        name = str(name).strip()
        if name == "合計":
            break
        if not name or name == "EOF":
            continue
        code = str(r[1]).strip() if r[1] is not None else ""
        units.append({"code": code, "name": name})
    wb.close()
    return units


def main() -> int:
    units = extract_units()
    print(f"從 {XLSX_PATH.name} 讀到 {len(units)} 個單位。")
    created = skipped = failed = 0
    for u in units:
        if DRY_RUN:
            print(f"  [dry-run] 會新增：{u['code'] or '(無代號)'}　{u['name']}")
            continue
        try:
            store.create_unit_master(u["code"], u["name"])
            created += 1
        except ValueError as exc:
            skipped += 1
            print(f"  跳過（已存在）：{u['code'] or '(無代號)'}　{u['name']}　— {exc}")
        except Exception as exc:  # noqa: BLE001
            failed += 1
            print(f"  ✗ 失敗：{u['code'] or '(無代號)'}　{u['name']}　— {exc}")
    if not DRY_RUN:
        print(f"完成：新增 {created}、跳過(已存在) {skipped}、失敗 {failed}。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

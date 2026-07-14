#!/usr/bin/env python3
"""一次性：把「docs/資訊架構部處級專案進度追蹤總表.xlsx」各專案分頁裡的「負責人」，灌進人員主檔（personnel_master）。
給案件/簽呈/預算/付款/專案/工作項這些表單的人員欄位下拉（或自動完成）選項，就是來自這裡。

用法：
    python scripts/seed_personnel_master.py            # 對 SQLITE_PATH（或預設路徑）灌資料
    python scripts/seed_personnel_master.py --dry-run   # 只印出會新增哪些，不寫入

冪等：同名已存在就跳過（不會報錯中斷、不會重複），可重複執行。

抓法：找每個分頁裡「負責人」那欄，一格可能是「陳昱杉/洪似妮」這種多人組合（用 / & 、 , ， 換行分隔），
逐一拆開登記成獨立姓名；順手濾掉看起來不是姓名的雜訊（含數字/標點、超過 6 個字，通常是誤填的備註文字）。
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl

from app import store

XLSX_PATH = Path(__file__).resolve().parents[1] / "docs" / "資訊架構部處級專案進度追蹤總表.xlsx"
PROJECT_SHEETS = ["資料庫組處級專案", "網路組處級專案", "主機組處級專案", "青埔機房搬遷建置案", "AI專案"]
SEP = re.compile(r"[/&、,，\n]")
NOISE = re.compile(r"[0-9。：「」（）():.]")  # 有數字/標點的多半是誤填的備註文字，不是姓名
DRY_RUN = "--dry-run" in sys.argv


def extract_names() -> list[str]:
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    names: set[str] = set()
    for sheet_name in PROJECT_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        header_idx = next(
            (i for i, r in enumerate(rows) if any(str(c).strip() == "負責人" for c in r if c is not None)), None
        )
        if header_idx is None:
            continue
        owner_col = next(j for j, c in enumerate(rows[header_idx]) if str(c).strip() == "負責人")
        for r in rows[header_idx + 1:]:
            if owner_col >= len(r):
                continue
            v = r[owner_col]
            if not v:
                continue
            for part in SEP.split(str(v)):
                name = part.strip()
                if not name or len(name) > 6 or NOISE.search(name):
                    continue
                names.add(name)
    wb.close()
    return sorted(names)


def main() -> int:
    names = extract_names()
    print(f"從 {XLSX_PATH.name} 讀到 {len(names)} 個人名。")
    created = skipped = failed = 0
    for name in names:
        if DRY_RUN:
            print(f"  [dry-run] 會新增：{name}")
            continue
        try:
            store.create_personnel_master(name)
            created += 1
        except ValueError as exc:
            skipped += 1
            print(f"  跳過（已存在）：{name}　— {exc}")
        except Exception as exc:  # noqa: BLE001
            failed += 1
            print(f"  ✗ 失敗：{name}　— {exc}")
    if not DRY_RUN:
        print(f"完成：新增 {created}、跳過(已存在) {skipped}、失敗 {failed}。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

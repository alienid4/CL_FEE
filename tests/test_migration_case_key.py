"""第四輪回饋 AC-07：Migration Case Key。

舊資料匯入前，業務端自己整理一份對照表，把同一件業務事項在預算／專案／合約各自
的匯入檔案裡標上同一把「案件關聯鍵」——名稱、代碼可能完全不一樣，但 Key 一樣就
歸同一個案件，不靠系統瞎猜名稱像不像。用合成 xlsx 測，不依賴任何真實檔案。
"""
import io
import os

import openpyxl
from fastapi.testclient import TestClient


def _client(tmp_path, login="ap02"):
    os.environ["SQLITE_PATH"] = str(tmp_path / "mck.db")
    from app.main import create_app

    client = TestClient(create_app())
    if login:
        client.post("/api/auth/login", json={"username": login, "password": "T3st!Pass"})
    return client


CONTRACT_HEADERS = ["詩芸備註", "已確認完成", "合約編號", "合約名稱", "合約系統之內容說明",
                     "合約狀態", "合約狀態詳細說明", "組別", "合約維護人", "廠商名稱",
                     "廠商統編或ID", "合約開始日", "合約到期日", "案件關聯鍵"]


def _contract_wb(code, name, key="") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "測試組"
    for i in range(3):
        ws.append([f"填寫說明第 {i + 1} 行"])
    ws.append(CONTRACT_HEADERS)
    ws.append(["", "True", code, name, "", "進行中", "", "測試組", "王小明",
               "測試廠商", "12345678", "2026-01-01", "2026-12-31", key])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _budget_wb(code, key="") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = code
    ws.append(["統籌預估表"])
    ws.append([None, "預算項目：", "測試費用"])
    ws.append([None, "填寫部門：", "資訊管理處"])
    ws.append([None, "案件關聯鍵：", key])
    ws.append([None, None, None, None, "全年度費用"])
    ws.append([None, None, None, "115年度", 500])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


PROJECT_HEADER = ["標號", "專案名稱", "執行必要性(下拉式選單)", "總進度預計%", "總進度實際%", "總進度燈號",
                   "案件關聯鍵",
                   "標號", "工作主項目", "負責人", "開始日期", "結束日期", "執行進度(下拉式選單)",
                   "子項目總數", "子項目完成數", "完成度（%)", "燈號", "關鍵風險點/備註說明",
                   "需決策項目", "需支援項目", "持續天數"]


def _project_wb(name, key="") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "測試組處級專案"
    ws.append(["部處級專案進度追蹤總表"])
    ws.append(PROJECT_HEADER)
    ws.append([1, name, "必要", 0.5, 0.4, "", key, 1, "需求盤點", "王小明",
               "2026-01-01", "2026-03-31", "已完成", 3, 3, 1.0, "綠", "", "", "", 90])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def test_盤點表匯入現在會自動配案不再是孤兒(tmp_path):
    """AC-01 延伸：合約盤點表匯入這條路徑之前完全沒有配案邏輯，一律是孤兒合約。"""
    with _client(tmp_path) as client:
        client.post("/api/contracts/import-xlsx?commit=true", content=_contract_wb("EF-1", "沒填關聯鍵的合約"))
        k = client.get("/api/contracts").json()["data"][0]
        assert k["case_id"] is not None


def test_合約與預算填同一把案件關聯鍵會配到同一個案件(tmp_path):
    with _client(tmp_path) as client:
        client.post("/api/contracts/import-xlsx?commit=true",
                     content=_contract_wb("EF-KEY1", "機房搬遷合約", key="MIG-001"))
        client.post("/api/budgets/import-xlsx?commit=true",
                     content=_budget_wb("BUD-KEY1", key="MIG-001"))

        k = client.get("/api/contracts").json()["data"][0]
        b = client.get("/api/budgets").json()["data"][0]
        assert k["case_id"] is not None and b["case_id"] is not None
        assert k["case_id"] == b["case_id"]     # 名稱完全不同，靠 Key 接到同一個案件


def test_合約與專案填同一把案件關聯鍵會配到同一個案件(tmp_path):
    with _client(tmp_path) as client:
        client.post("/api/contracts/import-xlsx?commit=true",
                     content=_contract_wb("EF-KEY2", "機房搬遷合約二", key="MIG-002"))
        client.post("/api/projects/import-xlsx?commit=true",
                     content=_project_wb("完全不同名稱的專案", key="MIG-002"))

        k = client.get("/api/contracts").json()["data"][0]
        p = client.get("/api/projects").json()["data"][0]
        assert k["case_id"] == p["case_id"]


def test_沒填關聯鍵時退回既有同名配案邏輯不受影響(tmp_path):
    with _client(tmp_path) as client:
        client.post("/api/contracts/import-xlsx?commit=true", content=_contract_wb("EF-A", "合約名稱甲"))
        client.post("/api/budgets/import-xlsx?commit=true", content=_budget_wb("預算代號乙"))
        k = client.get("/api/contracts").json()["data"][0]
        b = client.get("/api/budgets").json()["data"][0]
        # 沒有 Key 時各自照原本「同名」邏輯配案，合約名稱跟預算的 budget_code 不同名，
        # 不會被誤接成同一個案件
        assert k["case_id"] != b["case_id"]


def test_關聯鍵重匯不會重複建案(tmp_path):
    with _client(tmp_path) as client:
        client.post("/api/contracts/import-xlsx?commit=true",
                     content=_contract_wb("EF-KEY3", "合約甲", key="MIG-003"))
        first_case_id = client.get("/api/contracts").json()["data"][0]["case_id"]

        client.post("/api/budgets/import-xlsx?commit=true", content=_budget_wb("BUD-KEY3", key="MIG-003"))
        client.post("/api/projects/import-xlsx?commit=true", content=_project_wb("專案乙", key="MIG-003"))

        cases = client.get("/api/cases").json()["data"]
        same_key_cases = [c for c in cases if c["id"] == first_case_id]
        assert len(same_key_cases) == 1   # 三個模組都掛同一個既有案件，不會多生案件
        b = client.get("/api/budgets").json()["data"][0]
        p = client.get("/api/projects").json()["data"][0]
        assert b["case_id"] == first_case_id and p["case_id"] == first_case_id
